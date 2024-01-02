from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import redirect
from django.template import loader
from django import forms
from django.views.generic.edit import FormView
from .forms import UploadFilesForm, UploadFileForm, AsinForm, SkuForm, UploadShipmentFileForm ,\
                   UploadTransactionAdCurrencyForm, ConfirmPoHeadShippingForm, UploadFileCountryForm ,\
                   FcCodeCountryForm
from django.contrib.auth.decorators import login_required
from django.contrib.auth.decorators import user_passes_test
from io import BytesIO,StringIO
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
import xlrd
import datetime
import pytz
import csv
import re
import json
from django.db.models import Sum, Q
from django.http import JsonResponse
from salesMonitor.models import Product, TodayProductSales, Last7dayProductSales, \
                                DailySalesLastYear, FbaInventory, RemoteFulfillmentSku, \
                                ReceivablePurchasedQty, HistoryTodayProductSales,\
                                NearestReceivablePurchasedQty ,\
                                HistoryTodaySales, Inventory, FbaShipment, ShippedSkuQty , FulfillmentCenterCodeCountry, \
                                FbaShipmentPaidBill, ProductInventoryUnitValue, \
                                ReceivedSkuQty, Upc, SkuUpc, Supplier, SkuSupplier, UserSupplier ,\
                                SkuPurchasingPrice, SkuHeadShippingUnitCost, SkuAssetLiabilityTable ,\
                                SkuManagedBySalesPerson, SkuContributor, SkuWeight, FbaShipmentCost ,\
                                SkuPurchaseOrder, ProfitLossTable, \
                                CurrencyRate, ProductionPlanProgress, SkuProductionStageTypeParameter ,\
                                ProductionStageTypeParameter, ProductionStage

from salesMonitor.excelReadData import readSkuValue, readTransaction \
                                     ,readCurrencyRate \
                                     ,readFixedCost  \


from dateutil import parser as dateparser

# Build paths inside the project like this: os.path.join(BASE_DIR, ...)
ENCODINGS = ('CP932', 'utf-8','cp1252', 'windows-1252')
shenzhen_warehouse_name = '深圳A016'
SKU_UNIT_SHIPPING_DATE_MONTH_RANGE = [7,0]

UPDATE_PRODUCT_INFORMATION_IN_BULK_TABLE_HEADER = ['SKU', '长', '宽', '高', '重量', '是否强制使用实重']
UPDATE_SKU_PRODUCTION_STAGE_DEFAULT_TABLE_HEADER = ['SKU', '参考SKU']
UPDATE_SKU_PRODUCTION_STAGE_DETAILED_NUMBERS_DEFAULT_TABLE_HEADER = ['', '名称', '典型的天数', '每天可以完成该工序的件数']

PRODUCTION_TYPE_NAME_OPTIONS = ['皮粘款', '泡泡款及其他软包', '油边磁吸款', '油边无磁吸款']
PRODUCTION_STAGE_NAME_OPTIONS = ['买料' , \
                                    '过渡期(买料到贴合)', \
                                    '贴合',\
                                    '过渡期(贴合到开料压Logo)', \
                                    '开料压Logo', \
                                    '散味', \
                                    '过渡期(散味到油边1)', \
                                    '油边1', \
                                    '过渡期(油边1到车缝)',\
                                    '车缝',\
                                    '过渡期(车缝到油边2)',\
                                    '油边2',\
                                    '过渡期(油边2到质检)',\
                                    '质检',\
                                    '等待期(下单到买料)',\
                                    '过渡期(开料压Logo到塞磁铁)',\
                                    '塞磁铁',\
                                    '过渡期(散味到车缝)',\
                                    '过渡期(车缝到质检)',\
                                    '过渡期(买料到开料)',\
                                    '开料',\
                                    '过渡期(开料到车缝)'\
                                ]
PRODUCTION_STAGE_MONITOR =  ['买料' ,\
                                    '开料'\
                                    '贴合',\
                                    '开料压Logo',\
                                    '塞磁铁',\
                                    '散味',\
                                    '油边1',\
                                    '油边2',\
                                    '车缝',\
                                    '质检',\
                                ]
PRODUCTION_STAGE_PAUSE_STARTING_WORDS =  [ '过渡期(买料到贴合)', \
                                    '过渡期(贴合到开料压Logo)', \
                                    '过渡期(散味到油边1)', \
                                    '过渡期(油边1到车缝)',\
                                    '过渡期(车缝到油边2)',\
                                    '过渡期(油边2到质检)',\
                                    '等待期(下单到买料)',\
                                    '过渡期(开料压Logo到塞磁铁)',\
                                    '过渡期(散味到车缝)',\
                                    '过渡期(车缝到质检)',\
                                    '过渡期(买料到开料)',\
                                    '过渡期(开料到车缝)' \
                                ]

AGILE_MANUFACTURER_NUMBERS = ['GYS008 东莞市全辉塑胶制品有限公司', 'GYS010 东莞市国丰手袋制品有限公司']
@login_required
def index(request):
    if request.user.groups.filter(name = 'supplier').exists():
        template = loader.get_template('todaySalesForSupplier.html')
    else:
        template = loader.get_template('todaySales.html')
    default_country = 'US'
    todaySales = HistoryTodaySales.objects.filter(country=default_country).order_by('-date')[0]

    if request.user.groups.filter(name = 'supplier').exists():
        supplier_id_list = [user_supplier.supplier_id for user_supplier in UserSupplier.objects.filter(user = request.user).all()]
        supplier_sku_list = [i.sku for i in SkuSupplier.objects.filter(supplier_id__in = supplier_id_list).all()]
        supplier_sku_sales_data = TodayProductSales.objects.filter(product__sku__in=supplier_sku_list, country=default_country).order_by('-sold_qty')
        total_sold_unit = supplier_sku_sales_data.aggregate(Sum('sold_qty'))['sold_qty__sum']
        country = default_country
        context = {
            'date': todaySales.date.strftime('%Y/%m/%d') ,\
            'total_sold_unit': total_sold_unit ,\
            'supplier_sku_sales_data': supplier_sku_sales_data ,\
            'country': country ,\
            'today': True
        }
    else:
        class TopSku:
            def __init__(self, sku, sold_qty):
                self.sku = sku
                self.sold_qty = sold_qty
        top_sales_skus = TodayProductSales.objects.filter(country=default_country).order_by('-sales_amount')[:5]
        top_sold_qty_skus = TodayProductSales.objects.filter(country=default_country).order_by('-sold_qty')[:5]
        sales_today = todaySales.sales_today
        country = default_country
        countries = list(set([i.country for i in TodayProductSales.objects.all()]))
        countries = [c for c in countries if c != country]
        new_or_all = True
        if 'new_or_all' in request.GET and request.GET['new_or_all']:
            new_or_all = bool(int(request.GET['new_or_all']))
        if new_or_all:
            all_today_new_product_skus = TodayProductSales.objects.filter(product__new = True, country=default_country).order_by('-sales_amount')
            sales_new_product = TodayProductSales.objects.filter(product__new = True, country=default_country).aggregate(Sum('sales_amount'))['sales_amount__sum']
            if sales_today:
                sales_new_product_percent = int(sales_new_product / sales_today * 100)
            else:
                sales_new_product_percent = 0
            context = {
                'date': todaySales.date.strftime('%Y/%m/%d') ,\
                'sales_today':  sales_today,\
                'sales_same_day_last_year': todaySales.sales_same_day_last_year ,\
                'monthly_increase_on_sales': todaySales.monthly_increase_on_sales  * 100,\
                'ad_cost': todaySales.ad_cost ,\
                'acos': todaySales.acos * 100 ,\
                'ad_cost_on_sales': todaySales.ad_cost_on_sales * 100 ,\
                'top_sales_skus': top_sales_skus ,\
                'top_sold_qty_skus': top_sold_qty_skus ,\
                'all_today_new_product_skus' : all_today_new_product_skus ,\
                'sales_new_product_percent' : sales_new_product_percent ,\
                'country': country ,\
                'countries': countries ,\
                'today': True ,\
                'new_or_all': new_or_all \
            }
        else:
            all_today_product_skus = TodayProductSales.objects.filter(country=default_country).order_by('-sales_amount')
            context = {
                'date': todaySales.date.strftime('%Y/%m/%d') ,\
                'sales_today':  sales_today,\
                'sales_same_day_last_year': todaySales.sales_same_day_last_year ,\
                'monthly_increase_on_sales': todaySales.monthly_increase_on_sales  * 100,\
                'ad_cost': todaySales.ad_cost ,\
                'acos': todaySales.acos * 100 ,\
                'ad_cost_on_sales': todaySales.ad_cost_on_sales * 100 ,\
                'top_sales_skus': top_sales_skus ,\
                'top_sold_qty_skus': top_sold_qty_skus ,\
                'all_today_product_skus': all_today_product_skus ,\
                #'all_today_new_product_skus' : all_today_new_product_skus ,\
                #'sales_new_product_percent' : sales_new_product_percent ,\
                'country': country ,\
                'countries': countries ,\
                'today': True ,\
                'new_or_all': new_or_all \
            }
    return HttpResponse(template.render(context, request))

@login_required
def other_country_today_sales(request, country):
    template = loader.get_template('other_country_today_sales_todaySales.html')
    default_country = country
    todaySales = HistoryTodaySales.objects.filter(country=default_country).order_by('-date')[0]

    if request.user.groups.filter(name = 'supplier').exists():
        context = {}
        # user_supplier = UserSupplier.objects.filter(user = request.user).all()[0]
        # supplier = user_supplier.supplier
        # suplier_sku_list = [i.sku for i in SkuSupplier.objects.filter(supplier = supplier).all()]
        # supplier_sku_sales_data = TodayProductSales.objects.filter(product__sku__in=suplier_sku_list, country='US').order_by('-sold_qty')
        # total_sold_unit = supplier_sku_sales_data.aggregate(Sum('sold_qty'))['sold_qty__sum']
        # context = {
        #     'date': todaySales.date.strftime('%Y/%m/%d') ,\
        #     'total_sold_unit': total_sold_unit ,\
        #     'supplier_sku_sales_data': supplier_sku_sales_data ,\
        #     'today': True
        # }
    else:
        class TopSku:
            def __init__(self, sku, sold_qty):
                self.sku = sku
                self.sold_qty = sold_qty
        top_sales_skus = TodayProductSales.objects.filter(country=default_country).order_by('-sales_amount')[:5]
        top_sold_qty_skus = TodayProductSales.objects.filter(country=default_country).order_by('-sold_qty')[:5]
        all_today_new_product_skus = TodayProductSales.objects.filter(product__new = True, country=default_country).order_by('-sales_amount')
        sales_today = todaySales.sales_today
        all_today_product_skus = TodayProductSales.objects.filter(country=default_country).order_by('-sales_amount')
        countries = list(set([i.country for i in TodayProductSales.objects.all()]))
        countries = [c for c in countries if c != country]
        context = {
            'date': todaySales.date.strftime('%Y/%m/%d') ,\
            'sales_today':  sales_today,\
            'sales_same_day_last_year': todaySales.sales_same_day_last_year ,\
            'monthly_increase_on_sales': todaySales.monthly_increase_on_sales  * 100,\
            'ad_cost': todaySales.ad_cost ,\
            'acos': todaySales.acos * 100 ,\
            'ad_cost_on_sales': todaySales.ad_cost_on_sales * 100 ,\
            'top_sales_skus': top_sales_skus ,\
            'top_sold_qty_skus': top_sold_qty_skus ,\
            'all_today_product_skus' : all_today_product_skus ,\
            'country': country ,\
            'countries': countries ,\
            'today': True
        }
    return HttpResponse(template.render(context, request))



def date_split(date):
    if len(date) > 4:
        year = int(date[:4])
        month = int(date[4:6])
        day = int(date[6:])
    else:
        year = datetime.date.today().year
        month = int(date[:2])
        day = int(date[2:])
    return {'year':year,'month':month,'day':day}

@login_required
def get_top_sales_vialation(request, country):
    if request.user.groups.filter(name = 'sales').exists():
        template = loader.get_template('get_top_sales_vialation.html')
        countries = list(set([i.country for i in TodayProductSales.objects.all()]))
        countries = [c for c in countries if c != country]
        Min_Sales_Qty = 3
        first_history_today_product_sales = HistoryTodayProductSales.objects.filter(country = country).order_by('-date').first()
        seven_days_ago_date = first_history_today_product_sales.date - datetime.timedelta(days = 7)

        seven_days_ago_history_today_product_sales = HistoryTodayProductSales.objects.filter(date = seven_days_ago_date,  sold_qty_average_7d__gt= Min_Sales_Qty, country=country).all()
        today_product_sales = HistoryTodayProductSales.objects.filter(date = first_history_today_product_sales.date,  sold_qty_average_7d__gt= Min_Sales_Qty, country=country).all()
        skus_list = []
        for tps in today_product_sales:
            sku = tps.product.sku
            skus_list.append(sku)
        for tps in seven_days_ago_history_today_product_sales:
            sku = tps.product.sku
            skus_list.append(sku)
        skus_list = list(set(skus_list))
        skus_seven_days_sales_rates = []
        for sku in skus_list:
            if HistoryTodayProductSales.objects.filter(date = first_history_today_product_sales.date,  product__sku=sku, country=country).count() and HistoryTodayProductSales.objects.filter(date = seven_days_ago_date,  product__sku=sku, country=country).count():
                skus_seven_days_sales_rates.append({'sku': sku, 'seven_days_sales_rate': HistoryTodayProductSales.objects.filter(date = first_history_today_product_sales.date,  product__sku=sku, country=country).first().sold_qty_average_7d / HistoryTodayProductSales.objects.filter(date = seven_days_ago_date,  product__sku=sku, country=country).first().sold_qty_average_7d})
        skus_seven_days_sales_rates = sorted(skus_seven_days_sales_rates, key=lambda i:i['seven_days_sales_rate'])
        skus_seven_days_sales_rates_top = list(reversed(skus_seven_days_sales_rates))[:5]
        skus_seven_days_sales_rates_tail = skus_seven_days_sales_rates[:5]
        seven_days_sales_rates_top = [round(i['seven_days_sales_rate'],1) for i in skus_seven_days_sales_rates_top if i['seven_days_sales_rate'] > 1]
        seven_days_sales_rates_tail = [round(i['seven_days_sales_rate'],1) for i in skus_seven_days_sales_rates_tail if i['seven_days_sales_rate'] < 1]
        history_today_product_sales_top = [HistoryTodayProductSales.objects.filter(date = first_history_today_product_sales.date, product__sku = i['sku'],country = country).first() for i in skus_seven_days_sales_rates_top if i['seven_days_sales_rate'] > 1]
        history_today_product_sales_tail = [HistoryTodayProductSales.objects.filter(date = first_history_today_product_sales.date, product__sku = i['sku'],country = country).first() for i in skus_seven_days_sales_rates_tail if i['seven_days_sales_rate'] < 1]

        context = {
            'history_today_product_sales_top': history_today_product_sales_top ,\
            'history_today_product_sales_tail': history_today_product_sales_tail ,\
            'seven_days_sales_rates_top': seven_days_sales_rates_top ,\
            'seven_days_sales_rates_tail': seven_days_sales_rates_tail ,\
            'country': country ,\
            'countries': countries ,\
        }
        return HttpResponse(template.render(context, request))

@login_required
def get_excess_inventory(request, country):
    if request.user.groups.filter(name = 'sales').exists():
        template = loader.get_template('get_excess_inventory.html')
        import operator
        class ExcessProduct:
            def __init__(self, product, sold_qty_average_7d, sold_qty_today, total_unit, days_maintain, excess_qty, excess_value):
                self.product = product
                self.sold_qty_average_7d = sold_qty_average_7d
                self.sold_qty_today = sold_qty_today
                self.total_unit = total_unit
                self.days_maintain = days_maintain
                self.excess_qty = excess_qty
                self.excess_value = excess_value
        countries = list(set([i.country for i in TodayProductSales.objects.all()]))
        countries = [c for c in countries if c != country]
        Min_Days_Maintain = 180.0
        Default_Long_Days = 365.0
        excess_product_list = []
        excess_product_total_unit = 0.0
        excess_product_total_value = 0.0
        default_purchasing_price = 30.0
        for fba_inventory in FbaInventory.objects.filter(country = country).all():
            sku = fba_inventory.sku
            days_maintain = Default_Long_Days
            if TodayProductSales.objects.filter(country=country, product__sku = sku).count():
                today_product_sales = TodayProductSales.objects.filter(country=country, product__sku = sku).first()
                product = today_product_sales.product
                sold_qty_average_7d = today_product_sales.sold_qty_average_7d
                sold_qty_today = today_product_sales.sold_qty
                days_maintain = fba_inventory.total_unit / sold_qty_average_7d
                if days_maintain > Min_Days_Maintain:
                    excess_qty = fba_inventory.total_unit - sold_qty_average_7d * Min_Days_Maintain
                    if product.sku:

                        excess_product_total_unit += excess_qty
                        if SkuPurchasingPrice.objects.filter(sku = sku).count():
                            price = SkuPurchasingPrice.objects.filter(sku = sku).order_by('-date').first().purchasing_price
                        else:
                            price = default_purchasing_price
                        excess_value = excess_qty * price
                        excess_product_total_value += excess_value
                        excess_product = ExcessProduct(product, sold_qty_average_7d, sold_qty_today,  fba_inventory.total_unit, days_maintain, excess_qty,excess_value)
                        excess_product_list.append(excess_product)
            elif fba_inventory.total_unit > 0.0:
                days_maintain = Default_Long_Days
                excess_qty = fba_inventory.total_unit
                sold_qty_average_7d = 0.0
                sold_qty_today = 0.0
                if Product.objects.filter(sku = sku).count():
                    excess_product_total_unit += excess_qty
                    if SkuPurchasingPrice.objects.filter(sku = sku).count():
                        price = SkuPurchasingPrice.objects.filter(sku = sku).order_by('-date').first().purchasing_price
                    else:
                        price = default_purchasing_price
                    excess_value = excess_qty * price
                    excess_product_total_value += excess_value
                    excess_product = ExcessProduct(Product.objects.filter(sku = sku).first(), sold_qty_average_7d, sold_qty_today,  fba_inventory.total_unit, days_maintain, excess_qty,excess_value)
                    excess_product_list.append(excess_product)
        excess_product_list.sort(key=operator.attrgetter('excess_qty'),reverse = True)
        context = {
            'excess_product_list': excess_product_list ,\
            'excess_product_total_unit': excess_product_total_unit ,\
            'excess_product_total_value': excess_product_total_value ,\
            'country': country ,\
            'countries': countries ,\
        }
        return HttpResponse(template.render(context, request))




@login_required
def history_index(request, year, month, day):
    date = "%04d/%02d/%02d" % (year, month, day,)
    today_date_obj = datetime.date(year, month, day )
    country = 'US'
    if 'country' in request.GET and request.GET['country']:
        country = request.GET['country']
    country = country
    if request.user.groups.filter(name = 'supplier').exists():
        template = loader.get_template('todaySalesForSupplier.html')
    else:
        if country == 'US':
            template = loader.get_template('todaySales.html')
        else:
            template = loader.get_template('other_country_today_sales_todaySales.html')
    todaySales = HistoryTodaySales.objects.filter(date = today_date_obj, country = country)
    if todaySales.count():
        todaySales = todaySales.all()[0]
    else:
        todaySales = HistoryTodaySales.objects.filter(country = country).order_by('-date')[0]

    if request.user.groups.filter(name = 'supplier').exists():
        supplier_id_list = [user_supplier.supplier_id for user_supplier in UserSupplier.objects.filter(user = request.user).all()]
        supplier_sku_list = [i.sku for i in SkuSupplier.objects.filter(supplier_id__in = supplier_id_list).all()]
        supplier_sku_sales_data = HistoryTodayProductSales.objects.filter(date = today_date_obj, product__sku__in=supplier_sku_list, country=country).order_by('-sold_qty')

        total_sold_unit = supplier_sku_sales_data.aggregate(Sum('sold_qty'))['sold_qty__sum']
        context = {
            'date': todaySales.date.strftime('%Y/%m/%d') ,\
            'total_sold_unit': total_sold_unit ,\
            'supplier_sku_sales_data': supplier_sku_sales_data ,\
            'today': True ,\
            'country': country
        }
    else:
        class TopSku:
            def __init__(self, sku, sold_qty):
                self.sku = sku
                self.sold_qty = sold_qty
        top_sales_skus = HistoryTodayProductSales.objects.filter(date = today_date_obj, country=country).order_by('-sales_amount')[:5]
        top_sold_qty_skus = HistoryTodayProductSales.objects.filter(date = today_date_obj, country=country).order_by('-sold_qty')[:5]

        countries = list(set([i.country for i in TodayProductSales.objects.all()]))
        countries = [c for c in countries if c != country]
        new_or_all = True
        if 'new_or_all' in request.GET and request.GET['new_or_all']:
            new_or_all = bool(int(request.GET['new_or_all']))
        if country == 'US':
            if new_or_all:
                all_today_new_product_skus = HistoryTodayProductSales.objects.filter(date = today_date_obj, product__new = True, country = country).order_by('-sales_amount')
                sales_new_product = all_today_new_product_skus.aggregate(Sum('sales_amount'))['sales_amount__sum']
                sales_new_product_percent = int(sales_new_product / todaySales.sales_today * 100)
                context = {
                    'date': todaySales.date.strftime('%Y/%m/%d') ,\
                    'sales_today': todaySales.sales_today ,\
                    'sales_same_day_last_year': todaySales.sales_same_day_last_year ,\
                    'monthly_increase_on_sales': todaySales.monthly_increase_on_sales  * 100,\
                    'ad_cost': todaySales.ad_cost ,\
                    'acos': todaySales.acos * 100 ,\
                    'ad_cost_on_sales': todaySales.ad_cost_on_sales * 100 ,\
                    'top_sales_skus': top_sales_skus ,\
                    'top_sold_qty_skus': top_sold_qty_skus ,\
                    'all_today_new_product_skus' : all_today_new_product_skus ,\
                    'sales_new_product_percent' : sales_new_product_percent ,\
                    'today': False,\
                    'country': country ,\
                    'countries': countries ,\
                    'new_or_all': new_or_all \
                }
            else:
                all_today_product_skus =  HistoryTodayProductSales.objects.filter(date = today_date_obj, country = country).order_by('-sales_amount')
                context = {
                    'date': todaySales.date.strftime('%Y/%m/%d') ,\
                    'sales_today': todaySales.sales_today ,\
                    'sales_same_day_last_year': todaySales.sales_same_day_last_year ,\
                    'monthly_increase_on_sales': todaySales.monthly_increase_on_sales  * 100,\
                    'ad_cost': todaySales.ad_cost ,\
                    'acos': todaySales.acos * 100 ,\
                    'ad_cost_on_sales': todaySales.ad_cost_on_sales * 100 ,\
                    'top_sales_skus': top_sales_skus ,\
                    'top_sold_qty_skus': top_sold_qty_skus ,\
                    'all_today_product_skus': all_today_product_skus ,\
                    # 'all_today_new_product_skus' : all_today_new_product_skus ,\
                    # 'sales_new_product_percent' : sales_new_product_percent ,\
                    'today': False,\
                    'country': country ,\
                    'countries': countries ,\
                    'new_or_all': new_or_all \
                }
        else:
            all_today_product_skus =  HistoryTodayProductSales.objects.filter(date = today_date_obj, country = country).order_by('-sales_amount')
            context = {
                'date': todaySales.date.strftime('%Y/%m/%d') ,\
                'sales_today': todaySales.sales_today ,\
                'sales_same_day_last_year': todaySales.sales_same_day_last_year ,\
                'monthly_increase_on_sales': todaySales.monthly_increase_on_sales  * 100,\
                'ad_cost': todaySales.ad_cost ,\
                'acos': todaySales.acos * 100 ,\
                'ad_cost_on_sales': todaySales.ad_cost_on_sales * 100 ,\
                'top_sales_skus': top_sales_skus ,\
                'top_sold_qty_skus': top_sold_qty_skus ,\
                'all_today_product_skus': all_today_product_skus ,\
                # 'all_today_new_product_skus' : all_today_new_product_skus ,\
                # 'sales_new_product_percent' : sales_new_product_percent ,\
                'today': False,\
                'country': country ,\
                'countries': countries ,\
                'new_or_all': new_or_all \
            }
    return HttpResponse(template.render(context, request))

@login_required
def update_last_year_sales(request):
    if request.method == 'POST':
        start_line_list = ['Time', 'Selected date range (Ordered product sales)', 'Selected date range (Units ordered)']
        end_line = ['Compare Sales - Table view']
        start_read = False
        fileInMemory = request.FILES['file'].read().decode('utf-8')
        csv_data = csv.reader(StringIO(fileInMemory), delimiter=',')
        DailySalesLastYear.objects.all().delete()
        for row in csv_data:
            if start_line_list == row:

                start_read = True
                continue
            elif end_line == row:
                break
            if start_read and row != []:
                date_str = row[0]
                sales = float(row[1].replace(',','').replace('$',''))
                date_str_split = date_str.split('T')[0].split('-')
                day = date_str_split[2]
                month = date_str_split[1]
                dailySalesLastYear = DailySalesLastYear(day = int(day), month = int(month), sales = sales)
                dailySalesLastYear.save()
            elif start_read and row == []:
                start_read = False

    template = loader.get_template('updateLastYearSales.html')
    context = {
        'form':UploadFileForm()
    }

    return HttpResponse(template.render(context, request))

def check_if_within_7day(year_month_day,time_to_check):
    timezone = pytz.timezone('America/Los_Angeles')
    date_time_obj = datetime.datetime.strptime(time_to_check[:19], '%Y-%m-%dT%H:%M:%S')
    timezone_date_time_obj = date_time_obj.astimezone(pytz.timezone('America/Los_Angeles'))
    countable_day = timezone_date_time_obj.day
    # start_datetime = datetime.datetime(timezone_date_time_obj.year, \
    start_datetime = datetime.datetime(year_month_day['year'], \
                                        year_month_day['month'], year_month_day['day'],\
                                        0,0,0,0,pytz.timezone('America/Los_Angeles') \
                                        )
    compare_datetime = start_datetime + datetime.timedelta(days=1)
    compare_earliest_datetime = start_datetime - datetime.timedelta(days=6)
    timezone_date_time_obj = datetime.datetime(timezone_date_time_obj.year, \
                                        timezone_date_time_obj.month, timezone_date_time_obj.day,\
                                        timezone_date_time_obj.hour,timezone_date_time_obj.minute, \
                                        timezone_date_time_obj.second,0,pytz.timezone('America/Los_Angeles') \
                                        )
    if timezone_date_time_obj > compare_datetime or timezone_date_time_obj <= compare_earliest_datetime:
        return False
    else:
        return True

def check_if_today(end_day,time_to_check):
    timezone = pytz.timezone('America/Los_Angeles')
    date_time_obj = datetime.datetime.strptime(time_to_check[:19], '%Y-%m-%dT%H:%M:%S')
    timezone_date_time_obj = date_time_obj.astimezone(pytz.timezone('America/Los_Angeles'))
    countable_day = timezone_date_time_obj.day
    if countable_day == end_day:
        return True
    return False

def save_history_today_product_sales(todayProductSales,date,country):
    historyTodayProductSales = HistoryTodayProductSales.objects.filter(date = date,product = todayProductSales.product,country = country)
    if historyTodayProductSales.count():
        historyTodayProductSales = historyTodayProductSales.all()[0]
        historyTodayProductSales.sold_qty = todayProductSales.sold_qty
        historyTodayProductSales.sales_amount = todayProductSales.sales_amount
        historyTodayProductSales.sold_qty_average_7d = todayProductSales.sold_qty_average_7d
        historyTodayProductSales.average_price_7d =  todayProductSales.average_price_7d
        historyTodayProductSales.fba_inventory =  todayProductSales.fba_inventory
        historyTodayProductSales.lasting_day_estimated_by_us =  todayProductSales.lasting_day_estimated_by_us
        historyTodayProductSales.lasting_day_of_available_estimated_by_us =  todayProductSales.lasting_day_of_available_estimated_by_us
        historyTodayProductSales.lasting_day_of_available_fc_estimated_by_us =  todayProductSales.lasting_day_of_available_fc_estimated_by_us
        historyTodayProductSales.lasting_day_of_total_fba_unit_estimated_by_us =  todayProductSales.lasting_day_of_total_fba_unit_estimated_by_us
    else:
        historyTodayProductSales = HistoryTodayProductSales( \
            product = todayProductSales.product ,\
            date = date ,\
            sold_qty = todayProductSales.sold_qty ,\
            sales_amount = todayProductSales.sales_amount ,\
            sold_qty_average_7d = todayProductSales.sold_qty_average_7d ,\
            average_price_7d =  todayProductSales.average_price_7d ,\
            fba_inventory =  todayProductSales.fba_inventory ,\
            lasting_day_estimated_by_us =  todayProductSales.lasting_day_estimated_by_us ,\
            lasting_day_of_available_estimated_by_us =  todayProductSales.lasting_day_of_available_estimated_by_us ,\
            lasting_day_of_available_fc_estimated_by_us =  todayProductSales.lasting_day_of_available_fc_estimated_by_us ,\
            lasting_day_of_total_fba_unit_estimated_by_us =  todayProductSales.lasting_day_of_total_fba_unit_estimated_by_us ,\
            country = country )
    historyTodayProductSales.save()

def save_today_sales(todaySales, date):
    default_country = 'US'
    historyTodaySales = HistoryTodaySales.objects.filter(date = date, country = default_country)
    if historyTodaySales.count():
        historyTodaySales = historyTodaySales.all()[0]
        historyTodaySales.sales_today = todaySales.sales_today
        historyTodaySales.sales_same_day_last_year = todaySales.sales_same_day_last_year
        historyTodaySales.sales_month_to_date = todaySales.sales_month_to_date
        historyTodaySales.monthly_increase_on_sales = todaySales.monthly_increase_on_sales
        historyTodaySales.ad_cost = todaySales.ad_cost
        historyTodaySales.acos = todaySales.acos
        historyTodaySales.ad_cost_on_sales = todaySales.ad_cost_on_sales
    else:
        historyTodaySales = HistoryTodaySales( \
                                date = date ,\
                                sales_today = todaySales.sales_today ,\
                                sales_same_day_last_year = todaySales.sales_same_day_last_year ,\
                                sales_month_to_date = todaySales.sales_month_to_date ,\
                                monthly_increase_on_sales = todaySales.monthly_increase_on_sales ,\
                                ad_cost = todaySales.ad_cost ,\
                                acos = todaySales.acos ,\
                                ad_cost_on_sales = todaySales.ad_cost_on_sales ,\
                                            )
    historyTodaySales.save()

def sales_channel_converting(x):
    if x == 'Amazon.com':
        ret = 'US'
    elif x == 'Amazon.ca':
        ret = 'CA'
    elif x == 'Amazon.com.mx':
        ret = 'MX'
    elif x == 'Amazon.co.uk':
        ret = 'GB'
    elif x in ['Amazon.de', 'Amazon.it', 'Amazon.fr', 'Amazon.es', 'Amazon.se', 'Amazon.nl', 'Amazon.pl', 'Amazon.com.tr', 'Amazon.com.be']  :
        ret = 'EU'
    elif x == 'Amazon.ae':
        ret = 'AE'
    elif x == 'Amazon.co.jp':
        ret = 'JP'
    return ret

def harmonize_price(price, sales_channel):
    ret = price
    if sales_channel  == 'Amazon.pl':
        currency_rate = CurrencyRate.objects.filter(from_country = 'PL', to_country = 'EU').order_by('-date').first()
        ret = price * currency_rate.rate
    elif sales_channel  == 'Amazon.se':
        currency_rate = CurrencyRate.objects.filter(from_country = 'SE', to_country = 'EU').order_by('-date').first()
        ret = price * currency_rate.rate
    elif sales_channel  == 'Amazon.com.tr':
        currency_rate = CurrencyRate.objects.filter(from_country = 'TR', to_country = 'EU').order_by('-date').first()
        ret = price * currency_rate.rate
    return ret

def trim_sku(x):
    x = x.strip()
    if len(x) and x[:1] == '?':
        x = x[1:]
    return x.strip('\u200e')

def amzgrsku2sku(x):
    x = trim_sku(x)
    if x[:8] == 'amzn.gr.':
        str = x[8:-16]
        endWithDigit = re.search(r'-[A-Z]-\d+',str)
        if endWithDigit is None:
            if x == 'amzn.gr.v1.65xIJT-M-gciQxewPHFUPx2dUWI':
                return 'LS-LHL-09-13-A-1'
            else:
                for sku in Product.objects.all().values_list('sku', flat=True):
                    if sku in x:
                        return sku
            # elif x == 'amzn.gr.ERS-LHL-01-06-B-1-Jij43wc9PUt-VG':
            #     return 'ERS-LHL-01-06-B-1'
            # elif x.count('"') == 2:
            #     iter = re.finditer('"',x)
            #     indices = [m.start(0) for m in iter]
            #     return x[(indices[0]+1):indices[1]]
            return None
        else:
            if x.count('"') > 0:
                iter = re.finditer('"',x)
                indices = [m.start(0) for m in iter]
                return x[(indices[0]+1):indices[1]]
            else:
                return str[:endWithDigit.end()]
    return x

@login_required
def update_7d_orders(request):
    if request.method == 'POST':
        date = request.POST['date']
        year_month_day = date_split(date)
        year = year_month_day['year']
        month = year_month_day['month']
        day = year_month_day['day']
        #start_line_list = ['Time', 'Selected date range (OPS)', 'Selected date range (Units)']
        #start_read = False
        encodings = ENCODINGS
        for e in encodings:
            try:
                request.FILES['file'].seek(0,0)
                fileInMemory = request.FILES['file'].read().decode(e)
                break
            except UnicodeDecodeError:
                pass
        # fileInMemory = request.FILES['file'].read().decode('utf-8')
        csv_data = csv.reader(StringIO(fileInMemory), delimiter='\t')
        all_country_sku_qty_price = {}
        order_list = []
        countries_set = set()
        today_date_obj = datetime.date(year, month, day )
        for row in csv_data:
            # 判断是哪个渠道的订单
            order_country_cell = row[6]
            if order_country_cell not in ['Non-Amazon', 'sales-channel'] and order_country_cell[:3] != 'SI ':
                country = sales_channel_converting(order_country_cell)
                sku = amzgrsku2sku(row[11])
                qty = int(row[14])
                order_status = row[4]
                if qty and order_status != 'Cancelled':
                    if_today = check_if_today(day,row[2])
                    price = float(row[16])
                    price = harmonize_price(price, order_country_cell)
                    countries_set.add(country)
                    order_list.append({'country': country, 'sku': sku, 'qty': qty, 'price': price, 'if_today': if_today})
        for country in list(countries_set):
            all_country_sku_qty_price[country]= {'sku_qty':{},'sku_price':{},'sku_today_qty':{},'sku_today_price':{},'total_today_sales':0.0}
            Last7dayProductSales.objects.filter(country = country).delete()
            TodayProductSales.objects.filter(country = country).delete()
        for order in order_list:
            price = order['price']
            sku = order['sku']
            qty = order['qty']
            if_today = order['if_today']
            country = order['country']
            if sku in all_country_sku_qty_price[country]['sku_qty']:
                all_country_sku_qty_price[country]['sku_qty'][sku] += qty
                all_country_sku_qty_price[country]['sku_price'][sku] += price
                if sku in all_country_sku_qty_price[country]['sku_today_qty'] and if_today:
                    all_country_sku_qty_price[country]['sku_today_qty'][sku] += qty
                    all_country_sku_qty_price[country]['sku_today_price'][sku] += price
            else:
                all_country_sku_qty_price[country]['sku_qty'][sku] = qty
                all_country_sku_qty_price[country]['sku_price'][sku] = price
                if if_today:
                    all_country_sku_qty_price[country]['sku_today_qty'][sku] = qty
                    all_country_sku_qty_price[country]['sku_today_price'][sku] = price
                else:
                    if sku not in all_country_sku_qty_price[country]['sku_today_qty']:
                        all_country_sku_qty_price[country]['sku_today_qty'][sku] = 0
                        all_country_sku_qty_price[country]['sku_today_price'][sku] = 0.0
        for k, v in all_country_sku_qty_price.items():
            country = k
            sku_qty = v['sku_qty']
            sku_price = v['sku_price']
            sku_today_qty = v['sku_today_qty']
            sku_today_price = v['sku_today_price']
            total_today_sales = v['total_today_sales']
            for sku, qty in sku_qty.items():
                product, created = Product.objects.get_or_create(sku = sku)
                if created:
                    product.save()
                sales_amount = sku_price[sku]
                sold_qty_average_7d = qty / 7
                average_price_7d = sales_amount / qty
                last7dayProductSales = Last7dayProductSales(product = product ,\
                                                            sold_qty = qty ,\
                                                            sales_amount = sales_amount ,\
                                                            sold_qty_average_7d = sold_qty_average_7d ,\
                                                            average_price_7d = average_price_7d ,\
                                                            country = country)
                last7dayProductSales.save()
                if sku in sku_today_qty:
                    todayProductSales = TodayProductSales(product = product ,\
                                                          sold_qty = sku_today_qty[sku] ,\
                                                          sales_amount = sku_today_price[sku] ,\
                                                          sold_qty_average_7d = sold_qty_average_7d ,\
                                                          average_price_7d = average_price_7d ,\
                                                          country = country)
                    fbaInventory = FbaInventory.objects.filter(sku = sku,country =country)
                    if fbaInventory.count():
                        todayProductSales.fba_inventory = fbaInventory.all()[0]
                        todayProductSales.lasting_day_estimated_by_us = int(float(todayProductSales.fba_inventory.total_unit) / float(sold_qty_average_7d))
                    todayProductSales.save()

                    save_history_today_product_sales(todayProductSales,today_date_obj, country)
                    total_today_sales += sku_today_price[sku]
            save_today_sales(year_month_day = year_month_day, sales_today = total_today_sales, country=country)
        return redirect('today_sales')

    template = loader.get_template('update_7d_orders.html')
    today_date_str = ""
    if HistoryTodayProductSales.objects.first() != None:
        today_date_str = HistoryTodayProductSales.objects.order_by('-date').first().date.strftime('%m%d')
    class UploadFileAndDateForm(forms.Form):
        date = forms.CharField(max_length=8)
        file = forms.FileField()
    context = {
        'form':UploadFileAndDateForm(initial={'date': today_date_str})
    }

    return HttpResponse(template.render(context, request))


def country_harmonizing(x):
    if x in ['DE','FR','SE','ES','NL','PL','TR','IT','TR']:
        return 'EU'
    return x

@login_required
def update_restock_report(request):
    if request.method == 'POST':
        encodings = ENCODINGS
        for e in encodings:
            try:
                request.FILES['file'].seek(0,0)
                fileInMemory = request.FILES['file'].read().decode(e)
                break
            except UnicodeDecodeError:
                pass
        csv_data = csv.reader(StringIO(fileInMemory), delimiter='\t')
        for row in csv_data:
            sku = trim_sku(row[3])
            fnsku = row[2]
            if fnsku != 'FNSKU':
                country = country_harmonizing(row[0])
                fnsku = row[2]
                asin = row[4]
                working_unit = int(row[19])
                total_unit = int(row[12])# + working_unit
                available = int(row[14])
                inbound_fc_unit = int(row[13]) + int(row[15]) + int(row[16])
                inbound_unit = int(row[13])
                fc_unit = int(row[15]) + int(row[16])
                days_of_supply = row[20+3]
                try:
                    days_of_supply = int(days_of_supply)
                except:
                    days_of_supply = 365
                recommended_replenishment_qty = row[22+4]
                if recommended_replenishment_qty:
                    recommended_replenishment_qty = int(recommended_replenishment_qty)
                else:
                    recommended_replenishment_qty = 0
                recommended_ship_date = row[23+4]
                if not(recommended_ship_date):
                    recommended_ship_date = ''
                fbaInventory = FbaInventory.objects.filter(sku = sku,country = country)
                if fbaInventory.count():
                    fbaInventory = fbaInventory.all()[0]
                    fbaInventory.fnsku = fnsku
                    fbaInventory.asin = asin
                    fbaInventory.total_unit = total_unit
                    fbaInventory.available = available
                    fbaInventory.inbound_fc_unit = inbound_fc_unit
                    fbaInventory.inbound_unit = inbound_unit
                    fbaInventory.fc_unit = fc_unit
                    fbaInventory.days_of_supply = days_of_supply
                    fbaInventory.recommended_replenishment_qty = recommended_replenishment_qty
                    fbaInventory.recommended_ship_date = recommended_ship_date
                    fbaInventory.country = country
                else:
                    fbaInventory = FbaInventory( \
                                               sku = sku ,\
                                               fnsku = fnsku ,\
                                               asin = asin ,\
                                               total_unit = total_unit ,\
                                               available = available ,\
                                               inbound_unit = inbound_unit ,\
                                               fc_unit = fc_unit ,\
                                               inbound_fc_unit = inbound_fc_unit ,\
                                               days_of_supply = days_of_supply ,\
                                               recommended_replenishment_qty = recommended_replenishment_qty ,\
                                               recommended_ship_date = recommended_ship_date ,\
                                               country = country

                    )
                fbaInventory.save()
                todayProductSales = TodayProductSales.objects.filter(product__sku = sku,country = country)
                if todayProductSales.count():
                    todayProductSales = todayProductSales.all()[0]
                    todayProductSales.fba_inventory = fbaInventory
                    todayProductSales.save()
        return redirect('today_sales')
    template = loader.get_template('update_restock_report.html')

    context = {
        'form':UploadFileForm()
    }

    return HttpResponse(template.render(context, request))

@login_required
def update_remote_fulfillment_eligible_ASIN_Report(request):
    if request.method == 'POST':
        fileInMemory = request.FILES['file'].read()
        filePath = BytesIO(fileInMemory)
        wb2 = load_workbook(filePath, data_only=True)
        sheet = wb2['Enrollment']
        country = 'CA'
        if sheet.cell(3,4).value in ['加拿大','Canada'] and sheet.cell(4,4).value in ['商品状态','Offer status']:
            RemoteFulfillmentSku.objects.all().delete()
            for i_row in range(5, sheet.max_row + 1):
                sku = sheet.cell(i_row,1).value
                offer_status = sheet.cell(i_row,4).value
                if offer_status in ['Enabled', '已启用']:
                    RemoteFulfillmentSku.objects.get_or_create(sku = sku, country=country)


    template = loader.get_template('upload_a_single_xlsx_file.html')
    context = {
        'header_title' : '请上传Remote Fulfillment Eligible ASIN Report', \
        'form':UploadFileForm()
    }

    return HttpResponse(template.render(context, request))


@login_required
def update_currency_rate(request):
    if request.method == 'POST':
        fileInMemory = request.FILES['file'].read()
        filePath = BytesIO(fileInMemory)
        wb2 = load_workbook(filePath, data_only=True)
        sheet = wb2.worksheets[0]
        rate_dict = {}
        for row in range(1, sheet.max_row + 1):
            row_name = sheet.cell(row,1).value
            row_value = sheet.cell(row,2).value
            rate_dict[row_name] = row_value
            date_update_str = str(sheet.cell(row,5).value)
        date_update = datetime.datetime.strptime(date_update_str, "%Y%m%d").date()
        for k, v in rate_dict.items():
            if k == '1PLN=?RMB':
                rate = v / rate_dict['1EUR=?RMB']
                (currency_rate, found_or_not) = CurrencyRate.objects.get_or_create(from_country = 'PL', to_country = 'EU', date = date_update)
                currency_rate.rate = rate
                currency_rate.save()
            elif k == '1SEK=?RMB':
                rate = v / rate_dict['1EUR=?RMB']
                (currency_rate, found_or_not) = CurrencyRate.objects.get_or_create(from_country = 'SE', to_country = 'EU', date = date_update)
                currency_rate.rate = rate
                currency_rate.save()
            elif k == '1TRY=?RMB':
                rate = v / rate_dict['1EUR=?RMB']
                (currency_rate, found_or_not) = CurrencyRate.objects.get_or_create(from_country = 'TR', to_country = 'EU', date = date_update)
                currency_rate.rate = rate
                currency_rate.save()
            if k == '1PLN=?RMB':
                (currency_rate, found_or_not) = CurrencyRate.objects.get_or_create(from_country = 'PL', to_country = 'CN', date = date_update)
            if k == '1SEK=?RMB':
                (currency_rate, found_or_not) = CurrencyRate.objects.get_or_create(from_country = 'SE', to_country = 'CN', date = date_update)
            if k == '1TRY=?RMB':
                (currency_rate, found_or_not) = CurrencyRate.objects.get_or_create(from_country = 'TR', to_country = 'CN', date = date_update)
            if k == '1USD=?RMB':
                (currency_rate, found_or_not) = CurrencyRate.objects.get_or_create(from_country = 'US', to_country = 'CN', date = date_update)
            if k == '1CAD=?RMB':
                (currency_rate, found_or_not) = CurrencyRate.objects.get_or_create(from_country = 'CA', to_country = 'CN', date = date_update)
            if k == '1AUD=?RMB':
                (currency_rate, found_or_not) = CurrencyRate.objects.get_or_create(from_country = 'AU', to_country = 'CN', date = date_update)
            if k == '1AED=?RMB':
                (currency_rate, found_or_not) = CurrencyRate.objects.get_or_create(from_country = 'AE', to_country = 'CN', date = date_update)
            if k == '1GBP=?RMB':
                (currency_rate, found_or_not) = CurrencyRate.objects.get_or_create(from_country = 'GB', to_country = 'CN', date = date_update)
            if k == '1EUR=?RMB':
                (currency_rate, found_or_not) = CurrencyRate.objects.get_or_create(from_country = 'EU', to_country = 'CN', date = date_update)
            if k == '1SGD=?RMB':
                (currency_rate, found_or_not) = CurrencyRate.objects.get_or_create(from_country = 'SG', to_country = 'CN', date = date_update)
            if k == '1MXN=?RMB':
                (currency_rate, found_or_not) = CurrencyRate.objects.get_or_create(from_country = 'MX', to_country = 'CN', date = date_update)
            currency_rate.rate = v
            currency_rate.save()
    template = loader.get_template('update_currency_rate.html')
    context = {
        'form':UploadFileForm()
    }

    return HttpResponse(template.render(context, request))


@login_required
def update_shenzhen_inventory(request):
    if request.method == 'POST':
        fileInMemory = request.FILES['file'].read()
        filePath = BytesIO(fileInMemory)
        wb2 = load_workbook(filePath, data_only=True)
        sheet = wb2.worksheets[0]
        max_column = sheet.max_column
        if sheet.cell(4,max_column - 1).value == shenzhen_warehouse_name:
            Inventory.objects.filter(warehouse_name = shenzhen_warehouse_name).delete()
            for row in range(6, sheet.max_row):
                sku = sheet.cell(row,1).value
                qty = int(sheet.cell(row,max_column - 1).value)
                shenzhenInventory =Inventory(warehouse_name = shenzhen_warehouse_name,sku = sku, qty = qty)
                shenzhenInventory.save()
        if sheet.cell(3,max_column - 1).value == shenzhen_warehouse_name:
            Inventory.objects.filter(warehouse_name = shenzhen_warehouse_name).delete()
            for row in range(5, sheet.max_row):
                sku = sheet.cell(row,1).value
                qty = int(sheet.cell(row,max_column - 1).value)
                shenzhenInventory =Inventory(warehouse_name = shenzhen_warehouse_name,sku = sku, qty = qty)
                shenzhenInventory.save()
        if sheet.cell(3,max_column).value == shenzhen_warehouse_name:
            Inventory.objects.filter(warehouse_name = shenzhen_warehouse_name).delete()
            for row in range(5, sheet.max_row):
                sku = sheet.cell(row,1).value
                qty = int(sheet.cell(row,max_column).value)
                shenzhenInventory =Inventory(warehouse_name = shenzhen_warehouse_name,sku = sku, qty = qty)
                shenzhenInventory.save()
    template = loader.get_template('update_shenzhen_inventory.html')
    context = {
        'form':UploadFileForm()
    }

    return HttpResponse(template.render(context, request))

@user_passes_test(lambda u: u.groups.filter(name__in = ['logistics', 'accountant']).exists())
def update_purchasing_orders(request):
    if request.method == 'POST':
        wb2 = xlrd.open_workbook(file_contents=request.FILES['file'].read())
        sheet = wb2.sheet_by_index(0)
        max_column = sheet.ncols
        for col in range(0,max_column):
            title = sheet.cell(4,col).value
            if title == '未入库数量':
                qty_col = col
            if title == '状态':
                status_col = col
            if title == '预计交货日期':
                estimated_receiving_date_col = col
        ReceivablePurchasedQty.objects.all().delete()
        for row in range(5,sheet.nrows):
            sku_cell_value = sheet.cell(row,0).value
            if sku_cell_value:
                sku = sku_cell_value
            elif sheet.cell(row,status_col).value == '小计':
                receivable_qty = int(sheet.cell(row,qty_col).value)
                receivablePurchasedQty = ReceivablePurchasedQty(sku = sku, qty = receivable_qty)
                receivablePurchasedQty.save()
        # NearestReceivablePurchasedQty
        NearestReceivablePurchasedQty.objects.all().delete()
        date_qty_list = []
        for row in range(5,sheet.nrows):
            sku_cell_value = sheet.cell(row,0).value
            if sku_cell_value:
                sku = sku_cell_value
                receivable_qty = int(sheet.cell(row,qty_col).value)
                if receivable_qty > 50:
                    date_qty_list.append({'date': datetime.datetime.strptime(sheet.cell(row, estimated_receiving_date_col).value, "%Y-%m-%d").date(), 'receivable_qty': receivable_qty})
            elif sheet.cell(row,status_col).value == '小计':
                if date_qty_list != []:
                    date_qty_nearest = min(date_qty_list, key=lambda x: x['date'])
                    receivable_qty = date_qty_nearest['receivable_qty']
                    date = date_qty_nearest['date']
                    nearestReceivablePurchasedQty = NearestReceivablePurchasedQty(sku = sku, qty = receivable_qty, date = date)
                    nearestReceivablePurchasedQty.save()
                    date_qty_list = []
    template = loader.get_template('update_purchasing_orders.html')
    context = {
        'form':UploadFileForm()
    }

    return HttpResponse(template.render(context, request))

def month_to_date_last_year_month_to_date(year,month,day):
    # dailySalesLastYear = DailySalesLastYear.objects.filter(month = month, day = day).all()[0]
    # dailySalesLastYear = DailySalesLastYear.objects.get(pk = dailySalesLastYear.pk + 2)
    default_country = 'US'
    sales_month_to_date_last_year = DailySalesLastYear.objects.filter(month = month, day__lte = day).aggregate(Sum('sales'))['sales__sum']
    sales_month_to_date = HistoryTodaySales.objects.filter(country = default_country, date__range=(datetime.date(year, month, 1), datetime.date(year, month, day))).aggregate(Sum('sales_today'))['sales_today__sum']
    if sales_month_to_date == None:
        return {'sales_month_to_date': 0.0 ,\
                'monthly_increase_on_sales': 0.0}
    else:
        monthly_increase_on_sales = (sales_month_to_date - sales_month_to_date_last_year)/sales_month_to_date_last_year
        return {'sales_month_to_date': sales_month_to_date ,\
                'monthly_increase_on_sales': monthly_increase_on_sales}

def save_today_sales(**kwargs):
    year_month_day = kwargs['year_month_day']
    country = kwargs['country']
    year = year_month_day['year']
    month = year_month_day['month']
    day = year_month_day['day']
    dailySalesLastYear = DailySalesLastYear.objects.filter(month = month, day = day).all()[0]
    from django.core.exceptions import ObjectDoesNotExist
    try:
        dailySalesLastYear = DailySalesLastYear.objects.get(pk = dailySalesLastYear.pk + 1)
    except ObjectDoesNotExist:
        pass
    today_date_obj = datetime.date(year = year, month = month, day = day)
    #检查当天HistoryTodaySales 是否已经建立
    if not HistoryTodaySales.objects.filter(date=today_date_obj, country=country).count() and 'sales_today' in kwargs.keys():
        historyTodaySales = HistoryTodaySales(date = today_date_obj ,\
                                            sales_today = kwargs['sales_today'] ,\
                                            sales_same_day_last_year = 999.99 ,\
                                            sales_month_to_date = 999.99 ,\
                                            monthly_increase_on_sales = 999.99 ,\
                                            ad_cost = 999.99 ,\
                                            acos = 999.99 ,\
                                            ad_cost_on_sales =999.99 ,\
                                            country= country)
        historyTodaySales.save()
    else:
        historyTodaySales = HistoryTodaySales.objects.filter(date = today_date_obj,country= country).first()
        if 'sales_today' in kwargs.keys():
            historyTodaySales.sales_today = kwargs['sales_today']
    if(country == 'US'):
        sales_month_to_date_and_monthly_increase_on_sales = month_to_date_last_year_month_to_date(year, month, day)

        historyTodaySales.sales_same_day_last_year = dailySalesLastYear.sales
        historyTodaySales.sales_month_to_date = sales_month_to_date_and_monthly_increase_on_sales['sales_month_to_date']
        historyTodaySales.monthly_increase_on_sales = sales_month_to_date_and_monthly_increase_on_sales['monthly_increase_on_sales']

    if 'ad_cost' in kwargs.keys():
        ad_cost = kwargs['ad_cost']
        acos = kwargs['acos']
        historyTodaySales.ad_cost = ad_cost
        historyTodaySales.acos = acos
        historyTodaySales.ad_cost_on_sales = ad_cost / historyTodaySales.sales_today
    historyTodaySales.save()
@login_required
def update_today_sales(request):
    if request.method == 'POST':
        ad_cost = float(request.POST['ad_cost'])
        acos = float(request.POST['acos'])
        date = request.POST['date']
        year_month_day = date_split(date)
        country = 'US'
        save_today_sales(year_month_day=year_month_day, ad_cost= ad_cost,acos= acos,country=country)
    template = loader.get_template('updateTodaySales.html')
    class TodaySalesForm(forms.Form):
        date = forms.CharField(max_length=8)
        ad_cost = forms.FloatField()
        acos = forms.FloatField()
    context = {
        'form':TodaySalesForm()
    }

    return HttpResponse(template.render(context, request))

def findCountryByFcCode(code):
    if FulfillmentCenterCodeCountry.objects.filter(code = code).count():
        return FulfillmentCenterCodeCountry.objects.filter(code = code).first().country
    else:
        return None

def handle_uploaded_fba_shipment_file(file):
    encodings = ENCODINGS
    for e in encodings:
        try:
            file.seek(0,0)
            fileInMemory = file.read().decode(e)
            print(e)
            break
        except UnicodeDecodeError:
            pass
    csv_data = csv.reader(StringIO(fileInMemory), delimiter='\t')
    start_read = False
    shipment_id = False
    shipment_name = False
    shipped_sku_qties_list = []
    for i, row in enumerate(csv_data):
        if i == 0:
            if row[0] in ['Shipment ID', '货件编号']:
                shipment_id = row[1]
        if i == 1:
            if row[0] in ['Name', '名称']:
                shipment_name = row[1]
        if len(row) and row[0] in ['Ship To', '配送地址']:
            fc_code = row[1]
            country = findCountryByFcCode(fc_code)
            if country == None:
                return {'fc_code_unknown': True , \
                        'fc_code': fc_code}
        if row:
            if row[0] in ['Merchant SKU','卖家 SKU'] :
                for j,col in enumerate(row):
                    if col in ['Shipped', '已发货']:
                        shipped_sku_qty_col = j
                start_read = True
            elif start_read:
                shipped_sku_qties_list.append({'sku': trim_sku(row[0]), 'qty': int(row[shipped_sku_qty_col])})
    if len(shipped_sku_qties_list) and shipment_id and shipment_name:
        fba_shipment = FbaShipment.objects.filter(shipment_id = shipment_id)
        if fba_shipment.count():
            fba_shipment = fba_shipment.first()
            fba_shipment.shipment_name = shipment_name
            fba_shipment.save()
        else:
            fba_shipment = FbaShipment(shipment_id = shipment_id \
                                    , shipment_name = shipment_name \
                                    , fc_code = fc_code \
                                    , country = country)
            fba_shipment.save()
            for shipped_sku_qty_i in shipped_sku_qties_list:
                sku = shipped_sku_qty_i['sku']
                (product, found_or_not) = Product.objects.get_or_create(sku = sku)
                shipped_sku_qty = ShippedSkuQty(product = product, sku = sku, qty = shipped_sku_qty_i['qty'])
                shipped_sku_qty.save()
                fba_shipment.shipped_sku_qties.add(shipped_sku_qty)
        return  {'fc_code_unknown': False , 'fba_shipment_id': shipment_id}
# 还不需要输入运费等信息的版本
@login_required
def update_fba_shipment(request):
    success_updated_fba_shipment_id = ''
    if request.method == 'POST':
        form = UploadFilesForm(request.POST, request.FILES)
        if form.is_valid():
            files = request.FILES.getlist('file_field')
            for f in files:
                handle_uploaded_fba_shipment_file_result = handle_uploaded_fba_shipment_file(f)
                if handle_uploaded_fba_shipment_file_result['fc_code_unknown'] == True:
                    template = loader.get_template('update_fc_code_country.html')
                    context = {
                        'form':FcCodeCountryForm(initial={'fc_code': handle_uploaded_fba_shipment_file_result['fc_code']})
                    }
                    return HttpResponse(template.render(context, request))
                else:
                    success_updated_fba_shipment_id = handle_uploaded_fba_shipment_file_result['fba_shipment_id']
        else:
            form = FcCodeCountryForm(request.POST)
            if form.is_valid():
                fulfillment_center_code_country, created = FulfillmentCenterCodeCountry.objects.get_or_create(code = form.cleaned_data.get('fc_code'))
                fulfillment_center_code_country.country = form.cleaned_data.get('country')
                fulfillment_center_code_country.save()
                return redirect('update_fba_shipment')
    template = loader.get_template('update_fba_shipment.html')
    context = {
        'success_updated_fba_shipment_id': success_updated_fba_shipment_id ,\
        'form':UploadFilesForm()
    }
    return HttpResponse(template.render(context, request))

def get_sku_paid_unit_weight(sku, weight_volumn_factor):
    product = Product.objects.filter(sku = sku)
    if product.count():
        product = product.first()
        real_weight = product.package_weight / 1000.0
        volumn_weight = product.package_length * product.package_width * product.package_height / weight_volumn_factor
        if real_weight > volumn_weight:
            return real_weight
        return volumn_weight

def stat_sku_shipping_cost_in_a_shipment_id(shipment_id):
    sku_shipped_qty_dict = {}
    sku_shipping_cost_dict = {}
    sku_shipped_theoritical_weight_dict = {}
    fba_shipment_paid_bill = FbaShipmentPaidBill.objects.get(shipment_id = shipment_id)
    shipment_paid_amount = fba_shipment_paid_bill.paid_amount
    weight_volumn_factor = fba_shipment_paid_bill.weight_volumn_factor
    total_theoritical_weight = 0
    for shipped_sku_qty in FbaShipment.objects.get(shipment_id = shipment_id).shipped_sku_qties.all():
        sku = shipped_sku_qty.sku
        sku_paid_unit_weight = get_sku_paid_unit_weight(sku, weight_volumn_factor)
        sku_theoritical_total_weight = shipped_sku_qty.qty * sku_paid_unit_weight
        sku_shipped_theoritical_weight_dict[sku] = sku_theoritical_total_weight
        sku_shipped_qty_dict[sku] = shipped_sku_qty.qty
        total_theoritical_weight += sku_theoritical_total_weight
    for sku, sku_theoritical_total_weight in sku_shipped_theoritical_weight_dict.items():
        if total_theoritical_weight:
            sku_shipping_cost_dict[sku] = sku_theoritical_total_weight / total_theoritical_weight * shipment_paid_amount
        else:
            sku_shipping_cost_dict[sku] = 0.0
    return [sku_shipped_qty_dict, sku_shipping_cost_dict]

def stat_sku_unit_shipping_cost_in_a_shipment_id_list(shipment_ids):
    sku_shipped_qty_dict = {}
    sku_shipping_cost_dict = {}
    for shipment_id in shipment_ids:
        [sku_shipped_qty_in_a_shipment_dict, sku_shipping_cost_in_a_shipment_dict] = stat_sku_shipping_cost_in_a_shipment_id(shipment_id)
        for sku in sku_shipped_qty_in_a_shipment_dict.keys():
            if sku not in sku_shipped_qty_dict:
                sku_shipped_qty_dict[sku] = sku_shipped_qty_in_a_shipment_dict[sku]
                sku_shipping_cost_dict[sku] = sku_shipping_cost_in_a_shipment_dict[sku]
            else:
                sku_shipped_qty_dict[sku] += sku_shipped_qty_in_a_shipment_dict[sku]
                sku_shipping_cost_dict[sku] += sku_shipping_cost_in_a_shipment_dict[sku]
    sku_unit_shipping_cost_dict = {}
    for sku in sku_shipped_qty_dict.keys():
        sku_unit_shipping_cost_dict[sku] = sku_shipping_cost_dict[sku] / sku_shipped_qty_dict[sku]
    return sku_unit_shipping_cost_dict

@user_passes_test(lambda u: u.groups.filter(name__in = ['logistics', 'accountant']).exists())
def update_fba_shipment_paid_bills(request):
    template = loader.get_template('update_fba_shipment_paid_bills.html')
    error_tip = ''
    if request.method == 'POST':
        fileInMemory = request.FILES['file'].read()
        filePath = BytesIO(fileInMemory)
        wb2 = load_workbook(filePath, data_only=True)
        sheet = wb2.worksheets[0]
        max_column = sheet.max_column

        fba_shipment_id_titles = ['FBA ID（发货员填）']
        weight_volumn_factor_titles = ['体积因子（采购专员填）']
        weight_titles = ['双方确认后计费重kg（采购专员填）']
        fee_titles = ['费用（采购专员填）']
        extra_fee_titles = ['快递到物流仓库费用']
        reimbursement_titles = ['赔偿金额（采购专员填）']
        oversea_warehouse_inbound_fee_titles = ['海外仓出库费用（采购专员填）']
        inhouse_measured_weight_titles = ['发货员测量计费重kg（发货员填）']
        price_per_unit_weight_titles = ['单价（采购专员填）']
        extra_fee_at_agency_titles = ['附加费（采购专员填）']

        fba_shipment_id_col = 0
        weight_volumn_factor_col = 0
        weight_col = 0
        fee_col = 0
        extra_fee_col = 0
        reimbursement_col = 0
        oversea_warehouse_inbound_fee_col = 0
        inhouse_measured_weight_col = 0
        price_per_unit_weight_col = 0
        extra_fee_at_agency_col = 0
        none_existing_fba_shipment_ids = []
        STOP_CHECK_NONE_EXISTING_FBA_SHIPMENT = False
        STOP_CHECK_NONE_EXISTING_FBA_SHIPMENT_ID = 'FBA171FQ2ND3'
        for i_col in range(1, max_column + 1):
            if sheet.cell(1,i_col).value in fba_shipment_id_titles:
                fba_shipment_id_col = i_col
            elif sheet.cell(1,i_col).value in weight_volumn_factor_titles:
                weight_volumn_factor_col = i_col
            elif sheet.cell(1,i_col).value in weight_titles:
                weight_col = i_col
            elif sheet.cell(1,i_col).value in fee_titles:
                fee_col = i_col
            elif sheet.cell(1,i_col).value in extra_fee_titles:
                extra_fee_col = i_col
            elif sheet.cell(1,i_col).value in reimbursement_titles:
                reimbursement_col = i_col
            elif sheet.cell(1,i_col).value in oversea_warehouse_inbound_fee_titles:
                oversea_warehouse_inbound_fee_col = i_col
            elif sheet.cell(1,i_col).value in inhouse_measured_weight_titles:
                inhouse_measured_weight_col = i_col
            elif sheet.cell(1,i_col).value in price_per_unit_weight_titles:
                price_per_unit_weight_col = i_col
            elif sheet.cell(1,i_col).value in extra_fee_at_agency_titles:
                extra_fee_at_agency_col = i_col
        if fba_shipment_id_col * weight_volumn_factor_col * weight_col * fee_col \
           * extra_fee_col * reimbursement_col * oversea_warehouse_inbound_fee_col \
           * inhouse_measured_weight_col * price_per_unit_weight_col * extra_fee_at_agency_col == 0:
            context = {
                'form':UploadFileForm() ,\
                'error_tip': '上传文件标题有问题，请检查'
            }
        for i_row in range(2, sheet.max_row + 1):
            fba_shipment_id = sheet.cell(i_row, fba_shipment_id_col).value.strip()
            if fba_shipment_id:
                extra_fee = reimbursement = oversea_warehouse_inbound_fee = inhouse_measured_weight = price_per_unit_weight = extra_fee_at_agency = 0
                weight_volumn_factor = int(sheet.cell(i_row, weight_volumn_factor_col).value)
                weight = sheet.cell(i_row, weight_col).value
                fee = float(sheet.cell(i_row, fee_col).value)
                if sheet.cell(i_row, extra_fee_col).value:
                    extra_fee = float(sheet.cell(i_row, extra_fee_col).value)
                if sheet.cell(i_row, reimbursement_col).value:
                    reimbursement = float(sheet.cell(i_row, reimbursement_col).value)
                if sheet.cell(i_row, oversea_warehouse_inbound_fee_col).value:
                    if sheet.cell(i_row, oversea_warehouse_inbound_fee_col).value[:3] == 'US$':
                        oversea_warehouse_inbound_fee = float(sheet.cell(i_row, oversea_warehouse_inbound_fee_col).value[3:]) * 7
                if sheet.cell(i_row, inhouse_measured_weight_col).value:
                    inhouse_measured_weight = float(sheet.cell(i_row, inhouse_measured_weight_col).value)
                if sheet.cell(i_row, price_per_unit_weight_col).value:
                    price_per_unit_weight = float(sheet.cell(i_row, price_per_unit_weight_col).value)
                if sheet.cell(i_row, extra_fee_at_agency_col).value:
                    extra_fee_at_agency = float(sheet.cell(i_row, extra_fee_at_agency_col).value)

                fba_shipment_paid_bill, created = FbaShipmentPaidBill.objects.get_or_create( \
                                                        shipment_id = fba_shipment_id)
                if weight:
                    fba_shipment_paid_bill.weight = weight
                else:
                    fba_shipment_paid_bill.weight = inhouse_measured_weight
                if not fee:
                    fee = inhouse_measured_weight * price_per_unit_weight
                fee += extra_fee - reimbursement + oversea_warehouse_inbound_fee + extra_fee_at_agency
                fba_shipment_paid_bill.paid_amount = fee
                fba_shipment_paid_bill.weight_volumn_factor = weight_volumn_factor
                fba_shipment_paid_bill.save()
                if fba_shipment_id == STOP_CHECK_NONE_EXISTING_FBA_SHIPMENT_ID:
                    STOP_CHECK_NONE_EXISTING_FBA_SHIPMENT = True
                if not STOP_CHECK_NONE_EXISTING_FBA_SHIPMENT and not FbaShipment.objects.filter(shipment_id = fba_shipment_id).count():
                    none_existing_fba_shipment_ids.append(fba_shipment_id)
        shipment_ids_in_a_time_range = [fba_shipment.shipment_id for fba_shipment in FbaShipment.objects.filter(shipped_date__range = [datetime.datetime.now() - datetime.timedelta(days=30*SKU_UNIT_SHIPPING_DATE_MONTH_RANGE[0]), datetime.datetime.now() - datetime.timedelta(days=30*SKU_UNIT_SHIPPING_DATE_MONTH_RANGE[1])])]
        sku_list_lack_of_product_information = [k for k, v in stat_sku_unit_shipping_cost_in_a_shipment_id_list(shipment_ids_in_a_time_range).items() if v == 0.0]
        if sku_list_lack_of_product_information != []:
            error_tip = '以下sku缺少尺寸重量信息，请补充：' + ' '.join(sku_list_lack_of_product_information)
        if len(none_existing_fba_shipment_ids):
            error_tip += '另外，以下货件没有录入AOA，请补充：%s' %(', '.join(none_existing_fba_shipment_ids))
    if error_tip:
        context = {
            'error_tip':  error_tip,\
            'form':UploadFileForm()
        }
    else:
        context = {
            'form':UploadFileForm()
        }
    return HttpResponse(template.render(context, request))

def split_year_month(year_month):
    year_month = int(year_month)
    if year_month < 201801:
        return False
    return [int(year_month/100),int(year_month - 100 * int(year_month/100))]

def get_shipment_inventory_value(shipment_id):
    shipment_inventory_value = 0.0
    shipped_date = FbaShipment.objects.get(shipment_id = shipment_id).shipped_date
    for shipped_sku_qty in FbaShipment.objects.get(shipment_id = shipment_id).shipped_sku_qties.all():
        sku = shipped_sku_qty.sku
        qty = shipped_sku_qty.qty
        # inventory_value_date = (shipped_date.replace(day = 1) - datetime.timedelta(days = 1)).replace(day = 1)
        inventory_value_date = shipped_date.replace(day = 1)
        product_inventory_unit_value = ProductInventoryUnitValue.objects.filter(date = inventory_value_date, sku = sku)
        if product_inventory_unit_value.count():
            product_inventory_unit_value = product_inventory_unit_value.first()
        else:
            product_inventory_unit_value = ProductInventoryUnitValue.objects.filter(sku = sku).order_by('-date')
            if product_inventory_unit_value.count():
                product_inventory_unit_value = product_inventory_unit_value.first()
            else:
                return False
        product_inventory_total_value = product_inventory_unit_value.inventory_value_plus_additional_cost() * qty
        shipment_inventory_value += product_inventory_total_value
    return shipment_inventory_value

@user_passes_test(lambda u: u.groups.filter(name__in = ['logistics', 'sales', 'accountant']).exists())
def export_sku_inbound_shipping_cost(request):
    template = loader.get_template('export_sku_inbound_shipping_cost.html')
    error_tip = ''
    if request.method == 'POST':
        year_month = request.POST['year_month']
        [year, month] = split_year_month(year_month)
        if month == 12:
            reference_date = datetime.date(year= year + 1, month=1, day = 1)
            reference_date_for_inventory_value = datetime.date(year= year + 1, month=1, day = 4)
        else:
            reference_date = datetime.date(year= year, month=month+1, day = 1)
            reference_date_for_inventory_value = datetime.date(year= year, month=month+1, day = 4)
        shipment_ids_in_a_time_range = [fba_shipment.shipment_id for fba_shipment in FbaShipment.objects.filter(shipped_date__range = [reference_date - datetime.timedelta(days=30*SKU_UNIT_SHIPPING_DATE_MONTH_RANGE[0]), reference_date - datetime.timedelta(days=30*SKU_UNIT_SHIPPING_DATE_MONTH_RANGE[1])])]
        sku_list_lack_of_product_information = [k for k, v in stat_sku_unit_shipping_cost_in_a_shipment_id_list(shipment_ids_in_a_time_range).items() if v == 0.0]
        if sku_list_lack_of_product_information != []:
            error_tip = '以下sku缺少尺寸重量信息，请补充：' + ' '.join(sku_list_lack_of_product_information)
        else:
            countries = FbaShipment.objects.all().values_list('country',flat=True).distinct()
            wb = Workbook()
            for country in countries:
                #记录头程运费单价
                sheet = wb.create_sheet(country + '头程单价')
                shipment_ids_in_a_time_range = [fba_shipment.shipment_id for fba_shipment in FbaShipment.objects.filter(country = country, shipped_date__range = [reference_date - datetime.timedelta(days=30*SKU_UNIT_SHIPPING_DATE_MONTH_RANGE[0]), reference_date - datetime.timedelta(days=30*SKU_UNIT_SHIPPING_DATE_MONTH_RANGE[1])])]
                sku_unit_shipping_cost = stat_sku_unit_shipping_cost_in_a_shipment_id_list(shipment_ids_in_a_time_range)
                # sheet.title = country
                i_row = 1
                sheet.cell(1,1).value = 'SKU'
                sheet.cell(1,2).value = str(year_month) + '单位头程运费'
                for sku, unit_shipping_cost in sku_unit_shipping_cost.items():
                    i_row += 1
                    sheet.cell(i_row, 1).value = sku
                    sheet.cell(i_row, 2).value = unit_shipping_cost
                #记录头程运费货值
                sheet = wb.create_sheet(country + '货件货值')
                i_row = 1
                sheet.cell(1,1).value = 'Shipment ID'
                sheet.cell(1,2).value = 'Shipment Name'
                sheet.cell(1,3).value = '货件货值'
                shipment_ids_in_a_time_range = [fba_shipment.shipment_id for fba_shipment in FbaShipment.objects.filter(country = country, shipped_date__range = [reference_date_for_inventory_value - datetime.timedelta(days=30*1 + 8), reference_date_for_inventory_value]).order_by('-shipped_date')]
                for shipment_id in shipment_ids_in_a_time_range:
                    i_row += 1
                    sheet.cell(i_row, 1).value = shipment_id
                    sheet.cell(i_row, 2).value = FbaShipment.objects.get(shipment_id = shipment_id).shipment_name
                    sheet.cell(i_row, 3).value = get_shipment_inventory_value(shipment_id)
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename={year_month}-sku_unit_shipping_cost.xlsx'.format( \
                   year_month=year_month \
                   ,)
            sheet = wb['Sheet']
            sheet.title = '库存单位采购成本%s' %(year_month)
            sheet.cell(1,1).value = 'SKU'
            sheet.cell(1,2).value = '库存单位采购成本'
            i_row = 1
            for product_inventory_unit_value in ProductInventoryUnitValue.objects.filter(date = datetime.date(year= year, month=month, day = 1)).all():
                i_row += 1
                sheet.cell(i_row,1).value = product_inventory_unit_value.sku
                sheet.cell(i_row,2).value = product_inventory_unit_value.inventory_value_plus_additional_cost()
            wb.save(response)
            return response
    class YearMonthForm(forms.Form):
        year_month = forms.CharField(max_length=8)
    if error_tip:
        context = {
            'error_tip': error_tip ,\
            'form':YearMonthForm()
        }
    else:
        context = {
            'form':YearMonthForm()
        }
    return HttpResponse(template.render(context, request))

def find_logistic_cost_of_warehouse_name(sku, warehouse_name, districts_dict):
    ret = 0.0
    if warehouse_name == '北美仓':
        for country, sku_unit_cost_dict in districts_dict['beimei'].items():
            if sku in sku_unit_cost_dict:
                return sku_unit_cost_dict[sku]
    elif warehouse_name == '欧洲仓':
        for country, sku_unit_cost_dict in districts_dict['ouzhou'].items():
            if sku in sku_unit_cost_dict:
                return sku_unit_cost_dict[sku]
    else:
        for country, sku_unit_cost_dict in districts_dict['yuandong'].items():
            if sku in sku_unit_cost_dict:
                return sku_unit_cost_dict[sku]
    return ret

@user_passes_test(lambda u: u.groups.filter(name__in = ['logistics', 'sales', 'accountant']).exists())
def export_sku_inbound_shipping_cost_for_sellfox(request):
    MSKU_COLUMN = 4
    WAREHOUSE_COLUMN = 2
    PURCHASE_COST_COLUMN = 6
    LOGISTIC_COST_COLUMN = 8
    template = loader.get_template('export_sku_inbound_shipping_cost_for_sellfox.html')
    error_tip = ''
    if request.method == 'POST':
        year_month = request.POST['year_month']
        [year, month] = split_year_month(year_month)
        reference_date = datetime.date(year= year, month=month+1, day = 1)
        reference_date_for_inventory_value = datetime.date(year= year, month=month+1, day = 4)
        shipment_ids_in_a_time_range = [fba_shipment.shipment_id for fba_shipment in FbaShipment.objects.filter(shipped_date__range = [reference_date - datetime.timedelta(days=30*SKU_UNIT_SHIPPING_DATE_MONTH_RANGE[0]), reference_date - datetime.timedelta(days=30*SKU_UNIT_SHIPPING_DATE_MONTH_RANGE[1]) - datetime.timedelta(days=1)])]
        sku_list_lack_of_product_information = [k for k, v in stat_sku_unit_shipping_cost_in_a_shipment_id_list(shipment_ids_in_a_time_range).items() if v == 0.0]
        if sku_list_lack_of_product_information != []:
            error_tip = '以下sku缺少尺寸重量信息，请补充：' + ' '.join(sku_list_lack_of_product_information)
        else:
            countries = FbaShipment.objects.all().values_list('country',flat=True).distinct()
            districts_dict = {'ouzhou':{},'beimei':{},'yuandong':{}}
            sku_unit_shipping_cost = stat_sku_unit_shipping_cost_in_a_shipment_id_list(shipment_ids_in_a_time_range)
            for country in countries:
                shipment_ids_in_a_time_range = [fba_shipment.shipment_id for fba_shipment in FbaShipment.objects.filter(country = country, shipped_date__range = [reference_date - datetime.timedelta(days=30*SKU_UNIT_SHIPPING_DATE_MONTH_RANGE[0]), reference_date - datetime.timedelta(days=30*SKU_UNIT_SHIPPING_DATE_MONTH_RANGE[1]) - datetime.timedelta(days=1)])]
                sku_unit_shipping_cost = stat_sku_unit_shipping_cost_in_a_shipment_id_list(shipment_ids_in_a_time_range)
                if country in ['US', 'CA', 'MX']:
                    districts_dict['beimei'][country] = sku_unit_shipping_cost
                elif country in ['EU', 'GB']:
                    districts_dict['ouzhou'][country] = sku_unit_shipping_cost
                else:
                    districts_dict['yuandong'][country] = sku_unit_shipping_cost

            wb = Workbook()
            sheet_export = wb['Sheet']
            sheet_export.title = 'Sheet1'
            fileInMemory = request.FILES['file'].read()
            filePath = BytesIO(fileInMemory)
            wb2 = load_workbook(filePath, data_only=True)
            sheet = wb2.worksheets[0]
            max_column = sheet.max_column
            i_row = 1
            for i_column in range(1, max_column +1):
                sheet_export.cell(i_row, i_column).value = sheet.cell(i_row, i_column).value
            for i_row in range(2, sheet.max_row + 1):
                sku = sheet.cell(i_row, MSKU_COLUMN).value
                warehouse_name = sheet.cell(i_row, WAREHOUSE_COLUMN).value[-3:]
                for i_column in range(1, max_column +1):
                    sheet_export.cell(i_row, i_column).value = sheet.cell(i_row, i_column).value
                    if i_column in [PURCHASE_COST_COLUMN, LOGISTIC_COST_COLUMN]:
                        product_inventory_unit_value = ProductInventoryUnitValue.objects.filter(date = datetime.date(year= year, month=month, day = 1), sku = sku)
                        if product_inventory_unit_value.count():
                            sheet_export.cell(i_row, PURCHASE_COST_COLUMN).value = product_inventory_unit_value.first().inventory_value_plus_additional_cost()
                            sheet_export.cell(i_row, LOGISTIC_COST_COLUMN).value = find_logistic_cost_of_warehouse_name(sku, warehouse_name, districts_dict)
                        else:
                            sheet_export.cell(i_row, PURCHASE_COST_COLUMN).value = 0
                            sheet_export.cell(i_row, LOGISTIC_COST_COLUMN).value = 0
                    else:
                        sheet_export.cell(i_row, i_column).value = sheet.cell(i_row, i_column).value

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename={year_month}-sellfox-unit-cost.xlsx'.format( \
                   year_month=year_month \
                   ,)
            wb.save(response)
            return response
    class YearMonthForm(forms.Form):
        year_month = forms.CharField(max_length=8)
    if error_tip:
        context = {
            'error_tip': error_tip ,\
            'form':YearMonthForm() ,\
            'fileForm': UploadFileForm() \
        }
    else:
        context = {
            'form':YearMonthForm() ,\
            'fileForm': UploadFileForm() \
        }
    return HttpResponse(template.render(context, request))

# 需要输入运费等信息的版本
# @login_required
# def update_fba_shipment(request):
#     template = loader.get_template('update_fba_shipment.html')
#     if request.method == 'POST':
#         form = UploadShipmentFileForm(request.POST, request.FILES)
#         if form.is_valid():
#             file = request.FILES['file']
#             shipment_id_in_file = handle_uploaded_fba_shipment_file(file)
#             shipment_id = request.POST['shipment_id']
#             shipway = request.POST['shipway']
#             cost_per_kg = request.POST['cost_per_kg']
#             if shipment_id_in_file == shipment_id:
#                 fbaShipmentCost = FbaShipmentCost.objects.filter(shipment_id = shipment_id).first()
#                 if fbaShipmentCost:
#                     fbaShipmentCost.shipway = shipway
#                     fbaShipmentCost.cost_per_kg = cost_per_kg
#                 else:
#                     fbaShipmentCost = FbaShipmentCost(shipment_id = shipment_id, shipway = shipway ,\
#                                                         cost_per_kg = cost_per_kg, date = datetime.date.today())
#                 fbaShipmentCost.save()
#             else:
#                 return HttpResponse('上传的货件文件的Shipment ID和填写的shipment ID不符')
#         else:
#             context = {
#                 'form': form
#             }
#             return HttpResponse(template.render(context, request))
#     context = {
#         'form':UploadShipmentFileForm()
#     }
#     return HttpResponse(template.render(context, request))

@login_required
def update_sku_weight(request):
    if request.method == 'POST':
        fileInMemory = request.FILES['file'].read()
        filePath = BytesIO(fileInMemory)
        wb2 = load_workbook(filePath, data_only=True)
        sheet = wb2.worksheets[0]
        max_column = sheet.max_column

        # 检查title是否正确
        titles = sheet.cell(1, 1).value + sheet.cell(1, 2).value + sheet.cell(1, 3).value + sheet.cell(1, 4).value + sheet.cell(1, 5).value
        if titles != 'SKU产品中文实重否？实重22.5kg一箱的大小，装满该产品能装多少个？':
            return HttpResponse('请检查上传的模版')
        for row in range(2, sheet.max_row + 1):
            sku = sheet.cell(row, 1).value
            heavy_or_light = sheet.cell(row, 3).value
            real_weight = 0.0
            converted_weight = 0.0
            if heavy_or_light:
                heavy_or_light = True
                real_weight = sheet.cell(row, 4).value
            else:
                heavy_or_light = False
                converted_weight = sheet.cell(row, 5).value

            skuWeight = SkuWeight.objects.filter(sku = sku).first()
            if not skuWeight:
                skuWeight = SkuWeight(sku = sku, heavy_or_light = heavy_or_light, real_weight = real_weight, converted_weight = converted_weight)
            else:
                skuWeight.heavy_or_light = heavy_or_light
                skuWeight.real_weight = real_weight
                skuWeight.converted_weight = converted_weight
            skuWeight.save()
    template = loader.get_template('update_sku_weight.html')
    context = {
        'form':UploadFileForm()
    }
    return HttpResponse(template.render(context, request))
#
# @login_required
# def update_fba_shipment_received_sku_qty(request):
#     if request.method == 'POST':
#         ReceivedSkuQty.objects.all().delete()
#         encodings = ENCODINGS
#         for e in encodings:
#             try:
#                 request.FILES['file'].seek(0,0)
#                 fileInMemory = request.FILES['file'].read().decode(e)
#                 break
#             except UnicodeDecodeError:
#                 pass
#         fba_shipment_id_titles = ['fba-shipment-id', 'FBA Shipment ID', 'Reference ID']
#         sku_titles = ['sku', 'Merchant SKU','MSKU']
#         quantity_titles = ['quantity', 'Quantity']
#         # fileInMemory = request.FILES['file'].read().decode('windows-1252')
#         csv_data = csv.reader(StringIO(fileInMemory), delimiter=',')
#         for i, row in enumerate(csv_data):
#             if i == 0:
#                 for col_num, title in enumerate(row):
#                     if title in fba_shipment_id_titles:
#                         fba_shipment_id_col = col_num
#                     elif title in sku_titles:
#                         sku_col = col_num
#                     elif title in quantity_titles:
#                         quantity_col = col_num
#             else:
#                 received_sku_qty, created = ReceivedSkuQty.objects.get_or_create(shipment_id = row[fba_shipment_id_col] \
#                                                                                         ,sku = row[sku_col])
#                 received_sku_qty.qty += int(row[quantity_col])
#                 received_sku_qty.save()
#     template = loader.get_template('update_fba_shipment_received_sku_qty.html')
#     context = {
#         'form':UploadFileForm()
#     }
#     return HttpResponse(template.render(context, request))

@login_required
def update_fba_shipment_received_sku_qty(request):
    template = loader.get_template('update_fba_shipment_received_sku_qty.html')
    if request.method == 'POST':
        ReceivedSkuQty.objects.all().delete()
        encodings = ENCODINGS
        for e in encodings:
            try:
                if request.FILES['file'].name[-4:] != '.txt':
                    context = {
                        'form':UploadFileForm(), \
                        'error_tip': '请上传.txt格式的文件'
                    }
                    return HttpResponse(template.render(context, request))
                request.FILES['file'].seek(0,0)
                fileInMemory = request.FILES['file'].read().decode(e)
                break
            except UnicodeDecodeError:
                pass
        fba_shipment_id_titles = ['fba-shipment-id', 'FBA Shipment ID', 'Reference ID']
        sku_titles = ['sku', 'Merchant SKU','MSKU']
        quantity_titles = ['quantity', 'Quantity']
        # fileInMemory = request.FILES['file'].read().decode('windows-1252')
        csv_data = csv.reader(StringIO(fileInMemory), delimiter='\t')
        received_sku_qty_dict = {}
        for i, row in enumerate(csv_data):
            if i == 0:
                for col_num, title in enumerate(row):
                    if title in fba_shipment_id_titles:
                        fba_shipment_id_col = col_num
                    elif title in sku_titles:
                        sku_col = col_num
                    elif title in quantity_titles:
                        quantity_col = col_num
            else:
                sku = row[sku_col]
                fba_shipment_id = row[fba_shipment_id_col]
                qty = int(row[quantity_col])
                sku_fba_shipment_id = sku + fba_shipment_id
                if sku_fba_shipment_id in received_sku_qty_dict:
                    received_sku_qty_dict[sku_fba_shipment_id] += qty
                else:
                    received_sku_qty_dict[sku_fba_shipment_id] = qty
        for sku_fba_shipment_id, qty in received_sku_qty_dict.items():
            sku = sku_fba_shipment_id[:-12]
            fba_shipment_id = sku_fba_shipment_id[-12:]

            received_sku_qty = ReceivedSkuQty(shipment_id = fba_shipment_id, sku = sku,qty=qty)
            received_sku_qty.save()
    context = {
        'form':UploadFileForm()
    }
    return HttpResponse(template.render(context, request))

@login_required
def update_fba_shipment_estimated_receiving_date(request):
    if request.method == 'POST':
        import json
        post_json = json.loads(request.body)
        date_type = post_json['date_type']
        date_str = post_json['date']
        date_obj = datetime.datetime.strptime(date_str, '%Y/%m/%d').date()
        fba_shipment_id = post_json['fba_shipment_id']
        fba_shipment = FbaShipment.objects.filter(shipment_id = fba_shipment_id).first()
        if fba_shipment:
            if date_type == 'ship_date':
                fba_shipment.shipped_date = date_obj
            elif date_type == 'estimated_receiving_date':
                fba_shipment.estimated_receiving_date = date_obj
            fba_shipment.save()
        return HttpResponse(request.POST.get('a'))

    template = loader.get_template('update_fba_shipment_estimated_receiving_date.html')
    fba_shipments = FbaShipment.objects.filter(closed = False).all()

    context = {
        'fba_shipments': fba_shipments
    }
    return HttpResponse(template.render(context, request))

@login_required
def estimated_sku_qty_receiving_date(request):
    template = loader.get_template('estimated_sku_qty_receiving_date.html')
    if 'country' in request.GET:
        country = request.GET['country']
    else:
        country = 'US'
    countries = ['US','EU','GB','CA','JP', 'AE', 'SA','AU']
    fba_shipments = FbaShipment.objects.filter(closed = False, country = country).order_by('estimated_receiving_date')
    shipped_sku_qty_by_shipment = {}
    class ShippedSkuQtyByShipment():
        def __init__(self, qty, date, shipment_name, shipment_id):
            self.qty = qty
            self.date = date
            self.shipment_name = shipment_name
            self.shipment_id = shipment_id
    shipment_closed_dict = {}
    for fba_shipment in fba_shipments:
        shipment_id = fba_shipment.shipment_id
        shipment_name = fba_shipment.shipment_name
        shipment_closed_dict[shipment_id] = True
        if fba_shipment.estimated_receiving_date:
            estimated_receiving_date = fba_shipment.estimated_receiving_date.strftime('%m/%d')
        else:
            estimated_receiving_date = "无"
        shipped_sku_qties = fba_shipment.shipped_sku_qties.all()
        for shipped_sku_qty in shipped_sku_qties:
            sku = shipped_sku_qty.sku
            shipped_qty = shipped_sku_qty.qty
            received_sku_qty = ReceivedSkuQty.objects.filter(shipment_id = shipment_id \
                                                        ,sku = sku).aggregate(Sum('qty'))['qty__sum']

            if received_sku_qty == None:
                received_sku_qty = 0

            unreceived_qty = shipped_qty - received_sku_qty
            if unreceived_qty > 15 or (country != 'US' and unreceived_qty > 4):
                if sku in shipped_sku_qty_by_shipment.keys():
                    shipped_sku_qty_by_shipment[sku].append(ShippedSkuQtyByShipment(shipment_name = shipment_name \
                                                                                    , shipment_id = shipment_id \
                                                                                    ,qty =  unreceived_qty \
                                                                                    , date = estimated_receiving_date))
                else:
                    shipped_sku_qty_by_shipment[sku] = [ShippedSkuQtyByShipment(shipment_name = shipment_name \
                                                                                , shipment_id = shipment_id \
                                                                                , qty = unreceived_qty \
                                                                                , date = estimated_receiving_date)]
                shipment_closed_dict[shipment_id] = False

        for k,v in shipment_closed_dict.items():
            if v:
                fba_shipment = FbaShipment.objects.get(shipment_id = k)
                fba_shipment.closed = True
                fba_shipment.save()
    context = {
        'country' : country , \
        'countries' : countries ,\
        'to_be_received_skus':shipped_sku_qty_by_shipment
    }
    return HttpResponse(template.render(context, request))

@login_required
def get_estimated_sku_qty_receiving_date_of_a_sku(request):
    sku = request.GET['sku']
    if 'country' in request.GET:
        country = request.GET['country']
        if country == "":
            country = 'US'
    else:
        country = 'US'
    fba_shipments = FbaShipment.objects.filter(closed = False, shipped_sku_qties__sku=sku, country = country).order_by('estimated_receiving_date')
    receivable_qty_list = []
    for fba_shipment in fba_shipments:
        shipment_id = fba_shipment.shipment_id
        shipment_name = fba_shipment.shipment_name
        if fba_shipment.estimated_receiving_date:
            estimated_receiving_date = fba_shipment.estimated_receiving_date.strftime('%m/%d')
        else:
            estimated_receiving_date = "无"
        shipped_sku_qties = fba_shipment.shipped_sku_qties.all()
        for shipped_sku_qty in shipped_sku_qties:
            if shipped_sku_qty.sku == sku:
                shipped_qty = shipped_sku_qty.qty
                received_sku_qty = ReceivedSkuQty.objects.filter(shipment_id = shipment_id \
                                                            ,sku = sku).aggregate(Sum('qty'))['qty__sum']

                if received_sku_qty == None:
                    received_sku_qty = 0

                unreceived_qty = shipped_qty - received_sku_qty
                if unreceived_qty > 15 or (country != 'US' and unreceived_qty > 4):
                    receivable_qty_list.append({'shipment_id': shipment_id \
                                                                ,'date': estimated_receiving_date \
                                                                , 'qty': unreceived_qty \
                                                                , 'shipment_name':shipment_name})
    return JsonResponse(receivable_qty_list, safe=False)
@login_required
def get_history_sales_of_a_sku(request):
    sku = request.GET['sku']
    country = request.GET['country']
    history_today_product_sales = HistoryTodayProductSales.objects.filter(product__sku = sku, country = country).order_by('-date')[:7]
    history_today_product_sales_list = [{'date':i.date.strftime('%m/%d') \
                                        ,'qty':round(i.sold_qty_average_7d,1)} for i in history_today_product_sales]
    history_today_product_sales_list.reverse()
    return JsonResponse(history_today_product_sales_list, safe=False)


def check_receivable_purchased_qty(sku):
    receivablePurchasedQty = ReceivablePurchasedQty.objects.filter(sku = sku)
    if receivablePurchasedQty.count():
        receivablePurchasedQty = receivablePurchasedQty.all()[0].qty
    else:
        receivablePurchasedQty = 0
    return receivablePurchasedQty

def check_shenzhen_inventory(sku):
    shenzhen_inventory = Inventory.objects.filter(sku = sku, warehouse_name = shenzhen_warehouse_name)
    if shenzhen_inventory.count():
        shenzhen_inventory = shenzhen_inventory.first().qty
    else:
        shenzhen_inventory = 0
    return shenzhen_inventory

def if_count_for_po(todayProductSale):
    maximum_days_lasting = 200 # 100才是正确的，200是为了让LS-VIE-33-13-A-1出现
    sku = todayProductSale.product.sku
    amazon_inventory = todayProductSale.fba_inventory.total_unit
    shenzhen_inventory = check_shenzhen_inventory(sku)
    lasting_day_of_total_fba_and_shenzhen_inventory_estimated_by_us = (todayProductSale.fba_inventory.total_unit + shenzhen_inventory)/ todayProductSale.sold_qty_average_7d
    if todayProductSale.product.discontinued:
        return False
    if lasting_day_of_total_fba_and_shenzhen_inventory_estimated_by_us < maximum_days_lasting:
        return True
    if todayProductSale.fba_inventory.days_of_supply < maximum_days_lasting:
        return True
    return False

def get_sku_purchase_price(sku):
    price = 0.0
    skuPurchaseOrders = SkuPurchaseOrder.objects.filter(sku = sku).all()
    if skuPurchaseOrders.count():
        sum_of_transaction_amount = skuPurchaseOrders.aggregate(Sum('transaction_amount'))['transaction_amount__sum']
        sum_of_qty = float(skuPurchaseOrders.aggregate(Sum('qty'))['qty__sum'])
        price = sum_of_transaction_amount / sum_of_qty
        return price
    else:
        skuPurchasingPrice = SkuPurchasingPrice.objects.filter(sku = sku).order_by('-date').first()
        if skuPurchasingPrice:
            return skuPurchasingPrice.purchasing_price

#只考虑美国需求的版本
# @login_required
# def restock_today(request):
#     if request.method == 'POST':
#         abroad_warehouse_sku_qty = {}
#         abroad_warehouse_uploaded = False
#         form = UploadFileForm(request.POST,request.FILES)
#         if form.is_valid():
#             fileInMemory = request.FILES['file'].read()
#             filePath = BytesIO(fileInMemory)
#             wb2 = load_workbook(filePath, data_only=True)
#             sheet = wb2.worksheets[0]
#             for row in range(1, sheet.max_row + 1):
#                 sku = sheet.cell(row,1).value
#                 qty = int(sheet.cell(row,2).value)
#                 abroad_warehouse_sku_qty[sku]= qty
#             abroad_warehouse_uploaded = True
#
#         wb = Workbook()
#         sheet = wb.active
#         sheet.title = '今日生产单'
#         country = 'US'
#         row = 1
#         max_months_to_last_for_pos = [4, 5, 6]
#         moq = 200
#         min_days_lasting_for_notice = 60
#         max_days_to_last_for_pos = [max_months_to_last_for_po * 30 for max_months_to_last_for_po in max_months_to_last_for_pos]
#         if abroad_warehouse_uploaded:
#             titles = ['SKU', 'FBA 可售库存', '中转库存', '在途库存', '亚马逊总库存', 'F199库存', '海外仓', '今日销售', '过去7天平均日销售' \
#                     , '7日暴增比' \
#                     , '亚马逊预测能撑多少天', '我们预测能撑多少天', '单价', '亚马逊建议补货量', '已经下过购货订单' \
#                     ]
#         else:
#             titles = ['SKU', 'FBA 可售库存', '中转库存', '在途库存', '亚马逊总库存', 'F199库存', '今日销售', '过去7天平均日销售' \
#                     , '7日暴增比' \
#                     , '亚马逊预测能撑多少天', '我们预测能撑多少天', '单价', '亚马逊建议补货量', '已经下过购货订单' \
#                     ]
#         start_col_restock = len(titles) + 1
#         for max_months_to_last_for_po in max_months_to_last_for_pos:
#             titles.append('撑到%i个月生产单建议量' %max_months_to_last_for_po)
#         if request.user.is_superuser:
#             is_head_of_sales = True
#             #先不加复杂的财务部分数据
#             #这个是复杂版本titles.extend(['运营', '运营助理','净资产','单个采购成本','单个的海派头程运费','单个的空派头程运费','订购量', '货款+头程运费'])
#         else:
#             is_head_of_sales = False
#             #先不加复杂的财务部分数据
#             #这个是复杂版本titles.extend(['净资产','单个采购成本','单个的海派头程运费','单个的空派头程运费','订购量', '货款+头程运费'])
#         for i, title in enumerate(titles):
#             sheet.cell(row,i + 1).value = title
#
#         yellow_fill = PatternFill("solid", fgColor="FFFF00")
#         todayProductSales = TodayProductSales.objects.filter(country = country).all()
#
#         first_history_today_product_sales = HistoryTodayProductSales.objects.filter(country = country).order_by('-date').first()
#         seven_days_ago_date = first_history_today_product_sales.date - datetime.timedelta(days = 7)
#         processed_sku_list = []
#         row +=1
#
#         history_today_product_sales_filtered = HistoryTodayProductSales.objects.filter(date__gt=seven_days_ago_date,country = country).order_by('-date')
#         for todayProductSale in history_today_product_sales_filtered:
#             sku = todayProductSale.product.sku
#             # 让只显示运营负责的sku的，但是目前还不弄这种运营分sku的模式，所以先全部显示给运营
#             # if is_head_of_sales == False:
#             #     skuManagedBySalesPerson = SkuManagedBySalesPerson.objects.filter(sku = sku, sales_person_name = request.user.username).first()
#             #     if not skuManagedBySalesPerson:
#             #         skuManagedBySalesPerson = SkuManagedBySalesPerson.objects.filter(sku = sku, sales_assistant_name = request.user.username).first()
#             #     if not skuManagedBySalesPerson:
#             #         continue
#             # else:
#             #     skuManagedBySalesPerson = SkuManagedBySalesPerson.objects.filter(sku = sku).first()
#             if sku not in processed_sku_list and  if_count_for_po(todayProductSale):
#                 processed_sku_list.append(sku)
#                 sheet.cell(row,1).value = sku
#                 sheet.cell(row,2).value = todayProductSale.fba_inventory.available
#                 sheet.cell(row,3).value = todayProductSale.fba_inventory.fc_unit
#                 sheet.cell(row,4).value = todayProductSale.fba_inventory.inbound_unit
#                 fba_total_unit = todayProductSale.fba_inventory.total_unit
#                 sheet.cell(row,5).value = fba_total_unit
#                 shenzhen_inventory = check_shenzhen_inventory(sku)
#                 sheet.cell(row,6).value = shenzhen_inventory
#                 abroad_qty = 0
#                 if abroad_warehouse_uploaded :
#                     sheet.cell(row,7).value = 0
#                     if sku in abroad_warehouse_sku_qty:
#                         abroad_qty = abroad_warehouse_sku_qty[sku]
#                         sheet.cell(row,7).value = abroad_warehouse_sku_qty[sku]
#                 sheet.cell(row,7 + abroad_warehouse_uploaded).value = int(todayProductSale.sold_qty)
#                 sheet.cell(row,8 + abroad_warehouse_uploaded).value = round(todayProductSale.sold_qty_average_7d,1)
#                 increase_rate = 1.0
#                 seven_days_ago_history_today_product_sales = HistoryTodayProductSales.objects.filter(date = seven_days_ago_date, product__sku=sku, country=country).first()
#                 if seven_days_ago_history_today_product_sales != None:
#                     seven_days_ago_sold_qty_average_7d = seven_days_ago_history_today_product_sales.sold_qty_average_7d
#                     if seven_days_ago_sold_qty_average_7d != 0:
#                         increase_rate = todayProductSale.sold_qty_average_7d / seven_days_ago_sold_qty_average_7d
#                 sheet.cell(row,9 + abroad_warehouse_uploaded).value = round(increase_rate,1)# 暴增
#                 sheet.cell(row,10 + abroad_warehouse_uploaded).value = todayProductSale.fba_inventory.days_of_supply
#                 lasting_day_of_total_fba_and_shenzhen_inventory_estimated_by_us = (todayProductSale.fba_inventory.total_unit + shenzhen_inventory + abroad_qty)/ todayProductSale.sold_qty_average_7d
#                 sheet.cell(row,11 + abroad_warehouse_uploaded).value = int(lasting_day_of_total_fba_and_shenzhen_inventory_estimated_by_us)
#                 sheet.cell(row,12 + abroad_warehouse_uploaded).value = round(todayProductSale.average_price_7d,2)
#                 sheet.cell(row,12 + abroad_warehouse_uploaded).number_format = u'"$ "#,##0.00'
#                 amazon_recommended_replenishment_qty = todayProductSale.fba_inventory.recommended_replenishment_qty
#                 sheet.cell(row,13 + abroad_warehouse_uploaded).value = amazon_recommended_replenishment_qty
#                 receivable_purchased_qty = check_receivable_purchased_qty(sku)
#                 sheet.cell(row,14 + abroad_warehouse_uploaded).value = receivable_purchased_qty
#                 if todayProductSale.product.discontinued:
#                     continue #sheet.cell(row,13).value = "不再订货"
#                 else:
#                     for i, max_days_to_last_for_po in enumerate(max_days_to_last_for_pos):
#                         po_qty = 0
#                         if todayProductSale.fba_inventory.days_of_supply >= lasting_day_of_total_fba_and_shenzhen_inventory_estimated_by_us:
#
#                             days_vacancy = max_days_to_last_for_po - lasting_day_of_total_fba_and_shenzhen_inventory_estimated_by_us
#                             qty_needed = int(days_vacancy * todayProductSale.sold_qty_average_7d)
#                             if qty_needed > receivable_purchased_qty:
#                                 qty_needed = qty_needed - receivable_purchased_qty
#                                 if qty_needed > float(moq) / 2:
#                                     po_qty = max([qty_needed, moq])
#                         else:
#                             shenzhen_inventory_and_receivable_purchased_qty = shenzhen_inventory + receivable_purchased_qty
#                             if abroad_warehouse_uploaded and sku in abroad_warehouse_sku_qty:
#                                 shenzhen_inventory_and_receivable_purchased_qty += abroad_warehouse_sku_qty[sku]
#                             if i > 0:
#                                 if todayProductSale.fba_inventory.days_of_supply:
#                                     amazon_recommended_replenishment_qty_i = amazon_recommended_replenishment_qty + float(max_days_to_last_for_pos[i] - max_days_to_last_for_pos[0])/ todayProductSale.fba_inventory.days_of_supply * fba_total_unit
#                                 else:
#                                     amazon_recommended_replenishment_qty_i = amazon_recommended_replenishment_qty + fba_total_unit * i
#                             else:
#                                 amazon_recommended_replenishment_qty_i = amazon_recommended_replenishment_qty
#                             if amazon_recommended_replenishment_qty_i > shenzhen_inventory_and_receivable_purchased_qty:
#                                 qty_needed = amazon_recommended_replenishment_qty_i - shenzhen_inventory_and_receivable_purchased_qty
#                                 if qty_needed > float(moq) / 2:
#                                     po_qty = max([qty_needed, moq])
#                         if po_qty:
#                             sheet.cell(row,start_col_restock + i).value = int(po_qty)
#                     #先不加复杂的财务部分数据
#                     # 增加财务数据
#                     # if is_head_of_sales:
#                     #     sheet.cell(row, start_col_restock + len(max_days_to_last_for_pos) ).value = skuManagedBySalesPerson.sales_person_name
#                     #     sheet.cell(row, start_col_restock + len(max_days_to_last_for_pos) + 1).value = skuManagedBySalesPerson.sales_assistant_name
#                     #     skuAssetLiabilityTables = SkuAssetLiabilityTable.objects.filter(sku = sku, for_sales = True).order_by('-date').all()
#                     #     if skuAssetLiabilityTables.count() > 1:
#                     #         if skuAssetLiabilityTables[0].date == skuAssetLiabilityTables[1].date:
#                     #             skuAssetLiabilityTableForSalesPerson = SkuAssetLiabilityTable.objects.filter(sku = sku, for_sales = True, date = skuAssetLiabilityTables[0].date, initial = True).order_by('-date').first()
#                     #         else:
#                     #             skuAssetLiabilityTableForSalesPerson = SkuAssetLiabilityTable.objects.filter(sku = sku, for_sales = True).order_by('-date').first()
#                     #     elif skuAssetLiabilityTables.count() == 1:
#                     #         skuAssetLiabilityTableForSalesPerson = SkuAssetLiabilityTable.objects.filter(sku = sku, for_sales = True).order_by('-date').first()
#                     #
#                     #     sheet.cell(row, start_col_restock + len(max_days_to_last_for_pos) + 2).value = skuAssetLiabilityTableForSalesPerson.net_asset_amount
#                     #
#                     #     sheet.cell(row, start_col_restock + len(max_days_to_last_for_pos) + 3).value = get_sku_purchase_price(sku)
#                     #     skuHeadShippingUnitCostSea = SkuHeadShippingUnitCost.objects.filter(sku = sku, type = ('S', 'Sea')).first()
#                     #     if skuHeadShippingUnitCostSea:
#                     #         sheet.cell(row, start_col_restock + len(max_days_to_last_for_pos) + 4).value = skuHeadShippingUnitCostSea.head_shipping_unit_cost
#                     #     skuHeadShippingUnitCostAir = SkuHeadShippingUnitCost.objects.filter(sku = sku, type = ('A', 'Air')).first()
#                     #     if skuHeadShippingUnitCostAir:
#                     #         sheet.cell(row, start_col_restock + len(max_days_to_last_for_pos) + 5).value = skuHeadShippingUnitCostAir.head_shipping_unit_cost
#                     if min([todayProductSale.fba_inventory.days_of_supply, lasting_day_of_total_fba_and_shenzhen_inventory_estimated_by_us]) < min_days_lasting_for_notice and not(todayProductSale.product.discontinued):
#                         for cell in sheet["%i:%i" %(row,row)]:
#                             cell.fill = yellow_fill
#
#                 row +=1
#         response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#         response['Content-Disposition'] = 'attachment; filename={date}-PO.xlsx'.format( \
#                date=datetime.datetime.now().strftime('%Y-%m-%d') \
#                ,)
#         wb.save(response)
#         return response
#     template = loader.get_template('restock_today.html')
#     context = {
#         'form':UploadFileForm()
#     }
#     return HttpResponse(template.render(context, request))

def if_count_for_po_version2(sku, x):
    maximum_days_lasting = 200
    amazon_inventory = x['fba_total_unit']
    sold_qty_average_7d = x['sold_qty_average_7d']
    if sold_qty_average_7d <=  0:
        return False
    shenzhen_inventory = check_shenzhen_inventory(sku)
    lasting_day_of_total_fba_and_shenzhen_inventory_estimated_by_us = (amazon_inventory + shenzhen_inventory)/ sold_qty_average_7d
    if lasting_day_of_total_fba_and_shenzhen_inventory_estimated_by_us < maximum_days_lasting:
        return True
    return False

#考虑所有国家需求的版本
@login_required
def restock_today(request):
    if request.method == 'POST':
        abroad_warehouse_uploaded = False
        wb = Workbook()
        sheet = wb.active
        sheet.title = '今日生产单'
        country = 'US'
        countries_counted = ['US', 'CA', 'EU', 'GB', 'JP']
        row = 1
        max_months_to_last_for_pos = [4, 5, 6]
        moq = 200
        min_days_lasting_for_notice = 60
        max_days_to_last_for_pos = [max_months_to_last_for_po * 30 for max_months_to_last_for_po in max_months_to_last_for_pos]
        titles = ['SKU', 'FBA 可售库存', '中转库存', '在途库存', '亚马逊总库存', 'F199库存', '今日销售', '过去7天平均日销售' \
                , '我们预测能撑多少天', '已经下过购货订单' \
                ]

        start_col_restock = len(titles) + 1
        for max_months_to_last_for_po in max_months_to_last_for_pos:
            titles.append('撑到%i个月生产单建议量' %max_months_to_last_for_po)
        if request.user.is_superuser:
            is_head_of_sales = True
            #先不加复杂的财务部分数据
            #这个是复杂版本titles.extend(['运营', '运营助理','净资产','单个采购成本','单个的海派头程运费','单个的空派头程运费','订购量', '货款+头程运费'])
        else:
            is_head_of_sales = False
            #先不加复杂的财务部分数据
            #这个是复杂版本titles.extend(['净资产','单个采购成本','单个的海派头程运费','单个的空派头程运费','订购量', '货款+头程运费'])
        for i, title in enumerate(titles):
            sheet.cell(row,i + 1).value = title

        yellow_fill = PatternFill("solid", fgColor="FFFF00")
        todayProductSales = TodayProductSales.objects.all()

        first_history_today_product_sales = HistoryTodayProductSales.objects.filter(country = 'US').order_by('-date').first()
        seven_days_ago_date = first_history_today_product_sales.date - datetime.timedelta(days = 7)
        processed_sku_list_all_countries = {}
        sku_sales_dict_all_countries = {}
        row +=1
        history_today_product_sales_filtered = HistoryTodayProductSales.objects.filter(date__gt=seven_days_ago_date, country__in=countries_counted).order_by('-date')
        for todayProductSale in history_today_product_sales_filtered:
            if todayProductSale.product.discontinued:
                continue
            sku = todayProductSale.product.sku
            country = todayProductSale.country
            if country not in processed_sku_list_all_countries:
                processed_sku_list_all_countries[country] = []
            if country not in sku_sales_dict_all_countries:
                sku_sales_dict_all_countries[country] = {}
            if sku not in processed_sku_list_all_countries[country]:
                processed_sku_list_all_countries[country].append(sku)
                if sku not in sku_sales_dict_all_countries[country]:
                    # 启动了远程配送的情况，亚马逊那边的库存在该国家站点应该设置为0，因为都是美国的库存在支撑
                    if country == 'CA' and RemoteFulfillmentSku.objects.filter(sku = sku, country=country).count() > 0:
                        sku_sales_dict_all_countries[country][sku] = {'fba_inventory_available':0 ,\
                                            'fba_inventory_fc_unit':0 ,\
                                            'fba_inventory_inbound_unit': 0 ,\
                                            'fba_total_unit' : 0 ,\
                                            'sold_qty_today' : int(todayProductSale.sold_qty), \
                                            'sold_qty_average_7d': round(todayProductSale.sold_qty_average_7d,1)}
                    else:
                        sku_sales_dict_all_countries[country][sku] = {'fba_inventory_available':todayProductSale.fba_inventory.available ,\
                                            'fba_inventory_fc_unit':todayProductSale.fba_inventory.fc_unit ,\
                                            'fba_inventory_inbound_unit': todayProductSale.fba_inventory.inbound_unit ,\
                                            'fba_total_unit' : todayProductSale.fba_inventory.total_unit ,\
                                            'sold_qty_today' : int(todayProductSale.sold_qty), \
                                            'sold_qty_average_7d': round(todayProductSale.sold_qty_average_7d,1)}

        sku_sales_dict_all_countries_combined = {}
        for sku_sales_dict in sku_sales_dict_all_countries.values():
            for k,v in sku_sales_dict.items():
                if k not in sku_sales_dict_all_countries_combined:
                    sku_sales_dict_all_countries_combined[k] = { \
                        'fba_inventory_available': 0 ,\
                        'fba_inventory_fc_unit': 0 ,\
                        'fba_inventory_inbound_unit': 0,\
                        'fba_total_unit' : 0 ,\
                        'sold_qty_today' : 0, \
                        'sold_qty_average_7d': 0 \
                    }
                sku_sales_dict_all_countries_combined[k]['fba_inventory_available'] += v['fba_inventory_available']
                sku_sales_dict_all_countries_combined[k]['fba_inventory_fc_unit'] += v['fba_inventory_fc_unit']
                sku_sales_dict_all_countries_combined[k]['fba_inventory_inbound_unit'] += v['fba_inventory_inbound_unit']
                sku_sales_dict_all_countries_combined[k]['fba_total_unit'] += v['fba_total_unit']
                sku_sales_dict_all_countries_combined[k]['sold_qty_today'] += v['sold_qty_today']
                sku_sales_dict_all_countries_combined[k]['sold_qty_average_7d'] += v['sold_qty_average_7d']

        for sku, sku_sales_dict in sku_sales_dict_all_countries_combined.items():
            if if_count_for_po_version2(sku, sku_sales_dict):
                sheet.cell(row,1).value = sku
                sheet.cell(row,2).value = sku_sales_dict['fba_inventory_available']
                sheet.cell(row,3).value = sku_sales_dict['fba_inventory_fc_unit']
                sheet.cell(row,4).value = sku_sales_dict['fba_inventory_inbound_unit']
                sheet.cell(row,5).value = sku_sales_dict['fba_total_unit']
                shenzhen_inventory = check_shenzhen_inventory(sku)
                sheet.cell(row,6).value = shenzhen_inventory
                abroad_warehouse_uploaded = 0
                sheet.cell(row,7 + abroad_warehouse_uploaded).value = sku_sales_dict['sold_qty_today']
                sheet.cell(row,8 + abroad_warehouse_uploaded).value = sku_sales_dict['sold_qty_average_7d']
                lasting_day_of_total_fba_and_shenzhen_inventory_estimated_by_us = (sku_sales_dict['fba_total_unit'] + shenzhen_inventory)/ sku_sales_dict['sold_qty_average_7d']
                sheet.cell(row,9 + abroad_warehouse_uploaded).value = int(lasting_day_of_total_fba_and_shenzhen_inventory_estimated_by_us)
                receivable_purchased_qty = check_receivable_purchased_qty(sku)
                sheet.cell(row,10 + abroad_warehouse_uploaded).value = receivable_purchased_qty
                for i, max_days_to_last_for_po in enumerate(max_days_to_last_for_pos):
                    po_qty = 0
                    days_vacancy = max_days_to_last_for_po - lasting_day_of_total_fba_and_shenzhen_inventory_estimated_by_us
                    qty_needed = int(days_vacancy * sku_sales_dict['sold_qty_average_7d'])
                    if qty_needed > receivable_purchased_qty:
                        qty_needed = qty_needed - receivable_purchased_qty
                        if qty_needed > float(moq) / 2:
                            po_qty = max([qty_needed, moq])
                    if po_qty:
                        sheet.cell(row,start_col_restock + i).value = int(po_qty)

                if lasting_day_of_total_fba_and_shenzhen_inventory_estimated_by_us < min_days_lasting_for_notice:
                    for cell in sheet["%i:%i" %(row,row)]:
                        cell.fill = yellow_fill
                row +=1
        #
        # for todayProductSale in history_today_product_sales_filtered:
        #     sku = todayProductSale.product.sku
        #     if sku not in processed_sku_list and  if_count_for_po(todayProductSale):
        #         processed_sku_list.append(sku)
        #         sheet.cell(row,1).value = sku
        #         sheet.cell(row,2).value = todayProductSale.fba_inventory.available
        #         sheet.cell(row,3).value = todayProductSale.fba_inventory.fc_unit
        #         sheet.cell(row,4).value = todayProductSale.fba_inventory.inbound_unit
        #         fba_total_unit = todayProductSale.fba_inventory.total_unit
        #         sheet.cell(row,5).value = fba_total_unit
        #         shenzhen_inventory = check_shenzhen_inventory(sku)
        #         sheet.cell(row,6).value = shenzhen_inventory
        #         abroad_qty = 0
        #         if abroad_warehouse_uploaded :
        #             sheet.cell(row,7).value = 0
        #             if sku in abroad_warehouse_sku_qty:
        #                 abroad_qty = abroad_warehouse_sku_qty[sku]
        #                 sheet.cell(row,7).value = abroad_warehouse_sku_qty[sku]
        #         sheet.cell(row,7 + abroad_warehouse_uploaded).value = int(todayProductSale.sold_qty)
        #         sheet.cell(row,8 + abroad_warehouse_uploaded).value = round(todayProductSale.sold_qty_average_7d,1)
        #         increase_rate = 1.0
        #         seven_days_ago_history_today_product_sales = HistoryTodayProductSales.objects.filter(date = seven_days_ago_date, product__sku=sku, country=country).first()
        #         if seven_days_ago_history_today_product_sales != None:
        #             seven_days_ago_sold_qty_average_7d = seven_days_ago_history_today_product_sales.sold_qty_average_7d
        #             if seven_days_ago_sold_qty_average_7d != 0:
        #                 increase_rate = todayProductSale.sold_qty_average_7d / seven_days_ago_sold_qty_average_7d
        #         sheet.cell(row,9 + abroad_warehouse_uploaded).value = round(increase_rate,1)# 暴增
        #         sheet.cell(row,10 + abroad_warehouse_uploaded).value = todayProductSale.fba_inventory.days_of_supply
        #         lasting_day_of_total_fba_and_shenzhen_inventory_estimated_by_us = (todayProductSale.fba_inventory.total_unit + shenzhen_inventory + abroad_qty)/ todayProductSale.sold_qty_average_7d
        #         sheet.cell(row,11 + abroad_warehouse_uploaded).value = int(lasting_day_of_total_fba_and_shenzhen_inventory_estimated_by_us)
        #         sheet.cell(row,12 + abroad_warehouse_uploaded).value = round(todayProductSale.average_price_7d,2)
        #         sheet.cell(row,12 + abroad_warehouse_uploaded).number_format = u'"$ "#,##0.00'
        #         amazon_recommended_replenishment_qty = todayProductSale.fba_inventory.recommended_replenishment_qty
        #         sheet.cell(row,13 + abroad_warehouse_uploaded).value = amazon_recommended_replenishment_qty
        #         receivable_purchased_qty = check_receivable_purchased_qty(sku)
        #         sheet.cell(row,14 + abroad_warehouse_uploaded).value = receivable_purchased_qty
        #         if todayProductSale.product.discontinued:
        #             continue #sheet.cell(row,13).value = "不再订货"
        #         else:
        #             for i, max_days_to_last_for_po in enumerate(max_days_to_last_for_pos):
        #                 po_qty = 0
        #                 if todayProductSale.fba_inventory.days_of_supply >= lasting_day_of_total_fba_and_shenzhen_inventory_estimated_by_us:
        #
        #                     days_vacancy = max_days_to_last_for_po - lasting_day_of_total_fba_and_shenzhen_inventory_estimated_by_us
        #                     qty_needed = int(days_vacancy * todayProductSale.sold_qty_average_7d)
        #                     if qty_needed > receivable_purchased_qty:
        #                         qty_needed = qty_needed - receivable_purchased_qty
        #                         if qty_needed > float(moq) / 2:
        #                             po_qty = max([qty_needed, moq])
        #                 else:
        #                     shenzhen_inventory_and_receivable_purchased_qty = shenzhen_inventory + receivable_purchased_qty
        #                     if abroad_warehouse_uploaded and sku in abroad_warehouse_sku_qty:
        #                         shenzhen_inventory_and_receivable_purchased_qty += abroad_warehouse_sku_qty[sku]
        #                     if i > 0:
        #                         if todayProductSale.fba_inventory.days_of_supply:
        #                             amazon_recommended_replenishment_qty_i = amazon_recommended_replenishment_qty + float(max_days_to_last_for_pos[i] - max_days_to_last_for_pos[0])/ todayProductSale.fba_inventory.days_of_supply * fba_total_unit
        #                         else:
        #                             amazon_recommended_replenishment_qty_i = amazon_recommended_replenishment_qty + fba_total_unit * i
        #                     else:
        #                         amazon_recommended_replenishment_qty_i = amazon_recommended_replenishment_qty
        #                     if amazon_recommended_replenishment_qty_i > shenzhen_inventory_and_receivable_purchased_qty:
        #                         qty_needed = amazon_recommended_replenishment_qty_i - shenzhen_inventory_and_receivable_purchased_qty
        #                         if qty_needed > float(moq) / 2:
        #                             po_qty = max([qty_needed, moq])
        #                 if po_qty:
        #                     sheet.cell(row,start_col_restock + i).value = int(po_qty)
        #
        #             if min([todayProductSale.fba_inventory.days_of_supply, lasting_day_of_total_fba_and_shenzhen_inventory_estimated_by_us]) < min_days_lasting_for_notice and not(todayProductSale.product.discontinued):
        #                 for cell in sheet["%i:%i" %(row,row)]:
        #                     cell.fill = yellow_fill

                # row +=1
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename={date}-PO.xlsx'.format( \
               date=datetime.datetime.now().strftime('%Y-%m-%d') \
               ,)
        wb.save(response)
        return response
    template = loader.get_template('restock_today.html')
    context = {
        'form':UploadFileForm()
    }
    return HttpResponse(template.render(context, request))

@login_required
def shipment_today(request):
    if request.method == 'POST':
        inbound_sku_qty = {}
        form = UploadFileCountryForm(request.POST,request.FILES)
        if form.is_valid():
            if 'file' in request.FILES:
                fileInMemory = request.FILES['file'].read()
                filePath = BytesIO(fileInMemory)
                wb2 = load_workbook(filePath, data_only=True)
                sheet = wb2.worksheets[0]
                for row in range(1, sheet.max_row + 1):
                    sku = sheet.cell(row,1).value
                    qty = int(sheet.cell(row,2).value)
                    inbound_sku_qty[sku]= qty
            country = request.POST['country']
        wb = Workbook()
        sheet = wb.active
        sheet.title = '今日发货单'
        yellow_fill = PatternFill("solid", fgColor="FFFF00")
        min_shippable_qty = 20
        min_sold_qty_per_day_for_notice = 1
        max_days_lasting = 120
        row = 1
        titles = ['SKU', 'FNSKU', 'FBA可售库存', '中转库存', '在途库存', '亚马逊总库存', 'F199库存' \
                    , '过去7天平均日销售', '亚马逊预测能撑多少天', '我们预测总fba库存能撑多少天' \
                    , '可售库存能撑多少天','可售+中转库存能撑多少天' \
                    , '单价', '未入库数量', '预计入库日期', '决策' \
                    ]
        for i, title in enumerate(titles):
            sheet.cell(row,i + 1).value = title
        row += 1
        inventories = Inventory.objects.filter(warehouse_name = shenzhen_warehouse_name, qty__gte = min_shippable_qty)
        # country = 'US'
        for inventory in inventories.all():
            sku = inventory.sku
            fba_inventory = FbaInventory.objects.filter(sku = sku,country=country).first()
            if fba_inventory:
                fnsku = fba_inventory.fnsku
            else:
                fnsku = ""
            # 没启动远程配送的才计算发货量
            if country != 'CA' or RemoteFulfillmentSku.objects.filter(sku = sku, country=country).count() == 0:
                shenzhen_inventory = check_shenzhen_inventory(sku)
                if sku in inbound_sku_qty:
                    shenzhen_inventory += inbound_sku_qty[sku]
                    inbound_sku_qty.pop(sku, None)
                sheet.cell(row,1).value = sku
                if fnsku:
                    sheet.cell(row,2).value = fnsku
                todayProductSales = TodayProductSales.objects.filter(product__sku = sku,country=country)
                if todayProductSales.count():
                    todayProductSale = todayProductSales.all()[0]
                    sheet.cell(row,3).value = todayProductSale.fba_inventory.available
                    sheet.cell(row,4).value = todayProductSale.fba_inventory.fc_unit
                    sheet.cell(row,5).value = todayProductSale.fba_inventory.inbound_unit
                    fba_total_unit = todayProductSale.fba_inventory.total_unit
                    sheet.cell(row,6).value = fba_total_unit
                    sheet.cell(row,7).value = shenzhen_inventory
                    sheet.cell(row,8).value = round(todayProductSale.sold_qty_average_7d,1)
                    sheet.cell(row,9).value = todayProductSale.fba_inventory.days_of_supply
                    sheet.cell(row,10).value = int(todayProductSale.lasting_day_of_total_fba_unit_estimated_by_us)
                    sheet.cell(row,11).value = int(todayProductSale.lasting_day_of_available_estimated_by_us)
                    sheet.cell(row,12).value = int(todayProductSale.lasting_day_of_available_fc_estimated_by_us)
                    sheet.cell(row,13).value = round(todayProductSale.average_price_7d,2)
                    sheet.cell(row,13).number_format = u'"$ "#,##0.00'
                    if min([todayProductSale.fba_inventory.days_of_supply, todayProductSale.lasting_day_of_total_fba_unit_estimated_by_us]) < max_days_lasting and todayProductSale.sold_qty_average_7d > min_sold_qty_per_day_for_notice:
                        for cell in sheet["%i:%i" %(row,row)]:
                            cell.fill = yellow_fill
                else:
                    if fba_inventory:
                        sheet.cell(row,3).value = fba_inventory.available
                        sheet.cell(row,4).value = fba_inventory.fc_unit
                        sheet.cell(row,5).value = fba_inventory.inbound_unit
                        fba_total_unit = fba_inventory.total_unit
                        sheet.cell(row,6).value = fba_total_unit
                        shenzhen_inventory = check_shenzhen_inventory(sku)
                        sheet.cell(row,9).value = fba_inventory.days_of_supply
                    sheet.cell(row,7).value = shenzhen_inventory
                nearestReceivablePurchasedQty = NearestReceivablePurchasedQty.objects.filter(sku = sku)
                if nearestReceivablePurchasedQty.count():
                    nearestReceivablePurchasedQty = nearestReceivablePurchasedQty.first()
                    sheet.cell(row,14).value = nearestReceivablePurchasedQty.qty
                    sheet.cell(row,15).value = nearestReceivablePurchasedQty.date
                row +=1
        for sku, qty in inbound_sku_qty.items():
            if country != 'CA' or RemoteFulfillmentSku.objects.filter(sku = sku, country=country).count() == 0:
                sheet.cell(row,1).value = sku
                fba_inventory = FbaInventory.objects.filter(sku = sku,country=country).first()
                if fba_inventory:
                    sheet.cell(row,2).value = fba_inventory.fnsku
                todayProductSales = TodayProductSales.objects.filter(product__sku = sku,country=country)
                if todayProductSales.count():
                    todayProductSale = todayProductSales.all()[0]
                    sheet.cell(row,3).value = todayProductSale.fba_inventory.available
                    sheet.cell(row,4).value = todayProductSale.fba_inventory.fc_unit
                    sheet.cell(row,5).value = todayProductSale.fba_inventory.inbound_unit
                    fba_total_unit = todayProductSale.fba_inventory.total_unit
                    sheet.cell(row,6).value = fba_total_unit
                    shenzhen_inventory = check_shenzhen_inventory(sku) + qty
                    sheet.cell(row,7).value = shenzhen_inventory
                    sheet.cell(row,8).value = round(todayProductSale.sold_qty_average_7d,1)
                    sheet.cell(row,9).value = todayProductSale.fba_inventory.days_of_supply
                    sheet.cell(row,10).value = int(todayProductSale.lasting_day_of_total_fba_unit_estimated_by_us)
                    sheet.cell(row,11).value = int(todayProductSale.lasting_day_of_available_estimated_by_us)
                    sheet.cell(row,12).value = int(todayProductSale.lasting_day_of_available_fc_estimated_by_us)
                    sheet.cell(row,13).value = round(todayProductSale.average_price_7d,2)
                    sheet.cell(row,13).number_format = u'"$ "#,##0.00'
                    if min([todayProductSale.fba_inventory.days_of_supply, todayProductSale.lasting_day_of_total_fba_unit_estimated_by_us]) < max_days_lasting and todayProductSale.sold_qty_average_7d > min_sold_qty_per_day_for_notice:
                        for cell in sheet["%i:%i" %(row,row)]:
                            cell.fill = yellow_fill
                else:
                    fbaInventory = FbaInventory.objects.filter(sku = sku,country=country)
                    if fbaInventory.count():
                        fba_inventory = fbaInventory.all()[0]
                        sheet.cell(row,3).value = fba_inventory.available
                        sheet.cell(row,4).value = fba_inventory.fc_unit
                        sheet.cell(row,5).value = fba_inventory.inbound_unit
                        sheet.cell(row,6).value = fba_inventory.total_unit
                        shenzhen_inventory = check_shenzhen_inventory(sku) + qty
                        sheet.cell(row,7).value = shenzhen_inventory
                        sheet.cell(row,9).value = fba_inventory.days_of_supply
                    else:
                        sheet.cell(row,7).value = qty
                row +=1
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename={date}-fbaShipment-{country}.xlsx'.format( \
               date=datetime.datetime.now().strftime('%Y-%m-%d') \
               , country = country \
               ,)
        wb.save(response)
        return response

    template = loader.get_template('shipment_today.html')
    context = {
        'form':UploadFileCountryForm(initial={'country': 'US'})
    }
    return HttpResponse(template.render(context, request))


@login_required
def reprice_today(request):
    min_available_qty_for_adjust_price = 10
    standard_price = 25.99
    highest_price = 34.99
    country = 'US'
    template = loader.get_template('reprice_today.html')

    titles = ['产品图片', 'SKU', 'FBA可售库存', '中转库存', '在途库存', '亚马逊总库存', 'F199库存', '亚马逊预测撑天' \
            , '我们预测总fba库存撑天', '可售库存撑天','可售+中转库存撑天' \
            , '过去7天平均日销售', '今日销量', '正常价', '最高限价', '单价', '决策' \
            ]

    products_to_reprice_list = []

    #找出所有FBA有库存的商品
    for fba_inventory in FbaInventory.objects.filter(available__gt = min_available_qty_for_adjust_price, country=country):

        sku = fba_inventory.sku
        product_to_reprice_info = [""]
        product_to_reprice_info.append("")
        product_to_reprice_info.append(sku)
        fba_inventory_available = fba_inventory.available
        fba_inventory_fc_unit = fba_inventory.fc_unit
        fba_inventory_inbound_unit = fba_inventory.inbound_unit
        fba_total_unit = fba_inventory.total_unit
        shenzhen_inventory = check_shenzhen_inventory(sku)
        fba_inventory_days_of_supply = fba_inventory.days_of_supply
        product_to_reprice_info.append(fba_inventory_available)
        product_to_reprice_info.append(fba_inventory_fc_unit)
        product_to_reprice_info.append(fba_inventory_inbound_unit)
        product_to_reprice_info.append(fba_total_unit)
        product_to_reprice_info.append(shenzhen_inventory)
        product_to_reprice_info.append(fba_inventory_days_of_supply)


        #降价部分
        #判断价格是否高于正常价
        #先找到最新售价
        lastest_sales_having_this_sku = HistoryTodayProductSales.objects.filter(product__sku = sku, country=country).order_by('-date')

        if lastest_sales_having_this_sku.count():

            lastest_sales_having_this_sku = lastest_sales_having_this_sku.first()
            sold_qty_average_7d = lastest_sales_having_this_sku.sold_qty_average_7d
            average_price_today = lastest_sales_having_this_sku.average_price_7d
            if lastest_sales_having_this_sku.sales_amount != 0 and lastest_sales_having_this_sku.sold_qty != 0:
                average_price_today = lastest_sales_having_this_sku.sales_amount / lastest_sales_having_this_sku.sold_qty
            lasting_days_with_fba_inventory_available  = fba_inventory_available / sold_qty_average_7d
            lasting_days_with_fba_inventory_available_and_fc_unit = (fba_inventory_available + fba_inventory_fc_unit)/ sold_qty_average_7d
            lasting_days_with_fba_total_unit = fba_total_unit / sold_qty_average_7d
            sold_qty_average_7d = lastest_sales_having_this_sku.sold_qty_average_7d
            today_sold_qty = 0
            todayProductSales = TodayProductSales.objects.filter(product__sku = sku, country=country)
            if todayProductSales.count():
                todayProductSales = todayProductSales.first()
                today_sold_qty = todayProductSales.sold_qty
            product_to_reprice_info.append(int(lasting_days_with_fba_total_unit))
            product_to_reprice_info.append(int(lasting_days_with_fba_inventory_available))
            product_to_reprice_info.append(int(lasting_days_with_fba_inventory_available_and_fc_unit))
            product_to_reprice_info.append(round(sold_qty_average_7d,1))
            product_to_reprice_info.append(today_sold_qty)
            product_to_reprice_info.append(standard_price)
            product_to_reprice_info.append(highest_price)
            product_to_reprice_info.append(round(average_price_today, 2))

            if int(average_price_today) > int(standard_price):
                #高于正常价
                #判断销售是否在下降
                if today_sold_qty <= sold_qty_average_7d - 1:
                    #销售在下降，则要降价
                    #降价前还是要判断是否有断货风险
                    if lasting_days_with_fba_total_unit > 30 and lasting_days_with_fba_inventory_available_and_fc_unit > 20 :
                        #没有断货风险
                        product = Product.objects.get(sku = sku)
                        product_to_reprice_info[1] = product.image.url
                        product_to_reprice_info.append( '降价' )
                        product_to_reprice_info[0] = "blue"
                        products_to_reprice_list.append(product_to_reprice_info)
                elif lasting_days_with_fba_inventory_available_and_fc_unit < 20:
                    #可售+中转库存撑不到20天
                    if today_sold_qty >= sold_qty_average_7d:
                        product_to_reprice_info.append( '涨价' )
                        product_to_reprice_info[0] = "red"
                    product = Product.objects.get(sku = sku)
                    product_to_reprice_info[1] = product.image.url
                    products_to_reprice_list.append(product_to_reprice_info)
            elif lasting_days_with_fba_inventory_available_and_fc_unit < 20 and today_sold_qty > sold_qty_average_7d:
                #不高于正常价但是可售+中转库存撑不到20天，而且销售在上升
                product = Product.objects.get(sku = sku)
                product_to_reprice_info[1] = product.image.url
                product_to_reprice_info.append( '涨价' )
                product_to_reprice_info[0] = "red"
                products_to_reprice_list.append(product_to_reprice_info)
    context = {
        'titles': titles ,\
        'products_to_reprice_list': products_to_reprice_list \
    }
    return HttpResponse(template.render(context, request))

def scrape(url):
    headers = {
        'authority': 'www.amazon.com',
        'pragma': 'no-cache',
        'cache-control': 'no-cache',
        'dnt': '1',
        'upgrade-insecure-requests': '1',
        'user-agent': 'Mozilla/5.0 (X11; CrOS x86_64 8172.45.0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/51.0.2704.64 Safari/537.36',
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
        'sec-fetch-site': 'none',
        'sec-fetch-mode': 'navigate',
        'sec-fetch-dest': 'document',
        'accept-language': 'en-GB,en-US;q=0.9,en;q=0.8',
    }

    # Download the page using requests
    r = requests.get(url, headers=headers)
    # Simple check to check if page was blocked (Usually 503)
    if r.status_code > 500:
        if "To discuss automated access to Amazon data please contact" in r.text:
            print("Page %s was blocked by Amazon. Please try using better proxies\n"%url)
        else:
            print("Page %s must have been blocked by Amazon as the status code was %d"%(url,r.status_code))
        return None
    # Pass the HTML of the page and create
    f = open(os.path.join(BASE_DIR,'amazon_response.html'), "w")
    f.write(r.text)
    f.close()
    data = extractor.extract(r.text,base_url=url)
    reviews = []
    for r in data['reviews']:
        r["product"] = data["product_title"]
        r['url'] = url
        if 'verified_purchase' in r:
            if r['verified_purchase'] and 'Verified Purchase' in r['verified_purchase']:
                r['verified_purchase'] = True
            else:
                r['verified_purchase'] = False
        r['rating'] = r['rating'].split(' out of')[0]
        date_posted = r['date'].split('on ')[-1]
        if r['images']:
            r['images'] = "\n".join(r['images'])
        r['date'] = dateparser.parse(date_posted).strftime('%d %b %Y')
        reviews.append(r)
    data.pop('histogram', None)
    # histogram = {}
    # for h in data['histogram']:
    #     histogram[h['key']] = h['value']
    # data['histogram'] = histogram
    data['average_rating'] = float(data['average_rating'].split(' out')[0])
    data['reviews'] = reviews
    data['number_of_reviews'] = int(data['number_of_reviews'].split('  customer')[0])
    return data

def distinct_size_color(size_color_str):
    Size_Tip_Str = ['Size: ']
    Color_Tip_Str = ['Color: ', 'Colour: ']
    pos_dict = {}
    size_pos = color_pos =-1
    for k in Size_Tip_Str:
        size_pos = size_color_str.find(k)
        if size_pos > -1:
            pos_dict['size'] = {'tip_str':k, 'pos':size_pos}
            break
    for k in Color_Tip_Str:
        color_pos = size_color_str.find(k)
        if color_pos > -1:
            pos_dict['color'] = {'tip_str':k, 'pos':color_pos}
            break
    if len(pos_dict.keys()) == 1:
        return {list(pos_dict.keys())[0]: size_color_str[len(list(pos_dict.values())[0]['tip_str']):]}
    elif len(pos_dict.keys()) > 1:
        if pos_dict['size']['pos'] > pos_dict['color']['pos']:
            return {'color': size_color_str[len(pos_dict['color']['tip_str']):pos_dict['size']['pos']] \
                    ,'size': size_color_str[(pos_dict['size']['pos'] + len(pos_dict['size']['tip_str'])):]}
        else:
            return {'size': size_color_str[len(pos_dict['size']['tip_str']):pos_dict['color']['pos']] \
                    ,'color': size_color_str[(pos_dict['color']['pos'] + len(pos_dict['color']['tip_str'])):]}

def generate_pivot_table(data_list):
    variants = []
    for i in data_list:
        if 'variant' in i and i['variant'] != None:
            variants.append(i['variant'])
    return

def size_color_counter_dict_to_list(x):
    res1 = []
    res2 = []
    for k,v in sorted(x.items(), key=lambda item: item[1], reverse=True):
        res1.append(k)
        res2.append(v)
    return {'name': res1, 'value': res2}

@login_required
def get_reviews_stat(request):
    if request.method == 'POST':
        from collections import Counter
        reviews_by_size_color = []
        earliest_date = latest_date = None

        form = UploadFileForm(request.POST,request.FILES)
        if form.is_valid():
            # a-size-base 2 是日期列，a-size-mini的变量列
            col_nums = {'a-size-base 2':0, 'a-size-mini':0, 'a-size-mini 2':0}
            fileInMemory = request.FILES['file'].read().decode('utf-8')
            csv_data = csv.reader(StringIO(fileInMemory), delimiter=',')
            for i, row in enumerate(csv_data):
                if i == 0:
                    for col_num, title in enumerate(row):
                        if title in col_nums.keys():
                            col_nums[title] = col_num
                    col_nums['date'] = col_nums['a-size-base 2']
                    col_nums['variant'] = col_nums['a-size-mini'] or col_nums['a-size-mini 2']
                else:
                    if row[col_nums['a-size-mini']] == 'Verified Purchase':
                        col_nums['variant'] = col_nums['a-size-mini 2']
                    date_posted = dateparser.parse(row[col_nums['date']].split('on ')[-1])
                    variant_str = row[col_nums['variant']]
                    if variant_str:
                        reviews_by_size_color.append(distinct_size_color(variant_str))
                    if latest_date == None:
                        earliest_date = latest_date = date_posted
                    else:
                        if date_posted > latest_date:
                            latest_date = date_posted
                        if date_posted < earliest_date:
                            earliest_date = date_posted
            day_range_tip = '从%s到%s，共%i天%.1f月' %(earliest_date.strftime('%Y-%m-%d'), latest_date.strftime('%Y-%m-%d'), (latest_date - earliest_date).days, float((latest_date - earliest_date).days)/30)
            counter_size_color = {}
            if len(reviews_by_size_color[0].keys()) == 1:
                size_color = list(reviews_by_size_color[0].keys())[0]
                size_color_list = [i[size_color] for i in reviews_by_size_color]
                counter_size_color[size_color] = dict(Counter(size_color_list))
                counter_size_color[size_color] = size_color_counter_dict_to_list(counter_size_color[size_color])
            elif len(reviews_by_size_color[0].keys()) > 1:
                for i_size_color in list(reviews_by_size_color[0].keys()):
                    size_color_list = [i[i_size_color] for i in reviews_by_size_color]
                    counter_size_color[i_size_color] = dict(Counter(size_color_list))
                    counter_size_color[i_size_color] = size_color_counter_dict_to_list(counter_size_color[i_size_color])
            return JsonResponse({'counter_size_color': counter_size_color, 'day_range_tip':day_range_tip}, safe=False)
    template = loader.get_template('get_reviews.html')
    context = {
        'form':UploadFileForm()
    }
    return HttpResponse(template.render(context, request))

@login_required
def get_upcs(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST,request.FILES)
        if form.is_valid():
            fileInMemory = request.FILES['file'].read()
            filePath = BytesIO(fileInMemory)
            wb2 = load_workbook(filePath, data_only=True)
            sheet = wb2.worksheets[0]
            sku_upc_list = []
            for row in range(1, sheet.max_row + 1):
                sku = str(sheet.cell(row,1).value)
                if sku and sku != 'None' and sku.upper().strip():
                    sku = sku.upper().strip()
                    sku_upc, created = SkuUpc.objects.get_or_create(sku = sku)
                    if created:
                        upc = Upc.objects.filter(used = False).first()
                        sku_upc.upc = upc
                        sku_upc.save()
                        upc.used = True
                        upc.save()
                    sku_upc_list.append(sku_upc)
            wb = Workbook()
            sheet = wb.active
            sheet.title = 'sku_upc'
            for i,sku_upc in enumerate(sku_upc_list):
                sheet.cell(i + 1, 1).value = sku_upc.sku
                sheet.cell(i + 1, 2).value = sku_upc.upc.upc
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename={date}-sku-upc.xlsx'.format( \
                   date=datetime.datetime.now().strftime('%Y-%m-%d') \
                   ,)
            wb.save(response)
            return response
    template = loader.get_template('get_upcs.html')
    context = {
        'form':UploadFileForm()
    }
    return HttpResponse(template.render(context, request))

@login_required
def input_upcs(request):
    if request.method == 'POST' : # and request.get_host() == '209.97.151.168':
        form = UploadFileForm(request.POST,request.FILES)
        if form.is_valid():
            fileInMemory = request.FILES['file'].read()
            filePath = BytesIO(fileInMemory)
            wb2 = load_workbook(filePath, data_only=True)
            sheet = wb2.worksheets[0]
            for row in range(1, sheet.max_row + 1):
                upc = str(sheet.cell(row,1).value)
                if upc:
                    upc = upc.upper().strip()
                    upc, created = Upc.objects.get_or_create(upc = upc)
    template = loader.get_template('input_upcs.html')
    context = {
        'form':UploadFileForm()
    }
    return HttpResponse(template.render(context, request))

@login_required
def update_sku_supplier(request):
    if request.method == 'POST' : # and request.get_host() == '209.97.151.168':
        form = UploadFileForm(request.POST,request.FILES)
        if form.is_valid():
            fileInMemory = request.FILES['file'].read()
            filePath = BytesIO(fileInMemory)
            wb2 = load_workbook(filePath, data_only=True)
            sheet = wb2.worksheets[0]
            for row in range(1, sheet.max_row + 1):
                sku = str(sheet.cell(row,1).value)
                supplier_long_str = sheet.cell(row,2).value
                supplier_jingdouyun_id = supplier_long_str[:6]
                if sku:
                    sku = sku.upper().strip()
                    if not SkuSupplier.objects.filter(sku = sku, supplier_id = supplier_jingdouyun_id).exists():
                        sku_supplier = SkuSupplier(sku = sku, supplier_id = supplier_jingdouyun_id)
                        sku_supplier.save()
                    for sku_supplier in SkuSupplier.objects.filter(sku = sku).all():
                        if sku_supplier.supplier_id != supplier_jingdouyun_id:
                            sku_supplier.delete()

    template = loader.get_template('update_sku_supplier.html')
    context = {
        'form':UploadFileForm()
    }
    return HttpResponse(template.render(context, request))


@login_required
def product_information(request):
    template = loader.get_template('product_information.html')
    context = {
        'products': Product.objects.filter(discontinued = False).order_by('sku')
    }
    return HttpResponse(template.render(context, request))

@user_passes_test(lambda u: u.groups.filter(name__in = ['sales', 'accountant']).exists())
def product_information_export(request):
    wb = Workbook()
    sheet = wb.active
    for i_col in range(1,len(UPDATE_PRODUCT_INFORMATION_IN_BULK_TABLE_HEADER) + 1):
        sheet.cell(1,i_col).value = UPDATE_PRODUCT_INFORMATION_IN_BULK_TABLE_HEADER[i_col - 1]

    i_product = 1
    for product in Product.objects.filter(discontinued = False).order_by('sku'):
        sheet.cell(i_product + 1,1).value = product.sku
        sheet.cell(i_product + 1,2).value = product.package_length
        sheet.cell(i_product + 1,3).value = product.package_width
        sheet.cell(i_product + 1,4).value = product.package_height
        sheet.cell(i_product + 1,5).value = product.package_weight
        if product.actual_weight_forced:
            sheet.cell(i_product + 1,6).value = "是"
        i_product += 1
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename={date}-product_information.xlsx'.format( \
           date=datetime.datetime.now().strftime('%Y-%m-%d') \
           ,)
    wb.save(response)
    return response


@user_passes_test(lambda u: u.groups.filter(name='logistics').exists())
def update_product_information_in_bulk(request):
    correction_tip = ''
    update_completed_tip = ''
    if request.method == 'POST' : # and request.get_host() == '209.97.151.168':
        form = UploadFileForm(request.POST,request.FILES)
        if form.is_valid():
            fileInMemory = request.FILES['file'].read()
            filePath = BytesIO(fileInMemory)
            wb2 = load_workbook(filePath, data_only=True)
            sheet = wb2.worksheets[0]
            for i_col in range(1, sheet.max_column + 1):
                if sheet.cell(1,i_col).value != UPDATE_PRODUCT_INFORMATION_IN_BULK_TABLE_HEADER[i_col - 1]:
                    correction_tip = '出错啦！第一行标题需要修正，正确的是' + ' '.join(UPDATE_PRODUCT_INFORMATION_IN_BULK_TABLE_HEADER)
                    break;
            if not correction_tip:
                for row in range(1, sheet.max_row + 1):
                    sku = str(sheet.cell(row,1).value).split()[0]
                    product = Product.objects.filter(sku = sku)
                    if product.count():
                        product = product.first()
                        package_length = sheet.cell(row,2).value
                        package_width = sheet.cell(row,3).value
                        package_height = sheet.cell(row,4).value
                        package_weight = sheet.cell(row,5).value
                        actual_weight_forced = str(sheet.cell(row,6).value)
                        if package_length * package_width * package_height * package_weight == 0:
                            continue;
                        product.package_length = package_length
                        product.package_width = package_width
                        product.package_height = package_height
                        product.package_weight = package_weight

                        if actual_weight_forced in ['是','Yes', 'YES', '1']:
                            product.actual_weight_forced = True
                        else:
                            product.actual_weight_forced = False
                        product.save(force_update=True)
                update_completed_tip = '更新完成'

    template = loader.get_template('update_product_information_in_bulk.html')
    context = {
        'form':UploadFileForm(), \
        'correction_tip': correction_tip , \
        'update_completed_tip': update_completed_tip \
    }
    return HttpResponse(template.render(context, request))


@login_required
def update_all_product_purchase_price(request):
    if request.method == 'POST':
        if request.user.is_superuser:
            fileInMemory = request.FILES['file'].read()
            filePath = BytesIO(fileInMemory)
            wb2 = load_workbook(filePath, data_only=True)
            sheet = wb2.worksheets[0]
            max_column = sheet.max_column
            correct_price_title = '单位成本'
            correct_price_title_cell_row = 5
            correct_price_title_cell_columb = 11
            if sheet.cell(correct_price_title_cell_row, correct_price_title_cell_columb).value == correct_price_title:

                date = datetime.datetime.strptime(sheet.cell(3, 1).value[-10:], "%Y-%m-%d").date()
                for row in range(6, sheet.max_row):
                    sku = sheet.cell(row, 1).value
                    price = sheet.cell(row, correct_price_title_cell_columb).value
                    skuPurchasingPrice = SkuPurchasingPrice.objects.filter(sku = sku, date = date)
                    if skuPurchasingPrice.count():
                        skuPurchasingPrice = skuPurchasingPrice.first()
                        skuPurchasingPrice.purchasing_price = price
                        skuPurchasingPrice.save()
                    else:
                        (sku_purchasing_price, create) = SkuPurchasingPrice.objects.get_or_create(sku = sku ,\
                                                                                            purchasing_price = price ,\
                                                                                            date = date)

            else:
                return HttpResponse('请检查上传的文件格式是否正确')
    template = loader.get_template('update_all_product_purchase_price.html')
    context = {
        'form':UploadFileForm()
    }

    return HttpResponse(template.render(context, request))

@login_required
def update_all_product_head_shipping_unit_cost(request):
    if request.method == 'POST':
        fileInMemory = request.FILES['file'].read()
        filePath = BytesIO(fileInMemory)
        wb2 = load_workbook(filePath, data_only=True)
        sheet = wb2.worksheets[0]
        max_column = sheet.max_column

        year_month = sheet.cell(1, 1).value
        year_month_day = str(year_month) + '01'
        date = datetime.datetime.strptime(year_month_day, "%Y%m%d").date()

        country = sheet.cell(2, 1).value.upper()

        for row in range(4, sheet.max_row + 1):
            sku = sheet.cell(row, 1).value
            head_shipping_unit_cost = sheet.cell(row, 3).value
            skuHeadShippingUnitCost = SkuHeadShippingUnitCost.objects.filter(sku = sku ,\
                                                                                    type = ('G', 'General') ,\
                                                                                    country = country ,\
                                                                                    date = date)
            if skuHeadShippingUnitCost.count():
                skuHeadShippingUnitCost = skuHeadShippingUnitCost.first()
                skuHeadShippingUnitCost.head_shipping_unit_cost = head_shipping_unit_cost
                skuHeadShippingUnitCost.save()
            else:
                (sku_purchasing_price, create) = SkuHeadShippingUnitCost.objects.get_or_create(sku = sku ,\
                                                                                    type = ('G', 'General') ,\
                                                                                    country = country ,\
                                                                                    head_shipping_unit_cost = head_shipping_unit_cost ,\
                                                                                    date = date)

    template = loader.get_template('update_all_product_head_shipping_unit_cost.html')
    context = {
        'form':UploadFileForm()
    }

    return HttpResponse(template.render(context, request))

def setting_up_sku_managed_by_sales_person(sku, sales_person_name, sales_assistant_name):
    skuManagedBySalesPerson = SkuManagedBySalesPerson.objects.filter(sku = sku).first()
    if skuManagedBySalesPerson:
        skuManagedBySalesPerson.sales_person_name = sales_person_name
        skuManagedBySalesPerson.sales_assistant_name = sales_assistant_name
        skuManagedBySalesPerson.save()
    else:
        (skuManagedBySalesPerson, created) = SkuManagedBySalesPerson.objects.get_or_create(sku = sku ,\
                                                                                        sales_person_name = sales_person_name, \
                                                                                        sales_assistant_name = sales_assistant_name)

def setting_up_sku_contributor(sku, proposer_name, designer_name):
    skuContributor = SkuContributor.objects.filter(sku = sku).first()
    if skuContributor:
        if proposer_name:
            skuContributor.proposer_name = proposer_name
        if designer_name:
            skuContributor.designer_name = designer_name
        skuContributor.save()
    else:
        if proposer_name and designer_name:
            (skuContributor, created) = SkuContributor.objects.get_or_create(sku = sku,proposer_name = proposer_name, designer_name = designer_name)


def initialize_sku_asset_liability(sku):
    initial = True
    date = datetime.date.today()
    fba_inventory = 0
    shenzhen_inventory = 0
    fbaInventory = FbaInventory.objects.filter(sku = sku)
    if fbaInventory.count():
        fba_inventory = fbaInventory.first().total_unit

    inventory = Inventory.objects.filter(warehouse_name = shenzhen_warehouse_name,sku = sku)
    if inventory.count():
        shenzhen_inventory = inventory.first().qty
    initial_inventory_quantity = fba_inventory + shenzhen_inventory
    greater = SkuPurchasingPrice.objects.filter(sku = sku, date__gte=date).order_by('date').first()
    if greater:
        unit_purchasing_price = greater.purchasing_price
    else:
        less = SkuPurchasingPrice.objects.filter(sku = sku, date__lte=date).order_by('-date').first()
        if less:
            unit_purchasing_price = less.purchasing_price
        else:
            unit_purchasing_price = 0.0
    greater = SkuHeadShippingUnitCost.objects.filter(sku = sku, type = ('G', 'General'), country = 'US', date__gte=date).order_by('date').first()
    if greater:
        head_shipping_unit_price = greater.head_shipping_unit_cost
    else:
        less = SkuHeadShippingUnitCost.objects.filter(sku = sku, type = ('G', 'General'), country = 'US', date__lte=date).order_by('-date').first()
        if less:
            head_shipping_unit_price = less.head_shipping_unit_cost
        else:
            head_shipping_unit_price = 0.0
    initial_other_cost = 0.0
    initial_inventory_value = (fba_inventory + shenzhen_inventory) * unit_purchasing_price
    initial_liabilities = fba_inventory * (unit_purchasing_price + head_shipping_unit_price) + shenzhen_inventory * unit_purchasing_price
    liabilities = initial_liabilities
    initial_investment = 0.0
    history_inventment = 0.0
    cash_amount = 0.0
    net_asset_amount = -liabilities

    #先看看给销售外角色用的资产负债表是否已存在，存在的话，就不用管,不存在，则要读取当前的库存数量，对于已经有库存的产品，设置要放在月初1号，不是1号，则不会建立资产负债表，对于没有库存的产品，可以建立
    skuAssetLiabilityTable = SkuAssetLiabilityTable.objects.filter(sku = sku, for_sales = False).first()
    if not skuAssetLiabilityTable :
        if (initial_inventory_value > 0 and date.day == 15) or initial_inventory_value == 0:
            skuAssetLiabilityTable = SkuAssetLiabilityTable(sku = sku ,\
                                                            liabilities = liabilities ,\
                                                            date = date ,\
                                                            initial_inventory_quantity = initial_inventory_quantity ,\
                                                            unit_purchasing_price = unit_purchasing_price ,\
                                                            initial_inventory_value = initial_inventory_value ,\
                                                            head_shipping_unit_price = head_shipping_unit_price ,\
                                                            initial_other_cost = initial_other_cost ,\
                                                            initial_liabilities = initial_liabilities ,\
                                                            initial = initial ,\
                                                            initial_investment = initial_investment ,\
                                                            history_inventment = history_inventment ,\
                                                            cash_amount = cash_amount ,\
                                                            net_asset_amount = net_asset_amount)
            skuAssetLiabilityTable.save()
    #再建给销售夜色的资产负债表
        skuAssetLiabilityTable = SkuAssetLiabilityTable(sku = sku ,\
                                                        for_sales = True ,\
                                                        liabilities = liabilities ,\
                                                        date = date ,\
                                                        initial_inventory_quantity = initial_inventory_quantity ,\
                                                        unit_purchasing_price = unit_purchasing_price ,\
                                                        initial_inventory_value = initial_inventory_value ,\
                                                        head_shipping_unit_price = head_shipping_unit_price ,\
                                                        initial_other_cost = initial_other_cost ,\
                                                        initial_liabilities = initial_liabilities ,\
                                                        initial = initial ,\
                                                        initial_investment = initial_investment ,\
                                                        history_inventment = history_inventment ,\
                                                        cash_amount = cash_amount ,\
                                                        net_asset_amount = net_asset_amount)
        skuAssetLiabilityTable.save()
    return

@login_required
def update_sales_person_and_managing_sku_list(request):
    if request.method == 'POST':
        fileInMemory = request.FILES['file'].read()
        filePath = BytesIO(fileInMemory)
        wb2 = load_workbook(filePath, data_only=True)
        sheet = wb2.worksheets[0]
        max_column = sheet.max_column

        # 检查title是否正确
        titles = sheet.cell(1, 1).value + sheet.cell(1, 2).value + sheet.cell(1, 3).value + sheet.cell(1, 4).value + sheet.cell(1, 5).value + sheet.cell(1, 6).value
        if titles != 'SKU销售名销售助理名选品师名设计师名初始化销售的财务报表否':
            return HttpResponse('请检查上传的模版')
        for row in range(2, sheet.max_row + 1):
            sku = sheet.cell(row, 1).value
            sales_person_name = sheet.cell(row, 2).value
            sales_assistant_name = sheet.cell(row, 3).value
            proposer_name = sheet.cell(row, 4).value
            designer_name = sheet.cell(row, 5).value
            initial = sheet.cell(row, 6).value
            if initial:
                initial = True
            else:
                initial = False
            setting_up_sku_managed_by_sales_person(sku, sales_person_name, sales_assistant_name)
            if proposer_name or designer_name:
                setting_up_sku_contributor(sku, proposer_name, designer_name)
            if initial:
                initialize_sku_asset_liability(sku)
    template = loader.get_template('update_sales_person_and_managing_sku_list.html')
    context = {
        'form':UploadFileForm()
    }

    return HttpResponse(template.render(context, request))

@user_passes_test(lambda u: u.groups.filter(name__in = ['logistics', 'accountant']).exists())
def update_inventory_additional_value(request):
    template = loader.get_template('update_inventory_additional_value.html')
    error_tip = ''
    sucess_tip = ''
    if request.method == 'POST':
        fileInMemory = request.FILES['file'].read()
        filePath = BytesIO(fileInMemory)
        wb2 = load_workbook(filePath, data_only=True)
        sheet = wb2.worksheets[0]
        max_column = sheet.max_column

        sku_title = 'sku'
        unit_additional_cost_col = 4

        [year, month] = split_year_month(sheet.cell(1, 1).value)
        day = 1
        reference_date = datetime.date(year = year, month = month, day =day)

        if sheet.cell(2, 1).value.lower() != sku_title:
            error_tip = '第二行标题要设置为sku'
        else:
            sku_additional_cost_dict = {}
            for i_row in range(3, sheet.max_row + 1):
                sku = sheet.cell(i_row, 1).value
                additional_cost = sheet.cell(i_row, unit_additional_cost_col).value
                if sku in sku_additional_cost_dict:
                    sku_additional_cost_dict[sku] += additional_cost
                else:
                    sku_additional_cost_dict[sku] = additional_cost
            for sku, additional_cost in sku_additional_cost_dict.items():
                product_inventory_unit_value = ProductInventoryUnitValue.objects.filter(sku = sku, date = reference_date)
                if product_inventory_unit_value.count():
                    product_inventory_unit_value = product_inventory_unit_value.first()
                    product_inventory_unit_value.additional_cost = additional_cost
                    product_inventory_unit_value.save()
        sucess_tip = '完成录入'
    if error_tip:
        context = {
            'error_tip': error_tip ,\
            'form':UploadFileForm()
        }
    elif sucess_tip:
        context = {
            'sucess_tip': sucess_tip ,\
            'form':UploadFileForm()
        }
    else:
        context = {
            'form':UploadFileForm()
        }
    return HttpResponse(template.render(context, request))

@user_passes_test(lambda u: u.groups.filter(name__in = ['logistics', 'accountant']).exists())
def update_inventory_value(request):
    template = loader.get_template('update_inventory_value.html')
    error_tip = ''
    sucess_tip = ''
    if request.method == 'POST':
        fileInMemory = request.FILES['file'].read()
        filePath = BytesIO(fileInMemory)
        wb2 = load_workbook(filePath, data_only=True)
        sheet = wb2.worksheets[0]
        max_column = sheet.max_column

        sku_title = '商品编号'
        unit_cost = '单位成本'
        inventory_date_str = sheet.cell(2, 1).value[-10:]
        inventory_date_day = inventory_date_str.split('-')[2]

        for i_col in range(1, max_column):
            if sheet.cell(4, i_col).value == '单位成本':
                unit_cost_col = i_col
                break
        date = datetime.datetime.strptime(inventory_date_str, '%Y-%m-%d')
        the_first_day_of_the_last_month = (date.replace(day=1) - datetime.timedelta(days =1)).replace(day = 1)
        for i_row in range(5, sheet.max_row):
            sku = sheet.cell(i_row, 1).value
            if sku == '合计':
                break
            unit_cost = sheet.cell(i_row, unit_cost_col).value
            product_inventory_unit_value = ProductInventoryUnitValue.objects.filter(sku = sku, date = the_first_day_of_the_last_month)
            if product_inventory_unit_value.count():
                product_inventory_unit_value = product_inventory_unit_value.first()
            else:
                product_inventory_unit_value = ProductInventoryUnitValue(sku = sku, inventory_value = unit_cost, date = the_first_day_of_the_last_month)
            product_inventory_unit_value.save()
        sucess_tip = '完成录入'
    if error_tip:
        context = {
            'error_tip': error_tip ,\
            'form':UploadFileForm()
        }
    elif sucess_tip:
        context = {
            'sucess_tip': sucess_tip ,\
            'form':UploadFileForm()
        }
    else:
        context = {
            'form':UploadFileForm()
        }
    return HttpResponse(template.render(context, request))


@login_required
def update_paid_purchase_order(request):
    if request.method == 'POST':
        fileInMemory = request.FILES['file'].read()
        filePath = BytesIO(fileInMemory)
        wb2 = load_workbook(filePath, data_only=True)
        sheet = wb2.worksheets[0]
        max_column = sheet.max_column

        po_date_column = 1
        po_number_column = 2
        sku_column = 5
        qty_column = 11
        transaction_amount_column = 13

        po_date_title = '采购日期'
        po_number_title = '采购单据号'
        sku_title = '商品编号'
        qty_title = '数量'
        transaction_amount_title = '采购金额'

        # 先检查表头对不对

        if sheet.cell(5, po_date_column).value == po_date_title and \
            sheet.cell(5, po_number_column).value == po_number_title and \
            sheet.cell(5, sku_column).value == sku_title and \
            sheet.cell(5, qty_column).value == qty_title and \
            sheet.cell(5, transaction_amount_column).value == transaction_amount_title :
            for row in range(6, sheet.max_row):
                po_date = sheet.cell(row, po_date_column).value
                po_date = datetime.datetime.strptime(po_date, "%Y-%m-%d").date()
                po_number = sheet.cell(row, po_number_column).value
                sku = sheet.cell(row, sku_column).value
                qty = sheet.cell(row, qty_column).value
                transaction_amount = sheet.cell(row, transaction_amount_column).value
                skuPurchaseOrder = SkuPurchaseOrder.objects.filter(sku = sku, po_number = po_number).first()
                if skuPurchaseOrder:
                   skuPurchaseOrder.date = po_date
                   skuPurchaseOrder.qty = qty
                   skuPurchaseOrder.transaction_amount = transaction_amount
                else:
                    skuPurchaseOrder = SkuPurchaseOrder(sku = sku, po_number = po_number ,\
                                                        date = po_date, qty = qty, transaction_amount = transaction_amount)
                skuPurchaseOrder.save()
        else:
            return HttpResponse('请检查上传的文件格式是否正确')
    template = loader.get_template('update_paid_purchase_order.html')
    context = {
        'form':UploadFileForm()
    }

    return HttpResponse(template.render(context, request))

def notFound2Zero(k, dict, *args):
    if args:
        if k in dict:
            return dict[k][args[0]]
        return 0.0
    else:
        if k in dict:
            return dict[k]
        return 0.0

@login_required
def update_sales_transaction(request):
    template = loader.get_template('update_sales_transaction.html')
    if request.method == 'POST':
        form = UploadTransactionAdCurrencyForm(request.POST, request.FILES)
        if form.is_valid():
            START_ROW = 2
            # PRODUCT_AD_COLUMN_NUM_DICT = {'sku':7, 'value': 13}
            CHECK_KEYWORD = {'sku': 'Advertised SKU', 'value': 'Spend'}
            CURRENCY_KEYWORD = {'US':'usd2cny', 'CA': 'cad2cny', 'AU': 'aud2cny' \
                                , 'AE': 'aed2cny', 'GB': 'gbp2cny' , 'DE': 'eur2cny' \
                                , 'FR': 'eur2cny', 'IT': 'eur2cny', 'ES': 'eur2cny' \
                                , 'SG': 'sgd2cny'}
            adCosts = {}
            if 'file2' in request.FILES:
                fileInMemory = request.FILES['file2'].read()
                filePath = BytesIO(fileInMemory)
                adCosts = readSkuValue(filePath, CHECK_KEYWORD, START_ROW)
            if 'file3' in request.FILES:
                fileInMemory = request.FILES['file3'].read()
                filePath = BytesIO(fileInMemory)
                brandAdCosts = readSkuValue(filePath, {'sku': 'Campaign Name', 'value': 'Spend'}, 2)
                totalBrandAdCost = sum([v for k,v in brandAdCosts.items()])
            else:
                totalBrandAdCost = 0.0
            country = 'US'
            file2InMemory = request.FILES['file1'].read()
            sales = readTransaction(file2InMemory)
            totalQuantity = sum([v['quantity'] for k, v in sales.items() if k])
            fixedCost = readFixedCost(file2InMemory)
            currencyRatesDict, currencyRateUpdateTime = readCurrencyRate(BytesIO(request.FILES['file4'].read()))
            currencyRate = currencyRatesDict[CURRENCY_KEYWORD[country]]
            skus = list(set(list(sales.keys()) + list(adCosts.keys())))
            if "" in skus:
                skus.remove("")
            skus.sort()
            start_date = datetime.datetime.strptime(request.POST['start_date'], "%Y-%m-%d").date()
            end_date = datetime.datetime.strptime(request.POST['end_date'], "%Y-%m-%d").date()
            for sku in skus:
                adCost = notFound2Zero(sku,adCosts) * currencyRate
                quantity = notFound2Zero(sku,sales,'quantity')
                brandAdCost = 0.0
                fixedCostPerSKU = 0.0
                if totalQuantity:
                    brandAdCost = float(quantity) / float(totalQuantity) * totalBrandAdCost
                    fixedCostPerSKU = float(quantity) / float(totalQuantity) * fixedCost
                adCost += brandAdCost
                amazonFee = -1.0 * (notFound2Zero(sku,sales,'AmazonFee')  + fixedCostPerSKU)
                skuManagedBySalesPerson = SkuManagedBySalesPerson.objects.filter(sku = sku).first()
                profitLossTable = ProfitLossTable.objects.filter(sku = sku, start_date = start_date, end_date = end_date).first()
                if not profitLossTable:
                    profitLossTable = ProfitLossTable(sku = sku, start_date = start_date, end_date = end_date, status = ('T', '交易订单录入'))
                profitLossTable.sales_amount = notFound2Zero(sku,sales,'sales')
                profitLossTable.sales_quantity = quantity
                profitLossTable.amazon_fee = amazonFee
                profitLossTable.ad_fee = adCost
                profitLossTable.country = (country, country)
                profitLossTable.currency_rate = currencyRate
                if skuManagedBySalesPerson:
                    sales_person_name = skuManagedBySalesPerson.sales_person_name
                    profitLossTable.sales_person_name = sales_person_name
                profitLossTable.save()
        else:
            context = {
                'form': form
            }
            return HttpResponse(template.render(context, request))
    context = {
        'form':UploadTransactionAdCurrencyForm()
    }
    return HttpResponse(template.render(context, request))

@login_required
def confirm_po_head_shipping(request):
    template = loader.get_template('confirm_po_head_shipping.html')
    profitLossTable = ProfitLossTable.objects.filter(status = ('T', '交易订单录入')).order_by('-end_date').first()
    no_need_to_confirm = False
    if profitLossTable:
        skuProfitLossTableDict = {}
        skuHeadShippingCostDict = {}
        total_po_amount = 0.0
        start_date = profitLossTable.start_date
        end_date = profitLossTable.end_date
        # 先统计头程运费
        fbaShipmentCosts = FbaShipmentCost.objects.filter(date__gte=start_date,date__lte=end_date)
        if fbaShipmentCosts.count():
            for fbaShipmentCost in fbaShipmentCosts:
                shipment_id = fbaShipmentCost.shipment_id
                cost_per_kg = fbaShipmentCost.cost_per_kg
                fbaShipments = FbaShipment.objects.filter(shipment_id = shipment_id)
                for fbaShipment in fbaShipments:
                    for shippedSkuQty in fbaShipment.shipped_sku_qties.all():
                        sku = shippedSkuQty.sku
                        qty = shippedSkuQty.qty
                        skuWeight = SkuWeight.objects.filter(sku=sku).first()
                        if skuWeight.heavy_or_light:
                            weight = skuWeight.real_weight
                        elif fbaShipmentCost.shipway == ('E', '快递'):
                            weight = skuWeight.converted_weight * 6.0 / 5.0
                        else:
                            weight = skuWeight.converted_weight
                        head_shipping_cost = weight * qty * cost_per_kg
                        if sku in skuHeadShippingCostDict:
                            skuHeadShippingCostDict[sku] += head_shipping_cost
                        else:
                            skuHeadShippingCostDict[sku] = head_shipping_cost
        skus_in_profitLossTables = set()
        for profitLossTable in ProfitLossTable.objects.filter(status = ('T', '交易订单录入')).order_by('-end_date').all():
            sku = profitLossTable.sku
            skus_in_profitLossTables.add(sku)
            skuProfitLossTableDict[sku] = {'profitLossTable': profitLossTable, 'total_sku_po_amount':0.0, 'total_sku_head_shipping_cost':0.0}

            skuPurchaseOrders = SkuPurchaseOrder.objects.filter(sku = sku, date__gte=start_date,date__lte=end_date)
            total_sku_po_amount = 0.0
            if skuPurchaseOrders.count():
                total_sku_po_amount =  skuPurchaseOrders.aggregate(Sum('transaction_amount'))['transaction_amount__sum']
            skuProfitLossTableDict[sku]['total_sku_po_amount'] = total_sku_po_amount
            total_po_amount += total_sku_po_amount
            if sku in skuHeadShippingCostDict:
                skuProfitLossTableDict[sku]['total_sku_head_shipping_cost'] = skuHeadShippingCostDict[sku]
        for skuPurchaseOrder in SkuPurchaseOrder.objects.filter(date__gte=start_date,date__lte=end_date).all():
            if skuPurchaseOrder.sku not in skus_in_profitLossTables:
                sku = skuPurchaseOrder.sku
                skuProfitLossTableDict[sku] = {'profitLossTable': profitLossTable, 'total_sku_po_amount':0.0, 'total_sku_head_shipping_cost':0.0}
                skuPurchaseOrders = SkuPurchaseOrder.objects.filter(sku = sku, date__gte=start_date,date__lte=end_date)
                total_sku_po_amount =  skuPurchaseOrders.aggregate(Sum('transaction_amount'))['transaction_amount__sum']
                skuProfitLossTableDict[sku]['total_sku_po_amount'] = total_sku_po_amount

                total_po_amount += total_sku_po_amount
                if sku in skuHeadShippingCostDict:
                    skuProfitLossTableDict[sku]['total_sku_head_shipping_cost'] = skuHeadShippingCostDict[sku]
                # skuManagedBySalesPerson = SkuManagedBySalesPerson.objects.filter(sku = sku).first()
                # profitLossTable = ProfitLossTable.objects.filter(sku = sku, start_date = start_date, end_date = end_date).first()
                # if not profitLossTable:
                #     profitLossTable = ProfitLossTable(sku = sku, start_date = start_date, end_date = end_date, status = ('T', '交易订单录入'))
                # profitLossTable.country = ('US','US')
                # if skuManagedBySalesPerson:
                #     sales_person = skuManagedBySalesPerson.sales_person
                #     profitLossTable.sales_person = sales_person
                #     profitLossTable.sales_person_bonus_fee_percent = skuManagedBySalesPerson.bonus_percent
                # profitLossTable.save()

    else:
        no_need_to_confirm = True
    if request.method == 'POST':
        form = ConfirmPoHeadShippingForm(request.POST)
        if form.is_valid():
            if request.POST['yes_or_no'] == 'yes':
                # 对有销售数据的损益表更新购货信息以及头程运费信息
                for profitLossTable in ProfitLossTable.objects.filter(status = ('T', '交易订单录入')).order_by('-end_date').all():
                    sku = profitLossTable.sku
                    profitLossTable.product_purchasing_fee = skuProfitLossTableDict[sku]['total_sku_po_amount']
                    if sku in skuHeadShippingCostDict:
                        profitLossTable.head_shipping_fee = skuHeadShippingCostDict[sku]
                    profitLossTable.status = ('P', '购货和头程运费确认')
                    profitLossTable.save()
                #将没有销售数据但是有购货订单的损益表建立了
                sku_in_po_not_in_profitLossTables = set()

                for skuPurchaseOrder in SkuPurchaseOrder.objects.filter(date__gte=start_date,date__lte=end_date).all():
                    if skuPurchaseOrder.sku not in skus_in_profitLossTables:
                        sku = skuPurchaseOrder.sku
                        sku_in_po_not_in_profitLossTables.add(sku)
                        skuPurchaseOrders = SkuPurchaseOrder.objects.filter(sku = sku, date__gte=start_date,date__lte=end_date)
                        total_sku_po_amount =  skuPurchaseOrders.aggregate(Sum('transaction_amount'))['transaction_amount__sum']
                        skuManagedBySalesPerson = SkuManagedBySalesPerson.objects.filter(sku = sku).first()
                        profitLossTable = ProfitLossTable.objects.filter(sku = sku, start_date = start_date, end_date = end_date).first()
                        if not profitLossTable:
                            profitLossTable = ProfitLossTable(sku = sku, start_date = start_date, end_date = end_date)
                        profitLossTable.country = ('US','US')
                        profitLossTable.status = ('P', '购货和头程运费确认')
                        if skuManagedBySalesPerson:
                            sales_person_name = skuManagedBySalesPerson.sales_person_name
                            profitLossTable.sales_person_name = sales_person_name
                        profitLossTable.product_purchasing_fee = total_sku_po_amount
                        if sku in skuHeadShippingCostDict:
                            profitLossTable.head_shipping_fee = skuHeadShippingCostDict[sku]
                        profitLossTable.save()
                #将没有销售数据没有有购货订单，但是有头程运费的的损益表建立了
                for sku, sku_head_shipping_cost in skuHeadShippingCostDict.items():
                    if sku not in sku_in_po_not_in_profitLossTables and sku not in skus_in_profitLossTables:
                        skuManagedBySalesPerson = SkuManagedBySalesPerson.objects.filter(sku = sku).first()
                        profitLossTable = ProfitLossTable.objects.filter(sku = sku, start_date = start_date, end_date = end_date).first()
                        if not profitLossTable:
                            profitLossTable = ProfitLossTable(sku = sku, start_date = start_date, end_date = end_date)
                        profitLossTable.country = ('US','US')
                        profitLossTable.status = ('P', '购货和头程运费确认')
                        if skuManagedBySalesPerson:
                            sales_person_name = skuManagedBySalesPerson.sales_person_name
                            profitLossTable.sales_person_name = sales_person_name
                        profitLossTable.head_shipping_fee = skuHeadShippingCostDict[sku]
                        profitLossTable.save()
                no_need_to_confirm = True
        else:
            context = {
                'start_date' : start_date ,\
                'end_date' : end_date ,\
                'po_amount':total_po_amount ,\
                'head_shipping_cost': sum(skuHeadShippingCostDict.values()),\
                'form': form
            }
            return HttpResponse(template.render(context, request))
    if no_need_to_confirm:
        template = loader.get_template('no_need_confirm_po_head_shipping.html')
        context = {}
    else:
        context = {
            'start_date' : start_date ,\
            'end_date' : end_date ,\
            'po_amount':total_po_amount ,\
            'head_shipping_cost': sum(skuHeadShippingCostDict.values()),\
            'form':ConfirmPoHeadShippingForm()
        }
    return HttpResponse(template.render(context, request))

@login_required
def get_sku_pl_al_table(request,sku):
    if request.user.groups.filter(name='head_of_sales').exists() or \
      SkuManagedBySalesPerson.objects.filter(sku = sku, sales_person_name = request.user.username).count() or \
      SkuManagedBySalesPerson.objects.filter(sku = sku, sales_assistant_name = request.user.username).count() or \
      SkuContributor.objects.filter(sku = sku, proposer_name = request.user.username).count() or \
      SkuContributor.objects.filter(sku = sku, designer_name = request.user.username).count():

        template = loader.get_template('get_sku_pl_al_table.html')
        profitLossTable = ProfitLossTable.objects.filter(sku = sku, status = ('P', '购货和头程运费确认')).order_by('-end_date').first()
        if profitLossTable:
            start_date = profitLossTable.start_date
            end_date = profitLossTable.end_date
        #先判断用户是哪类？1 运营，无该产品的其他贡献 2:运营，且有其他产品的贡献 3 非运营，但是有其他产品的贡献
        if SkuManagedBySalesPerson.objects.filter(sku = sku, sales_person_name = request.user.username).count() or \
            SkuManagedBySalesPerson.objects.filter(sku = sku, sales_assistant_name = request.user.username).count() and \
            SkuContributor.objects.filter(sku = sku, proposer_name = request.user.username).count() == 0 and \
            SkuContributor.objects.filter(sku = sku, designer_name = request.user.username).count() == 0:
            skuAssetLiabilityTables = SkuAssetLiabilityTable.objects.filter(sku = sku, for_sales = True, date__lt = end_date).order_by('-date').all()
            if skuAssetLiabilityTables.count() > 1:
                if skuAssetLiabilityTables[0].date == skuAssetLiabilityTables[1].date:
                    skuAssetLiabilityTableForSalesPerson = SkuAssetLiabilityTable.objects.filter(sku = sku, for_sales = True, date = skuAssetLiabilityTables[0].date, initial = True).order_by('-date').first()
                else:
                    skuAssetLiabilityTableForSalesPerson = SkuAssetLiabilityTable.objects.filter(sku = sku, for_sales = True, date__lt = end_date).order_by('-date').first()
            elif skuAssetLiabilityTables.count() == 1:
                skuAssetLiabilityTableForSalesPerson = SkuAssetLiabilityTable.objects.filter(sku = sku, for_sales = True, date__lt = end_date).order_by('-date').first()
            else:
                return HttpResponse('%s 还没有初始化资产负债表' %sku)
            user_type = 'SalesOnly'
        elif (SkuManagedBySalesPerson.objects.filter(sku = sku, sales_person_name = request.user.username).count() or \
            SkuManagedBySalesPerson.objects.filter(sku = sku, sales_assistant_name = request.user.username).count()) and \
            (SkuContributor.objects.filter(sku = sku, proposer_name = request.user.username).count() or \
            SkuContributor.objects.filter(sku = sku, designer_name = request.user.username).count()):
            skuAssetLiabilityTables = SkuAssetLiabilityTable.objects.filter(sku = sku, for_sales = True, date__lt = end_date).order_by('-date').all()
            if skuAssetLiabilityTables.count() > 1:
                if skuAssetLiabilityTables[0].date == skuAssetLiabilityTables[1].date:
                    skuAssetLiabilityTableForSalesPerson = SkuAssetLiabilityTable.objects.filter(sku = sku, for_sales = True, date = skuAssetLiabilityTables[0].date, initial = True).order_by('-date').first()
                else:
                    skuAssetLiabilityTableForSalesPerson = SkuAssetLiabilityTable.objects.filter(sku = sku, for_sales = True, date__lt = end_date).order_by('-date').first()
            elif skuAssetLiabilityTables.count() == 1:
                skuAssetLiabilityTableForSalesPerson = SkuAssetLiabilityTable.objects.filter(sku = sku, for_sales = True, date__lt = end_date).order_by('-date').first()
            else:
                return HttpResponse('%s 还没有初始化资产负债表' %sku)
            skuAssetLiabilityTable = SkuAssetLiabilityTable.objects.filter(sku = sku, for_sales = False, date__lt = end_date).order_by('-date').first()
            user_type = 'SalesAndProductContributor'
        elif (SkuManagedBySalesPerson.objects.filter(sku = sku, sales_person_name = request.user.username).count() == 0 and \
            SkuManagedBySalesPerson.objects.filter(sku = sku, sales_assistant_name = request.user.username).count() == 0) and \
            (SkuContributor.objects.filter(sku = sku, proposer_name = request.user.username).count() or \
            SkuContributor.objects.filter(sku = sku, designer_name = request.user.username).count()):
            skuAssetLiabilityTable = SkuAssetLiabilityTable.objects.filter(sku = sku, for_sales = False, date__lt = end_date).order_by('-date').first()
            user_type = 'ProductContributorOnly'

        skuAssetLiabilityTable = SkuAssetLiabilityTable.objects.filter(sku = sku, date__lt = end_date).order_by('-date').first()
        cash_before = skuAssetLiabilityTable.cash_amount
        po_head_shipping_cost = profitLossTable.product_purchasing_fee + profitLossTable.head_shipping_fee
        profit = (profitLossTable.sales_amount - profitLossTable.amazon_fee - profitLossTable.ad_fee) * profitLossTable.currency_rate - po_head_shipping_cost
        sales_amount_in_cny = profitLossTable.sales_amount * profitLossTable.currency_rate
        amazon_fee_in_cny = profitLossTable.amazon_fee * profitLossTable.currency_rate
        ad_fee_in_cny = profitLossTable.ad_fee * profitLossTable.currency_rate
        cash_after = cash_before + profit
        if cash_after < 0:
            liabilities_after = skuAssetLiabilityTable.liabilities - cash_after
            cash_after = 0.0
        else:
            if cash_after >= skuAssetLiabilityTable.liabilities:
                liabilities_after = 0.0
                cash_after = cash_after - skuAssetLiabilityTable.liabilities
            else:
                liabilities_after = skuAssetLiabilityTable.liabilities - cash_after
                cash_after = 0.0
        historyTodayProductSales = HistoryTodayProductSales.objects.filter(product__sku =sku).order_by('-date').first()
        sold_qty_average_7d = historyTodayProductSales.sold_qty_average_7d
        fba_shenzhen_inventory = historyTodayProductSales.fba_inventory.total_unit + check_shenzhen_inventory(sku)
        sold_out_day_number = int(float(fba_shenzhen_inventory) / sold_qty_average_7d)
        receive_money_per_unit_sold = (sales_amount_in_cny-amazon_fee_in_cny-ad_fee_in_cny) / float(profitLossTable.sales_quantity)
        gross_profit_sold_out = int(receive_money_per_unit_sold*fba_shenzhen_inventory - liabilities_after)
        liabilities_zero_day_number = 0
        roi = 0.0
        roi_annual = 0.0
        if liabilities_after > 0:
            liabilities_zero_day_number = int(liabilities_after/receive_money_per_unit_sold/sold_qty_average_7d)
            roi = gross_profit_sold_out / liabilities_after
            roi_annual = (1.0 + roi) ** (float(365) / float(sold_out_day_number)) - 1
            roi_commission_deducted = roi*0.8
            roi_commission_deducted_annual = float(365) / float(sold_out_day_number) * gross_profit_sold_out * 0.8 /liabilities_after #(1.0 + roi_commission_deducted) ** (float(365) / float(sold_out_day_number)) - 1
        context = {
            'sku':sku,
            'start_date': start_date,
            'end_date': end_date ,
            'liabilities_before': skuAssetLiabilityTable.liabilities,
            'liabilities': liabilities_after,
            'cash': cash_after ,
            'net_asset': cash_after - liabilities_after ,
            'profit': profit,
            'profitLossTable': profitLossTable,
            'sales_amount_in_cny':sales_amount_in_cny,
            'amazon_fee_in_cny':amazon_fee_in_cny,
            'ad_fee_in_cny':ad_fee_in_cny,
            'receive_money_per_unit_sold': receive_money_per_unit_sold,
            'fba_shenzhen_inventory' : fba_shenzhen_inventory ,
            'sold_qty_average_7d': sold_qty_average_7d,
            'sold_out_day_number': sold_out_day_number,
            'liabilities_zero_day_number':liabilities_zero_day_number,
            'gross_profit_sold_out': gross_profit_sold_out,
            'gross_profit_sold_out_5_percent': gross_profit_sold_out * 0.05,
            'roi': int(roi * 100),
            'roi_annual':int(roi_annual * 100) ,
            'roi_commission_deducted': int(roi_commission_deducted * 100),
            'roi_commission_deducted_annual':int(roi_commission_deducted_annual * 100),
        }

        return HttpResponse(template.render(context, request))
    else:
        return HttpResponse('你没有权限查看这个产品的财务报表')

def get_sku_list_pl_al_table(sku_list):
    profitLossTable = ProfitLossTable.objects.filter(status = ('P', '购货和头程运费确认'), sku__in = sku_list).order_by('-end_date').first()
    start_date = profitLossTable.start_date
    end_date = profitLossTable.end_date
    skuAssetLiabilityTables = SkuAssetLiabilityTable.objects.filter(for_sales=False, date__gte=start_date, sku__in = sku_list).all()
    liabilities_before_total = 0.0
    liabilities_after_total = 0.0
    cash_after_total = 0.0
    profit_total = 0.0
    sales_amount_in_cny_total =0.0
    amazon_fee_in_cny_total = 0.0
    ad_fee_in_cny_total = 0.0
    receive_money_total = 0.0
    fba_shenzhen_inventory_total = 0
    sold_qty_average_7d_total = 0.0
    liabilities_zero_day_number_total = 0
    gross_profit_sold_out_total = 0.0
    sold_qty_total = 0.0
    po_cost_total = 0.0
    head_shipping_cost_total = 0.0
    for skuAssetLiabilityTable in skuAssetLiabilityTables:
        liabilities_before_total += skuAssetLiabilityTable.liabilities
        cash_before = skuAssetLiabilityTable.cash_amount
        sku = skuAssetLiabilityTable.sku
        profitLossTable = ProfitLossTable.objects.filter(sku = sku, status = ('P', '购货和头程运费确认')).order_by('-end_date').first()
        if profitLossTable:
            sold_qty_total += profitLossTable.sales_quantity
            po_cost_total += profitLossTable.product_purchasing_fee
            head_shipping_cost_total += profitLossTable.head_shipping_fee
            po_head_shipping_cost = profitLossTable.product_purchasing_fee + profitLossTable.head_shipping_fee
            profit = (profitLossTable.sales_amount - profitLossTable.amazon_fee - profitLossTable.ad_fee) * profitLossTable.currency_rate - po_head_shipping_cost
            sales_amount_in_cny = profitLossTable.sales_amount * profitLossTable.currency_rate
            amazon_fee_in_cny = profitLossTable.amazon_fee * profitLossTable.currency_rate
            ad_fee_in_cny = profitLossTable.ad_fee * profitLossTable.currency_rate
            cash_after = cash_before + profit
            if cash_after < 0:
                liabilities_after = skuAssetLiabilityTable.liabilities - cash_after
                cash_after = 0.0
            else:
                if cash_after >= skuAssetLiabilityTable.liabilities:
                    liabilities_after = 0.0
                    cash_after = cash_after - skuAssetLiabilityTable.liabilities
                else:
                    liabilities_after = skuAssetLiabilityTable.liabilities - cash_after
                    cash_after = 0.0
            historyTodayProductSale = HistoryTodayProductSales.objects.filter(product__sku =sku).order_by('-date').first()
            if historyTodayProductSale:
                sold_qty_average_7d = historyTodayProductSale.sold_qty_average_7d
                fba_shenzhen_inventory = historyTodayProductSale.fba_inventory.total_unit + check_shenzhen_inventory(sku)
            else:
                sold_qty_average_7d = 0.0
                fba_shenzhen_inventory = 0.0 + check_shenzhen_inventory(sku)
            # sold_out_day_number = int(float(fba_shenzhen_inventory) / sold_qty_average_7d)

            if profitLossTable.sales_quantity:
                receive_money_by_this_sku = (sales_amount_in_cny-amazon_fee_in_cny-ad_fee_in_cny)
                receive_money_per_unit_sold = (sales_amount_in_cny-amazon_fee_in_cny-ad_fee_in_cny) / float(profitLossTable.sales_quantity)
            else:
                receive_money_by_this_sku = (sales_amount_in_cny-amazon_fee_in_cny-ad_fee_in_cny)
                receive_money_per_unit_sold = (sales_amount_in_cny-amazon_fee_in_cny-ad_fee_in_cny) / 100.0
            receive_money_total += receive_money_by_this_sku
            gross_profit_sold_out = int(receive_money_per_unit_sold*fba_shenzhen_inventory - liabilities_after)

            liabilities_after_total += liabilities_after
            cash_after_total += cash_after
            profit_total += profit
            sales_amount_in_cny_total += sales_amount_in_cny
            amazon_fee_in_cny_total += amazon_fee_in_cny
            ad_fee_in_cny_total += ad_fee_in_cny
            fba_shenzhen_inventory_total += fba_shenzhen_inventory
            sold_qty_average_7d_total += sold_qty_average_7d
            gross_profit_sold_out_total += gross_profit_sold_out
        else:
            liabilities_after_total += skuAssetLiabilityTable.liabilities
            cash_after_total += cash_before
    sold_out_day_number_total = int(float(fba_shenzhen_inventory_total) / sold_qty_average_7d_total)
    receive_money_per_unit_sold_average = receive_money_total / sold_qty_total
    if liabilities_after_total > 0:
        liabilities_zero_day_number_total = int(liabilities_after_total/receive_money_per_unit_sold_average/sold_qty_average_7d_total)
    gross_profit_sold_out_total = int(receive_money_per_unit_sold_average*fba_shenzhen_inventory_total - liabilities_after_total)

    context = {
        # 'sku':sku,
        'start_date': start_date,
        'end_date': end_date ,
        'liabilities_before': liabilities_before_total,
        'liabilities': liabilities_after_total,
        'cash': cash_after_total ,
        'net_asset': cash_after_total - liabilities_after_total ,
        'profit': profit_total,
        'sold_qty_total': sold_qty_total,
        'po_cost_total': po_cost_total,
        'head_shipping_cost_total': head_shipping_cost_total,
        # 'profitLossTable': profitLossTable,
        'sales_amount_in_cny':sales_amount_in_cny_total,
        'amazon_fee_in_cny':amazon_fee_in_cny_total,
        'ad_fee_in_cny':ad_fee_in_cny_total,
        'receive_money_per_unit_sold': receive_money_per_unit_sold_average,
        'fba_shenzhen_inventory' : fba_shenzhen_inventory_total ,
        'sold_qty_average_7d': sold_qty_average_7d_total,
        'sold_out_day_number': sold_out_day_number_total,
        'liabilities_zero_day_number':liabilities_zero_day_number_total,
        'gross_profit_sold_out': gross_profit_sold_out_total,
        'gross_profit_sold_out_5_percent': gross_profit_sold_out_total * 0.05
    }
    return context

@login_required
def finance_dashboard(request):
    template = loader.get_template('finance_dashboard.html')
    if request.user.groups.filter(name='sales').exists() or \
        request.user.groups.filter(name='designer').exists() or \
        request.user.groups.filter(name='proposer').exists():
        roles_sku_list_dict = {}
        if request.user.groups.filter(name='head_of_sales').exists():
            profitLossTable = ProfitLossTable.objects.filter(status = ('P', '购货和头程运费确认')).order_by('-end_date').first()
            start_date = profitLossTable.start_date
            end_date = profitLossTable.end_date
            skuAssetLiabilityTables = SkuAssetLiabilityTable.objects.filter(for_sales=False, date__gte=start_date).all()
            sku_list_as_head_of_sales = [skuAssetLiabilityTables.sku for skuAssetLiabilityTables in skuAssetLiabilityTables]
        sku_list_as_sales_person = [skuManagedBySalesPerson.sku for skuManagedBySalesPerson in SkuManagedBySalesPerson.objects.filter(sales_person_name=request.user.username).all()]
        sku_list_as_sales_assistant = [skuManagedBySalesPerson.sku for skuManagedBySalesPerson in SkuManagedBySalesPerson.objects.filter(sales_assistant_name=request.user.username).all()]
        sku_list_as_proposer = [skuContributor.sku for skuContributor in SkuContributor.objects.filter(proposer_name = request.user.username).all()]
        sku_list_as_designer = [skuContributor.sku for skuContributor in SkuContributor.objects.filter(designer_name = request.user.username).all()]
        if(len(sku_list_as_head_of_sales)):
            roles_sku_list_dict['head_of_sales'] = sku_list_as_head_of_sales
        if(len(sku_list_as_sales_person)):
            roles_sku_list_dict['sales_person'] = sku_list_as_sales_person
        if(len(sku_list_as_sales_assistant)):
            roles_sku_list_dict['sales_assistant'] = sku_list_as_sales_assistant
        if(len(sku_list_as_proposer)):
            roles_sku_list_dict['proposer'] = sku_list_as_proposer
        if(len(sku_list_as_designer)):
            roles_sku_list_dict['designer'] = sku_list_as_designer
        if request.method == 'POST':
            role = request.POST['role']
        else:
            role = next(iter(roles_sku_list_dict))
        context = get_sku_list_pl_al_table(roles_sku_list_dict[role])
        ROLE_NAMES_DICT = {'head_of_sales': '销售总监', 'sales_person': '销售', 'sales_assistant': '销售助理', 'proposer': '选品师', 'designer': '设计师'}
        context['role_name'] = ROLE_NAMES_DICT[role]
        context['roles_sku_list_dict'] = roles_sku_list_dict
    return HttpResponse(template.render(context, request))

@user_passes_test(lambda u: u.groups.filter(name__in = ['logistics', 'supplier']).exists())
def get_ongoing_production_plan_progresses(request):
    template = loader.get_template('get_ongoing_production_plan_progresses.html')
    all_ongoing_production_plan_progresses = ProductionPlanProgress.objects.filter(ongoing = True).order_by('deadline_date')
    total_qty = all_ongoing_production_plan_progresses.aggregate(Sum('qty'))['qty__sum']
    context = {
        'production_plan_progresses': all_ongoing_production_plan_progresses ,\
        'total_qty' : total_qty
    }
    return HttpResponse(template.render(context, request))

@user_passes_test(lambda u: u.groups.filter(name__in = ['logistics', 'supplier']).exists())
def get_ongoing_production_plan_progresses_by_sku(request):
    template = loader.get_template('get_ongoing_production_plan_progresses.html')
    sku = request.GET['sku']
    all_ongoing_production_plan_progresses = ProductionPlanProgress.objects.filter(ongoing = True,sku__icontains = sku).order_by('deadline_date')
    total_qty = all_ongoing_production_plan_progresses.aggregate(Sum('qty'))['qty__sum']
    context = {
        'production_plan_progresses': all_ongoing_production_plan_progresses ,\
        'total_qty' : total_qty
    }
    return HttpResponse(template.render(context, request))

def remove_empty_elements(lst, nt_lst):
    new_lst = []
    new_nt_lst = []
    for i in range(len(lst)):
        if lst[i] != []:
            new_lst.append(lst[i])
            new_nt_lst.append(nt_lst[i])
    return new_lst, new_nt_lst

@user_passes_test(lambda u: u.groups.filter(name__in = ['logistics', 'supplier']).exists())
def production_plan_today(request):
    template = loader.get_template('production_plan_today.html')
    production_combined_stage_combined_index_dict = {'买料':0, '开料':1, '贴合':2, '塞磁铁':3, '散味':4, '油边':5, '车缝':6, '质检':7}
    production_combined_stage_combined_name_list = ['买料', '开料', '贴合', '塞磁铁', '散味', '油边', '车缝', '质检']
    production_plan_today_list = [[],[],[],[],[],[],[],[]]
    for ppp in ProductionPlanProgress.objects.filter(ongoing = True, qty__gt = 0).order_by('deadline_date').all():
        if ppp.soonest_finishing_date > ppp.deadline_date:
            if ppp.current_stage_id is None:
                for production_stage in ppp.production_stages.all().order_by('order_number'):
                    should_break = False
                    for k, v in production_combined_stage_combined_index_dict.items():
                        if production_stage.name.startswith(k):
                            production_plan_today_list[v].append(ppp)
                            should_break = True
                            break
                    if should_break:
                        break
            else:
                current_production_stage = ProductionStage.objects.get(id = ppp.current_stage_id)
                # 判断是否已经超过该步骤的时间
                if current_production_stage.start_date_actually + datetime.timedelta(days = current_production_stage.duration_days) < datetime.date.today():
                    if current_production_stage.name != '质检':
                        next_production_stage = ppp.production_stages.filter(order_number__gt = current_production_stage.order_number).order_by('order_number').first()
                        if next_production_stage.start_date_estimated is not None:
                            if next_production_stage.start_date_estimated < datetime.date.today():
                                if next_production_stage.name.startswith('过渡期') or next_production_stage.name.startswith('等待期'):
                                    next_production_stage_order_number = next_production_stage.order_number
                                    next_production_stage = ppp.production_stages.filter(order_number__gt = next_production_stage_order_number).order_by('order_number').first()
                                for k, v in production_combined_stage_combined_index_dict.items():
                                    if next_production_stage.name.startswith(k):
                                        production_plan_today_list[v].append(ppp)
                                        break
                        else:
                            if next_production_stage.name.startswith('过渡期') or next_production_stage.name.startswith('等待期'):
                                next_production_stage_order_number = next_production_stage.order_number
                                next_production_stage = ppp.production_stages.filter(order_number__gt = next_production_stage_order_number).order_by('order_number').first()
                                if datetime.timedelta(days = ppp.production_stages.filter(order_number__gt = next_production_stage_order_number).aggregate(Sum('duration_days'))['duration_days__sum']) + datetime.date.today() < ppp.deadline_date:
                                    continue
                            for k, v in production_combined_stage_combined_index_dict.items():
                                if next_production_stage.name.startswith(k):
                                    production_plan_today_list[v].append(ppp)
                                    break
        else:
            #最快完工在截止日期内
            if ppp.current_stage_id is not None:
                current_production_stage = ProductionStage.objects.get(id = ppp.current_stage_id)
                if current_production_stage.name != '质检':
                    next_estimated_production_stage = ppp.production_stages.filter(order_number__gt = current_production_stage.order_number, start_date_estimated__isnull=False)
                    if next_estimated_production_stage.count():
                        next_estimated_production_stage = next_estimated_production_stage.order_by('-order_number').first()
                        # 但是下一个预计工序日期已经超过今天的
                        if next_estimated_production_stage.start_date_estimated < datetime.date.today():
                            if next_estimated_production_stage.name.startswith('过渡期') or next_estimated_production_stage.name.startswith('等待期'):
                                next_estimated_production_stage_order_number = next_estimated_production_stage.order_number
                                next_estimated_production_stage = ppp.production_stages.filter(order_number__gt = next_estimated_production_stage_order_number).order_by('order_number').first()
                            for k, v in production_combined_stage_combined_index_dict.items():
                                if next_estimated_production_stage.name.startswith(k):
                                    production_plan_today_list[v].append(ppp)
                                    break
                        # 下一个预计工序日期虽然没有超过今天，但是当前工序已经滞后，需要尽早启动当前工序的下一个邻近工序
                        else:
                            # 当前工序实际启动日期 + 当前工序总共所花时长 : 理论上下一个工序起始日期
                            # estimated_left_days： 下一个预计工序启动日期 - 预计当天工序的下一个工序起始日期 ： 剩下的天数
                            # theoritical_min_required_days： 下一个工序到下一个预计工序前所需最小理论天数
                            # 如果 estimated_left_days > theoritical_min_required_days：，而且今天到下一个预计工序启动时间的天数也大于 theoritical_min_required_days，则安全，不用提醒
                            # 如果estimated_left_days > theoritical_min_required_days，但是今天到下一个预计工序启动时间天数已经小于theoritical_min_required_days，则需要提醒
                            # 如果 estimated_left_days < theoritical_min_required_days，则也需要提醒提前进入下一个工序

                            estimated_end_date_of_current_pp = current_production_stage.start_date_actually + datetime.timedelta(days = current_production_stage.duration_days)
                            estimated_left_days = (next_estimated_production_stage.start_date_estimated - estimated_end_date_of_current_pp).days
                            if ppp.production_stages.filter(Q(order_number__gt = current_production_stage.order_number) & Q(order_number__lt = next_estimated_production_stage.order_number)).count():
                                theoritical_min_required_days = ppp.production_stages.filter(Q(order_number__gt = current_production_stage.order_number) & Q(order_number__lt = next_estimated_production_stage.order_number)).aggregate(Sum('duration_days'))['duration_days__sum']
                            else:
                                theoritical_min_required_days = 0
                            if (estimated_left_days > theoritical_min_required_days and (next_estimated_production_stage.start_date_estimated - datetime.date.today()).days < theoritical_min_required_days) or estimated_left_days < theoritical_min_required_days:
                                ps_next_current_ps = ppp.production_stages.filter(order_number__gt = current_production_stage.order_number).order_by('order_number').first()
                                for k, v in production_combined_stage_combined_index_dict.items():
                                    if ps_next_current_ps.name.startswith(k):
                                        production_plan_today_list[v].append(ppp)
                                        break
    production_plan_today_list, production_combined_stage_combined_name_list = remove_empty_elements(production_plan_today_list, production_combined_stage_combined_name_list)
    context = {
        'production_combined_stage_combined_name_list' : production_combined_stage_combined_name_list ,\
        'production_plan_today_list': production_plan_today_list \
    }
    return HttpResponse(template.render(context, request))

@user_passes_test(lambda u: u.groups.filter(name__in = ['logistics', 'accountant']).exists())
def update_purchasing_orders_including_completed(request):
    correction_tip = ''
    not_save_ppp = False
    update_completed_tip = ''
    if request.method == 'POST':
        CLOSE_PLAN_QTY_THRESHOLD = 50
        wb2 = xlrd.open_workbook(file_contents=request.FILES['file'].read())
        sheet = wb2.sheet_by_index(0)
        max_column = sheet.ncols

        for col in range(0,max_column):
            title = sheet.cell(4,col).value
            if title == '未入库数量':
                unreceived_qty_col = col
            elif title == '状态':
                status_col = col
            elif title == '采购订单编号':
                po_number_col = col
            elif title == '预计交货日期':
                deadline_date_col = col
            elif title == '供应商':
                manufacturer_number_col = col

        for row in range(5,sheet.nrows):
            sku_cell_value = sheet.cell(row,0).value
            if sku_cell_value:
                sku = sku_cell_value
                po_number = sheet.cell(row,po_number_col).value
                status = sheet.cell(row,status_col).value
                unreceived_qty = sheet.cell(row, unreceived_qty_col).value
                deadline_date =  datetime.datetime.strptime(sheet.cell(row, deadline_date_col).value, "%Y-%m-%d").date()
                manufacturer_number = sheet.cell(row, manufacturer_number_col).value
                if manufacturer_number in AGILE_MANUFACTURER_NUMBERS:
                    if status != '已关闭':
                        production_plan_progress = ProductionPlanProgress.objects.filter(production_plan_number = po_number)
                        if production_plan_progress.count():
                            created = False
                            production_plan_progress = production_plan_progress.first()
                            production_plan_progress.manufacturer_number = manufacturer_number
                        else:
                            production_plan_progress = ProductionPlanProgress(production_plan_number = po_number, deadline_date = deadline_date, manufacturer_number = manufacturer_number)
                            created = True
                        production_plan_progress.sku = sku
                        if created:
                            if status == '部分入库 ' and unreceived_qty > CLOSE_PLAN_QTY_THRESHOLD:
                                production_plan_progress.qty = unreceived_qty
                            elif (status == '部分入库 ' and unreceived_qty <= CLOSE_PLAN_QTY_THRESHOLD) or status == '已入库':
                                production_plan_progress.ongoing = False
                            else:
                                production_plan_progress.qty = unreceived_qty
                            production_plan_progress.save()

                        else:
                            ppps = ProductionPlanProgress.objects.filter(production_plan_number__startswith=po_number)
                            if ppps.count() <2:
                                if status == '部分入库 ' and unreceived_qty > CLOSE_PLAN_QTY_THRESHOLD:
                                    production_plan_progress.qty = unreceived_qty
                                elif (status == '部分入库 ' and unreceived_qty <= CLOSE_PLAN_QTY_THRESHOLD) or status == '已入库':
                                    production_plan_progress.ongoing = False
                                else:
                                    production_plan_progress.qty = unreceived_qty
                                # production_plan_progress.deadline_date = deadline_date
                            else:
                                if (status == '部分入库 ' and unreceived_qty <= CLOSE_PLAN_QTY_THRESHOLD) or status == '已入库':
                                    production_plan_progress.ongoing = False
                                    for ppp in ppps.all():
                                        ppp.ongoing = False
                                        ppp.save()
                        if production_plan_progress.ongoing == True and production_plan_progress.production_stages.all().count() == 0:
                            sku_production_stage_type_parameters = SkuProductionStageTypeParameter.objects.filter(sku = sku)
                            if sku_production_stage_type_parameters.count():
                                sku_production_stage_type_parameter = sku_production_stage_type_parameters.first()
                                for production_stage in sku_production_stage_type_parameter.production_stages.all():
                                    production_stage.pk = None
                                    if production_stage.name == '等待期(下单到买料)':
                                        production_stage.start_date_actually = datetime.date.today()
                                        production_plan_progress.current_stage_id = production_stage.id
                                    production_stage.save()
                                    production_plan_progress.production_stages.add(production_stage)
                                not_save_ppp = False
                            else:
                                if correction_tip.endswith('的生产流程还未设置，请先设置'):
                                    correction_tip = sku + ' ' + correction_tip
                                else:
                                    correction_tip = '%s的生产流程还未设置，请先设置' %sku
                                not_save_ppp = True
                        if not not_save_ppp:
                            production_plan_progress.save()
        update_completed_tip = '更新完成'
    template = loader.get_template('update_purchasing_orders_including_completed.html')
    context = {
        'correction_tip': correction_tip , \
        'update_completed_tip': update_completed_tip ,\
        'form':UploadFileForm()
    }

    return HttpResponse(template.render(context, request))

@user_passes_test(lambda u: u.groups.filter(name__in = ['logistics', 'accountant']).exists())
def update_sku_production_stage_by_a_reference(request):
    correction_tip = ''
    update_completed_tip = ''
    if request.method == 'POST' : # and request.get_host() == '209.97.151.168':
        form = UploadFileForm(request.POST,request.FILES)
        if form.is_valid():
            fileInMemory = request.FILES['file'].read()
            filePath = BytesIO(fileInMemory)
            wb2 = load_workbook(filePath, data_only=True)
            sheet = wb2.worksheets[0]
            for i_col in range(1, sheet.max_column + 1):
                if sheet.cell(1,i_col).value != UPDATE_SKU_PRODUCTION_STAGE_DEFAULT_TABLE_HEADER[i_col - 1]:
                    correction_tip = '出错啦！第一行标题需要修正，正确的是' + ' '.join(UPDATE_SKU_PRODUCTION_STAGE_DEFAULT_TABLE_HEADER)
                    break;
            if not correction_tip:
                for row in range(2, sheet.max_row + 1):
                    if not sheet.cell(row,1).value:
                        continue
                    sku = str(sheet.cell(row,1).value).split()[0]
                    if not sku:
                        continue
                    reference_sku = sheet.cell(row,2).value
                    sku_pstp_to_update = SkuProductionStageTypeParameter.objects.filter(sku = sku)
                    if sku_pstp_to_update.count():
                        for i_sku_pstp_to_update in sku_pstp_to_update.all():
                            for production_stage in i_sku_pstp_to_update.production_stages.all():
                                production_stage.delete()
                            i_sku_pstp_to_update.delete()
                        sku_pstp_to_update.all().delete()
                    if reference_sku:
                        sku_production_stage_type_parameters = SkuProductionStageTypeParameter.objects.filter(sku = reference_sku)
                        if sku_production_stage_type_parameters.count():
                            sku_pstp_to_refer= sku_production_stage_type_parameters.first()
                        else:
                            correction_tip = '出错啦！%s 还没提前设置过生产流程' %reference_sku
                            break
                        sku_pstp_to_update = SkuProductionStageTypeParameter(sku = sku, production_type_name = sku_pstp_to_refer.production_type_name)
                        sku_pstp_to_update.save()
                        for obj in sku_pstp_to_refer.production_stages.all():
                            obj.pk = None
                            obj.save()
                            sku_pstp_to_update.production_stages.add(obj)
                        sku_pstp_to_update.save()

                update_completed_tip = '设置完成'

    template = loader.get_template('update_sku_production_stage_by_a_reference.html')
    context = {
        'form':UploadFileForm(), \
        'correction_tip': correction_tip , \
        'update_completed_tip': update_completed_tip \
    }
    return HttpResponse(template.render(context, request))

@user_passes_test(lambda u: u.groups.filter(name__in = ['logistics', 'accountant']).exists())
def update_sku_production_stage_detailed_numbers_default(request):
    correction_tip = ''
    update_completed_tip = ''
    template = loader.get_template('update_sku_production_stage_detailed_numbers_default.html')

    if request.method == 'POST' : # and request.get_host() == '209.97.151.168':
        form = UploadFileForm(request.POST,request.FILES)
        if form.is_valid():
            fileInMemory = request.FILES['file'].read()
            filePath = BytesIO(fileInMemory)
            wb2 = load_workbook(filePath, data_only=True)
            sheet = wb2.worksheets[0]
            for i_col in range(1, sheet.max_column + 1):
                if sheet.cell(1,i_col).value != UPDATE_SKU_PRODUCTION_STAGE_DETAILED_NUMBERS_DEFAULT_TABLE_HEADER[i_col - 1]:
                    if not(i_col == 1 and sheet.cell(1,i_col).value == None):
                        correction_tip = '出错啦！第一行标题需要修正，正确的是' + ' '.join(UPDATE_SKU_PRODUCTION_STAGE_DETAILED_NUMBERS_DEFAULT_TABLE_HEADER)
                        break
            if not correction_tip:
                sku = sheet.cell(2,2).value
                production_type_name = sheet.cell(3,2).value
                if production_type_name not in PRODUCTION_TYPE_NAME_OPTIONS:
                    correction_tip = '出错啦！请检查你的款式分类名字' + production_type_name
                    context = {
                        'form':UploadFileForm(), \
                        'correction_tip': correction_tip , \
                        'update_completed_tip': update_completed_tip \
                    }
                    return HttpResponse(template.render(context, request))
                sku_pstp_to_update = SkuProductionStageTypeParameter.objects.filter(sku = sku)
                if sku_pstp_to_update.count():
                    for i_sku_pstp_to_update in sku_pstp_to_update.all():
                        for production_stage in i_sku_pstp_to_update.production_stages.all():
                            production_stage.delete()
                        i_sku_pstp_to_update.delete()
                sku_pstp_to_update = SkuProductionStageTypeParameter(production_type_name = production_type_name, \
                                                                     sku = sku )
                sku_pstp_to_update.save()
                order_number = 0
                for row in range(4, sheet.max_row + 1):
                    if sheet.cell(row,1).value:
                        order_number += 1
                        stage_name = sheet.cell(row,1).value
                        if stage_name not in PRODUCTION_STAGE_NAME_OPTIONS:
                            correction_tip = '出错啦！请检查你的工序名字' + stage_name
                            context = {
                                'form':UploadFileForm(), \
                                'correction_tip': correction_tip , \
                                'update_completed_tip': update_completed_tip \
                            }
                            return HttpResponse(template.render(context, request))
                        stage_duration_days = sheet.cell(row,3).value or 0
                        production_units = sheet.cell(row,4).value
                        if production_units and stage_duration_days:
                            production_stage = ProductionStage(name = stage_name, production_stage_type_name = production_type_name ,\
                                                            minimum_days = stage_duration_days, duration_days = stage_duration_days, order_number = order_number)
                        else:
                            production_stage = ProductionStage(name = stage_name, production_stage_type_name = production_type_name ,\
                                                            duration_days = stage_duration_days, order_number = order_number)
                        if production_units:
                            production_stage.daily_production_units = production_units
                        production_stage.save()
                        sku_pstp_to_update.production_stages.add(production_stage)
                update_completed_tip = '设置完成'

    context = {
        'form':UploadFileForm(), \
        'correction_tip': correction_tip , \
        'update_completed_tip': update_completed_tip \
    }
    return HttpResponse(template.render(context, request))


@user_passes_test(lambda u: u.groups.filter(name__in = ['supplier', 'logistics']).exists())
def update_ongoing_production_plan_progress(request, production_plan_progress_id):
    if 'deadline_date' in request.GET:
        deadline_date = request.GET['deadline_date']
        if deadline_date:
            production_plan_progress_id = ProductionPlanProgress.objects.get(id = production_plan_progress_id)
            production_plan_progress_id.deadline_date = datetime.datetime.strptime(deadline_date, "%Y/%m/%d").date()
            production_plan_progress_id.save()
            response = {'response':'成功'}

        else:
            response = {'response':'deadline_date是空的'}
        return JsonResponse(response, safe=False)
    else:
        template = loader.get_template('update_ongoing_production_plan_progress.html')
        context = {
            'production_plan_progress': ProductionPlanProgress.objects.get(id=production_plan_progress_id) , \
        }
        return HttpResponse(template.render(context, request))

@user_passes_test(lambda u: u.groups.filter(name__in = ['supplier', 'logistics']).exists())
def update_ongoing_production_stage(request):
    production_stage_id = request.GET['production_stage_id']
    production_stage = ProductionStage.objects.get(id = production_stage_id)
    if 'daily_production_units' in request.GET:
        daily_production_units = request.GET['daily_production_units']
        if daily_production_units == '':
            production_stage.daily_production_units = None
        else:
            production_stage.daily_production_units = daily_production_units
        production_stage.save()
        production_plan_progress = production_stage.productionplanprogress_set.first()
        production_plan_progress.save()
    if 'duration_days' in request.GET:
        duration_days = request.GET['duration_days']
        production_stage.duration_days = duration_days
        production_stage.save()
        production_plan_progress = production_stage.productionplanprogress_set.first()
        production_plan_progress.save()
    if 'start_date_actually' in request.GET:
        start_date_actually = request.GET['start_date_actually']
        if start_date_actually == "":
            production_stage.start_date_actually = None
            production_stage.save()
            production_plan_progress = production_stage.productionplanprogress_set.first()
            lastest_production_stage = production_plan_progress.production_stages.all().order_by('-start_date_actually').first()
            if lastest_production_stage.start_date_actually == None:
                production_plan_progress.current_stage_id = None
                production_plan_progress.save()
            else:
                production_plan_progress.current_stage_id = lastest_production_stage.id
                production_plan_progress.save()
        else:
            production_stage.start_date_actually = datetime.datetime.strptime(start_date_actually, "%Y/%m/%d").date()
            production_stage.save()
            production_plan_progress = production_stage.productionplanprogress_set.first()
            lastest_production_stage = production_plan_progress.production_stages.all().order_by('-start_date_actually').first()
            if lastest_production_stage.start_date_actually <= production_stage.start_date_actually:
                production_plan_progress.current_stage_id = production_stage.id
                production_plan_progress.save()
            else:
                production_plan_progress.current_stage_id = lastest_production_stage.id
                production_plan_progress.save()
    if 'start_date_estimated' in request.GET:
        start_date_estimated = request.GET['start_date_estimated']
        if start_date_estimated == "":
            production_stage.start_date_estimated = None
        else:
            production_stage.start_date_estimated = datetime.datetime.strptime(start_date_estimated, "%Y/%m/%d").date()
        production_stage.save()
        production_plan_progress = production_stage.productionplanprogress_set.first()
        production_plan_progress.save()
    response = {'response':'成功'}
    return JsonResponse(response, safe=False)

@user_passes_test(lambda u: u.groups.filter(name__in = ['supplier', 'logistics']).exists())
def split_production_plan(request, production_plan_progress_id):
    max_subcontractors = 4
    if 'subcontractor_name2' in request.GET:
        production_plan_progress_id = int(production_plan_progress_id)
        ppps = ProductionPlanProgress.objects.filter(id = production_plan_progress_id)
        if ppps.count():
            ppp = ppps.first()
            if request.GET['subcontractor_name2'] and request.GET['qty2']:
                for i in range(2, max_subcontractors + 1):
                    i_subcontractor_name = 'subcontractor_name%i'  %i
                    subcontractor_name = request.GET[i_subcontractor_name]
                    i_qty = 'qty%i'  %i
                    if request.GET[i_qty]:
                        qty = int(request.GET[i_qty])
                        i_ppp_id = 'ppp_id%i' %i
                        ppp_id = int(request.GET[i_ppp_id])
                        if subcontractor_name and qty:
                            ppp_subcontractors = ProductionPlanProgress.objects.filter(production_plan_number__startswith = ppp.production_plan_number, manufacturer_number = "", subcontractor_name = subcontractor_name)
                            if ppp_subcontractors.count():
                                ppp_subcontractor = ppp_subcontractors.first()
                                ppp_subcontractor.qty = qty
                                ppp_subcontractor.save()
                            else:
                                ppp_tmp = ProductionPlanProgress.objects.filter(id = production_plan_progress_id).first()
                                ppp_tmp.pk = None
                                ppp_tmp.production_plan_number = ppp.production_plan_number + subcontractor_name
                                ppp_tmp.save()
                                ppp_tmp.manufacturer_number = ""
                                ppp_tmp.subcontractor_name = subcontractor_name
                                ppp_tmp.qty= qty

                                ppp_tmp.save()
                                for production_stage in ppp.production_stages.all():
                                    production_stage.pk = None
                                    production_stage.save()
                                    ppp_tmp.production_stages.add(production_stage)
            elif not(request.GET['subcontractor_name2']) and not(request.GET['qty2']):
                ppp_subcontractors = ProductionPlanProgress.objects.filter(production_plan_number__startswith = ppp.production_plan_number).all()
                for ppp_subcontractor in ppp_subcontractors:
                    if ppp_subcontractor.id != ppp.id:
                        for ps in ppp_subcontractor.production_stages.all():
                            ps.delete()
                        ppp_subcontractor.delete()
            ppp.qty = int(request.GET['qty1'])
            ppp.save()
            response = {'response':'成功'}
            return JsonResponse(response, safe=False)

    else:
        template = loader.get_template('split_production_plan.html')
        ppps = ProductionPlanProgress.objects.filter(id = production_plan_progress_id)

        subcontractors_list = []
        if ppps.count():
            ppp = ppps.first()
            ppp_subcontractors = ProductionPlanProgress.objects.filter(production_plan_number__startswith = ppp.production_plan_number)
            for ppp_subcontractor in ppp_subcontractors.all():
                subcontractors_list.append(ppp_subcontractor)
            for i in range(max_subcontractors - len(subcontractors_list)):
                subcontractors_list.append("")
        context = {
            'subcontractors_list': subcontractors_list ,\
            'production_plan_progress': ProductionPlanProgress.objects.get(id=production_plan_progress_id) , \
        }
        return HttpResponse(template.render(context, request))

@user_passes_test(lambda u: u.groups.filter(name__in = ['supplier', 'logistics']).exists())
def delete_production_plan(request, production_plan_progress_id):
    template = loader.get_template('delete_production_plan.html')
    ppp = ProductionPlanProgress.objects.filter(id = production_plan_progress_id)
    if ppp.count():
        is_a_parent_production_plan = False
        ppp = ppp.first()
        if not ppp.subcontractor_name:
            production_plan_number = ppp.production_plan_number
            if ProductionPlanProgress.objects.filter(production_plan_number__startswith = production_plan_number).count() > 1:
                execution_completed_tip =  '不能删除%s，因为这是一个分单的母单' %ppp.production_plan_number
                is_a_parent_production_plan = True
        if not is_a_parent_production_plan:
            for production_stage in ppp.production_stages.all():
                production_stage.delete()
            ppp.delete()
            execution_completed_tip =  '删除%s完毕' %ppp.production_plan_number
    else:
        execution_completed_tip = '不存在id%s' %production_plan_progress_id
    context = {
        'execution_completed_tip': execution_completed_tip ,\
    }
    return HttpResponse(template.render(context, request))

@user_passes_test(lambda u: u.groups.filter(name__in = ['supplier', 'logistics', 'sales']).exists())
def sewing_planner_calendar(request):
    steps_after_sewing = ['车缝', '过渡期(车缝到质检)', '质检','过渡期(车缝到油边2)', '油边2', '过渡期(油边2到质检)']
    template = loader.get_template('sewing_planner_calendar.html')
    ppps = ProductionPlanProgress.objects.filter(subcontractor_name="", qty__gt=0, ongoing=True).all()
    ppps_sewing = ProductionPlanProgress.objects.filter(subcontractor_name="", qty__gt=0, ongoing=True, current_stage_name = '车缝').all()
    recent_available_date = datetime.date.today()
    end_sewing_dates = []
    for ppp in ppps_sewing:
        ps = ProductionStage.objects.get(id = ppp.current_stage_id)
        started_date = ps.start_date_actually
        end_date = started_date + datetime.timedelta(days = ps.duration_days)
        end_sewing_dates.append(end_date)
    if len(end_sewing_dates):
        recent_available_date = max([recent_available_date, max(end_sewing_dates)])
    # 筛选出还未进入到车缝阶段的ppps
    ppps_sewing_not_started = ProductionPlanProgress.objects.filter(subcontractor_name="", qty__gt=0, ongoing=True).exclude(current_stage_name__in = steps_after_sewing).all()
    # 筛选出还未进入到车缝阶段，也没有预计车缝日期的ppps
    ppps_sewing_not_started_without_estimated_date = ppps_sewing_not_started.filter(Q(production_stages__name = '车缝') & Q(production_stages__start_date_estimated__isnull=True))
    # 筛选出还未进入到车缝阶段，但是有预计车缝日期的ppps
    #ppps_sewing_not_started_with_estimated_date = ppps_sewing_not_started.filter(Q(production_stages__name = '车缝') & Q(production_stages__start_date_estimated__isnull=False))
    # 帅选出有预计车缝日期的ppps，包括已经开始的
    ppps_sewing_with_estimated_date = ProductionPlanProgress.objects.filter(subcontractor_name="", qty__gt=0, ongoing=True).all().filter(Q(production_stages__name = '车缝') & Q(production_stages__start_date_estimated__isnull=False))
    context = {'ppps_sewing_not_started_without_estimated_date': ppps_sewing_not_started_without_estimated_date , \
               'ppps_sewing_with_estimated_date': ppps_sewing_with_estimated_date ,\
               #'ppps_sewing_not_started_with_estimated_date' : ppps_sewing_not_started_with_estimated_date ,\
               'recent_available_date': recent_available_date }
    return HttpResponse(template.render(context, request))

@csrf_exempt
@user_passes_test(lambda u: u.groups.filter(name__in = ['supplier', 'logistics']).exists())
def set_sewing_start_dates_ajax(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        for production_plan_number, start_date_string in data.items():
             start_date = datetime.datetime.strptime(start_date_string, '%Y-%m-%d')
             start_date += datetime.timedelta(days = 1)
             ppp = ProductionPlanProgress.objects.get(production_plan_number = production_plan_number)
             sewing_ps = ppp.production_stages.all().get(name = '车缝')
             sewing_ps.start_date_estimated = start_date
             sewing_ps.save()
        return JsonResponse({'success': True})
    else:
        return JsonResponse({'success': False})

@csrf_exempt
@user_passes_test(lambda u: u.groups.filter(name__in = ['supplier', 'logistics']).exists())
def delete_sewing_start_dates_ajax(request):
    data = json.loads(request.body)
    ppps = ProductionPlanProgress.objects.filter(production_plan_number = data['production_plan_progress_name'])
    if ppps.count():
        ppp = ppps.first()
        sewing_ps = ppp.production_stages.all().get(name = '车缝')
        sewing_ps.start_date_estimated = None
        sewing_ps.save()
        response = {'response':'成功'}
        return JsonResponse(response, safe=False)
    response = {'response':'失败'}
    return JsonResponse(response, safe=False)

@csrf_exempt
@user_passes_test(lambda u: u.groups.filter(name__in = ['supplier', 'logistics']).exists())
def get_ppp_3_statuses_ajax(request):
    data = json.loads(request.body)
    ppps = ProductionPlanProgress.objects.filter(production_plan_number = data['production_plan_progress_name'])
    response = {'checkbox':{},'inventory':{}}
    if ppps.count():
        ppp = ppps.first()
        ppps.filter(Q(production_stages__name = '车缝') & Q(production_stages__start_date_estimated__isnull=True))
        bought_material_ps = ppps.filter(Q(production_stages__name = '买料') & Q(production_stages__start_date_actually__isnull=False))
        cut_material_ps = ppps.filter(Q(production_stages__name__startswith = '开料') & Q(production_stages__start_date_actually__isnull=False))
        sewing_ps = ppps.filter(Q(production_stages__name = '车缝') & Q(production_stages__start_date_actually__isnull=False))
        if bought_material_ps.count():
            response['checkbox']['hasBoughtMaterial'] = True
        else:
            response['checkbox']['hasBoughtMaterial'] = False
        if cut_material_ps.count():
            response['checkbox']['hasCutMaterial'] = True
        else:
            response['checkbox']['hasCutMaterial'] = False
        if sewing_ps.count():
            response['checkbox']['hasStartedSewing'] = True
        else:
            response['checkbox']['hasStartedSewing'] = False
        sku = ppp.sku
        product = Product.objects.filter(sku = sku)
        if product.count():
            product = product.first()
            inventory_shenzhen = Inventory.objects.filter(warehouse_name = shenzhen_warehouse_name, sku = sku)
            if inventory_shenzhen.count():
                inventory_shenzhen = inventory_shenzhen.first().qty
            else:
                inventory_shenzhen = 0
            todayProductSalesUS = TodayProductSales.objects.filter(product= product,country = 'US').first()
            todayProductSalesEU = TodayProductSales.objects.filter(product= product,country = 'EU').first()
            todayProductSalesGB = TodayProductSales.objects.filter(product= product,country = 'GB').first()
            if todayProductSalesUS:
                response['inventory'].update({
                                          'shenzhenInventory':inventory_shenzhen ,\
                                          'USFbaInventory': todayProductSalesUS.fba_inventory.total_unit ,\
                                          'US7dSoldQty': todayProductSalesUS.sold_qty_average_7d ,\
                                          'USSustainDays': todayProductSalesUS.lasting_day_of_total_fba_unit_estimated_by_us ,\
                                          })
            if todayProductSalesEU:
                response['inventory'].update({
                                          'shenzhenInventory':inventory_shenzhen ,\
                                          'EUFbaInventory': todayProductSalesEU.fba_inventory.total_unit ,\
                                          'EU7dSoldQty': todayProductSalesEU.sold_qty_average_7d ,\
                                          'EUSustainDays': todayProductSalesEU.lasting_day_of_total_fba_unit_estimated_by_us ,\
                                          })
            if todayProductSalesGB:
                response['inventory'].update({
                                          'shenzhenInventory':inventory_shenzhen ,\
                                          'GBFbaInventory': todayProductSalesGB.fba_inventory.total_unit ,\
                                          'GB7dSoldQty': todayProductSalesGB.sold_qty_average_7d ,\
                                          'GBSustainDays': todayProductSalesGB.lasting_day_of_total_fba_unit_estimated_by_us \
                                          })
        return JsonResponse(response, safe=False)
    response = {'response':'失败'}
    return JsonResponse(response, safe=False)


@csrf_exempt
@user_passes_test(lambda u: u.groups.filter(name__in = ['supplier', 'logistics']).exists())
def set_ppp_3_statuses_ajax(request):
    # print(request.POST)
    # data = json.loads(request.body)
    data = request.POST
    ppps = ProductionPlanProgress.objects.filter(production_plan_number = data['production_plan_progress_name'])
    response = {}
    def string_to_bool(x):
        if x == 'false':
            return False
        return True
    if ppps.count():
        ppp = ppps.first()
        ppps.filter(Q(production_stages__name = '车缝') & Q(production_stages__start_date_estimated__isnull=True))
        bought_material_ps = ppps.filter(Q(production_stages__name = '买料') & Q(production_stages__start_date_actually__isnull=False))
        cut_material_ps = ppps.filter(Q(production_stages__name__startswith = '开料') & Q(production_stages__start_date_actually__isnull=False))
        sewing_ps = ppps.filter(Q(production_stages__name = '车缝') & Q(production_stages__start_date_actually__isnull=False))
        stage = ''
        if bought_material_ps.count() != string_to_bool(data['hasBoughtMaterial']):
            ps = ppp.production_stages.get(name = '买料')
            if string_to_bool(data['hasBoughtMaterial']):
                # 设定今天为该步骤实际开始时间
                ps.start_date_actually = datetime.date.today()
            else:
                #  取消该步骤的实际开始时间
                ps.start_date_actually = None
            ps.save()
        if cut_material_ps.count() != string_to_bool(data['hasCutMaterial']):
            ps = ppp.production_stages.get(name__startswith = '开料')
            if string_to_bool(data['hasCutMaterial']):
                # 设定今天为该步骤实际开始时间
                ps.start_date_actually = datetime.date.today()
            else:
                #  取消该步骤的实际开始时间
                ps.start_date_actually = None
            ps.save()
        if sewing_ps.count() != string_to_bool(data['hasStartedSewing']):
            ps = ppp.production_stages.get(name = '车缝')
            if string_to_bool(data['hasStartedSewing']):
                # 设定今天为该步骤实际开始时间
                ps.start_date_actually = datetime.date.today()
            else:
                #  取消该步骤的实际开始时间
                ps.start_date_actually = None
            ps.save()
        if sewing_ps.count():
            response['backgroundColor'] = 'pink'
            response['textColor'] = 'purple'
        elif cut_material_ps.count():
            response['backgroundColor'] = 'green'
            response['textColor'] = 'white'
        elif bought_material_ps.count():
            response['backgroundColor'] = 'yellow'
            response['textColor'] = 'black'
        else:
            response['backgroundColor'] = 'blue'
            response['textColor'] = 'white'
        return JsonResponse(response, safe=False)
    response = {'response':'失败'}
    return JsonResponse(response, safe=False)

def find_product_information_by_fnsku(request):
    template = loader.get_template('find_product_information_by_fnsku.html')
    fba_inventory = product = None
    if 'fnsku' in request.GET:
        fnsku = request.GET['fnsku']
        fba_inventory = FbaInventory.objects.filter(fnsku = fnsku)
        if fba_inventory.count():
            fba_inventory = fba_inventory.first()
            sku = fba_inventory.sku
            product = Product.objects.filter(sku = sku)
            if product.count():
                product = product.first()
            else:
                product = None
    context = {
        'fba_inventory': fba_inventory ,\
        'product': product ,\
    }
    return HttpResponse(template.render(context, request))

@csrf_exempt
@user_passes_test(lambda u: u.groups.filter(name__in = ['sales', 'logistics']).exists())
def update_product_chinese_name_by_uploading(request):
    template = loader.get_template('update_product_chinese_name_by_uploading.html')
    error_tip = ''
    sucess_tip = ''
    if request.method == 'POST':
        fileInMemory = request.FILES['file'].read()
        filePath = BytesIO(fileInMemory)
        wb2 = load_workbook(filePath, data_only=True)
        sheet = wb2.worksheets[0]
        max_column = sheet.max_column

        sku_title = '*商品编号'
        name_in_chinese_title = '商品名称'
        for i_col in range(1, max_column):
            if sheet.cell(2, i_col).value == sku_title:
                sku_title_col = i_col
            elif sheet.cell(2, i_col).value == name_in_chinese_title:
                name_in_chinese_col = i_col
                break
        for i_row in range(3, sheet.max_row):
            sku = sheet.cell(i_row, sku_title_col).value
            product = Product.objects.filter(sku = sku)
            if product.count():
                product = product.first()
                name_in_chinese = sheet.cell(i_row, name_in_chinese_col).value
                product.name_in_chinese = name_in_chinese
                product.save()
        sucess_tip = '完成录入'
    if error_tip:
        context = {
            'error_tip': error_tip ,\
            'form':UploadFileForm()
        }
    elif sucess_tip:
        context = {
            'sucess_tip': sucess_tip ,\
            'form':UploadFileForm()
        }
    else:
        context = {
            'form':UploadFileForm()
        }
    return HttpResponse(template.render(context, request))

@csrf_exempt
@user_passes_test(lambda u: u.groups.filter(name__in = ['logistics']).exists())
def update_transparency_label_required_sku_by_uploading(request):
    template = loader.get_template('update_transparency_label_required_sku_by_uploading.html')
    error_tip = ''
    sucess_tip = ''
    if request.method == 'POST':
        fileInMemory = request.FILES['file'].read()
        filePath = BytesIO(fileInMemory)
        wb2 = load_workbook(filePath, data_only=True)
        sheet = wb2.worksheets[0]
        for product in Product.objects.all():
            product.transparency = False
            product.save()
        for i_row in range(1, sheet.max_row):
            sku = sheet.cell(i_row, 1).value
            product = Product.objects.filter(sku = sku)
            if product.count():
                product = product.first()
                product.transparency = True
                product.save()
        sucess_tip = '完成录入'
    if error_tip:
        context = {
            'error_tip': error_tip ,\
            'form':UploadFileForm()
        }
    elif sucess_tip:
        context = {
            'sucess_tip': sucess_tip ,\
            'form':UploadFileForm()
        }
    else:
        context = {
            'form':UploadFileForm()
        }
    return HttpResponse(template.render(context, request))
