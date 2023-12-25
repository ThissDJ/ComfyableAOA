from openpyxl import load_workbook
import csv
ORDER_KEYWORDS = ['Order', 'Bestellung', 'Commande', 'Ordine', 'Pedido']
REFUND_KEYWORDS = ['Refund', 'Erstattung', 'Rimborso','Remboursement','Reembolso']
ADJUSTMENT_KEYWORDS = ['Adjustment']
OTHER_ADJUSTMENT_KEYWORDS =['Other','Subscription Fee Adjustment']
TRANSFER_KEYWORDS = ['Transfer','Transfert','Übertrag','Trasferimento','Transferir']
COST_OF_ADVERTISING_KEYWORDS = ['Cost of Advertising']
FREIGHT_SHIPPING_CHARGE = ['FBA International Freight Shipping Charge', 'FBA International Freight Duties and Taxes Charge']
SERVICE_FEE_KEYWORDS = ['Service Fee','Frais de service','Servicegebühr']
IDENTIFY_COUNTRY_WORDS = {"All amounts in AED, unless specified":"AE" \
                            ,"All amounts in USD, unless specified":"US" \
                            ,"All amounts in CAD, unless specified":"CA" \
                            ,"All amounts in AUD, unless specified":"AU" \
                            ,"All amounts in SGD, unless specified":"SG" \
                            ,"All amounts in GBP, unless specified":"UK" \
                            ,"Alle Beträge in Euro sofern nicht anders gekennzeichnet":"DE" \
                            ,"Todos los importes en EUR, a menos que se especifique lo contrario" :"ES" \
                            ,"Tutti gli importi sono espressi in EUR, se non diversamente specificato." :"IT" \
                            ,"Tous les montants sont en EUR, sauf mention contraire." : "FR" \
                            }
def readTransaction(fileInMemory):

    productInfoDict = {}
    decodedFile = fileInMemory.decode('utf-8').splitlines()
    csvReader = csv.reader(decodedFile)
    # find start_line
    lineCount = 0
    for row in csvReader:
        # identify which country
        if lineCount == 1:
            COUNTRY = IDENTIFY_COUNTRY_WORDS[row[0]]
        lineCount += 1
        if len(row) > 1:
            START_LINE = lineCount
            if row[1].isdigit():
                break
    def correctFloatString(x):
        if COUNTRY in ["ES","DE","IT","FR"]:
            return x.replace('.','').replace(',','.')
        else:
            return x.replace(',','')
    if COUNTRY in ["US"]:
        COLUMN_NUM_DICT = {'type':3, 'orderId':4, 'sku':5, 'sales': 15, 'quantity': 7, 'AmazonFee':16, 'description': 6}
    elif COUNTRY in ["CA","UK","DE"]:
        COLUMN_NUM_DICT = {'type':3, 'orderId':4, 'sku':5, 'sales': 14, 'quantity': 7, 'AmazonFee':15, 'description': 6}
    else:
        COLUMN_NUM_DICT = {'type':3, 'orderId':4, 'sku':5, 'sales': 13, 'quantity': 7, 'AmazonFee':14, 'description': 6}
        # COLUMN_NUM_DICT = {'type':3, 'orderId':4, 'sku':5, 'sales': 13, 'quantity': 7, 'AmazonFee':14, 'description': 6}
    lineCount = 0
    csvReader = csv.reader(decodedFile)
    for row in csvReader:
        lineCount += 1
        if lineCount >= START_LINE:
            # sum of sales by sku
            if row[COLUMN_NUM_DICT['sku'] - 1] not in productInfoDict:
                productInfoDict[row[COLUMN_NUM_DICT['sku'] - 1].strip()] = {'sales': float(correctFloatString(row[COLUMN_NUM_DICT['sales'] - 1])) \
                                                                    ,'quantity': 0 \
                                                                    , 'AmazonFee': float(0)}

            else:
                productInfoDict[row[COLUMN_NUM_DICT['sku'] - 1].strip()]['sales'] += float(correctFloatString(row[COLUMN_NUM_DICT['sales'] - 1]))
            # sum of quantity by sku
            if row[COLUMN_NUM_DICT['type'] - 1] in REFUND_KEYWORDS:
                productInfoDict[row[COLUMN_NUM_DICT['sku'] - 1].strip()]['quantity'] -= int(row[COLUMN_NUM_DICT['quantity'] - 1])
            elif row[COLUMN_NUM_DICT['type'] - 1] in ORDER_KEYWORDS or row[COLUMN_NUM_DICT['type'] - 1] in ADJUSTMENT_KEYWORDS:
                if row[COLUMN_NUM_DICT['description'] - 1] not in ['FBA transportation fee', 'Tarifa de transporte de Logística de Amazon'] and row[COLUMN_NUM_DICT['description'] - 1] not in OTHER_ADJUSTMENT_KEYWORDS:
                    productInfoDict[row[COLUMN_NUM_DICT['sku'] - 1].strip()]['quantity'] += int(row[COLUMN_NUM_DICT['quantity'] - 1])
            # sum of fee by sku
            if row[COLUMN_NUM_DICT['type'] - 1] in ORDER_KEYWORDS +ADJUSTMENT_KEYWORDS + REFUND_KEYWORDS:
                for col in range(COLUMN_NUM_DICT['AmazonFee'], len(row) ) :
                    productInfoDict[row[COLUMN_NUM_DICT['sku'] - 1].strip()]['AmazonFee'] += float(correctFloatString(row[col - 1]))
    return productInfoDict

def readFixedCost(fileInMemory):
    TEST_LINE = 9
    COLUMN_NUM_DICT = {'type':3, 'description': 6}

    EXCEPT_KEYWORDS = ORDER_KEYWORDS + REFUND_KEYWORDS + ADJUSTMENT_KEYWORDS + TRANSFER_KEYWORDS
    fixedCost = 0.0

    decodedFile = fileInMemory.decode('utf-8').splitlines()
    csvReader = csv.reader(decodedFile)
    # find start_line
    lineCount = 0
    for row in csvReader:
        if lineCount == 1:
            COUNTRY = IDENTIFY_COUNTRY_WORDS[row[0]]
        lineCount += 1
        if len(row) > 1:
            START_LINE = lineCount
            if row[1].isdigit():
                break
    def correctFloatString(x):
        if COUNTRY in ["ES","DE","IT","FR"]:
            return x.replace('.','').replace(',','.')
        else:
            return x.replace(',','')
    csvReader = csv.reader(decodedFile)
    lineCount = 0
    for row in csvReader:
        lineCount += 1
        if lineCount >= START_LINE:
            # sum of fixed cost
            if row[COLUMN_NUM_DICT['type'] - 1] not in EXCEPT_KEYWORDS:
                if row[COLUMN_NUM_DICT['type'] - 1] in SERVICE_FEE_KEYWORDS:
                    if row[COLUMN_NUM_DICT['description'] - 1] not in COST_OF_ADVERTISING_KEYWORDS and row[COLUMN_NUM_DICT['description'] - 1] not in FREIGHT_SHIPPING_CHARGE:
                        fixedCost += float(correctFloatString(row[len(row) - 1]))
                else:
                    fixedCost += float(correctFloatString(row[len(row) - 1]))
            elif row[COLUMN_NUM_DICT['type'] - 1] in ADJUSTMENT_KEYWORDS and row[COLUMN_NUM_DICT['description'] - 1] in OTHER_ADJUSTMENT_KEYWORDS:
                fixedCost += float(correctFloatString(row[len(row) - 1]))
    return fixedCost

def readSkuValue(filePath, CHECK_KEYWORD, START_ROW):
    wb2 = load_workbook(filePath)
    ws = wb2[wb2.sheetnames[0]]
    costDict = {}
    COLUMN_NUM_DICT = {}
    # def validateFile(ws):
    #     for k, v in COLUMN_NUM_DICT.items():
    #         if ws.cell(START_ROW - 1,v).value != CHECK_KEYWORD[k]:
    #             return False
    #     return True
    def findCol(ws):
        for col in range(ws.max_column):
            for k,v in CHECK_KEYWORD.items():
                if ws.cell(START_ROW - 1,col + 1).value == CHECK_KEYWORD[k]:
                    COLUMN_NUM_DICT[k] = col + 1
        return COLUMN_NUM_DICT
    COLUMN_NUM_DICT = findCol(ws)
    for row in range(START_ROW, ws.max_row + 1):
        value = float(ws.cell(row, COLUMN_NUM_DICT['value']).value)
        sku = ws.cell(row, COLUMN_NUM_DICT['sku']).value.strip()
        if sku not in costDict.keys():
            costDict[sku] = 0.0
        costDict[sku] += value
    return costDict

def readCurrencyRate(filePath):
    CURRENCY_KEYWORD = {'usd2cny': '1USD=?RMB' \
                ,'cad2cny': '1CAD=?RMB' \
                ,'aud2cny': '1AUD=?RMB' \
                ,'aed2cny': '1AED=?RMB' \
                ,'gbp2cny': '1GBP=?RMB' \
                ,'eur2cny': '1EUR=?RMB' \
                ,'sgd2cny': '1SGD=?RMB'}
    UPDATE_HINT_CELL_NUM = (1, 4)
    UPDATE_COL_NUM = 5
    wb2 = load_workbook(filePath)
    ws = wb2[wb2.sheetnames[0]]
    currencyRateDict = {}
    currencyRateUpdateTime = {}
    def validateFile(ws):
        if ws.cell(UPDATE_HINT_CELL_NUM[0], UPDATE_HINT_CELL_NUM[1]).value != '更新时间':
            return False
        return True
    if validateFile(ws):
        for row in range(1,ws.max_row+1):
            currencyRateName = ws.cell(row,1).value
            if [key for key, value in CURRENCY_KEYWORD.items() if value == currencyRateName]:
                currencyRateDict[[key for key, value in CURRENCY_KEYWORD.items() if value == currencyRateName][0]] = float(ws.cell(row,2).value)
            else:
                return 'Unknown: %s' %currencyRateName
        for row in range(1,ws.max_row+1):
            currencyRateName = ws.cell(row,1).value
            if [key for key, value in CURRENCY_KEYWORD.items() if value == currencyRateName]:
                currencyRateUpdateTime[[key for key, value in CURRENCY_KEYWORD.items() if value == currencyRateName][0]] = float(ws.cell(row,UPDATE_COL_NUM).value)
            else:
                return 'Unknown: %s' %currencyRateName
    else:
        return '缺少更新时间'
    return (currencyRateDict,currencyRateUpdateTime)

def readPurchaseInboundCost(filePath):

    START_LINE = 4
    COLUMN_NUM_DICT = {'sku':1, 'purchaseCost': 2, 'inboundCost': 3}
    COUNTRY_ROW = 2
    TIME_ROW = 1
    CHECK_KEYWORD = {'SKU': 'SKU', 'purchaseCost': '单位采购成本', 'inboundCost': '单位头程运费'}

    wb2 = load_workbook(filePath)
    ws = wb2[wb2.sheetnames[0]]
    costDict = {}
    def validateFile(ws):
        if ws.max_row < 4:
            return False
        if ws.cell(3,1).value != CHECK_KEYWORD['SKU']  \
            or ws.cell(3,2).value != CHECK_KEYWORD['purchaseCost'] \
            or ws.cell(3,3).value != CHECK_KEYWORD['inboundCost'] :
            return False

        return True
    if validateFile(ws):
        for row in range(START_LINE,ws.max_row+1):
            costDict[ws.cell(row,COLUMN_NUM_DICT['sku']).value.strip()] = {'purchaseCost': float(ws.cell(row,COLUMN_NUM_DICT['purchaseCost']).value) \
                                                                ,'inboundCost': float(ws.cell(row,COLUMN_NUM_DICT['inboundCost']).value)}
        return {'purchaseInboundCost': costDict \
            ,'country': ws.cell(COUNTRY_ROW,1).value \
            , 'time': ws.cell(TIME_ROW,1).value}
    else:
        return False

def readSharePlan(filePath):
    class SkuOfDesigner:
        def __init__(self, shareRate):
            self.shareRate = shareRate
    class SkuOfSalesPerson:
        def __init__(self, shareRate, thresholdUS,thresholdCA, thresholdEU):
            self.shareRate = shareRate
            self.thresholdUS = thresholdUS
            self.thresholdCA = thresholdCA
            self.thresholdEU = thresholdEU
        def getThreshold(self,country):
            if  country in ['US', 'AE', 'AU','SG']:
                threshold = self.thresholdUS
            elif country == 'CA':
                threshold = self.thresholdCA
            elif country in ['UK', 'DE', 'IT', 'FR', 'ES']:
                threshold = self.thresholdEU
            return threshold
    class SkuOfSupplier:
        def __init__(self, shareRate, returnRate):
            self.shareRate = shareRate
            self.returnRate = returnRate

    class Sharer:
        def __init__(self, name, type, shareRateColNum, thresholdUSColNum,thresholdCAColNum, thresholdEUColNum, returnRateColNum):
            self.name = name
            self.type = type
            self.shareRateColNum = shareRateColNum
            self.thresholdUSColNum = thresholdUSColNum
            self.thresholdCAColNum = thresholdCAColNum
            self.thresholdEUColNum = thresholdEUColNum
            self.returnRateColNum = returnRateColNum
            self.skuDetail = {}

        def getType(self):
            return self.type

        def getSKUDetail(self, row, ws):
            sku = wsSKU.cell(row, 1).value.strip()
            shareRate = wsSKU.cell(row, self.shareRateColNum).value
            if self.type == '设计师':
                self.skuDetail[sku] = SkuOfDesigner(shareRate)
            elif self.type == '运营':
                thresholdUS = wsSKU.cell(row, self.thresholdUSColNum).value
                thresholdCA = wsSKU.cell(row, self.thresholdCAColNum).value
                thresholdEU = wsSKU.cell(row, self.thresholdEUColNum).value
                self.skuDetail[sku] = SkuOfSalesPerson(shareRate, thresholdUS, thresholdCA, thresholdEU)
            elif self.type == '供应商':
                returnRate = wsSKU.cell(row, self.returnRateColNum).value
                self.skuDetail[sku] = SkuOfSupplier(shareRate, returnRate)
        def getProfitShare(self, sku, country, purchaseCost, inboundCost, profit, profitMargin):
            shareRate = self.skuDetail[sku].shareRate
            if self.type == '设计师':
                return profit * shareRate
            elif self.type == '供应商':
                returnRate = self.skuDetail[sku].returnRate
                return {'profitShare': profit * shareRate, 'returnAmount': purchaseCost * returnRate }
            elif self.type == '运营':
                if profit < 0:
                    # 亏损时候运营负担多少？目前是0.15
                    return profit * 0.15
                else:
                    threshold = self.skuDetail[sku].getThreshold(country)
                    if profitMargin < threshold or profitMargin == 0:
                        return 0.0
                    else:
                        return profit/profitMargin * (profitMargin - threshold) * self.skuDetail[sku].shareRate


    DETAILS_SHEET_NAME = 'details'
    ROLE_DESCRIPTION_SHEET_NAME = 'role description'
    START_LINE = 2
    COLUMN_NUM_DICT = {'name': 1, 'type':2, 'shareRateColNum': 3 \
                    , 'thresholdUSColNum': 4, 'thresholdCAColNum': 5 \
                    , 'thresholdEUColNum': 6, 'returnRateColNum': 7}

    wb2 = load_workbook(filePath, data_only=True)
    ws = wb2.get_sheet_by_name(ROLE_DESCRIPTION_SHEET_NAME)

    # 读取各个角色设定
    def notNone2Int(x):
        if x is not None:
            return int(x)
        return x
    sharers = []
    for row in range(START_LINE,ws.max_row+1):
        name = ws.cell(row,COLUMN_NUM_DICT['name']).value.strip()
        type = ws.cell(row,COLUMN_NUM_DICT['type']).value.strip()
        shareRateColNum = notNone2Int(ws.cell(row,COLUMN_NUM_DICT['shareRateColNum']).value)
        thresholdUSColNum = notNone2Int(ws.cell(row,COLUMN_NUM_DICT['thresholdUSColNum']).value)
        thresholdCAColNum = notNone2Int(ws.cell(row,COLUMN_NUM_DICT['thresholdCAColNum']).value)
        thresholdEUColNum = notNone2Int(ws.cell(row,COLUMN_NUM_DICT['thresholdEUColNum']).value)
        returnRateColNum = notNone2Int(ws.cell(row,COLUMN_NUM_DICT['returnRateColNum']).value)

        sharer = Sharer(name,type, shareRateColNum, thresholdUSColNum,thresholdCAColNum, thresholdEUColNum, returnRateColNum)
        sharers.append(sharer)

    wsSKU = wb2.get_sheet_by_name(DETAILS_SHEET_NAME)
    for sharer in sharers:
        for row in range(START_LINE,wsSKU.max_row+1):
            sharer.getSKUDetail(row, wsSKU)

    return sharers

def profitSummaryBySkuProfitReport(filePath):
    wb2 = load_workbook(filePath)
    ws = wb2.get_sheet_by_name('profit')

    profitTime = ws.cell(1,1).value
    country = ws.cell(1,2).value

    purchaseCostBySku = readSkuValue(filePath, {'sku':'SKU', 'value':'总采购价格'}, 3)
    totalPurchaseCost = sum([v for k,v in purchaseCostBySku.items()])

    inboundCostBySku = readSkuValue(filePath, {'sku':'SKU', 'value':'总头程运费'}, 3)
    totalInboundCost = sum([v for k,v in inboundCostBySku.items()])

    adCostBySku = readSkuValue(filePath, {'sku':'SKU', 'value':'广告费用'}, 3)
    totalAdCost = sum([v for k,v in adCostBySku.items()])

    amazonFeeBySku = readSkuValue(filePath, {'sku':'SKU', 'value':'亚马逊费用'}, 3)
    totalamazonFeeCost = sum([v for k,v in amazonFeeBySku.items()])

    salesBySku = readSkuValue(filePath, {'sku':'SKU', 'value':'销售金额'}, 3)
    totalSales = sum([v for k,v in salesBySku.items()])

    profit = totalSales - totalPurchaseCost - totalInboundCost - totalAdCost - totalamazonFeeCost
    profitMargin = profit / totalSales

    purchaseCostPercent = totalPurchaseCost / totalSales
    inboundCostPercent = totalInboundCost / totalSales
    adCostPercent = totalAdCost / totalSales
    amazonFeePercent = totalamazonFeeCost / totalSales

    return {'profitTime': profitTime, 'country': country, 'tatal Sales': totalSales \
            , 'totalPurchaseCost': totalPurchaseCost, 'purchaseCostPercent': purchaseCostPercent \
            , 'totalInboundCost': totalInboundCost, 'inboundCostPercent': inboundCostPercent \
            , 'totalAdCost': totalAdCost, 'adCostPercent': adCostPercent \
            , 'totalamazonFeeCost':  totalamazonFeeCost, 'amazonFeePercent': amazonFeePercent \
            , 'profit': profit, 'profitMargin': profitMargin
            }

def readProfitBySkuAndRole(filePath):
    wb2 = load_workbook(filePath)
    class ProfitBySkuOrRole:
        def __init__(self, roleName, content):
            self.roleName = roleName
            self.content = content
    profitBySkuOfRoles= []
    for sheetName in wb2.sheetnames:
        if sheetName != 'profit':
            ws = wb2.get_sheet_by_name(sheetName)
            rowsContent = []
            for row in range(1, ws.max_row + 1):
                rowContent = []
                for col in range(1, ws.max_column + 1):
                    rowContent.append(ws.cell(row, col).value)
                rowsContent.append(rowContent)

            profitBySkuOfRole = ProfitBySkuOrRole(sheetName, rowsContent)
            profitBySkuOfRoles.append(profitBySkuOfRole)
    return profitBySkuOfRoles
