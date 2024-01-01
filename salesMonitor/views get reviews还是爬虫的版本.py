from django.http import HttpResponse
from django.shortcuts import redirect
from django.template import loader
from django import forms
from django.views.generic.edit import FormView
from .forms import UploadFilesForm, UploadFileForm, AsinForm
from django.contrib.auth.decorators import login_required
from io import BytesIO,StringIO
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
import xlrd
import datetime
import pytz
import csv
from django.db.models import Sum
from django.http import JsonResponse
from salesMonitor.models import Product, TodayProductSales, Last7dayProductSales, \
                                DailySalesLastYear, FbaInventory, \
                                ReceivablePurchasedQty, HistoryTodayProductSales,\
                                HistoryTodaySales, Inventory, FbaShipment, ShippedSkuQty ,\
                                ReceivedSkuQty


import selectorlib
import requests
from dateutil import parser as dateparser
import os

# Build paths inside the project like this: os.path.join(BASE_DIR, ...)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
extractor = selectorlib.Extractor.from_yaml_file(os.path.join(BASE_DIR,'selectors.yml'))

shenzhen_warehouse_name = '深圳F199'
@login_required
def index(request):
    template = loader.get_template('todaySales.html')
    todaySales = HistoryTodaySales.objects.order_by('-date')[0]
    class TopSku:
        def __init__(self, sku, sold_qty):
            self.sku = sku
            self.sold_qty = sold_qty
    top_sales_skus = TodayProductSales.objects.order_by('-sales_amount')[:5]
    top_sold_qty_skus = TodayProductSales.objects.order_by('-sold_qty')[:5]
    all_today_new_product_skus = TodayProductSales.objects.filter(product__new = True).all()
    context = {
        'date': todaySales.date.strftime('%Y/%m/%d') ,\
        'sales_today': todaySales.sales_today ,\
        'sales_same_day_last_year': todaySales.sales_same_day_last_year ,\
        'monthly_increase_on_sales': todaySales.monthly_increase_on_sales  * 100,\
        'ad_cost': todaySales.ad_cost ,\
        'acos': todaySales.acos * 100 ,\
        'ad_cost_on_sales': todaySales.ad_cost_on_sales * 100 ,\
        'top_sales_skus': top_sales_skus ,\
        'top_sold_qty_skus': top_sold_qty_skus ,\
        'all_today_new_product_skus' : all_today_new_product_skus ,\
        'today': True
    }
    return HttpResponse(template.render(context, request))

def date_split(date):
    if len(date) > 4:
        year = int(date[:4])
        month = int(date[4:6])
        day = int(date[6:])
    else:
        year = datetime.date.today().year
        month = int(date[:2])
        day = int(date[2:])
    return {'year':year,'month':month,'day':day}

@login_required
def history_index(request, year, month, day):
    date = "%04d/%02d/%02d" % (year, month, day,)
    today_date_obj = datetime.date(year, month, day )
    template = loader.get_template('todaySales.html')
    todaySales = HistoryTodaySales.objects.filter(date = today_date_obj)
    if todaySales.count():
        todaySales = todaySales.all()[0]
    else:
        todaySales = HistoryTodaySales.objects.order_by('-date')[0]
    class TopSku:
        def __init__(self, sku, sold_qty):
            self.sku = sku
            self.sold_qty = sold_qty
    top_sales_skus = HistoryTodayProductSales.objects.filter(date = today_date_obj).order_by('-sales_amount')[:5]
    top_sold_qty_skus = HistoryTodayProductSales.objects.filter(date = today_date_obj).order_by('-sold_qty')[:5]
    all_today_new_product_skus = HistoryTodayProductSales.objects.filter(date = today_date_obj, product__new = True).all()
    context = {
        'date': todaySales.date.strftime('%Y/%m/%d') ,\
        'sales_today': todaySales.sales_today ,\
        'sales_same_day_last_year': todaySales.sales_same_day_last_year ,\
        'monthly_increase_on_sales': todaySales.monthly_increase_on_sales  * 100,\
        'ad_cost': todaySales.ad_cost ,\
        'acos': todaySales.acos * 100 ,\
        'ad_cost_on_sales': todaySales.ad_cost_on_sales * 100 ,\
        'top_sales_skus': top_sales_skus ,\
        'top_sold_qty_skus': top_sold_qty_skus ,\
        'all_today_new_product_skus' : all_today_new_product_skus ,\
        'today': False
    }
    return HttpResponse(template.render(context, request))

@login_required
def update_last_year_sales(request):
    if request.method == 'POST':
        start_line_list = ['Time', 'Selected date range (OPS)', 'Selected date range (Units)']
        start_read = False
        fileInMemory = request.FILES['file'].read().decode('utf-8')
        csv_data = csv.reader(StringIO(fileInMemory), delimiter=',')
        DailySalesLastYear.objects.all().delete()
        for row in csv_data:
            if start_line_list == row:
                start_read = True
                continue
            if start_read and row != []:
                date_str = row[0]
                sales = float(row[1].replace(',','').replace('$',''))
                date_str_split = date_str.split('/')
                day = date_str_split[1]
                month = date_str_split[0]
                dailySalesLastYear = DailySalesLastYear(day = int(day), month = int(month), sales = sales)
                dailySalesLastYear.save()
            elif start_read and row == []:
                start_read = False

    template = loader.get_template('updateLastYearSales.html')
    context = {
        'form':UploadFileForm()
    }

    return HttpResponse(template.render(context, request))

def check_if_within_7day(year_month_day,time_to_check):
    timezone = pytz.timezone('America/Los_Angeles')
    date_time_obj = datetime.datetime.strptime(time_to_check[:19], '%Y-%m-%dT%H:%M:%S')
    timezone_date_time_obj = date_time_obj.astimezone(pytz.timezone('America/Los_Angeles'))
    countable_day = timezone_date_time_obj.day
    start_datetime = datetime.datetime(timezone_date_time_obj.year, \
                                        year_month_day['month'], year_month_day['day'],\
                                        0,0,0,0,pytz.timezone('America/Los_Angeles') \
                                        )
    compare_datetime = start_datetime + datetime.timedelta(days=1)
    compare_earliest_datetime = start_datetime - datetime.timedelta(days=6)
    timezone_date_time_obj = datetime.datetime(timezone_date_time_obj.year, \
                                        timezone_date_time_obj.month, timezone_date_time_obj.day,\
                                        timezone_date_time_obj.hour,timezone_date_time_obj.minute, \
                                        timezone_date_time_obj.second,0,pytz.timezone('America/Los_Angeles') \
                                        )
    if timezone_date_time_obj > compare_datetime or timezone_date_time_obj <= compare_earliest_datetime:
        return False
    else:
        return True

def check_if_today(end_day,time_to_check):
    timezone = pytz.timezone('America/Los_Angeles')
    date_time_obj = datetime.datetime.strptime(time_to_check[:19], '%Y-%m-%dT%H:%M:%S')
    timezone_date_time_obj = date_time_obj.astimezone(pytz.timezone('America/Los_Angeles'))
    countable_day = timezone_date_time_obj.day
    if countable_day == end_day:
        return True
    return False

def save_history_today_product_sales(todayProductSales,date):
    histroyTodayProductSales = HistoryTodayProductSales.objects.filter(date = date,product = todayProductSales.product)
    if histroyTodayProductSales.count():
        histroyTodayProductSales = histroyTodayProductSales.all()[0]
        histroyTodayProductSales.sold_qty = todayProductSales.sold_qty
        histroyTodayProductSales.sales_amount = todayProductSales.sales_amount
        histroyTodayProductSales.sold_qty_average_7d = todayProductSales.sold_qty_average_7d
        histroyTodayProductSales.average_price_7d =  todayProductSales.average_price_7d
        histroyTodayProductSales.fba_inventory =  todayProductSales.fba_inventory
        histroyTodayProductSales.lasting_day_estimated_by_us =  todayProductSales.lasting_day_estimated_by_us
        histroyTodayProductSales.lasting_day_of_available_estimated_by_us =  todayProductSales.lasting_day_of_available_estimated_by_us
        histroyTodayProductSales.lasting_day_of_available_fc_estimated_by_us =  todayProductSales.lasting_day_of_available_fc_estimated_by_us
        histroyTodayProductSales.lasting_day_of_total_fba_unit_estimated_by_us =  todayProductSales.lasting_day_of_total_fba_unit_estimated_by_us
    else:
        histroyTodayProductSales = HistoryTodayProductSales( \
            product = todayProductSales.product ,\
            date = date ,\
            sold_qty = todayProductSales.sold_qty ,\
            sales_amount = todayProductSales.sales_amount ,\
            sold_qty_average_7d = todayProductSales.sold_qty_average_7d ,\
            average_price_7d =  todayProductSales.average_price_7d ,\
            fba_inventory =  todayProductSales.fba_inventory ,\
            lasting_day_estimated_by_us =  todayProductSales.lasting_day_estimated_by_us ,\
            lasting_day_of_available_estimated_by_us =  todayProductSales.lasting_day_of_available_estimated_by_us ,\
            lasting_day_of_available_fc_estimated_by_us =  todayProductSales.lasting_day_of_available_fc_estimated_by_us ,\
            lasting_day_of_total_fba_unit_estimated_by_us =  todayProductSales.lasting_day_of_total_fba_unit_estimated_by_us \
                                                                )
    histroyTodayProductSales.save()

def save_today_sales(todaySales, date):
    historyTodaySales = HistoryTodaySales.objects.filter(date = date)
    if historyTodaySales.count():
        historyTodaySales = historyTodaySales.all()[0]
        historyTodaySales.sales_today = todaySales.sales_today
        historyTodaySales.sales_same_day_last_year = todaySales.sales_same_day_last_year
        historyTodaySales.sales_month_to_date = todaySales.sales_month_to_date
        historyTodaySales.monthly_increase_on_sales = todaySales.monthly_increase_on_sales
        historyTodaySales.ad_cost = todaySales.ad_cost
        historyTodaySales.acos = todaySales.acos
        historyTodaySales.ad_cost_on_sales = todaySales.ad_cost_on_sales
    else:
        historyTodaySales = HistoryTodaySales( \
                                date = date ,\
                                sales_today = todaySales.sales_today ,\
                                sales_same_day_last_year = todaySales.sales_same_day_last_year ,\
                                sales_month_to_date = todaySales.sales_month_to_date ,\
                                monthly_increase_on_sales = todaySales.monthly_increase_on_sales ,\
                                ad_cost = todaySales.ad_cost ,\
                                acos = todaySales.acos ,\
                                ad_cost_on_sales = todaySales.ad_cost_on_sales ,\
                                            )
    historyTodaySales.save()

@login_required
def update_7d_orders(request):
    if request.method == 'POST':
        date = request.POST['date']
        year_month_day = date_split(date)
        year = year_month_day['year']
        month = year_month_day['month']
        day = year_month_day['day']
        #start_line_list = ['Time', 'Selected date range (OPS)', 'Selected date range (Units)']
        #start_read = False
        fileInMemory = request.FILES['file'].read().decode('utf-8')
        csv_data = csv.reader(StringIO(fileInMemory), delimiter='\t')
        Last7dayProductSales.objects.all().delete()
        TodayProductSales.objects.all().delete()
        sku_qty = {}
        sku_price = {}
        sku_today_qty = {}
        sku_today_price = {}

        today_date_obj = datetime.date(year, month, day )
        for row in csv_data:
            order_id = row[0]
            if order_id[:2] == '11':
                if check_if_within_7day(year_month_day,row[2]):
                    if_today = check_if_today(day,row[2])
                    sku = row[11]
                    qty = int(row[14])
                    order_status = row[4]
                    if qty and order_status != 'Cancelled':
                        price = float(row[16])
                        if sku in sku_qty:
                            sku_qty[sku] += qty
                            sku_price[sku] += price
                            if sku in sku_today_qty and if_today:
                                sku_today_qty[sku] += qty
                                sku_today_price[sku] += price
                        else:
                            sku_qty[sku] = qty
                            sku_price[sku] = price
                            if if_today:
                                sku_today_qty[sku] = qty
                                sku_today_price[sku] = price
        total_today_sales = 0.0
        for sku, qty in sku_qty.items():
            product, created = Product.objects.get_or_create(sku = sku)
            if created:
                product.save()
            sales_amount = sku_price[sku]
            sold_qty_average_7d = qty / 7
            average_price_7d = sales_amount / qty
            last7dayProductSales = Last7dayProductSales(product = product ,\
                                                        sold_qty = qty ,\
                                                        sales_amount = sales_amount ,\
                                                        sold_qty_average_7d = sold_qty_average_7d ,\
                                                        average_price_7d = average_price_7d)
            last7dayProductSales.save()
            if sku in sku_today_qty:
                todayProductSales = TodayProductSales(product = product ,\
                                                      sold_qty = sku_today_qty[sku] ,\
                                                      sales_amount = sku_today_price[sku] ,\
                                                      sold_qty_average_7d = sold_qty_average_7d ,\
                                                      average_price_7d = average_price_7d)
                fbaInventory = FbaInventory.objects.filter(sku = sku)
                if fbaInventory.count():
                    todayProductSales.fba_inventory = fbaInventory.all()[0]
                    todayProductSales.lasting_day_estimated_by_us = int(float(todayProductSales.fba_inventory.total_unit) / float(sold_qty_average_7d))
                todayProductSales.save()
                save_history_today_product_sales(todayProductSales,today_date_obj)
                total_today_sales += sku_today_price[sku]

        save_today_sales(year_month_day = year_month_day, sales_today = total_today_sales)
        return redirect('today_sales')

    template = loader.get_template('update_7d_orders.html')
    today_date_str = ""
    if HistoryTodayProductSales.objects.first() != None:
        today_date_str = HistoryTodayProductSales.objects.order_by('-date').first().date.strftime('%m%d')
    class UploadFileAndDateForm(forms.Form):
        date = forms.CharField(max_length=8)
        file = forms.FileField()
    context = {
        'form':UploadFileAndDateForm(initial={'date': today_date_str})
    }

    return HttpResponse(template.render(context, request))

@login_required
def update_restock_report(request):
    if request.method == 'POST':
        fileInMemory = request.FILES['file'].read().decode('cp1252')
        csv_data = csv.reader(StringIO(fileInMemory), delimiter='\t')
        for row in csv_data:
            sku = row[3]
            if sku != 'Merchant SKU':
                fnsku = row[2]
                asin = row[4]
                total_unit = int(row[12])
                available = int(row[14])
                inbound_fc_unit = int(row[13]) + int(row[15]) + int(row[16])
                inbound_unit = int(row[13])
                fc_unit = int(row[15]) + int(row[16])
                days_of_supply = row[20]
                try:
                    days_of_supply = int(days_of_supply)
                except:
                    days_of_supply = 365
                recommended_replenishment_qty = row[22]
                if recommended_replenishment_qty:
                    recommended_replenishment_qty = int(recommended_replenishment_qty)
                else:
                    recommended_replenishment_qty = 0
                recommended_ship_date = row[23]
                if not(recommended_ship_date):
                    recommended_ship_date = ''
                fbaInventory = FbaInventory.objects.filter(sku = sku)
                if fbaInventory.count():
                    fbaInventory = fbaInventory.all()[0]
                    fbaInventory.fnsku = fnsku
                    fbaInventory.asin = asin
                    fbaInventory.total_unit = total_unit
                    fbaInventory.available = available
                    fbaInventory.inbound_fc_unit = inbound_fc_unit
                    fbaInventory.inbound_unit = inbound_unit
                    fbaInventory.fc_unit = fc_unit
                    fbaInventory.days_of_supply = days_of_supply
                    fbaInventory.recommended_replenishment_qty = recommended_replenishment_qty
                    fbaInventory.recommended_ship_date = recommended_ship_date
                else:
                    fbaInventory = FbaInventory( \
                                               sku = sku ,\
                                               fnsku = fnsku ,\
                                               asin = asin ,\
                                               total_unit = total_unit ,\
                                               available = available ,\
                                               inbound_unit = inbound_unit ,\
                                               fc_unit = fc_unit ,\
                                               inbound_fc_unit = inbound_fc_unit ,\
                                               days_of_supply = days_of_supply ,\
                                               recommended_replenishment_qty = recommended_replenishment_qty ,\
                                               recommended_ship_date = recommended_ship_date

                    )
                fbaInventory.save()
                todayProductSales = TodayProductSales.objects.filter(product__sku = sku)
                if todayProductSales.count():
                    todayProductSales = todayProductSales.all()[0]
                    todayProductSales.fba_inventory = fbaInventory
                    todayProductSales.save()
        return redirect('today_sales')
    template = loader.get_template('update_restock_report.html')

    context = {
        'form':UploadFileForm()
    }

    return HttpResponse(template.render(context, request))
@login_required
def update_shenzhen_inventory(request):
    if request.method == 'POST':
        fileInMemory = request.FILES['file'].read()
        filePath = BytesIO(fileInMemory)
        wb2 = load_workbook(filePath)
        sheet = wb2.worksheets[0]
        max_column = sheet.max_column
        if sheet.cell(4,max_column - 1).value == shenzhen_warehouse_name:
            Inventory.objects.filter(warehouse_name = shenzhen_warehouse_name).delete()
            for row in range(6, sheet.max_row):
                sku = sheet.cell(row,1).value
                qty = int(sheet.cell(row,max_column - 1).value)
                shenzhenInventory =Inventory(warehouse_name = shenzhen_warehouse_name,sku = sku, qty = qty)
                shenzhenInventory.save()
    template = loader.get_template('update_shenzhen_inventory.html')
    context = {
        'form':UploadFileForm()
    }

    return HttpResponse(template.render(context, request))

@login_required
def update_purchasing_orders(request):
    if request.method == 'POST':
        wb2 = xlrd.open_workbook(file_contents=request.FILES['file'].read())
        sheet = wb2.sheet_by_index(0)
        max_column = sheet.ncols
        for col in range(0,max_column):
            title = sheet.cell(4,col).value
            if title == '未入库数量':
                qty_col = col
            if title == '状态':
                status_col = col
        ReceivablePurchasedQty.objects.all().delete()
        for row in range(5,sheet.nrows):
            sku_cell_value = sheet.cell(row,0).value
            if sku_cell_value:
                sku = sku_cell_value
            elif sheet.cell(row,status_col).value == '小计':
                receivable_qty = int(sheet.cell(row,qty_col).value)
                receivablePurchasedQty = ReceivablePurchasedQty(sku = sku, qty = receivable_qty)
                receivablePurchasedQty.save()
    template = loader.get_template('update_purchasing_orders.html')
    context = {
        'form':UploadFileForm()
    }

    return HttpResponse(template.render(context, request))

def month_to_date_last_year_month_to_date(year,month,day):
    dailySalesLastYear = DailySalesLastYear.objects.filter(month = month, day = day).all()[0]
    dailySalesLastYear = DailySalesLastYear.objects.get(pk = dailySalesLastYear.pk + 2)
    sales_month_to_date_last_year = DailySalesLastYear.objects.filter(month = month, day__lte = day).aggregate(Sum('sales'))['sales__sum']
    sales_month_to_date = HistoryTodaySales.objects.filter(date__range=(datetime.date(year, month, 1), datetime.date(year, month, day))).aggregate(Sum('sales_today'))['sales_today__sum']
    monthly_increase_on_sales = (sales_month_to_date - sales_month_to_date_last_year)/sales_month_to_date_last_year
    return {'sales_month_to_date': sales_month_to_date ,\
            'monthly_increase_on_sales': monthly_increase_on_sales}

def save_today_sales(**kwargs):
    year_month_day = kwargs['year_month_day']
    year = year_month_day['year']
    month = year_month_day['month']
    day = year_month_day['day']
    dailySalesLastYear = DailySalesLastYear.objects.filter(month = month, day = day).all()[0]
    dailySalesLastYear = DailySalesLastYear.objects.get(pk = dailySalesLastYear.pk + 2)
    today_date_obj = datetime.date(year = year, month = month, day = day)
    #检查当天HistoryTodaySales 是否已经建立
    if not HistoryTodaySales.objects.filter(date=today_date_obj).count() and 'sales_today' in kwargs.keys():
        historyTodaySales = HistoryTodaySales(date = today_date_obj ,\
                                            sales_today = kwargs['sales_today'] ,\
                                            sales_same_day_last_year = 999.99 ,\
                                            sales_month_to_date = 999.99 ,\
                                            monthly_increase_on_sales = 999.99 ,\
                                            ad_cost = 999.99 ,\
                                            acos = 999.99 ,\
                                            ad_cost_on_sales =999.99)
        historyTodaySales.save()
    else:
        historyTodaySales = HistoryTodaySales.objects.filter(date = today_date_obj).first()
        if 'sales_today' in kwargs.keys():
            historyTodaySales.sales_today = kwargs['sales_today']
    sales_month_to_date_and_monthly_increase_on_sales = month_to_date_last_year_month_to_date(year, month, day)

    historyTodaySales.sales_same_day_last_year = dailySalesLastYear.sales
    historyTodaySales.sales_month_to_date = sales_month_to_date_and_monthly_increase_on_sales['sales_month_to_date']
    historyTodaySales.monthly_increase_on_sales = sales_month_to_date_and_monthly_increase_on_sales['monthly_increase_on_sales']

    if 'ad_cost' in kwargs.keys():
        ad_cost = kwargs['ad_cost']
        acos = kwargs['acos']
        historyTodaySales.ad_cost = ad_cost
        historyTodaySales.acos = acos
        historyTodaySales.ad_cost_on_sales = ad_cost / historyTodaySales.sales_today
    historyTodaySales.save()
@login_required
def update_today_sales(request):
    if request.method == 'POST':
        ad_cost = float(request.POST['ad_cost'])
        acos = float(request.POST['acos'])
        date = request.POST['date']
        year_month_day = date_split(date)
        save_today_sales(year_month_day=year_month_day, ad_cost= ad_cost,acos= acos)
    template = loader.get_template('updateTodaySales.html')
    class TodaySalesForm(forms.Form):
        date = forms.CharField(max_length=8)
        ad_cost = forms.FloatField()
        acos = forms.FloatField()
    context = {
        'form':TodaySalesForm()
    }

    return HttpResponse(template.render(context, request))

def handle_uploaded_fba_shipment_file(file):
    fileInMemory = file.read().decode('utf-8')
    csv_data = csv.reader(StringIO(fileInMemory), delimiter='\t')
    start_read = False
    shipment_id = False
    shipment_name = False
    shipped_sku_qties_list = []
    for i, row in enumerate(csv_data):
        if i == 0:
            if row[0] in ['Shipment ID', '货件编号']:
                shipment_id = row[1]
        if i == 1:
            if row[0] in ['Name', '名称']:
                shipment_name = row[1]
        if row:
            if row[0] in ['Merchant SKU','卖家 SKU'] :
                for j,col in enumerate(row):
                    if col in ['Shipped', '已发货']:
                        shipped_sku_qty_col = j
                start_read = True
            elif start_read:
                shipped_sku_qties_list.append({'sku': row[0], 'qty': int(row[shipped_sku_qty_col])})
    if len(shipped_sku_qties_list) and shipment_id and shipment_name:
        fba_shipment = FbaShipment.objects.filter(shipment_id = shipment_id)
        if fba_shipment.count():
            fba_shipment = fba_shipment.first()
            fba_shipment.shipment_name = shipment_name
            fba_shipment.save()
        else:
            fba_shipment = FbaShipment(shipment_id = shipment_id \
                                    ,shipment_name = shipment_name)
            fba_shipment.save()
            for shipped_sku_qty_i in shipped_sku_qties_list:
                sku = shipped_sku_qty_i['sku']
                (product, found_or_not) = Product.objects.get_or_create(sku = sku)
                shipped_sku_qty = ShippedSkuQty(product = product, sku = sku, qty = shipped_sku_qty_i['qty'])
                shipped_sku_qty.save()
                fba_shipment.shipped_sku_qties.add(shipped_sku_qty)

@login_required
def update_fba_shipment(request):
    if request.method == 'POST':
        form = UploadFilesForm(request.POST, request.FILES)
        if form.is_valid():
            files = request.FILES.getlist('file_field')
            for f in files:
                handle_uploaded_fba_shipment_file(f)
    template = loader.get_template('update_fba_shipment.html')
    context = {
        'form':UploadFilesForm()
    }
    return HttpResponse(template.render(context, request))

@login_required
def update_fba_shipment_received_sku_qty(request):
    if request.method == 'POST':
        ReceivedSkuQty.objects.all().delete()
        col_nums = {'sku':0, 'quantity':0, 'fba-shipment-id': 0}
        fileInMemory = request.FILES['file'].read().decode('utf-8')
        csv_data = csv.reader(StringIO(fileInMemory), delimiter=',')
        for i, row in enumerate(csv_data):
            if i == 0:
                for col_num, title in enumerate(row):
                    if title in col_nums.keys():
                        col_nums[title] = col_num
            else:
                received_sku_qty, created = ReceivedSkuQty.objects.get_or_create(shipment_id = row[col_nums['fba-shipment-id']] \
                                                                                        ,sku = row[col_nums['sku']])
                received_sku_qty.qty += int(row[col_nums['quantity']])
                received_sku_qty.save()

    template = loader.get_template('update_fba_shipment_received_sku_qty.html')
    context = {
        'form':UploadFileForm()
    }
    return HttpResponse(template.render(context, request))

@login_required
def update_fba_shipment_estimated_receiving_date(request):
    if request.method == 'POST':
        import json
        post_json = json.loads(request.body)
        date_type = post_json['date_type']
        date_str = post_json['date']
        date_obj = datetime.datetime.strptime(date_str, '%Y/%m/%d').date()
        fba_shipment_id = post_json['fba_shipment_id']
        fba_shipment = FbaShipment.objects.filter(shipment_id = fba_shipment_id).first()
        if fba_shipment:
            if date_type == 'ship_date':
                fba_shipment.shipped_date = date_obj
            elif date_type == 'estimated_receiving_date':
                fba_shipment.estimated_receiving_date = date_obj
            fba_shipment.save()
        return HttpResponse(request.POST.get('a'))

    template = loader.get_template('update_fba_shipment_estimated_receiving_date.html')
    fba_shipments = FbaShipment.objects.filter(closed = False).all()

    context = {
        'fba_shipments': fba_shipments
    }
    return HttpResponse(template.render(context, request))

@login_required
def estimated_sku_qty_receiving_date(request):
    template = loader.get_template('estimated_sku_qty_receiving_date.html')

    fba_shipments = FbaShipment.objects.filter(closed = False).order_by('estimated_receiving_date')
    shipped_sku_qty_by_shipment = {}
    class ShippedSkuQtyByShipment():
        def __init__(self, qty, date, shipment_name, shipment_id):
            self.qty = qty
            self.date = date
            self.shipment_name = shipment_name
            self.shipment_id = shipment_id
    shipment_closed_dict = {}
    for fba_shipment in fba_shipments:
        shipment_id = fba_shipment.shipment_id
        shipment_name = fba_shipment.shipment_name
        shipment_closed_dict[shipment_id] = True
        if fba_shipment.estimated_receiving_date:
            estimated_receiving_date = fba_shipment.estimated_receiving_date.strftime('%m/%d')
        else:
            estimated_receiving_date = "无"
        shipped_sku_qties = fba_shipment.shipped_sku_qties.all()
        for shipped_sku_qty in shipped_sku_qties:
            sku = shipped_sku_qty.sku
            shipped_qty = shipped_sku_qty.qty
            received_sku_qty = ReceivedSkuQty.objects.filter(shipment_id = shipment_id \
                                                        ,sku = sku).aggregate(Sum('qty'))['qty__sum']

            if received_sku_qty == None:
                received_sku_qty = 0

            unreceived_qty = shipped_qty - received_sku_qty
            if unreceived_qty > 15:
                if sku in shipped_sku_qty_by_shipment.keys():
                    shipped_sku_qty_by_shipment[sku].append(ShippedSkuQtyByShipment(shipment_name = shipment_name \
                                                                                    , shipment_id = shipment_id \
                                                                                    ,qty =  unreceived_qty \
                                                                                    , date = estimated_receiving_date))
                else:
                    shipped_sku_qty_by_shipment[sku] = [ShippedSkuQtyByShipment(shipment_name = shipment_name \
                                                                                , shipment_id = shipment_id \
                                                                                , qty = unreceived_qty \
                                                                                , date = estimated_receiving_date)]
                shipment_closed_dict[shipment_id] = False

        for k,v in shipment_closed_dict.items():
            if v:
                fba_shipment = FbaShipment.objects.get(shipment_id = k)
                fba_shipment.closed = True
                fba_shipment.save()
    context = {
        'to_be_received_skus':shipped_sku_qty_by_shipment
    }
    return HttpResponse(template.render(context, request))

@login_required
def get_estimated_sku_qty_receiving_date_of_a_sku(request):
    sku = request.GET['sku']
    fba_shipments = FbaShipment.objects.filter(closed = False, shipped_sku_qties__sku=sku).order_by('estimated_receiving_date')
    receivable_qty_list = []
    for fba_shipment in fba_shipments:
        shipment_id = fba_shipment.shipment_id
        shipment_name = fba_shipment.shipment_name
        if fba_shipment.estimated_receiving_date:
            estimated_receiving_date = fba_shipment.estimated_receiving_date.strftime('%m/%d')
        else:
            estimated_receiving_date = "无"
        shipped_sku_qties = fba_shipment.shipped_sku_qties.all()
        for shipped_sku_qty in shipped_sku_qties:
            if shipped_sku_qty.sku == sku:
                shipped_qty = shipped_sku_qty.qty
                received_sku_qty = ReceivedSkuQty.objects.filter(shipment_id = shipment_id \
                                                            ,sku = sku).aggregate(Sum('qty'))['qty__sum']

                if received_sku_qty == None:
                    received_sku_qty = 0

                unreceived_qty = shipped_qty - received_sku_qty
                if unreceived_qty > 15:
                    receivable_qty_list.append({'shipment_id': shipment_id \
                                                                ,'date': estimated_receiving_date \
                                                                , 'qty': unreceived_qty \
                                                                , 'shipment_name':shipment_name})
    return JsonResponse(receivable_qty_list, safe=False)
@login_required
def get_history_sales_of_a_sku(request):
    sku = request.GET['sku']
    history_today_product_sales = HistoryTodayProductSales.objects.filter(product__sku = sku).order_by('-date')[:7]
    history_today_product_sales_list = [{'date':i.date.strftime('%m/%d') \
                                        ,'qty':round(i.sold_qty_average_7d,1)} for i in history_today_product_sales]
    history_today_product_sales_list.reverse()
    return JsonResponse(history_today_product_sales_list, safe=False)


def check_receivable_purchased_qty(sku):
    receivablePurchasedQty = ReceivablePurchasedQty.objects.filter(sku = sku)
    if receivablePurchasedQty.count():
        receivablePurchasedQty = receivablePurchasedQty.all()[0].qty
    else:
        receivablePurchasedQty = 0
    return receivablePurchasedQty

def check_shenzhen_inventory(sku):
    shenzhen_inventory = Inventory.objects.filter(sku = sku, warehouse_name = shenzhen_warehouse_name)
    if shenzhen_inventory.count():
        shenzhen_inventory = shenzhen_inventory.first().qty
    else:
        shenzhen_inventory = 0
    return shenzhen_inventory

def if_count_for_po(todayProductSale):
    maximum_days_lasting = 100
    sku = todayProductSale.product.sku
    amazon_inventory = todayProductSale.fba_inventory.total_unit
    shenzhen_inventory = check_shenzhen_inventory(sku)
    lasting_day_of_total_fba_and_shenzhen_inventory_estimated_by_us = (todayProductSale.fba_inventory.total_unit + shenzhen_inventory)/ todayProductSale.sold_qty_average_7d
    if lasting_day_of_total_fba_and_shenzhen_inventory_estimated_by_us < maximum_days_lasting:
        return True
    if todayProductSale.fba_inventory.days_of_supply < maximum_days_lasting:
        return True
    return False

@login_required
def restock_today(request):

    if request.method == 'POST':
        wb = Workbook()
        sheet = wb.active
        sheet.title = '今日生产单'
        row = 1
        max_months_to_last_for_pos = [4, 5, 6]
        moq = 200
        min_days_lasting_for_notice = 60
        max_days_to_last_for_pos = [max_months_to_last_for_po * 30 for max_months_to_last_for_po in max_months_to_last_for_pos]
        titles = ['SKU', 'FBA 可售库存', '中转库存', '在途库存', '亚马逊总库存', 'F199库存', '过去7天平均日销售' \
                , '亚马逊预测能撑多少天', '我们预测能撑多少天', '单价', '亚马逊建议补货量', '已经下过购货订单' \
                ]
        start_col_restock = len(titles) + 1
        for max_months_to_last_for_po in max_months_to_last_for_pos:
            titles.append('撑到%i个月生产单建议量' %max_months_to_last_for_po)
        for i, title in enumerate(titles):
            sheet.cell(row,i + 1).value = title

        yellow_fill = PatternFill("solid", fgColor="FFFF00")
        todayProductSales = TodayProductSales.objects.all()
        row +=1
        for todayProductSale in todayProductSales:
            if if_count_for_po(todayProductSale):
                sku = todayProductSale.product.sku
                sheet.cell(row,1).value = sku
                sheet.cell(row,2).value = todayProductSale.fba_inventory.available
                sheet.cell(row,3).value = todayProductSale.fba_inventory.fc_unit
                sheet.cell(row,4).value = todayProductSale.fba_inventory.inbound_unit
                fba_total_unit = todayProductSale.fba_inventory.total_unit
                sheet.cell(row,5).value = fba_total_unit
                shenzhen_inventory = check_shenzhen_inventory(sku)
                sheet.cell(row,6).value = shenzhen_inventory
                sheet.cell(row,7).value = round(todayProductSale.sold_qty_average_7d,1)
                sheet.cell(row,8).value = todayProductSale.fba_inventory.days_of_supply
                lasting_day_of_total_fba_and_shenzhen_inventory_estimated_by_us = (todayProductSale.fba_inventory.total_unit + shenzhen_inventory)/ todayProductSale.sold_qty_average_7d
                sheet.cell(row,9).value = int(lasting_day_of_total_fba_and_shenzhen_inventory_estimated_by_us)
                sheet.cell(row,10).value = round(todayProductSale.average_price_7d,2)
                sheet.cell(row,10).number_format = u'"$ "#,##0.00'
                amazon_recommended_replenishment_qty = todayProductSale.fba_inventory.recommended_replenishment_qty
                sheet.cell(row,11).value = amazon_recommended_replenishment_qty
                receivable_purchased_qty = check_receivable_purchased_qty(sku)
                sheet.cell(row,12).value = receivable_purchased_qty
                if todayProductSale.product.discontinued:
                    continue #sheet.cell(row,13).value = "不再订货"
                else:
                    for i, max_days_to_last_for_po in enumerate(max_days_to_last_for_pos):
                        po_qty = 0
                        if todayProductSale.fba_inventory.days_of_supply >= lasting_day_of_total_fba_and_shenzhen_inventory_estimated_by_us:

                            days_vacancy = max_days_to_last_for_po - lasting_day_of_total_fba_and_shenzhen_inventory_estimated_by_us
                            qty_needed = int(days_vacancy * todayProductSale.sold_qty_average_7d)
                            if qty_needed > receivable_purchased_qty:
                                qty_needed = qty_needed - receivable_purchased_qty
                                if qty_needed > float(moq) / 2:
                                    po_qty = max([qty_needed, moq])
                        else:
                            shenzhen_inventory_and_receivable_purchased_qty = shenzhen_inventory + receivable_purchased_qty
                            if i > 0:
                                amazon_recommended_replenishment_qty_i = amazon_recommended_replenishment_qty + float(max_days_to_last_for_pos[i] - max_days_to_last_for_pos[0])/ todayProductSale.fba_inventory.days_of_supply * fba_total_unit
                            else:
                                amazon_recommended_replenishment_qty_i = amazon_recommended_replenishment_qty
                            if amazon_recommended_replenishment_qty_i > shenzhen_inventory_and_receivable_purchased_qty:
                                qty_needed = amazon_recommended_replenishment_qty_i - shenzhen_inventory_and_receivable_purchased_qty
                                if qty_needed > float(moq) / 2:
                                    po_qty = max([qty_needed, moq])
                        if po_qty:
                            sheet.cell(row,start_col_restock + i).value = int(po_qty)

                    if min([todayProductSale.fba_inventory.days_of_supply, lasting_day_of_total_fba_and_shenzhen_inventory_estimated_by_us]) < min_days_lasting_for_notice and not(todayProductSale.product.discontinued):
                        for cell in sheet["%i:%i" %(row,row)]:
                            cell.fill = yellow_fill

                row +=1
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename={date}-PO.xlsx'.format( \
               date=datetime.datetime.now().strftime('%Y-%m-%d') \
               ,)
        wb.save(response)
        return response
    template = loader.get_template('restock_today.html')
    context = {
    }
    return HttpResponse(template.render(context, request))

@login_required
def shipment_today(request):
    if request.method == 'POST':
        inbound_sku_qty = {}
        form = UploadFileForm(request.POST,request.FILES)
        if form.is_valid():
            fileInMemory = request.FILES['file'].read()
            filePath = BytesIO(fileInMemory)
            wb2 = load_workbook(filePath)
            sheet = wb2.worksheets[0]
            for row in range(1, sheet.max_row + 1):
                sku = sheet.cell(row,1).value
                qty = int(sheet.cell(row,2).value)
                inbound_sku_qty[sku]= qty
            print(inbound_sku_qty)
        else:
            print('is not valid')
        wb = Workbook()
        sheet = wb.active
        sheet.title = '今日发货单'
        yellow_fill = PatternFill("solid", fgColor="FFFF00")
        min_shippable_qty = 20
        min_sold_qty_per_day_for_notice = 1
        max_days_lasting = 120
        row = 1
        titles = ['SKU', 'FBA可售库存', '中转库存', '在途库存', '亚马逊总库存', 'F199库存' \
                    , '过去7天平均日销售', '亚马逊预测能撑多少天', '我们预测总fba库存能撑多少天' \
                    , '可售库存能撑多少天','可售+中转库存能撑多少天' \
                    , '单价', '决策' \
                    ]
        for i, title in enumerate(titles):
            sheet.cell(row,i + 1).value = title
        row += 1
        inventories = Inventory.objects.filter(warehouse_name = shenzhen_warehouse_name, qty__gte = min_shippable_qty)
        for inventory in inventories.all():
            sku = inventory.sku
            shenzhen_inventory = check_shenzhen_inventory(sku)
            if sku in inbound_sku_qty:
                shenzhen_inventory += inbound_sku_qty[sku]
                inbound_sku_qty.pop(sku, None)
            sheet.cell(row,1).value = sku
            todayProductSales = TodayProductSales.objects.filter(product__sku = sku)
            if todayProductSales.count():
                todayProductSale = todayProductSales.all()[0]
                sheet.cell(row,2).value = todayProductSale.fba_inventory.available
                sheet.cell(row,3).value = todayProductSale.fba_inventory.fc_unit
                sheet.cell(row,4).value = todayProductSale.fba_inventory.inbound_unit
                fba_total_unit = todayProductSale.fba_inventory.total_unit
                sheet.cell(row,5).value = fba_total_unit
                sheet.cell(row,6).value = shenzhen_inventory
                sheet.cell(row,7).value = round(todayProductSale.sold_qty_average_7d,1)
                sheet.cell(row,8).value = todayProductSale.fba_inventory.days_of_supply
                sheet.cell(row,9).value = int(todayProductSale.lasting_day_of_total_fba_unit_estimated_by_us)
                sheet.cell(row,10).value = int(todayProductSale.lasting_day_of_available_estimated_by_us)
                sheet.cell(row,11).value = int(todayProductSale.lasting_day_of_available_fc_estimated_by_us)
                sheet.cell(row,12).value = round(todayProductSale.average_price_7d,2)
                sheet.cell(row,12).number_format = u'"$ "#,##0.00'
                if min([todayProductSale.fba_inventory.days_of_supply, todayProductSale.lasting_day_of_total_fba_unit_estimated_by_us]) < max_days_lasting and todayProductSale.sold_qty_average_7d > min_sold_qty_per_day_for_notice:
                    for cell in sheet["%i:%i" %(row,row)]:
                        cell.fill = yellow_fill
            else:
                fbaInventory = FbaInventory.objects.filter(sku = sku)
                if fbaInventory.count():
                    fba_inventory = fbaInventory.all()[0]
                    sheet.cell(row,2).value = fba_inventory.available
                    sheet.cell(row,3).value = fba_inventory.fc_unit
                    sheet.cell(row,4).value = fba_inventory.inbound_unit
                    fba_total_unit = fba_inventory.total_unit
                    sheet.cell(row,5).value = fba_total_unit
                    shenzhen_inventory = check_shenzhen_inventory(sku)
                    sheet.cell(row,6).value = shenzhen_inventory
                    sheet.cell(row,8).value = fba_inventory.days_of_supply
            row +=1
        for sku, qty in inbound_sku_qty.items():
            sheet.cell(row,1).value = sku
            todayProductSales = TodayProductSales.objects.filter(product__sku = sku)
            if todayProductSales.count():
                todayProductSale = todayProductSales.all()[0]
                sheet.cell(row,2).value = todayProductSale.fba_inventory.available
                sheet.cell(row,3).value = todayProductSale.fba_inventory.fc_unit
                sheet.cell(row,4).value = todayProductSale.fba_inventory.inbound_unit
                fba_total_unit = todayProductSale.fba_inventory.total_unit
                sheet.cell(row,5).value = fba_total_unit
                shenzhen_inventory = check_shenzhen_inventory(sku) + qty
                sheet.cell(row,6).value = shenzhen_inventory
                sheet.cell(row,7).value = round(todayProductSale.sold_qty_average_7d,1)
                sheet.cell(row,8).value = todayProductSale.fba_inventory.days_of_supply
                sheet.cell(row,9).value = int(todayProductSale.lasting_day_of_total_fba_unit_estimated_by_us)
                sheet.cell(row,10).value = int(todayProductSale.lasting_day_of_available_estimated_by_us)
                sheet.cell(row,11).value = int(todayProductSale.lasting_day_of_available_fc_estimated_by_us)
                sheet.cell(row,12).value = round(todayProductSale.average_price_7d,2)
                sheet.cell(row,12).number_format = u'"$ "#,##0.00'
                if min([todayProductSale.fba_inventory.days_of_supply, todayProductSale.lasting_day_of_total_fba_unit_estimated_by_us]) < max_days_lasting and todayProductSale.sold_qty_average_7d > min_sold_qty_per_day_for_notice:
                    for cell in sheet["%i:%i" %(row,row)]:
                        cell.fill = yellow_fill
            else:
                fbaInventory = FbaInventory.objects.filter(sku = sku)
                if fbaInventory.count():
                    fba_inventory = fbaInventory.all()[0]
                    sheet.cell(row,2).value = fba_inventory.available
                    sheet.cell(row,3).value = fba_inventory.fc_unit
                    sheet.cell(row,4).value = fba_inventory.inbound_unit
                    sheet.cell(row,5).value = fba_inventory.total_unit
                    shenzhen_inventory = check_shenzhen_inventory(sku) + qty
                    sheet.cell(row,6).value = shenzhen_inventory
                    sheet.cell(row,8).value = fba_inventory.days_of_supply
                else:
                    sheet.cell(row,6).value = qty
            row +=1
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename={date}-fbaShipment.xlsx'.format( \
               date=datetime.datetime.now().strftime('%Y-%m-%d') \
               ,)
        wb.save(response)
        return response

    template = loader.get_template('shipment_today.html')
    context = {
        'form':UploadFileForm()
    }
    return HttpResponse(template.render(context, request))


@login_required
def reprice_today(request):
    min_available_qty_for_adjust_price = 10
    standard_price = 25.99
    highest_price = 34.99

    template = loader.get_template('reprice_today.html')

    titles = ['产品图片', 'SKU', 'FBA可售库存', '中转库存', '在途库存', '亚马逊总库存', 'F199库存', '亚马逊预测撑天' \
            , '我们预测总fba库存撑天', '可售库存撑天','可售+中转库存撑天' \
            , '过去7天平均日销售', '今日销量', '正常价', '最高限价', '单价', '决策' \
            ]

    products_to_reprice_list = []

    #找出所有FBA有库存的商品
    for fba_inventory in FbaInventory.objects.filter(available__gt = min_available_qty_for_adjust_price):

        sku = fba_inventory.sku
        product_to_reprice_info = [""]
        product_to_reprice_info.append("")
        product_to_reprice_info.append(sku)
        fba_inventory_available = fba_inventory.available
        fba_inventory_fc_unit = fba_inventory.fc_unit
        fba_inventory_inbound_unit = fba_inventory.inbound_unit
        fba_total_unit = fba_inventory.total_unit
        shenzhen_inventory = check_shenzhen_inventory(sku)
        fba_inventory_days_of_supply = fba_inventory.days_of_supply
        product_to_reprice_info.append(fba_inventory_available)
        product_to_reprice_info.append(fba_inventory_fc_unit)
        product_to_reprice_info.append(fba_inventory_inbound_unit)
        product_to_reprice_info.append(fba_total_unit)
        product_to_reprice_info.append(shenzhen_inventory)
        product_to_reprice_info.append(fba_inventory_days_of_supply)


        #降价部分
        #判断价格是否高于正常价
        #先找到最新售价
        lastest_sales_having_this_sku = HistoryTodayProductSales.objects.filter(product__sku = sku).order_by('-date')

        if lastest_sales_having_this_sku.count():

            lastest_sales_having_this_sku = lastest_sales_having_this_sku.first()
            sold_qty_average_7d = lastest_sales_having_this_sku.sold_qty_average_7d
            average_price_today = lastest_sales_having_this_sku.average_price_7d
            if lastest_sales_having_this_sku.sales_amount != 0 and lastest_sales_having_this_sku.sold_qty != 0:
                average_price_today = lastest_sales_having_this_sku.sales_amount / lastest_sales_having_this_sku.sold_qty
            lasting_days_with_fba_inventory_available  = fba_inventory_available / sold_qty_average_7d
            lasting_days_with_fba_inventory_available_and_fc_unit = (fba_inventory_available + fba_inventory_fc_unit)/ sold_qty_average_7d
            lasting_days_with_fba_total_unit = fba_total_unit / sold_qty_average_7d
            sold_qty_average_7d = lastest_sales_having_this_sku.sold_qty_average_7d
            today_sold_qty = 0
            todayProductSales = TodayProductSales.objects.filter(product__sku = sku)
            if todayProductSales.count():
                todayProductSales = todayProductSales.first()
                today_sold_qty = todayProductSales.sold_qty
            product_to_reprice_info.append(int(lasting_days_with_fba_total_unit))
            product_to_reprice_info.append(int(lasting_days_with_fba_inventory_available))
            product_to_reprice_info.append(int(lasting_days_with_fba_inventory_available_and_fc_unit))
            product_to_reprice_info.append(round(sold_qty_average_7d,1))
            product_to_reprice_info.append(today_sold_qty)
            product_to_reprice_info.append(standard_price)
            product_to_reprice_info.append(highest_price)
            product_to_reprice_info.append(round(average_price_today, 2))

            if int(average_price_today) > int(standard_price):
                #高于正常价
                #判断销售是否在下降
                if today_sold_qty <= sold_qty_average_7d - 1:
                    #销售在下降，则要降价
                    #降价前还是要判断是否有断货风险
                    if lasting_days_with_fba_total_unit > 30 and lasting_days_with_fba_inventory_available_and_fc_unit > 20 :
                        #没有断货风险
                        product = Product.objects.get(sku = sku)
                        product_to_reprice_info[1] = product.image.url
                        product_to_reprice_info.append( '降价' )
                        product_to_reprice_info[0] = "blue"
                        products_to_reprice_list.append(product_to_reprice_info)
                elif lasting_days_with_fba_inventory_available_and_fc_unit < 20:
                    #可售+中转库存撑不到20天
                    if today_sold_qty >= sold_qty_average_7d:
                        product_to_reprice_info.append( '涨价' )
                        product_to_reprice_info[0] = "red"
                    product = Product.objects.get(sku = sku)
                    product_to_reprice_info[1] = product.image.url
                    products_to_reprice_list.append(product_to_reprice_info)
            elif lasting_days_with_fba_inventory_available_and_fc_unit < 20 and today_sold_qty > sold_qty_average_7d:
                #不高于正常价但是可售+中转库存撑不到20天，而且销售在上升
                product = Product.objects.get(sku = sku)
                product_to_reprice_info[1] = product.image.url
                product_to_reprice_info.append( '涨价' )
                product_to_reprice_info[0] = "red"
                products_to_reprice_list.append(product_to_reprice_info)
    context = {
        'titles': titles ,\
        'products_to_reprice_list': products_to_reprice_list \
    }
    return HttpResponse(template.render(context, request))

def scrape(url):
    headers = {
        'authority': 'www.amazon.com',
        'pragma': 'no-cache',
        'cache-control': 'no-cache',
        'dnt': '1',
        'upgrade-insecure-requests': '1',
        'user-agent': 'Mozilla/5.0 (X11; CrOS x86_64 8172.45.0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/51.0.2704.64 Safari/537.36',
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
        'sec-fetch-site': 'none',
        'sec-fetch-mode': 'navigate',
        'sec-fetch-dest': 'document',
        'accept-language': 'en-GB,en-US;q=0.9,en;q=0.8',
    }

    # Download the page using requests
    print("Downloading %s"%url)
    r = requests.get(url, headers=headers)
    # Simple check to check if page was blocked (Usually 503)
    if r.status_code > 500:
        if "To discuss automated access to Amazon data please contact" in r.text:
            print("Page %s was blocked by Amazon. Please try using better proxies\n"%url)
        else:
            print("Page %s must have been blocked by Amazon as the status code was %d"%(url,r.status_code))
        return None
    # Pass the HTML of the page and create
    f = open(os.path.join(BASE_DIR,'amazon_response.html'), "w")
    f.write(r.text)
    f.close()
    data = extractor.extract(r.text,base_url=url)
    reviews = []
    for r in data['reviews']:
        r["product"] = data["product_title"]
        r['url'] = url
        if 'verified_purchase' in r:
            if r['verified_purchase'] and 'Verified Purchase' in r['verified_purchase']:
                r['verified_purchase'] = True
            else:
                r['verified_purchase'] = False
        r['rating'] = r['rating'].split(' out of')[0]
        date_posted = r['date'].split('on ')[-1]
        if r['images']:
            r['images'] = "\n".join(r['images'])
        r['date'] = dateparser.parse(date_posted).strftime('%d %b %Y')
        reviews.append(r)
    data.pop('histogram', None)
    # histogram = {}
    # for h in data['histogram']:
    #     histogram[h['key']] = h['value']
    # data['histogram'] = histogram
    data['average_rating'] = float(data['average_rating'].split(' out')[0])
    data['reviews'] = reviews
    data['number_of_reviews'] = int(data['number_of_reviews'].split('  customer')[0])
    return data

def distinct_size_color(size_color_str):
    Size_Tip_Str = ['Size: ']
    Color_Tip_Str = ['Color: ', 'Colour: ']
    pos_dict = {}
    size_pos = color_pos =-1
    for k in Size_Tip_Str:
        size_pos = size_color_str.find(k)
        if size_pos > -1:
            pos_dict['size'] = {'tip_str':k, 'pos':size_pos}
            break
    for k in Color_Tip_Str:
        color_pos = size_color_str.find(k)
        if color_pos > -1:
            pos_dict['color'] = {'tip_str':k, 'pos':color_pos}
            break
    if len(pos_dict.keys()) == 1:
        return {list(pos_dict.keys())[0]: size_color_str[len(list(pos_dict.values())[0]['tip_str']):]}
    elif len(pos_dict.keys()) > 1:
        if pos_dict['size']['pos'] > pos_dict['color']['pos']:
            return {'color': size_color_str[len(pos_dict['color']['tip_str']):pos_dict['size']['pos']] \
                    ,'size': size_color_str[(pos_dict['size']['pos'] + len(pos_dict['size']['tip_str'])):]}
        else:
            return {'size': size_color_str[len(pos_dict['size']['tip_str']):pos_dict['color']['pos']] \
                    ,'color': size_color_str[(pos_dict['color']['pos'] + len(pos_dict['color']['tip_str'])):]}

def generate_pivot_table(data_list):
    variants = []
    for i in data_list:
        if 'variant' in i and i['variant'] != None:
            variants.append(i['variant'])
    return

def size_color_counter_dict_to_list(x):
    res1 = []
    res2 = []
    for k,v in x.items():
        res1.append(k)
        res2.append(v)
    return {'name': res1, 'value': res2}
@login_required
def get_reviews(request):
    if request.method == 'GET' and 'asin' in request.GET:
        import math
        from collections import Counter
        asin = request.GET['asin']
        max_page_number = 100
        reviews_by_size_color = []
        for page_number in range(1, max_page_number):
            url = 'https://www.amazon.com/product-reviews/%s/?ie=UTF8&reviewerType=all_reviews&sortBy=recent&pageNumber=%i' %(asin, page_number)
            scraped_data = scrape(url)
            for review in scraped_data['reviews']:
                if 'variant' in review and review['variant']:
                    reviews_by_size_color.append(distinct_size_color(review['variant']))
            counter_size_color = {}
            if len(reviews_by_size_color[0].keys()) == 1:
                size_color = list(reviews_by_size_color[0].keys())[0]
                size_color_list = [i[size_color] for i in reviews_by_size_color]
                counter_size_color[size_color] = dict(Counter(size_color_list))
                counter_size_color[size_color] = size_color_counter_dict_to_list(counter_size_color[size_color])

            elif len(reviews_by_size_color[0].keys()) > 1:

                for i_size_color in list(reviews_by_size_color[0].keys()):
                    size_color_list = [i[i_size_color] for i in reviews_by_size_color]
                    counter_size_color[i_size_color] = dict(Counter(size_color_list))
                    counter_size_color[i_size_color] = size_color_counter_dict_to_list(counter_size_color[i_size_color])
            next_page = scraped_data['next_page']
            if next_page == None:
                break

        return JsonResponse(counter_size_color, safe=False)
    template = loader.get_template('get_reviews.html')
    context = {
        'form':AsinForm()
    }
    return HttpResponse(template.render(context, request))
