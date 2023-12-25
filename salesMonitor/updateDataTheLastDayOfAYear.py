from django.http import HttpResponse
from django.shortcuts import redirect
from django.template import loader
from django import forms
from django.views.generic.edit import FormView
from .forms import UploadFilesForm, UploadFileForm, AsinForm, SkuForm, UploadShipmentFileForm ,\
                   UploadTransactionAdCurrencyForm, ConfirmPoHeadShippingForm, UploadFileCountryForm
from django.contrib.auth.decorators import login_required
from io import BytesIO,StringIO
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
import xlrd
import datetime
import pytz
import csv
from django.db.models import Sum
from django.http import JsonResponse
from salesMonitor.models import Product, TodayProductSales, Last7dayProductSales, \
                                DailySalesLastYear, FbaInventory, \
                                ReceivablePurchasedQty, HistoryTodayProductSales,\
                                HistoryTodaySales, Inventory, FbaShipment, ShippedSkuQty ,\
                                ReceivedSkuQty, Upc, SkuUpc, Supplier, SkuSupplier, UserSupplier ,\
                                SkuPurchasingPrice, SkuHeadShippingUnitCost, SkuAssetLiabilityTable ,\
                                SkuManagedBySalesPerson, SkuContributor, SkuWeight, FbaShipmentCost ,\
                                SkuPurchaseOrder, ProfitLossTable, \
                                CurrencyRate

from salesMonitor.excelReadData import readSkuValue, readTransaction \
                                     ,readCurrencyRate \
                                     ,readFixedCost  \


from dateutil import parser as dateparser

# Build paths inside the project like this: os.path.join(BASE_DIR, ...)

shenzhen_warehouse_name = '深圳A016'
@login_required
def index(request):
    if request.user.groups.filter(name = 'supplier').exists():
        template = loader.get_template('todaySalesForSupplier.html')
    else:
        template = loader.get_template('todaySales.html')
    default_country = 'US'
    todaySales = HistoryTodaySales.objects.filter(country=default_country).order_by('-date')[0]

    if request.user.groups.filter(name = 'supplier').exists():
        supplier_id_list = [user_supplier.supplier_id for user_supplier in UserSupplier.objects.filter(user = request.user).all()]
        supplier_sku_list = [i.sku for i in SkuSupplier.objects.filter(supplier_id__in = supplier_id_list).all()]
        supplier_sku_sales_data = TodayProductSales.objects.filter(product__sku__in=supplier_sku_list, country=default_country).order_by('-sold_qty')
        total_sold_unit = supplier_sku_sales_data.aggregate(Sum('sold_qty'))['sold_qty__sum']
        country = default_country
        context = {
            'date': todaySales.date.strftime('%Y/%m/%d') ,\
            'total_sold_unit': total_sold_unit ,\
            'supplier_sku_sales_data': supplier_sku_sales_data ,\
            'country': country ,\
            'today': True
        }
    else:
        class TopSku:
            def __init__(self, sku, sold_qty):
                self.sku = sku
                self.sold_qty = sold_qty
        top_sales_skus = TodayProductSales.objects.filter(country=default_country).order_by('-sales_amount')[:5]
        top_sold_qty_skus = TodayProductSales.objects.filter(country=default_country).order_by('-sold_qty')[:5]
        sales_today = todaySales.sales_today
        country = default_country
        countries = list(set([i.country for i in TodayProductSales.objects.all()]))
        countries = [c for c in countries if c != country]
        new_or_all = True
        if 'new_or_all' in request.GET and request.GET['new_or_all']:
            new_or_all = bool(int(request.GET['new_or_all']))
        if new_or_all:
            all_today_new_product_skus = TodayProductSales.objects.filter(product__new = True, country=default_country).order_by('-sales_amount')
            sales_new_product = TodayProductSales.objects.filter(product__new = True, country=default_country).aggregate(Sum('sales_amount'))['sales_amount__sum']
            sales_new_product_percent = int(sales_new_product / sales_today * 100)
            context = {
                'date': todaySales.date.strftime('%Y/%m/%d') ,\
                'sales_today':  sales_today,\
                'sales_same_day_last_year': todaySales.sales_same_day_last_year ,\
                'monthly_increase_on_sales': todaySales.monthly_increase_on_sales  * 100,\
                'ad_cost': todaySales.ad_cost ,\
                'acos': todaySales.acos * 100 ,\
                'ad_cost_on_sales': todaySales.ad_cost_on_sales * 100 ,\
                'top_sales_skus': top_sales_skus ,\
                'top_sold_qty_skus': top_sold_qty_skus ,\
                'all_today_new_product_skus' : all_today_new_product_skus ,\
                'sales_new_product_percent' : sales_new_product_percent ,\
                'country': country ,\
                'countries': countries ,\
                'today': True ,\
                'new_or_all': new_or_all \
            }
        else:
            all_today_product_skus = TodayProductSales.objects.filter(country=default_country).order_by('-sales_amount')
            context = {
                'date': todaySales.date.strftime('%Y/%m/%d') ,\
                'sales_today':  sales_today,\
                'sales_same_day_last_year': todaySales.sales_same_day_last_year ,\
                'monthly_increase_on_sales': todaySales.monthly_increase_on_sales  * 100,\
                'ad_cost': todaySales.ad_cost ,\
                'acos': todaySales.acos * 100 ,\
                'ad_cost_on_sales': todaySales.ad_cost_on_sales * 100 ,\
                'top_sales_skus': top_sales_skus ,\
                'top_sold_qty_skus': top_sold_qty_skus ,\
                'all_today_product_skus': all_today_product_skus ,\
                #'all_today_new_product_skus' : all_today_new_product_skus ,\
                #'sales_new_product_percent' : sales_new_product_percent ,\
                'country': country ,\
                'countries': countries ,\
                'today': True ,\
                'new_or_all': new_or_all \
            }
    return HttpResponse(template.render(context, request))

@login_required
def other_country_today_sales(request, country):
    template = loader.get_template('other_country_today_sales_todaySales.html')
    default_country = country
    todaySales = HistoryTodaySales.objects.filter(country=default_country).order_by('-date')[0]

    if request.user.groups.filter(name = 'supplier').exists():
        context = {}
        # user_supplier = UserSupplier.objects.filter(user = request.user).all()[0]
        # supplier = user_supplier.supplier
        # suplier_sku_list = [i.sku for i in SkuSupplier.objects.filter(supplier = supplier).all()]
        # supplier_sku_sales_data = TodayProductSales.objects.filter(product__sku__in=suplier_sku_list, country='US').order_by('-sold_qty')
        # total_sold_unit = supplier_sku_sales_data.aggregate(Sum('sold_qty'))['sold_qty__sum']
        # context = {
        #     'date': todaySales.date.strftime('%Y/%m/%d') ,\
        #     'total_sold_unit': total_sold_unit ,\
        #     'supplier_sku_sales_data': supplier_sku_sales_data ,\
        #     'today': True
        # }
    else:
        class TopSku:
            def __init__(self, sku, sold_qty):
                self.sku = sku
                self.sold_qty = sold_qty
        top_sales_skus = TodayProductSales.objects.filter(country=default_country).order_by('-sales_amount')[:5]
        top_sold_qty_skus = TodayProductSales.objects.filter(country=default_country).order_by('-sold_qty')[:5]
        all_today_new_product_skus = TodayProductSales.objects.filter(product__new = True, country=default_country).order_by('-sales_amount')
        sales_today = todaySales.sales_today
        all_today_product_skus = TodayProductSales.objects.filter(country=default_country).order_by('-sales_amount')
        countries = list(set([i.country for i in TodayProductSales.objects.all()]))
        countries = [c for c in countries if c != country]
        context = {
            'date': todaySales.date.strftime('%Y/%m/%d') ,\
            'sales_today':  sales_today,\
            'sales_same_day_last_year': todaySales.sales_same_day_last_year ,\
            'monthly_increase_on_sales': todaySales.monthly_increase_on_sales  * 100,\
            'ad_cost': todaySales.ad_cost ,\
            'acos': todaySales.acos * 100 ,\
            'ad_cost_on_sales': todaySales.ad_cost_on_sales * 100 ,\
            'top_sales_skus': top_sales_skus ,\
            'top_sold_qty_skus': top_sold_qty_skus ,\
            'all_today_product_skus' : all_today_product_skus ,\
            'country': country ,\
            'countries': countries ,\
            'today': True
        }
    return HttpResponse(template.render(context, request))



def date_split(date):
    if len(date) > 4:
        year = int(date[:4])
        month = int(date[4:6])
        day = int(date[6:])
    else:
        year = datetime.date.today().year
        month = int(date[:2])
        day = int(date[2:])
    return {'year':year,'month':month,'day':day}

@login_required
def get_top_sales_vialation(request, country):
    if request.user.groups.filter(name = 'sales').exists():
        template = loader.get_template('get_top_sales_vialation.html')
        countries = list(set([i.country for i in TodayProductSales.objects.all()]))
        countries = [c for c in countries if c != country]
        Min_Sales_Qty = 3
        first_history_today_product_sales = HistoryTodayProductSales.objects.filter(country = country).order_by('-date').first()
        seven_days_ago_date = first_history_today_product_sales.date - datetime.timedelta(days = 7)

        seven_days_ago_history_today_product_sales = HistoryTodayProductSales.objects.filter(date = seven_days_ago_date,  sold_qty_average_7d__gt= Min_Sales_Qty, country=country).all()
        today_product_sales = HistoryTodayProductSales.objects.filter(date = first_history_today_product_sales.date,  sold_qty_average_7d__gt= Min_Sales_Qty, country=country).all()
        skus_list = []
        for tps in today_product_sales:
            sku = tps.product.sku
            skus_list.append(sku)
        for tps in seven_days_ago_history_today_product_sales:
            sku = tps.product.sku
            skus_list.append(sku)
        skus_list = list(set(skus_list))
        skus_seven_days_sales_rates = []
        for sku in skus_list:
            if HistoryTodayProductSales.objects.filter(date = first_history_today_product_sales.date,  product__sku=sku, country=country).count() and HistoryTodayProductSales.objects.filter(date = seven_days_ago_date,  product__sku=sku, country=country).count():
                skus_seven_days_sales_rates.append({'sku': sku, 'seven_days_sales_rate': HistoryTodayProductSales.objects.filter(date = first_history_today_product_sales.date,  product__sku=sku, country=country).first().sold_qty_average_7d / HistoryTodayProductSales.objects.filter(date = seven_days_ago_date,  product__sku=sku, country=country).first().sold_qty_average_7d})
        skus_seven_days_sales_rates = sorted(skus_seven_days_sales_rates, key=lambda i:i['seven_days_sales_rate'])
        skus_seven_days_sales_rates_top = list(reversed(skus_seven_days_sales_rates))[:5]
        skus_seven_days_sales_rates_tail = skus_seven_days_sales_rates[:5]
        seven_days_sales_rates_top = [round(i['seven_days_sales_rate'],1) for i in skus_seven_days_sales_rates_top if i['seven_days_sales_rate'] > 1]
        seven_days_sales_rates_tail = [round(i['seven_days_sales_rate'],1) for i in skus_seven_days_sales_rates_tail if i['seven_days_sales_rate'] < 1]
        history_today_product_sales_top = [HistoryTodayProductSales.objects.filter(date = first_history_today_product_sales.date, product__sku = i['sku'],country = country).first() for i in skus_seven_days_sales_rates_top if i['seven_days_sales_rate'] > 1]
        history_today_product_sales_tail = [HistoryTodayProductSales.objects.filter(date = first_history_today_product_sales.date, product__sku = i['sku'],country = country).first() for i in skus_seven_days_sales_rates_tail if i['seven_days_sales_rate'] < 1]

        context = {
            'history_today_product_sales_top': history_today_product_sales_top ,\
            'history_today_product_sales_tail': history_today_product_sales_tail ,\
            'seven_days_sales_rates_top': seven_days_sales_rates_top ,\
            'seven_days_sales_rates_tail': seven_days_sales_rates_tail ,\
            'country': country ,\
            'countries': countries ,\
        }
        return HttpResponse(template.render(context, request))

@login_required
def get_excess_inventory(request, country):
    if request.user.groups.filter(name = 'sales').exists():
        template = loader.get_template('get_excess_inventory.html')
        import operator
        class ExcessProduct:
            def __init__(self, product, sold_qty_average_7d, sold_qty_today, total_unit, days_maintain, excess_qty, excess_value):
                self.product = product
                self.sold_qty_average_7d = sold_qty_average_7d
                self.sold_qty_today = sold_qty_today
                self.total_unit = total_unit
                self.days_maintain = days_maintain
                self.excess_qty = excess_qty
                self.excess_value = excess_value
        countries = list(set([i.country for i in TodayProductSales.objects.all()]))
        countries = [c for c in countries if c != country]
        Min_Days_Maintain = 180.0
        Default_Long_Days = 365.0
        excess_product_list = []
        excess_product_total_unit = 0.0
        excess_product_total_value = 0.0
        default_purchasing_price = 30.0
        for fba_inventory in FbaInventory.objects.filter(country = country).all():
            sku = fba_inventory.sku
            days_maintain = Default_Long_Days
            if TodayProductSales.objects.filter(country=country, product__sku = sku).count():
                today_product_sales = TodayProductSales.objects.filter(country=country, product__sku = sku).first()
                product = today_product_sales.product
                sold_qty_average_7d = today_product_sales.sold_qty_average_7d
                sold_qty_today = today_product_sales.sold_qty
                days_maintain = fba_inventory.total_unit / sold_qty_average_7d
                if days_maintain > Min_Days_Maintain:
                    excess_qty = fba_inventory.total_unit - sold_qty_average_7d * Min_Days_Maintain
                    if product.sku:

                        excess_product_total_unit += excess_qty
                        if SkuPurchasingPrice.objects.filter(sku = sku).count():
                            price = SkuPurchasingPrice.objects.filter(sku = sku).order_by('-date').first().purchasing_price
                        else:
                            price = default_purchasing_price
                        excess_value = excess_qty * price
                        excess_product_total_value += excess_value
                        excess_product = ExcessProduct(product, sold_qty_average_7d, sold_qty_today,  fba_inventory.total_unit, days_maintain, excess_qty,excess_value)
                        excess_product_list.append(excess_product)
            elif fba_inventory.total_unit > 0.0:
                days_maintain = Default_Long_Days
                excess_qty = fba_inventory.total_unit
                sold_qty_average_7d = 0.0
                sold_qty_today = 0.0
                if Product.objects.filter(sku = sku).count():
                    excess_product_total_unit += excess_qty
                    if SkuPurchasingPrice.objects.filter(sku = sku).count():
                        price = SkuPurchasingPrice.objects.filter(sku = sku).order_by('-date').first().purchasing_price
                    else:
                        price = default_purchasing_price
                    excess_value = excess_qty * price
                    excess_product_total_value += excess_value
                    excess_product = ExcessProduct(Product.objects.filter(sku = sku).first(), sold_qty_average_7d, sold_qty_today,  fba_inventory.total_unit, days_maintain, excess_qty,excess_value)
                    excess_product_list.append(excess_product)
        excess_product_list.sort(key=operator.attrgetter('excess_qty'),reverse = True)
        context = {
            'excess_product_list': excess_product_list ,\
            'excess_product_total_unit': excess_product_total_unit ,\
            'excess_product_total_value': excess_product_total_value ,\
            'country': country ,\
            'countries': countries ,\
        }
        return HttpResponse(template.render(context, request))




@login_required
def history_index(request, year, month, day):
    date = "%04d/%02d/%02d" % (year, month, day,)
    today_date_obj = datetime.date(year, month, day )
    country = 'US'
    if 'country' in request.GET and request.GET['country']:
        country = request.GET['country']
    country = country
    if request.user.groups.filter(name = 'supplier').exists():
        template = loader.get_template('todaySalesForSupplier.html')
    else:
        if country == 'US':
            template = loader.get_template('todaySales.html')
        else:
            template = loader.get_template('other_country_today_sales_todaySales.html')
    todaySales = HistoryTodaySales.objects.filter(date = today_date_obj, country = country)
    if todaySales.count():
        todaySales = todaySales.all()[0]
    else:
        todaySales = HistoryTodaySales.objects.filter(country = country).order_by('-date')[0]

    if request.user.groups.filter(name = 'supplier').exists():
        supplier_id_list = [user_supplier.supplier_id for user_supplier in UserSupplier.objects.filter(user = request.user).all()]
        supplier_sku_list = [i.sku for i in SkuSupplier.objects.filter(supplier_id__in = supplier_id_list).all()]
        supplier_sku_sales_data = HistoryTodayProductSales.objects.filter(date = today_date_obj, product__sku__in=supplier_sku_list, country=country).order_by('-sold_qty')

        total_sold_unit = supplier_sku_sales_data.aggregate(Sum('sold_qty'))['sold_qty__sum']
        context = {
            'date': todaySales.date.strftime('%Y/%m/%d') ,\
            'total_sold_unit': total_sold_unit ,\
            'supplier_sku_sales_data': supplier_sku_sales_data ,\
            'today': True ,\
            'country': country
        }
    else:
        class TopSku:
            def __init__(self, sku, sold_qty):
                self.sku = sku
                self.sold_qty = sold_qty
        top_sales_skus = HistoryTodayProductSales.objects.filter(date = today_date_obj, country=country).order_by('-sales_amount')[:5]
        top_sold_qty_skus = HistoryTodayProductSales.objects.filter(date = today_date_obj, country=country).order_by('-sold_qty')[:5]

        countries = list(set([i.country for i in TodayProductSales.objects.all()]))
        countries = [c for c in countries if c != country]
        new_or_all = True
        if 'new_or_all' in request.GET and request.GET['new_or_all']:
            new_or_all = bool(int(request.GET['new_or_all']))
        if country == 'US':
            if new_or_all:
                all_today_new_product_skus = HistoryTodayProductSales.objects.filter(date = today_date_obj, product__new = True, country = country).order_by('-sales_amount')
                sales_new_product = all_today_new_product_skus.aggregate(Sum('sales_amount'))['sales_amount__sum']
                sales_new_product_percent = int(sales_new_product / todaySales.sales_today * 100)
                context = {
                    'date': todaySales.date.strftime('%Y/%m/%d') ,\
                    'sales_today': todaySales.sales_today ,\
                    'sales_same_day_last_year': todaySales.sales_same_day_last_year ,\
                    'monthly_increase_on_sales': todaySales.monthly_increase_on_sales  * 100,\
                    'ad_cost': todaySales.ad_cost ,\
                    'acos': todaySales.acos * 100 ,\
                    'ad_cost_on_sales': todaySales.ad_cost_on_sales * 100 ,\
                    'top_sales_skus': top_sales_skus ,\
                    'top_sold_qty_skus': top_sold_qty_skus ,\
                    'all_today_new_product_skus' : all_today_new_product_skus ,\
                    'sales_new_product_percent' : sales_new_product_percent ,\
                    'today': False,\
                    'country': country ,\
                    'countries': countries ,\
                    'new_or_all': new_or_all \
                }
            else:
                all_today_product_skus =  HistoryTodayProductSales.objects.filter(date = today_date_obj, country = country).order_by('-sales_amount')
                context = {
                    'date': todaySales.date.strftime('%Y/%m/%d') ,\
                    'sales_today': todaySales.sales_today ,\
                    'sales_same_day_last_year': todaySales.sales_same_day_last_year ,\
                    'monthly_increase_on_sales': todaySales.monthly_increase_on_sales  * 100,\
                    'ad_cost': todaySales.ad_cost ,\
                    'acos': todaySales.acos * 100 ,\
                    'ad_cost_on_sales': todaySales.ad_cost_on_sales * 100 ,\
                    'top_sales_skus': top_sales_skus ,\
                    'top_sold_qty_skus': top_sold_qty_skus ,\
                    'all_today_product_skus': all_today_product_skus ,\
                    # 'all_today_new_product_skus' : all_today_new_product_skus ,\
                    # 'sales_new_product_percent' : sales_new_product_percent ,\
                    'today': False,\
                    'country': country ,\
                    'countries': countries ,\
                    'new_or_all': new_or_all \
                }
        else:
            all_today_product_skus =  HistoryTodayProductSales.objects.filter(date = today_date_obj, country = country).order_by('-sales_amount')
            context = {
                'date': todaySales.date.strftime('%Y/%m/%d') ,\
                'sales_today': todaySales.sales_today ,\
                'sales_same_day_last_year': todaySales.sales_same_day_last_year ,\
                'monthly_increase_on_sales': todaySales.monthly_increase_on_sales  * 100,\
                'ad_cost': todaySales.ad_cost ,\
                'acos': todaySales.acos * 100 ,\
                'ad_cost_on_sales': todaySales.ad_cost_on_sales * 100 ,\
                'top_sales_skus': top_sales_skus ,\
                'top_sold_qty_skus': top_sold_qty_skus ,\
                'all_today_product_skus': all_today_product_skus ,\
                # 'all_today_new_product_skus' : all_today_new_product_skus ,\
                # 'sales_new_product_percent' : sales_new_product_percent ,\
                'today': False,\
                'country': country ,\
                'countries': countries ,\
                'new_or_all': new_or_all \
            }
    return HttpResponse(template.render(context, request))

@login_required
def update_last_year_sales(request):
    if request.method == 'POST':
        start_line_list = ['Time', 'Selected date range (Ordered product sales)', 'Selected date range (Units ordered)']
        end_line = ['Compare Sales - Table view']
        start_read = False
        fileInMemory = request.FILES['file'].read().decode('utf-8')
        csv_data = csv.reader(StringIO(fileInMemory), delimiter=',')
        DailySalesLastYear.objects.all().delete()
        for row in csv_data:
            if start_line_list == row:

                start_read = True
                continue
            elif end_line == row:
                break
            if start_read and row != []:
                date_str = row[0]
                sales = float(row[1].replace(',','').replace('$',''))
                date_str_split = date_str.split('T')[0].split('-')
                day = date_str_split[2]
                month = date_str_split[1]
                dailySalesLastYear = DailySalesLastYear(day = int(day), month = int(month), sales = sales)
                dailySalesLastYear.save()
            elif start_read and row == []:
                start_read = False

    template = loader.get_template('updateLastYearSales.html')
    context = {
        'form':UploadFileForm()
    }

    return HttpResponse(template.render(context, request))

def check_if_within_7day(year_month_day,time_to_check):
    timezone = pytz.timezone('America/Los_Angeles')
    date_time_obj = datetime.datetime.strptime(time_to_check[:19], '%Y-%m-%dT%H:%M:%S')
    timezone_date_time_obj = date_time_obj.astimezone(pytz.timezone('America/Los_Angeles'))
    countable_day = timezone_date_time_obj.day
    # start_datetime = datetime.datetime(timezone_date_time_obj.year, \
    start_datetime = datetime.datetime(year_month_day['year'], \
                                        year_month_day['month'], year_month_day['day'],\
                                        0,0,0,0,pytz.timezone('America/Los_Angeles') \
                                        )
    compare_datetime = start_datetime + datetime.timedelta(days=1)
    compare_earliest_datetime = start_datetime - datetime.timedelta(days=6)
    timezone_date_time_obj = datetime.datetime(timezone_date_time_obj.year, \
                                        timezone_date_time_obj.month, timezone_date_time_obj.day,\
                                        timezone_date_time_obj.hour,timezone_date_time_obj.minute, \
                                        timezone_date_time_obj.second,0,pytz.timezone('America/Los_Angeles') \
                                        )
    if timezone_date_time_obj > compare_datetime or timezone_date_time_obj <= compare_earliest_datetime:
        return False
    else:
        return True

def check_if_today(end_day,time_to_check):
    timezone = pytz.timezone('America/Los_Angeles')
    date_time_obj = datetime.datetime.strptime(time_to_check[:19], '%Y-%m-%dT%H:%M:%S')
    timezone_date_time_obj = date_time_obj.astimezone(pytz.timezone('America/Los_Angeles'))
    countable_day = timezone_date_time_obj.day
    if countable_day == end_day:
        return True
    return False

def save_history_today_product_sales(todayProductSales,date,country):
    historyTodayProductSales = HistoryTodayProductSales.objects.filter(date = date,product = todayProductSales.product,country = country)
    if historyTodayProductSales.count():
        historyTodayProductSales = historyTodayProductSales.all()[0]
        historyTodayProductSales.sold_qty = todayProductSales.sold_qty
        historyTodayProductSales.sales_amount = todayProductSales.sales_amount
        historyTodayProductSales.sold_qty_average_7d = todayProductSales.sold_qty_average_7d
        historyTodayProductSales.average_price_7d =  todayProductSales.average_price_7d
        historyTodayProductSales.fba_inventory =  todayProductSales.fba_inventory
        historyTodayProductSales.lasting_day_estimated_by_us =  todayProductSales.lasting_day_estimated_by_us
        historyTodayProductSales.lasting_day_of_available_estimated_by_us =  todayProductSales.lasting_day_of_available_estimated_by_us
        historyTodayProductSales.lasting_day_of_available_fc_estimated_by_us =  todayProductSales.lasting_day_of_available_fc_estimated_by_us
        historyTodayProductSales.lasting_day_of_total_fba_unit_estimated_by_us =  todayProductSales.lasting_day_of_total_fba_unit_estimated_by_us
    else:
        historyTodayProductSales = HistoryTodayProductSales( \
            product = todayProductSales.product ,\
            date = date ,\
            sold_qty = todayProductSales.sold_qty ,\
            sales_amount = todayProductSales.sales_amount ,\
            sold_qty_average_7d = todayProductSales.sold_qty_average_7d ,\
            average_price_7d =  todayProductSales.average_price_7d ,\
            fba_inventory =  todayProductSales.fba_inventory ,\
            lasting_day_estimated_by_us =  todayProductSales.lasting_day_estimated_by_us ,\
            lasting_day_of_available_estimated_by_us =  todayProductSales.lasting_day_of_available_estimated_by_us ,\
            lasting_day_of_available_fc_estimated_by_us =  todayProductSales.lasting_day_of_available_fc_estimated_by_us ,\
            lasting_day_of_total_fba_unit_estimated_by_us =  todayProductSales.lasting_day_of_total_fba_unit_estimated_by_us ,\
            country = country )
    historyTodayProductSales.save()

def save_today_sales(todaySales, date):
    default_country = 'US'
    historyTodaySales = HistoryTodaySales.objects.filter(date = date, country = default_country)
    if historyTodaySales.count():
        historyTodaySales = historyTodaySales.all()[0]
        historyTodaySales.sales_today = todaySales.sales_today
        historyTodaySales.sales_same_day_last_year = todaySales.sales_same_day_last_year
        historyTodaySales.sales_month_to_date = todaySales.sales_month_to_date
        historyTodaySales.monthly_increase_on_sales = todaySales.monthly_increase_on_sales
        historyTodaySales.ad_cost = todaySales.ad_cost
        historyTodaySales.acos = todaySales.acos
        historyTodaySales.ad_cost_on_sales = todaySales.ad_cost_on_sales
    else:
        historyTodaySales = HistoryTodaySales( \
                                date = date ,\
                                sales_today = todaySales.sales_today ,\
                                sales_same_day_last_year = todaySales.sales_same_day_last_year ,\
                                sales_month_to_date = todaySales.sales_month_to_date ,\
                                monthly_increase_on_sales = todaySales.monthly_increase_on_sales ,\
                                ad_cost = todaySales.ad_cost ,\
                                acos = todaySales.acos ,\
                                ad_cost_on_sales = todaySales.ad_cost_on_sales ,\
                                            )
    historyTodaySales.save()

def sales_channel_converting(x):
    if x == 'Amazon.com':
        ret = 'US'
    elif x == 'Amazon.ca':
        ret = 'CA'
    elif x == 'Amazon.com.mx':
        ret = 'MX'
    elif x == 'Amazon.co.uk':
        ret = 'GB'
    elif x in ['Amazon.de', 'Amazon.it', 'Amazon.fr', 'Amazon.es', 'Amazon.se', 'Amazon.nl', 'Amazon.pl', 'Amazon.com.tr']  :
        ret = 'EU'
    elif x == 'Amazon.ae':
        ret = 'AE'
    return ret

def harmonize_price(price, sales_channel):
    ret = price
    if sales_channel  == 'Amazon.pl':
        currency_rate = CurrencyRate.objects.get(from_country = 'PL', to_country = 'EU')
        ret = price * currency_rate.rate
    elif sales_channel  == 'Amazon.se':
        currency_rate = CurrencyRate.objects.get(from_country = 'SE', to_country = 'EU')
        ret = price * currency_rate.rate
    elif sales_channel  == 'Amazon.com.tr':
        currency_rate = CurrencyRate.objects.get(from_country = 'TR', to_country = 'EU')
        ret = price * currency_rate.rate
    return ret
@login_required
def update_7d_orders(request):
    if request.method == 'POST':
        date = request.POST['date']
        year_month_day = date_split(date)
        year = year_month_day['year']
        month = year_month_day['month']
        day = year_month_day['day']
        #start_line_list = ['Time', 'Selected date range (OPS)', 'Selected date range (Units)']
        #start_read = False
        fileInMemory = request.FILES['file'].read().decode('utf-8')
        csv_data = csv.reader(StringIO(fileInMemory), delimiter='\t')
        all_country_sku_qty_price = {}
        order_list = []
        countries_set = set()
        today_date_obj = datetime.date(year, month, day )
        for row in csv_data:
            # 判断是哪个渠道的订单
            order_country_cell = row[6]
            if order_country_cell not in ['Non-Amazon', 'SI UK Prod Marketplace', 'sales-channel']:
                country = sales_channel_converting(order_country_cell)
                sku = row[11]
                qty = int(row[14])
                order_status = row[4]
                if qty and order_status != 'Cancelled':
                    if_today = check_if_today(day,row[2])
                    price = float(row[16])
                    price = harmonize_price(price, order_country_cell)
                    countries_set.add(country)
                    order_list.append({'country': country, 'sku': sku, 'qty': qty, 'price': price, 'if_today': if_today})
        for country in list(countries_set):
            all_country_sku_qty_price[country]= {'sku_qty':{},'sku_price':{},'sku_today_qty':{},'sku_today_price':{},'total_today_sales':0.0}
            Last7dayProductSales.objects.filter(country = country).delete()
            TodayProductSales.objects.filter(country = country).delete()
        for order in order_list:
            price = order['price']
            sku = order['sku']
            qty = order['qty']
            if_today = order['if_today']
            country = order['country']
            if sku in all_country_sku_qty_price[country]['sku_qty']:
                all_country_sku_qty_price[country]['sku_qty'][sku] += qty
                all_country_sku_qty_price[country]['sku_price'][sku] += price
                if sku in all_country_sku_qty_price[country]['sku_today_qty'] and if_today:
                    all_country_sku_qty_price[country]['sku_today_qty'][sku] += qty
                    all_country_sku_qty_price[country]['sku_today_price'][sku] += price
            else:
                all_country_sku_qty_price[country]['sku_qty'][sku] = qty
                all_country_sku_qty_price[country]['sku_price'][sku] = price
                if if_today:
                    all_country_sku_qty_price[country]['sku_today_qty'][sku] = qty
                    all_country_sku_qty_price[country]['sku_today_price'][sku] = price
                else:
                    if sku not in all_country_sku_qty_price[country]['sku_today_qty']:
                        all_country_sku_qty_price[country]['sku_today_qty'][sku] = 0
                        all_country_sku_qty_price[country]['sku_today_price'][sku] = 0.0
        for k, v in all_country_sku_qty_price.items():
            country = k
            sku_qty = v['sku_qty']
            sku_price = v['sku_price']
            sku_today_qty = v['sku_today_qty']
            sku_today_price = v['sku_today_price']
            total_today_sales = v['total_today_sales']
            for sku, qty in sku_qty.items():
                product, created = Product.objects.get_or_create(sku = sku)
                if created:
                    product.save()
                sales_amount = sku_price[sku]
                sold_qty_average_7d = qty / 7
                average_price_7d = sales_amount / qty
                last7dayProductSales = Last7dayProductSales(product = product ,\
                                                            sold_qty = qty ,\
                                                            sales_amount = sales_amount ,\
                                                            sold_qty_average_7d = sold_qty_average_7d ,\
                                                            average_price_7d = average_price_7d ,\
                                                            country = country)
                last7dayProductSales.save()
                if sku in sku_today_qty:
                    todayProductSales = TodayProductSales(product = product ,\
                                                          sold_qty = sku_today_qty[sku] ,\
                                                          sales_amount = sku_today_price[sku] ,\
                                                          sold_qty_average_7d = sold_qty_average_7d ,\
                                                          average_price_7d = average_price_7d ,\
                                                          country = country)
                    fbaInventory = FbaInventory.objects.filter(sku = sku,country =country)
                    if fbaInventory.count():
                        todayProductSales.fba_inventory = fbaInventory.all()[0]
                        todayProductSales.lasting_day_estimated_by_us = int(float(todayProductSales.fba_inventory.total_unit) / float(sold_qty_average_7d))
                    todayProductSales.save()

                    save_history_today_product_sales(todayProductSales,today_date_obj, country)
                    total_today_sales += sku_today_price[sku]
            save_today_sales(year_month_day = year_month_day, sales_today = total_today_sales, country=country)
        return redirect('today_sales')

    template = loader.get_template('update_7d_orders.html')
    today_date_str = ""
    if HistoryTodayProductSales.objects.first() != None:
        today_date_str = HistoryTodayProductSales.objects.order_by('-date').first().date.strftime('%m%d')
    class UploadFileAndDateForm(forms.Form):
        date = forms.CharField(max_length=8)
        file = forms.FileField()
    context = {
        'form':UploadFileAndDateForm(initial={'date': today_date_str})
    }

    return HttpResponse(template.render(context, request))


def country_harmonizing(x):
    if x in ['DE','FR','SE','ES','NL','PL','TR','IT','TR']:
        return 'EU'
    return x
@login_required
def update_restock_report(request):
    if request.method == 'POST':
        encodings = ('cp1252', 'utf-8')
        for e in encodings:
            try:
                request.FILES['file'].seek(0,0)
                fileInMemory = request.FILES['file'].read().decode(e)
                break
            except UnicodeDecodeError:
                pass
        csv_data = csv.reader(StringIO(fileInMemory), delimiter='\t')
        for row in csv_data:
            sku = row[3]
            if sku != 'Merchant SKU':
                country = country_harmonizing(row[0])
                fnsku = row[2]
                asin = row[4]
                working_unit = int(row[19])
                total_unit = int(row[12]) + working_unit
                available = int(row[14])
                inbound_fc_unit = int(row[13]) + int(row[15]) + int(row[16])  + working_unit
                inbound_unit = int(row[13]) + working_unit
                fc_unit = int(row[15]) + int(row[16])
                days_of_supply = row[20+3]
                try:
                    days_of_supply = int(days_of_supply)
                except:
                    days_of_supply = 365
                recommended_replenishment_qty = row[22+4]
                if recommended_replenishment_qty:
                    recommended_replenishment_qty = int(recommended_replenishment_qty)
                else:
                    recommended_replenishment_qty = 0
                recommended_ship_date = row[23+4]
                if not(recommended_ship_date):
                    recommended_ship_date = ''
                fbaInventory = FbaInventory.objects.filter(sku = sku,country = country)
                if fbaInventory.count():
                    fbaInventory = fbaInventory.all()[0]
                    fbaInventory.fnsku = fnsku
                    fbaInventory.asin = asin
                    fbaInventory.total_unit = total_unit
                    fbaInventory.available = available
                    fbaInventory.inbound_fc_unit = inbound_fc_unit
                    fbaInventory.inbound_unit = inbound_unit
                    fbaInventory.fc_unit = fc_unit
                    fbaInventory.days_of_supply = days_of_supply
                    fbaInventory.recommended_replenishment_qty = recommended_replenishment_qty
                    fbaInventory.recommended_ship_date = recommended_ship_date
                    fbaInventory.country = country
                else:
                    fbaInventory = FbaInventory( \
                                               sku = sku ,\
                                               fnsku = fnsku ,\
                                               asin = asin ,\
                                               total_unit = total_unit ,\
                                               available = available ,\
                                               inbound_unit = inbound_unit ,\
                                               fc_unit = fc_unit ,\
                                               inbound_fc_unit = inbound_fc_unit ,\
                                               days_of_supply = days_of_supply ,\
                                               recommended_replenishment_qty = recommended_replenishment_qty ,\
                                               recommended_ship_date = recommended_ship_date ,\
                                               country = country

                    )
                fbaInventory.save()
                todayProductSales = TodayProductSales.objects.filter(product__sku = sku,country = country)
                if todayProductSales.count():
                    todayProductSales = todayProductSales.all()[0]
                    todayProductSales.fba_inventory = fbaInventory
                    todayProductSales.save()
        return redirect('today_sales')
    template = loader.get_template('update_restock_report.html')

    context = {
        'form':UploadFileForm()
    }

    return HttpResponse(template.render(context, request))

@login_required
def update_currency_rate(request):
    if request.method == 'POST':
        fileInMemory = request.FILES['file'].read()
        filePath = BytesIO(fileInMemory)
        wb2 = load_workbook(filePath)
        sheet = wb2.worksheets[0]
        rate_dict = {}
        for row in range(1, sheet.max_row + 1):
            row_name = sheet.cell(row,1).value
            row_value = sheet.cell(row,2).value
            rate_dict[row_name] = row_value
        for k, v in rate_dict.items():
            if k == '1PLN=?RMB':
                rate = v / rate_dict['1EUR=?RMB']
                (currency_rate, found_or_not) = CurrencyRate.objects.get_or_create(from_country = 'PL', to_country = 'EU')
                currency_rate.rate = rate
                currency_rate.save()
            elif k == '1SEK=?RMB':
                rate = v / rate_dict['1EUR=?RMB']
                (currency_rate, found_or_not) = CurrencyRate.objects.get_or_create(from_country = 'SE', to_country = 'EU')
                currency_rate.rate = rate
                currency_rate.save()
            elif k == '1TRY=?RMB':
                rate = v / rate_dict['1EUR=?RMB']
                (currency_rate, found_or_not) = CurrencyRate.objects.get_or_create(from_country = 'TR', to_country = 'EU')
                currency_rate.rate = rate
                currency_rate.save()
    template = loader.get_template('update_currency_rate.html')
    context = {
        'form':UploadFileForm()
    }

    return HttpResponse(template.render(context, request))


@login_required
def update_shenzhen_inventory(request):
    if request.method == 'POST':
        fileInMemory = request.FILES['file'].read()
        filePath = BytesIO(fileInMemory)
        wb2 = load_workbook(filePath)
        sheet = wb2.worksheets[0]
        max_column = sheet.max_column
        if sheet.cell(4,max_column - 1).value == shenzhen_warehouse_name:
            Inventory.objects.filter(warehouse_name = shenzhen_warehouse_name).delete()
            for row in range(6, sheet.max_row):
                sku = sheet.cell(row,1).value
                qty = int(sheet.cell(row,max_column - 1).value)
                shenzhenInventory =Inventory(warehouse_name = shenzhen_warehouse_name,sku = sku, qty = qty)
                shenzhenInventory.save()
    template = loader.get_template('update_shenzhen_inventory.html')
    context = {
        'form':UploadFileForm()
    }

    return HttpResponse(template.render(context, request))

@login_required
def update_purchasing_orders(request):
    if request.method == 'POST':
        wb2 = xlrd.open_workbook(file_contents=request.FILES['file'].read())
        sheet = wb2.sheet_by_index(0)
        max_column = sheet.ncols
        for col in range(0,max_column):
            title = sheet.cell(4,col).value
            if title == '未入库数量':
                qty_col = col
            if title == '状态':
                status_col = col
        ReceivablePurchasedQty.objects.all().delete()
        for row in range(5,sheet.nrows):
            sku_cell_value = sheet.cell(row,0).value
            if sku_cell_value:
                sku = sku_cell_value
            elif sheet.cell(row,status_col).value == '小计':
                receivable_qty = int(sheet.cell(row,qty_col).value)
                receivablePurchasedQty = ReceivablePurchasedQty(sku = sku, qty = receivable_qty)
                receivablePurchasedQty.save()
    template = loader.get_template('update_purchasing_orders.html')
    context = {
        'form':UploadFileForm()
    }

    return HttpResponse(template.render(context, request))

def month_to_date_last_year_month_to_date(year,month,day):
    # dailySalesLastYear = DailySalesLastYear.objects.filter(month = month, day = day).all()[0]
    # dailySalesLastYear = DailySalesLastYear.objects.get(pk = dailySalesLastYear.pk + 2)
    default_country = 'US'
    sales_month_to_date_last_year = DailySalesLastYear.objects.filter(month = month, day__lte = day).aggregate(Sum('sales'))['sales__sum']
    sales_month_to_date = HistoryTodaySales.objects.filter(country = default_country, date__range=(datetime.date(year, month, 1), datetime.date(year, month, day))).aggregate(Sum('sales_today'))['sales_today__sum']
    if sales_month_to_date == None:
        return {'sales_month_to_date': 0.0 ,\
                'monthly_increase_on_sales': 0.0}
    else:
        monthly_increase_on_sales = (sales_month_to_date - sales_month_to_date_last_year)/sales_month_to_date_last_year
        return {'sales_month_to_date': sales_month_to_date ,\
                'monthly_increase_on_sales': monthly_increase_on_sales}

def save_today_sales(**kwargs):
    year_month_day = kwargs['year_month_day']
    country = kwargs['country']
    year = year_month_day['year']
    month = year_month_day['month']
    day = year_month_day['day']
    dailySalesLastYear = DailySalesLastYear.objects.filter(month = month, day = day).all()[0]
    from django.core.exceptions import ObjectDoesNotExist
    try:
        dailySalesLastYear = DailySalesLastYear.objects.get(pk = dailySalesLastYear.pk + 1)
    except ObjectDoesNotExist:
        pass
    today_date_obj = datetime.date(year = year, month = month, day = day)
    #检查当天HistoryTodaySales 是否已经建立
    if not HistoryTodaySales.objects.filter(date=today_date_obj, country=country).count() and 'sales_today' in kwargs.keys():
        historyTodaySales = HistoryTodaySales(date = today_date_obj ,\
                                            sales_today = kwargs['sales_today'] ,\
                                            sales_same_day_last_year = 999.99 ,\
                                            sales_month_to_date = 999.99 ,\
                                            monthly_increase_on_sales = 999.99 ,\
                                            ad_cost = 999.99 ,\
                                            acos = 999.99 ,\
                                            ad_cost_on_sales =999.99 ,\
                                            country= country)
        historyTodaySales.save()
    else:
        historyTodaySales = HistoryTodaySales.objects.filter(date = today_date_obj,country= country).first()
        if 'sales_today' in kwargs.keys():
            historyTodaySales.sales_today = kwargs['sales_today']
    sales_month_to_date_and_monthly_increase_on_sales = month_to_date_last_year_month_to_date(year, month, day)

    historyTodaySales.sales_same_day_last_year = dailySalesLastYear.sales
    historyTodaySales.sales_month_to_date = sales_month_to_date_and_monthly_increase_on_sales['sales_month_to_date']
    historyTodaySales.monthly_increase_on_sales = sales_month_to_date_and_monthly_increase_on_sales['monthly_increase_on_sales']

    if 'ad_cost' in kwargs.keys():
        ad_cost = kwargs['ad_cost']
        acos = kwargs['acos']
        historyTodaySales.ad_cost = ad_cost
        historyTodaySales.acos = acos
        historyTodaySales.ad_cost_on_sales = ad_cost / historyTodaySales.sales_today
    historyTodaySales.save()
@login_required
def update_today_sales(request):
    if request.method == 'POST':
        ad_cost = float(request.POST['ad_cost'])
        acos = float(request.POST['acos'])
        date = request.POST['date']
        year_month_day = date_split(date)
        country = 'US'
        save_today_sales(year_month_day=year_month_day, ad_cost= ad_cost,acos= acos,country=country)
    template = loader.get_template('updateTodaySales.html')
    class TodaySalesForm(forms.Form):
        date = forms.CharField(max_length=8)
        ad_cost = forms.FloatField()
        acos = forms.FloatField()
    context = {
        'form':TodaySalesForm()
    }

    return HttpResponse(template.render(context, request))

def handle_uploaded_fba_shipment_file(file):
    fileInMemory = file.read().decode('utf-8')
    csv_data = csv.reader(StringIO(fileInMemory), delimiter='\t')
    start_read = False
    shipment_id = False
    shipment_name = False
    shipped_sku_qties_list = []
    for i, row in enumerate(csv_data):
        if i == 0:
            if row[0] in ['Shipment ID', '货件编号']:
                shipment_id = row[1]
        if i == 1:
            if row[0] in ['Name', '名称']:
                shipment_name = row[1]
        if row:
            if row[0] in ['Merchant SKU','卖家 SKU'] :
                for j,col in enumerate(row):
                    if col in ['Shipped', '已发货']:
                        shipped_sku_qty_col = j
                start_read = True
            elif start_read:
                shipped_sku_qties_list.append({'sku': row[0], 'qty': int(row[shipped_sku_qty_col])})
    if len(shipped_sku_qties_list) and shipment_id and shipment_name:
        fba_shipment = FbaShipment.objects.filter(shipment_id = shipment_id)
        if fba_shipment.count():
            fba_shipment = fba_shipment.first()
            fba_shipment.shipment_name = shipment_name
            fba_shipment.save()
        else:
            fba_shipment = FbaShipment(shipment_id = shipment_id \
                                    ,shipment_name = shipment_name)
            fba_shipment.save()
            for shipped_sku_qty_i in shipped_sku_qties_list:
                sku = shipped_sku_qty_i['sku']
                (product, found_or_not) = Product.objects.get_or_create(sku = sku)
                shipped_sku_qty = ShippedSkuQty(product = product, sku = sku, qty = shipped_sku_qty_i['qty'])
                shipped_sku_qty.save()
                fba_shipment.shipped_sku_qties.add(shipped_sku_qty)
        return shipment_id
# 还不需要输入运费等信息的版本
# @login_required
# def update_fba_shipment(request):
#     if request.method == 'POST':
#         form = UploadFilesForm(request.POST, request.FILES)
#         if form.is_valid():
#             files = request.FILES.getlist('file_field')
#             for f in files:
#                 handle_uploaded_fba_shipment_file(f)
#     template = loader.get_template('update_fba_shipment.html')
#     context = {
#         'form':UploadFilesForm()
#     }
#     return HttpResponse(template.render(context, request))

@login_required
def update_fba_shipment(request):
    template = loader.get_template('update_fba_shipment.html')
    if request.method == 'POST':
        form = UploadShipmentFileForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            shipment_id_in_file = handle_uploaded_fba_shipment_file(file)
            shipment_id = request.POST['shipment_id']
            shipway = request.POST['shipway']
            cost_per_kg = request.POST['cost_per_kg']
            if shipment_id_in_file == shipment_id:
                fbaShipmentCost = FbaShipmentCost.objects.filter(shipment_id = shipment_id).first()
                if fbaShipmentCost:
                    fbaShipmentCost.shipway = shipway
                    fbaShipmentCost.cost_per_kg = cost_per_kg
                else:
                    fbaShipmentCost = FbaShipmentCost(shipment_id = shipment_id, shipway = shipway ,\
                                                        cost_per_kg = cost_per_kg, date = datetime.date.today())
                fbaShipmentCost.save()
            else:
                return HttpResponse('上传的货件文件的Shipment ID和填写的shipment ID不符')
        else:
            context = {
                'form': form
            }
            return HttpResponse(template.render(context, request))
    context = {
        'form':UploadShipmentFileForm()
    }
    return HttpResponse(template.render(context, request))

@login_required
def update_sku_weight(request):
    if request.method == 'POST':
        fileInMemory = request.FILES['file'].read()
        filePath = BytesIO(fileInMemory)
        wb2 = load_workbook(filePath)
        sheet = wb2.worksheets[0]
        max_column = sheet.max_column

        # 检查title是否正确
        titles = sheet.cell(1, 1).value + sheet.cell(1, 2).value + sheet.cell(1, 3).value + sheet.cell(1, 4).value + sheet.cell(1, 5).value
        if titles != 'SKU产品中文实重否？实重22.5kg一箱的大小，装满该产品能装多少个？':
            return HttpResponse('请检查上传的模版')
        for row in range(2, sheet.max_row + 1):
            sku = sheet.cell(row, 1).value
            heavy_or_light = sheet.cell(row, 3).value
            real_weight = 0.0
            converted_weight = 0.0
            if heavy_or_light:
                heavy_or_light = True
                real_weight = sheet.cell(row, 4).value
            else:
                heavy_or_light = False
                converted_weight = sheet.cell(row, 5).value

            skuWeight = SkuWeight.objects.filter(sku = sku).first()
            if not skuWeight:
                skuWeight = SkuWeight(sku = sku, heavy_or_light = heavy_or_light, real_weight = real_weight, converted_weight = converted_weight)
            else:
                skuWeight.heavy_or_light = heavy_or_light
                skuWeight.real_weight = real_weight
                skuWeight.converted_weight = converted_weight
            skuWeight.save()
    template = loader.get_template('update_sku_weight.html')
    context = {
        'form':UploadFileForm()
    }
    return HttpResponse(template.render(context, request))

@login_required
def update_fba_shipment_received_sku_qty(request):
    if request.method == 'POST':
        ReceivedSkuQty.objects.all().delete()
        col_nums = {'sku':0, 'quantity':0, 'fba-shipment-id': 0}
        fileInMemory = request.FILES['file'].read().decode('windows-1252')
        csv_data = csv.reader(StringIO(fileInMemory), delimiter=',')
        for i, row in enumerate(csv_data):
            if i == 0:
                for col_num, title in enumerate(row):
                    if title in col_nums.keys():
                        col_nums[title] = col_num
            else:
                received_sku_qty, created = ReceivedSkuQty.objects.get_or_create(shipment_id = row[col_nums['fba-shipment-id']] \
                                                                                        ,sku = row[col_nums['sku']])
                received_sku_qty.qty += int(row[col_nums['quantity']])
                received_sku_qty.save()
    template = loader.get_template('update_fba_shipment_received_sku_qty.html')
    context = {
        'form':UploadFileForm()
    }
    return HttpResponse(template.render(context, request))

@login_required
def update_fba_shipment_estimated_receiving_date(request):
    if request.method == 'POST':
        import json
        post_json = json.loads(request.body)
        date_type = post_json['date_type']
        date_str = post_json['date']
        date_obj = datetime.datetime.strptime(date_str, '%Y/%m/%d').date()
        fba_shipment_id = post_json['fba_shipment_id']
        fba_shipment = FbaShipment.objects.filter(shipment_id = fba_shipment_id).first()
        if fba_shipment:
            if date_type == 'ship_date':
                fba_shipment.shipped_date = date_obj
            elif date_type == 'estimated_receiving_date':
                fba_shipment.estimated_receiving_date = date_obj
            fba_shipment.save()
        return HttpResponse(request.POST.get('a'))

    template = loader.get_template('update_fba_shipment_estimated_receiving_date.html')
    fba_shipments = FbaShipment.objects.filter(closed = False).all()

    context = {
        'fba_shipments': fba_shipments
    }
    return HttpResponse(template.render(context, request))

@login_required
def estimated_sku_qty_receiving_date(request):
    template = loader.get_template('estimated_sku_qty_receiving_date.html')

    fba_shipments = FbaShipment.objects.filter(closed = False).order_by('estimated_receiving_date')
    shipped_sku_qty_by_shipment = {}
    class ShippedSkuQtyByShipment():
        def __init__(self, qty, date, shipment_name, shipment_id):
            self.qty = qty
            self.date = date
            self.shipment_name = shipment_name
            self.shipment_id = shipment_id
    shipment_closed_dict = {}
    for fba_shipment in fba_shipments:
        shipment_id = fba_shipment.shipment_id
        shipment_name = fba_shipment.shipment_name
        shipment_closed_dict[shipment_id] = True
        if fba_shipment.estimated_receiving_date:
            estimated_receiving_date = fba_shipment.estimated_receiving_date.strftime('%m/%d')
        else:
            estimated_receiving_date = "无"
        shipped_sku_qties = fba_shipment.shipped_sku_qties.all()
        for shipped_sku_qty in shipped_sku_qties:
            sku = shipped_sku_qty.sku
            shipped_qty = shipped_sku_qty.qty
            received_sku_qty = ReceivedSkuQty.objects.filter(shipment_id = shipment_id \
                                                        ,sku = sku).aggregate(Sum('qty'))['qty__sum']

            if received_sku_qty == None:
                received_sku_qty = 0

            unreceived_qty = shipped_qty - received_sku_qty
            if unreceived_qty > 15:
                if sku in shipped_sku_qty_by_shipment.keys():
                    shipped_sku_qty_by_shipment[sku].append(ShippedSkuQtyByShipment(shipment_name = shipment_name \
                                                                                    , shipment_id = shipment_id \
                                                                                    ,qty =  unreceived_qty \
                                                                                    , date = estimated_receiving_date))
                else:
                    shipped_sku_qty_by_shipment[sku] = [ShippedSkuQtyByShipment(shipment_name = shipment_name \
                                                                                , shipment_id = shipment_id \
                                                                                , qty = unreceived_qty \
                                                                                , date = estimated_receiving_date)]
                shipment_closed_dict[shipment_id] = False

        for k,v in shipment_closed_dict.items():
            if v:
                fba_shipment = FbaShipment.objects.get(shipment_id = k)
                fba_shipment.closed = True
                fba_shipment.save()
    context = {
        'to_be_received_skus':shipped_sku_qty_by_shipment
    }
    return HttpResponse(template.render(context, request))

@login_required
def get_estimated_sku_qty_receiving_date_of_a_sku(request):
    sku = request.GET['sku']
    fba_shipments = FbaShipment.objects.filter(closed = False, shipped_sku_qties__sku=sku).order_by('estimated_receiving_date')
    receivable_qty_list = []
    for fba_shipment in fba_shipments:
        shipment_id = fba_shipment.shipment_id
        shipment_name = fba_shipment.shipment_name
        if fba_shipment.estimated_receiving_date:
            estimated_receiving_date = fba_shipment.estimated_receiving_date.strftime('%m/%d')
        else:
            estimated_receiving_date = "无"
        shipped_sku_qties = fba_shipment.shipped_sku_qties.all()
        for shipped_sku_qty in shipped_sku_qties:
            if shipped_sku_qty.sku == sku:
                shipped_qty = shipped_sku_qty.qty
                received_sku_qty = ReceivedSkuQty.objects.filter(shipment_id = shipment_id \
                                                            ,sku = sku).aggregate(Sum('qty'))['qty__sum']

                if received_sku_qty == None:
                    received_sku_qty = 0

                unreceived_qty = shipped_qty - received_sku_qty
                if unreceived_qty > 15:
                    receivable_qty_list.append({'shipment_id': shipment_id \
                                                                ,'date': estimated_receiving_date \
                                                                , 'qty': unreceived_qty \
                                                                , 'shipment_name':shipment_name})
    return JsonResponse(receivable_qty_list, safe=False)
@login_required
def get_history_sales_of_a_sku(request):
    sku = request.GET['sku']
    country = request.GET['country']
    history_today_product_sales = HistoryTodayProductSales.objects.filter(product__sku = sku, country = country).order_by('-date')[:7]
    history_today_product_sales_list = [{'date':i.date.strftime('%m/%d') \
                                        ,'qty':round(i.sold_qty_average_7d,1)} for i in history_today_product_sales]
    history_today_product_sales_list.reverse()
    return JsonResponse(history_today_product_sales_list, safe=False)


def check_receivable_purchased_qty(sku):
    receivablePurchasedQty = ReceivablePurchasedQty.objects.filter(sku = sku)
    if receivablePurchasedQty.count():
        receivablePurchasedQty = receivablePurchasedQty.all()[0].qty
    else:
        receivablePurchasedQty = 0
    return receivablePurchasedQty

def check_shenzhen_inventory(sku):
    shenzhen_inventory = Inventory.objects.filter(sku = sku, warehouse_name = shenzhen_warehouse_name)
    if shenzhen_inventory.count():
        shenzhen_inventory = shenzhen_inventory.first().qty
    else:
        shenzhen_inventory = 0
    return shenzhen_inventory

def if_count_for_po(todayProductSale):
    maximum_days_lasting = 200 # 100才是正确的，200是为了让LS-VIE-33-13-A-1出现
    sku = todayProductSale.product.sku
    amazon_inventory = todayProductSale.fba_inventory.total_unit
    shenzhen_inventory = check_shenzhen_inventory(sku)
    lasting_day_of_total_fba_and_shenzhen_inventory_estimated_by_us = (todayProductSale.fba_inventory.total_unit + shenzhen_inventory)/ todayProductSale.sold_qty_average_7d
    if todayProductSale.product.discontinued:
        return False
    if lasting_day_of_total_fba_and_shenzhen_inventory_estimated_by_us < maximum_days_lasting:
        return True
    if todayProductSale.fba_inventory.days_of_supply < maximum_days_lasting:
        return True
    return False

def get_sku_purchase_price(sku):
    price = 0.0
    skuPurchaseOrders = SkuPurchaseOrder.objects.filter(sku = sku).all()
    if skuPurchaseOrders.count():
        sum_of_transaction_amount = skuPurchaseOrders.aggregate(Sum('transaction_amount'))['transaction_amount__sum']
        sum_of_qty = float(skuPurchaseOrders.aggregate(Sum('qty'))['qty__sum'])
        price = sum_of_transaction_amount / sum_of_qty
        return price
    else:
        skuPurchasingPrice = SkuPurchasingPrice.objects.filter(sku = sku).order_by('-date').first()
        if skuPurchasingPrice:
            return skuPurchasingPrice.purchasing_price
@login_required
def restock_today(request):
    if request.method == 'POST':
        abroad_warehouse_sku_qty = {}
        abroad_warehouse_uploaded = False
        form = UploadFileForm(request.POST,request.FILES)
        if form.is_valid():
            fileInMemory = request.FILES['file'].read()
            filePath = BytesIO(fileInMemory)
            wb2 = load_workbook(filePath)
            sheet = wb2.worksheets[0]
            for row in range(1, sheet.max_row + 1):
                sku = sheet.cell(row,1).value
                qty = int(sheet.cell(row,2).value)
                abroad_warehouse_sku_qty[sku]= qty
            abroad_warehouse_uploaded = True

        wb = Workbook()
        sheet = wb.active
        sheet.title = '今日生产单'
        country = 'US'
        row = 1
        max_months_to_last_for_pos = [4, 5, 6]
        moq = 200
        min_days_lasting_for_notice = 60
        max_days_to_last_for_pos = [max_months_to_last_for_po * 30 for max_months_to_last_for_po in max_months_to_last_for_pos]
        if abroad_warehouse_uploaded:
            titles = ['SKU', 'FBA 可售库存', '中转库存', '在途库存', '亚马逊总库存', 'F199库存', '海外仓', '今日销售', '过去7天平均日销售' \
                    , '7日暴增比' \
                    , '亚马逊预测能撑多少天', '我们预测能撑多少天', '单价', '亚马逊建议补货量', '已经下过购货订单' \
                    ]
        else:
            titles = ['SKU', 'FBA 可售库存', '中转库存', '在途库存', '亚马逊总库存', 'F199库存', '今日销售', '过去7天平均日销售' \
                    , '7日暴增比' \
                    , '亚马逊预测能撑多少天', '我们预测能撑多少天', '单价', '亚马逊建议补货量', '已经下过购货订单' \
                    ]
        start_col_restock = len(titles) + 1
        for max_months_to_last_for_po in max_months_to_last_for_pos:
            titles.append('撑到%i个月生产单建议量' %max_months_to_last_for_po)
        if request.user.is_superuser:
            is_head_of_sales = True
            #先不加复杂的财务部分数据
            #这个是复杂版本titles.extend(['运营', '运营助理','净资产','单个采购成本','单个的海派头程运费','单个的空派头程运费','订购量', '货款+头程运费'])
        else:
            is_head_of_sales = False
            #先不加复杂的财务部分数据
            #这个是复杂版本titles.extend(['净资产','单个采购成本','单个的海派头程运费','单个的空派头程运费','订购量', '货款+头程运费'])
        for i, title in enumerate(titles):
            sheet.cell(row,i + 1).value = title

        yellow_fill = PatternFill("solid", fgColor="FFFF00")
        todayProductSales = TodayProductSales.objects.filter(country = country).all()

        first_history_today_product_sales = HistoryTodayProductSales.objects.filter(country = country).order_by('-date').first()
        seven_days_ago_date = first_history_today_product_sales.date - datetime.timedelta(days = 7)
        processed_sku_list = []
        row +=1

        history_today_product_sales_filtered = HistoryTodayProductSales.objects.filter(date__gt=seven_days_ago_date,country = country).order_by('-date')
        for todayProductSale in history_today_product_sales_filtered:
            sku = todayProductSale.product.sku
            # 让只显示运营负责的sku的，但是目前还不弄这种运营分sku的模式，所以先全部显示给运营
            # if is_head_of_sales == False:
            #     skuManagedBySalesPerson = SkuManagedBySalesPerson.objects.filter(sku = sku, sales_person_name = request.user.username).first()
            #     if not skuManagedBySalesPerson:
            #         skuManagedBySalesPerson = SkuManagedBySalesPerson.objects.filter(sku = sku, sales_assistant_name = request.user.username).first()
            #     if not skuManagedBySalesPerson:
            #         continue
            # else:
            #     skuManagedBySalesPerson = SkuManagedBySalesPerson.objects.filter(sku = sku).first()
            if sku not in processed_sku_list and  if_count_for_po(todayProductSale):
                processed_sku_list.append(sku)
                sheet.cell(row,1).value = sku
                sheet.cell(row,2).value = todayProductSale.fba_inventory.available
                sheet.cell(row,3).value = todayProductSale.fba_inventory.fc_unit
                sheet.cell(row,4).value = todayProductSale.fba_inventory.inbound_unit
                fba_total_unit = todayProductSale.fba_inventory.total_unit
                sheet.cell(row,5).value = fba_total_unit
                shenzhen_inventory = check_shenzhen_inventory(sku)
                sheet.cell(row,6).value = shenzhen_inventory
                abroad_qty = 0
                if abroad_warehouse_uploaded :
                    sheet.cell(row,7).value = 0
                    if sku in abroad_warehouse_sku_qty:
                        abroad_qty = abroad_warehouse_sku_qty[sku]
                        sheet.cell(row,7).value = abroad_warehouse_sku_qty[sku]
                sheet.cell(row,7 + abroad_warehouse_uploaded).value = int(todayProductSale.sold_qty)
                sheet.cell(row,8 + abroad_warehouse_uploaded).value = round(todayProductSale.sold_qty_average_7d,1)
                increase_rate = 1.0
                seven_days_ago_history_today_product_sales = HistoryTodayProductSales.objects.filter(date = seven_days_ago_date, product__sku=sku, country=country).first()
                if seven_days_ago_history_today_product_sales != None:
                    seven_days_ago_sold_qty_average_7d = seven_days_ago_history_today_product_sales.sold_qty_average_7d
                    if seven_days_ago_sold_qty_average_7d != 0:
                        increase_rate = todayProductSale.sold_qty_average_7d / seven_days_ago_sold_qty_average_7d
                sheet.cell(row,9 + abroad_warehouse_uploaded).value = round(increase_rate,1)# 暴增
                sheet.cell(row,10 + abroad_warehouse_uploaded).value = todayProductSale.fba_inventory.days_of_supply
                lasting_day_of_total_fba_and_shenzhen_inventory_estimated_by_us = (todayProductSale.fba_inventory.total_unit + shenzhen_inventory + abroad_qty)/ todayProductSale.sold_qty_average_7d
                sheet.cell(row,11 + abroad_warehouse_uploaded).value = int(lasting_day_of_total_fba_and_shenzhen_inventory_estimated_by_us)
                sheet.cell(row,12 + abroad_warehouse_uploaded).value = round(todayProductSale.average_price_7d,2)
                sheet.cell(row,12 + abroad_warehouse_uploaded).number_format = u'"$ "#,##0.00'
                amazon_recommended_replenishment_qty = todayProductSale.fba_inventory.recommended_replenishment_qty
                sheet.cell(row,13 + abroad_warehouse_uploaded).value = amazon_recommended_replenishment_qty
                receivable_purchased_qty = check_receivable_purchased_qty(sku)
                sheet.cell(row,14 + abroad_warehouse_uploaded).value = receivable_purchased_qty
                if todayProductSale.product.discontinued:
                    continue #sheet.cell(row,13).value = "不再订货"
                else:
                    for i, max_days_to_last_for_po in enumerate(max_days_to_last_for_pos):
                        po_qty = 0
                        if todayProductSale.fba_inventory.days_of_supply >= lasting_day_of_total_fba_and_shenzhen_inventory_estimated_by_us:

                            days_vacancy = max_days_to_last_for_po - lasting_day_of_total_fba_and_shenzhen_inventory_estimated_by_us
                            qty_needed = int(days_vacancy * todayProductSale.sold_qty_average_7d)
                            if qty_needed > receivable_purchased_qty:
                                qty_needed = qty_needed - receivable_purchased_qty
                                if qty_needed > float(moq) / 2:
                                    po_qty = max([qty_needed, moq])
                        else:
                            shenzhen_inventory_and_receivable_purchased_qty = shenzhen_inventory + receivable_purchased_qty
                            if abroad_warehouse_uploaded and sku in abroad_warehouse_sku_qty:
                                shenzhen_inventory_and_receivable_purchased_qty += abroad_warehouse_sku_qty[sku]
                            if i > 0:
                                if todayProductSale.fba_inventory.days_of_supply:
                                    amazon_recommended_replenishment_qty_i = amazon_recommended_replenishment_qty + float(max_days_to_last_for_pos[i] - max_days_to_last_for_pos[0])/ todayProductSale.fba_inventory.days_of_supply * fba_total_unit
                                else:
                                    amazon_recommended_replenishment_qty_i = amazon_recommended_replenishment_qty + fba_total_unit * i
                            else:
                                amazon_recommended_replenishment_qty_i = amazon_recommended_replenishment_qty
                            if amazon_recommended_replenishment_qty_i > shenzhen_inventory_and_receivable_purchased_qty:
                                qty_needed = amazon_recommended_replenishment_qty_i - shenzhen_inventory_and_receivable_purchased_qty
                                if qty_needed > float(moq) / 2:
                                    po_qty = max([qty_needed, moq])
                        if po_qty:
                            sheet.cell(row,start_col_restock + i).value = int(po_qty)
                    #先不加复杂的财务部分数据
                    # 增加财务数据
                    # if is_head_of_sales:
                    #     sheet.cell(row, start_col_restock + len(max_days_to_last_for_pos) ).value = skuManagedBySalesPerson.sales_person_name
                    #     sheet.cell(row, start_col_restock + len(max_days_to_last_for_pos) + 1).value = skuManagedBySalesPerson.sales_assistant_name
                    #     skuAssetLiabilityTables = SkuAssetLiabilityTable.objects.filter(sku = sku, for_sales = True).order_by('-date').all()
                    #     if skuAssetLiabilityTables.count() > 1:
                    #         if skuAssetLiabilityTables[0].date == skuAssetLiabilityTables[1].date:
                    #             skuAssetLiabilityTableForSalesPerson = SkuAssetLiabilityTable.objects.filter(sku = sku, for_sales = True, date = skuAssetLiabilityTables[0].date, initial = True).order_by('-date').first()
                    #         else:
                    #             skuAssetLiabilityTableForSalesPerson = SkuAssetLiabilityTable.objects.filter(sku = sku, for_sales = True).order_by('-date').first()
                    #     elif skuAssetLiabilityTables.count() == 1:
                    #         skuAssetLiabilityTableForSalesPerson = SkuAssetLiabilityTable.objects.filter(sku = sku, for_sales = True).order_by('-date').first()
                    #
                    #     sheet.cell(row, start_col_restock + len(max_days_to_last_for_pos) + 2).value = skuAssetLiabilityTableForSalesPerson.net_asset_amount
                    #
                    #     sheet.cell(row, start_col_restock + len(max_days_to_last_for_pos) + 3).value = get_sku_purchase_price(sku)
                    #     skuHeadShippingUnitCostSea = SkuHeadShippingUnitCost.objects.filter(sku = sku, type = ('S', 'Sea')).first()
                    #     if skuHeadShippingUnitCostSea:
                    #         sheet.cell(row, start_col_restock + len(max_days_to_last_for_pos) + 4).value = skuHeadShippingUnitCostSea.head_shipping_unit_cost
                    #     skuHeadShippingUnitCostAir = SkuHeadShippingUnitCost.objects.filter(sku = sku, type = ('A', 'Air')).first()
                    #     if skuHeadShippingUnitCostAir:
                    #         sheet.cell(row, start_col_restock + len(max_days_to_last_for_pos) + 5).value = skuHeadShippingUnitCostAir.head_shipping_unit_cost
                    if min([todayProductSale.fba_inventory.days_of_supply, lasting_day_of_total_fba_and_shenzhen_inventory_estimated_by_us]) < min_days_lasting_for_notice and not(todayProductSale.product.discontinued):
                        for cell in sheet["%i:%i" %(row,row)]:
                            cell.fill = yellow_fill

                row +=1
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename={date}-PO.xlsx'.format( \
               date=datetime.datetime.now().strftime('%Y-%m-%d') \
               ,)
        wb.save(response)
        return response
    template = loader.get_template('restock_today.html')
    context = {
        'form':UploadFileForm()
    }
    return HttpResponse(template.render(context, request))

@login_required
def shipment_today(request):
    if request.method == 'POST':
        inbound_sku_qty = {}
        form = UploadFileCountryForm(request.POST,request.FILES)
        if form.is_valid():
            if 'file' in request.FILES:
                fileInMemory = request.FILES['file'].read()
                filePath = BytesIO(fileInMemory)
                wb2 = load_workbook(filePath)
                sheet = wb2.worksheets[0]
                for row in range(1, sheet.max_row + 1):
                    sku = sheet.cell(row,1).value
                    qty = int(sheet.cell(row,2).value)
                    inbound_sku_qty[sku]= qty
            country = request.POST['country']
        wb = Workbook()
        sheet = wb.active
        sheet.title = '今日发货单'
        yellow_fill = PatternFill("solid", fgColor="FFFF00")
        min_shippable_qty = 20
        min_sold_qty_per_day_for_notice = 1
        max_days_lasting = 120
        row = 1
        titles = ['SKU', 'FBA可售库存', '中转库存', '在途库存', '亚马逊总库存', 'F199库存' \
                    , '过去7天平均日销售', '亚马逊预测能撑多少天', '我们预测总fba库存能撑多少天' \
                    , '可售库存能撑多少天','可售+中转库存能撑多少天' \
                    , '单价', '决策' \
                    ]
        for i, title in enumerate(titles):
            sheet.cell(row,i + 1).value = title
        row += 1
        inventories = Inventory.objects.filter(warehouse_name = shenzhen_warehouse_name, qty__gte = min_shippable_qty)
        # country = 'US'
        for inventory in inventories.all():
            sku = inventory.sku
            shenzhen_inventory = check_shenzhen_inventory(sku)
            if sku in inbound_sku_qty:
                shenzhen_inventory += inbound_sku_qty[sku]
                inbound_sku_qty.pop(sku, None)
            sheet.cell(row,1).value = sku
            todayProductSales = TodayProductSales.objects.filter(product__sku = sku,country=country)
            if todayProductSales.count():
                todayProductSale = todayProductSales.all()[0]
                sheet.cell(row,2).value = todayProductSale.fba_inventory.available
                sheet.cell(row,3).value = todayProductSale.fba_inventory.fc_unit
                sheet.cell(row,4).value = todayProductSale.fba_inventory.inbound_unit
                fba_total_unit = todayProductSale.fba_inventory.total_unit
                sheet.cell(row,5).value = fba_total_unit
                sheet.cell(row,6).value = shenzhen_inventory
                sheet.cell(row,7).value = round(todayProductSale.sold_qty_average_7d,1)
                sheet.cell(row,8).value = todayProductSale.fba_inventory.days_of_supply
                sheet.cell(row,9).value = int(todayProductSale.lasting_day_of_total_fba_unit_estimated_by_us)
                sheet.cell(row,10).value = int(todayProductSale.lasting_day_of_available_estimated_by_us)
                sheet.cell(row,11).value = int(todayProductSale.lasting_day_of_available_fc_estimated_by_us)
                sheet.cell(row,12).value = round(todayProductSale.average_price_7d,2)
                sheet.cell(row,12).number_format = u'"$ "#,##0.00'
                if min([todayProductSale.fba_inventory.days_of_supply, todayProductSale.lasting_day_of_total_fba_unit_estimated_by_us]) < max_days_lasting and todayProductSale.sold_qty_average_7d > min_sold_qty_per_day_for_notice:
                    for cell in sheet["%i:%i" %(row,row)]:
                        cell.fill = yellow_fill
            else:
                fba_inventory = FbaInventory.objects.filter(sku = sku,country=country).first()
                if fba_inventory:
                    sheet.cell(row,2).value = fba_inventory.available
                    sheet.cell(row,3).value = fba_inventory.fc_unit
                    sheet.cell(row,4).value = fba_inventory.inbound_unit
                    fba_total_unit = fba_inventory.total_unit
                    sheet.cell(row,5).value = fba_total_unit
                    shenzhen_inventory = check_shenzhen_inventory(sku)
                    sheet.cell(row,8).value = fba_inventory.days_of_supply
                sheet.cell(row,6).value = shenzhen_inventory
            row +=1
        for sku, qty in inbound_sku_qty.items():
            sheet.cell(row,1).value = sku
            todayProductSales = TodayProductSales.objects.filter(product__sku = sku,country=country)
            if todayProductSales.count():
                todayProductSale = todayProductSales.all()[0]
                sheet.cell(row,2).value = todayProductSale.fba_inventory.available
                sheet.cell(row,3).value = todayProductSale.fba_inventory.fc_unit
                sheet.cell(row,4).value = todayProductSale.fba_inventory.inbound_unit
                fba_total_unit = todayProductSale.fba_inventory.total_unit
                sheet.cell(row,5).value = fba_total_unit
                shenzhen_inventory = check_shenzhen_inventory(sku) + qty
                sheet.cell(row,6).value = shenzhen_inventory
                sheet.cell(row,7).value = round(todayProductSale.sold_qty_average_7d,1)
                sheet.cell(row,8).value = todayProductSale.fba_inventory.days_of_supply
                sheet.cell(row,9).value = int(todayProductSale.lasting_day_of_total_fba_unit_estimated_by_us)
                sheet.cell(row,10).value = int(todayProductSale.lasting_day_of_available_estimated_by_us)
                sheet.cell(row,11).value = int(todayProductSale.lasting_day_of_available_fc_estimated_by_us)
                sheet.cell(row,12).value = round(todayProductSale.average_price_7d,2)
                sheet.cell(row,12).number_format = u'"$ "#,##0.00'
                if min([todayProductSale.fba_inventory.days_of_supply, todayProductSale.lasting_day_of_total_fba_unit_estimated_by_us]) < max_days_lasting and todayProductSale.sold_qty_average_7d > min_sold_qty_per_day_for_notice:
                    for cell in sheet["%i:%i" %(row,row)]:
                        cell.fill = yellow_fill
            else:
                fbaInventory = FbaInventory.objects.filter(sku = sku,country=country)
                if fbaInventory.count():
                    fba_inventory = fbaInventory.all()[0]
                    sheet.cell(row,2).value = fba_inventory.available
                    sheet.cell(row,3).value = fba_inventory.fc_unit
                    sheet.cell(row,4).value = fba_inventory.inbound_unit
                    sheet.cell(row,5).value = fba_inventory.total_unit
                    shenzhen_inventory = check_shenzhen_inventory(sku) + qty
                    sheet.cell(row,6).value = shenzhen_inventory
                    sheet.cell(row,8).value = fba_inventory.days_of_supply
                else:
                    sheet.cell(row,6).value = qty
            row +=1
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename={date}-fbaShipment.xlsx'.format( \
               date=datetime.datetime.now().strftime('%Y-%m-%d') \
               ,)
        wb.save(response)
        return response

    template = loader.get_template('shipment_today.html')
    context = {
        'form':UploadFileCountryForm(initial={'country': 'US'})
    }
    return HttpResponse(template.render(context, request))


@login_required
def reprice_today(request):
    min_available_qty_for_adjust_price = 10
    standard_price = 25.99
    highest_price = 34.99
    country = 'US'
    template = loader.get_template('reprice_today.html')

    titles = ['产品图片', 'SKU', 'FBA可售库存', '中转库存', '在途库存', '亚马逊总库存', 'F199库存', '亚马逊预测撑天' \
            , '我们预测总fba库存撑天', '可售库存撑天','可售+中转库存撑天' \
            , '过去7天平均日销售', '今日销量', '正常价', '最高限价', '单价', '决策' \
            ]

    products_to_reprice_list = []

    #找出所有FBA有库存的商品
    for fba_inventory in FbaInventory.objects.filter(available__gt = min_available_qty_for_adjust_price, country=country):

        sku = fba_inventory.sku
        product_to_reprice_info = [""]
        product_to_reprice_info.append("")
        product_to_reprice_info.append(sku)
        fba_inventory_available = fba_inventory.available
        fba_inventory_fc_unit = fba_inventory.fc_unit
        fba_inventory_inbound_unit = fba_inventory.inbound_unit
        fba_total_unit = fba_inventory.total_unit
        shenzhen_inventory = check_shenzhen_inventory(sku)
        fba_inventory_days_of_supply = fba_inventory.days_of_supply
        product_to_reprice_info.append(fba_inventory_available)
        product_to_reprice_info.append(fba_inventory_fc_unit)
        product_to_reprice_info.append(fba_inventory_inbound_unit)
        product_to_reprice_info.append(fba_total_unit)
        product_to_reprice_info.append(shenzhen_inventory)
        product_to_reprice_info.append(fba_inventory_days_of_supply)


        #降价部分
        #判断价格是否高于正常价
        #先找到最新售价
        lastest_sales_having_this_sku = HistoryTodayProductSales.objects.filter(product__sku = sku, country=country).order_by('-date')

        if lastest_sales_having_this_sku.count():

            lastest_sales_having_this_sku = lastest_sales_having_this_sku.first()
            sold_qty_average_7d = lastest_sales_having_this_sku.sold_qty_average_7d
            average_price_today = lastest_sales_having_this_sku.average_price_7d
            if lastest_sales_having_this_sku.sales_amount != 0 and lastest_sales_having_this_sku.sold_qty != 0:
                average_price_today = lastest_sales_having_this_sku.sales_amount / lastest_sales_having_this_sku.sold_qty
            lasting_days_with_fba_inventory_available  = fba_inventory_available / sold_qty_average_7d
            lasting_days_with_fba_inventory_available_and_fc_unit = (fba_inventory_available + fba_inventory_fc_unit)/ sold_qty_average_7d
            lasting_days_with_fba_total_unit = fba_total_unit / sold_qty_average_7d
            sold_qty_average_7d = lastest_sales_having_this_sku.sold_qty_average_7d
            today_sold_qty = 0
            todayProductSales = TodayProductSales.objects.filter(product__sku = sku, country=country)
            if todayProductSales.count():
                todayProductSales = todayProductSales.first()
                today_sold_qty = todayProductSales.sold_qty
            product_to_reprice_info.append(int(lasting_days_with_fba_total_unit))
            product_to_reprice_info.append(int(lasting_days_with_fba_inventory_available))
            product_to_reprice_info.append(int(lasting_days_with_fba_inventory_available_and_fc_unit))
            product_to_reprice_info.append(round(sold_qty_average_7d,1))
            product_to_reprice_info.append(today_sold_qty)
            product_to_reprice_info.append(standard_price)
            product_to_reprice_info.append(highest_price)
            product_to_reprice_info.append(round(average_price_today, 2))

            if int(average_price_today) > int(standard_price):
                #高于正常价
                #判断销售是否在下降
                if today_sold_qty <= sold_qty_average_7d - 1:
                    #销售在下降，则要降价
                    #降价前还是要判断是否有断货风险
                    if lasting_days_with_fba_total_unit > 30 and lasting_days_with_fba_inventory_available_and_fc_unit > 20 :
                        #没有断货风险
                        product = Product.objects.get(sku = sku)
                        product_to_reprice_info[1] = product.image.url
                        product_to_reprice_info.append( '降价' )
                        product_to_reprice_info[0] = "blue"
                        products_to_reprice_list.append(product_to_reprice_info)
                elif lasting_days_with_fba_inventory_available_and_fc_unit < 20:
                    #可售+中转库存撑不到20天
                    if today_sold_qty >= sold_qty_average_7d:
                        product_to_reprice_info.append( '涨价' )
                        product_to_reprice_info[0] = "red"
                    product = Product.objects.get(sku = sku)
                    product_to_reprice_info[1] = product.image.url
                    products_to_reprice_list.append(product_to_reprice_info)
            elif lasting_days_with_fba_inventory_available_and_fc_unit < 20 and today_sold_qty > sold_qty_average_7d:
                #不高于正常价但是可售+中转库存撑不到20天，而且销售在上升
                product = Product.objects.get(sku = sku)
                product_to_reprice_info[1] = product.image.url
                product_to_reprice_info.append( '涨价' )
                product_to_reprice_info[0] = "red"
                products_to_reprice_list.append(product_to_reprice_info)
    context = {
        'titles': titles ,\
        'products_to_reprice_list': products_to_reprice_list \
    }
    return HttpResponse(template.render(context, request))

def scrape(url):
    headers = {
        'authority': 'www.amazon.com',
        'pragma': 'no-cache',
        'cache-control': 'no-cache',
        'dnt': '1',
        'upgrade-insecure-requests': '1',
        'user-agent': 'Mozilla/5.0 (X11; CrOS x86_64 8172.45.0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/51.0.2704.64 Safari/537.36',
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
        'sec-fetch-site': 'none',
        'sec-fetch-mode': 'navigate',
        'sec-fetch-dest': 'document',
        'accept-language': 'en-GB,en-US;q=0.9,en;q=0.8',
    }

    # Download the page using requests
    print("Downloading %s"%url)
    r = requests.get(url, headers=headers)
    # Simple check to check if page was blocked (Usually 503)
    if r.status_code > 500:
        if "To discuss automated access to Amazon data please contact" in r.text:
            print("Page %s was blocked by Amazon. Please try using better proxies\n"%url)
        else:
            print("Page %s must have been blocked by Amazon as the status code was %d"%(url,r.status_code))
        return None
    # Pass the HTML of the page and create
    f = open(os.path.join(BASE_DIR,'amazon_response.html'), "w")
    f.write(r.text)
    f.close()
    data = extractor.extract(r.text,base_url=url)
    reviews = []
    for r in data['reviews']:
        r["product"] = data["product_title"]
        r['url'] = url
        if 'verified_purchase' in r:
            if r['verified_purchase'] and 'Verified Purchase' in r['verified_purchase']:
                r['verified_purchase'] = True
            else:
                r['verified_purchase'] = False
        r['rating'] = r['rating'].split(' out of')[0]
        date_posted = r['date'].split('on ')[-1]
        if r['images']:
            r['images'] = "\n".join(r['images'])
        r['date'] = dateparser.parse(date_posted).strftime('%d %b %Y')
        reviews.append(r)
    data.pop('histogram', None)
    # histogram = {}
    # for h in data['histogram']:
    #     histogram[h['key']] = h['value']
    # data['histogram'] = histogram
    data['average_rating'] = float(data['average_rating'].split(' out')[0])
    data['reviews'] = reviews
    data['number_of_reviews'] = int(data['number_of_reviews'].split('  customer')[0])
    return data

def distinct_size_color(size_color_str):
    Size_Tip_Str = ['Size: ']
    Color_Tip_Str = ['Color: ', 'Colour: ']
    pos_dict = {}
    size_pos = color_pos =-1
    for k in Size_Tip_Str:
        size_pos = size_color_str.find(k)
        if size_pos > -1:
            pos_dict['size'] = {'tip_str':k, 'pos':size_pos}
            break
    for k in Color_Tip_Str:
        color_pos = size_color_str.find(k)
        if color_pos > -1:
            pos_dict['color'] = {'tip_str':k, 'pos':color_pos}
            break
    if len(pos_dict.keys()) == 1:
        return {list(pos_dict.keys())[0]: size_color_str[len(list(pos_dict.values())[0]['tip_str']):]}
    elif len(pos_dict.keys()) > 1:
        if pos_dict['size']['pos'] > pos_dict['color']['pos']:
            return {'color': size_color_str[len(pos_dict['color']['tip_str']):pos_dict['size']['pos']] \
                    ,'size': size_color_str[(pos_dict['size']['pos'] + len(pos_dict['size']['tip_str'])):]}
        else:
            return {'size': size_color_str[len(pos_dict['size']['tip_str']):pos_dict['color']['pos']] \
                    ,'color': size_color_str[(pos_dict['color']['pos'] + len(pos_dict['color']['tip_str'])):]}

def generate_pivot_table(data_list):
    variants = []
    for i in data_list:
        if 'variant' in i and i['variant'] != None:
            variants.append(i['variant'])
    return

def size_color_counter_dict_to_list(x):
    res1 = []
    res2 = []
    for k,v in sorted(x.items(), key=lambda item: item[1], reverse=True):
        res1.append(k)
        res2.append(v)
    return {'name': res1, 'value': res2}

@login_required
def get_reviews_stat(request):
    if request.method == 'POST':
        from collections import Counter
        reviews_by_size_color = []
        earliest_date = latest_date = None

        form = UploadFileForm(request.POST,request.FILES)
        if form.is_valid():
            # a-size-base 2 是日期列，a-size-mini的变量列
            col_nums = {'a-size-base 2':0, 'a-size-mini':0, 'a-size-mini 2':0}
            fileInMemory = request.FILES['file'].read().decode('utf-8')
            csv_data = csv.reader(StringIO(fileInMemory), delimiter=',')
            for i, row in enumerate(csv_data):
                if i == 0:
                    for col_num, title in enumerate(row):
                        if title in col_nums.keys():
                            col_nums[title] = col_num
                    col_nums['date'] = col_nums['a-size-base 2']
                    col_nums['variant'] = col_nums['a-size-mini'] or col_nums['a-size-mini 2']
                else:
                    if row[col_nums['a-size-mini']] == 'Verified Purchase':
                        col_nums['variant'] = col_nums['a-size-mini 2']
                    date_posted = dateparser.parse(row[col_nums['date']].split('on ')[-1])
                    variant_str = row[col_nums['variant']]
                    if variant_str:
                        reviews_by_size_color.append(distinct_size_color(variant_str))
                    if latest_date == None:
                        earliest_date = latest_date = date_posted
                    else:
                        if date_posted > latest_date:
                            latest_date = date_posted
                        if date_posted < earliest_date:
                            earliest_date = date_posted
            day_range_tip = '从%s到%s，共%i天%.1f月' %(earliest_date.strftime('%Y-%m-%d'), latest_date.strftime('%Y-%m-%d'), (latest_date - earliest_date).days, float((latest_date - earliest_date).days)/30)
            counter_size_color = {}
            if len(reviews_by_size_color[0].keys()) == 1:
                size_color = list(reviews_by_size_color[0].keys())[0]
                size_color_list = [i[size_color] for i in reviews_by_size_color]
                counter_size_color[size_color] = dict(Counter(size_color_list))
                counter_size_color[size_color] = size_color_counter_dict_to_list(counter_size_color[size_color])
            elif len(reviews_by_size_color[0].keys()) > 1:
                for i_size_color in list(reviews_by_size_color[0].keys()):
                    size_color_list = [i[i_size_color] for i in reviews_by_size_color]
                    counter_size_color[i_size_color] = dict(Counter(size_color_list))
                    counter_size_color[i_size_color] = size_color_counter_dict_to_list(counter_size_color[i_size_color])
            return JsonResponse({'counter_size_color': counter_size_color, 'day_range_tip':day_range_tip}, safe=False)
    template = loader.get_template('get_reviews.html')
    context = {
        'form':UploadFileForm()
    }
    return HttpResponse(template.render(context, request))

@login_required
def get_upcs(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST,request.FILES)
        if form.is_valid():
            fileInMemory = request.FILES['file'].read()
            filePath = BytesIO(fileInMemory)
            wb2 = load_workbook(filePath)
            sheet = wb2.worksheets[0]
            sku_upc_list = []
            for row in range(1, sheet.max_row + 1):
                sku = str(sheet.cell(row,1).value)
                if sku and sku != 'None' and sku.upper().strip():
                    sku = sku.upper().strip()
                    sku_upc, created = SkuUpc.objects.get_or_create(sku = sku)
                    if created:
                        upc = Upc.objects.filter(used = False).first()
                        sku_upc.upc = upc
                        sku_upc.save()
                        upc.used = True
                        upc.save()
                    sku_upc_list.append(sku_upc)
            wb = Workbook()
            sheet = wb.active
            sheet.title = 'sku_upc'
            for i,sku_upc in enumerate(sku_upc_list):
                sheet.cell(i + 1, 1).value = sku_upc.sku
                sheet.cell(i + 1, 2).value = sku_upc.upc.upc
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename={date}-sku-upc.xlsx'.format( \
                   date=datetime.datetime.now().strftime('%Y-%m-%d') \
                   ,)
            wb.save(response)
            return response
    template = loader.get_template('get_upcs.html')
    context = {
        'form':UploadFileForm()
    }
    return HttpResponse(template.render(context, request))

@login_required
def input_upcs(request):
    if request.method == 'POST' : # and request.get_host() == '209.97.151.168':
        form = UploadFileForm(request.POST,request.FILES)
        if form.is_valid():
            fileInMemory = request.FILES['file'].read()
            filePath = BytesIO(fileInMemory)
            wb2 = load_workbook(filePath)
            sheet = wb2.worksheets[0]
            for row in range(1, sheet.max_row + 1):
                upc = str(sheet.cell(row,1).value)
                if upc:
                    upc = upc.upper().strip()
                    upc, created = Upc.objects.get_or_create(upc = upc)
    template = loader.get_template('input_upcs.html')
    context = {
        'form':UploadFileForm()
    }
    return HttpResponse(template.render(context, request))

@login_required
def update_sku_supplier(request):
    if request.method == 'POST' : # and request.get_host() == '209.97.151.168':
        form = UploadFileForm(request.POST,request.FILES)
        if form.is_valid():
            fileInMemory = request.FILES['file'].read()
            filePath = BytesIO(fileInMemory)
            wb2 = load_workbook(filePath)
            sheet = wb2.worksheets[0]
            for row in range(1, sheet.max_row + 1):
                sku = str(sheet.cell(row,1).value)
                supplier_long_str = sheet.cell(row,2).value
                supplier_jingdouyun_id = supplier_long_str[:6]
                if sku:
                    sku = sku.upper().strip()
                    if not SkuSupplier.objects.filter(sku = sku, supplier_id = supplier_jingdouyun_id).exists():
                        sku_supplier = SkuSupplier(sku = sku, supplier_id = supplier_jingdouyun_id)
                        sku_supplier.save()
                    for sku_supplier in SkuSupplier.objects.filter(sku = sku).all():
                        if sku_supplier.supplier_id != supplier_jingdouyun_id:
                            sku_supplier.delete()

    template = loader.get_template('update_sku_supplier.html')
    context = {
        'form':UploadFileForm()
    }
    return HttpResponse(template.render(context, request))

@login_required
def update_all_product_purchase_price(request):
    if request.method == 'POST':
        if request.user.is_superuser:
            fileInMemory = request.FILES['file'].read()
            filePath = BytesIO(fileInMemory)
            wb2 = load_workbook(filePath)
            sheet = wb2.worksheets[0]
            max_column = sheet.max_column
            correct_price_title = '单位成本'
            correct_price_title_cell_row = 5
            correct_price_title_cell_columb = 11
            if sheet.cell(correct_price_title_cell_row, correct_price_title_cell_columb).value == correct_price_title:

                date = datetime.datetime.strptime(sheet.cell(3, 1).value[-10:], "%Y-%m-%d").date()
                for row in range(6, sheet.max_row):
                    sku = sheet.cell(row, 1).value
                    price = sheet.cell(row, correct_price_title_cell_columb).value
                    skuPurchasingPrice = SkuPurchasingPrice.objects.filter(sku = sku, date = date)
                    if skuPurchasingPrice.count():
                        skuPurchasingPrice = skuPurchasingPrice.first()
                        skuPurchasingPrice.purchasing_price = price
                        skuPurchasingPrice.save()
                    else:
                        (sku_purchasing_price, create) = SkuPurchasingPrice.objects.get_or_create(sku = sku ,\
                                                                                            purchasing_price = price ,\
                                                                                            date = date)

            else:
                return HttpResponse('请检查上传的文件格式是否正确')
    template = loader.get_template('update_all_product_purchase_price.html')
    context = {
        'form':UploadFileForm()
    }

    return HttpResponse(template.render(context, request))

@login_required
def update_all_product_head_shipping_unit_cost(request):
    if request.method == 'POST':
        fileInMemory = request.FILES['file'].read()
        filePath = BytesIO(fileInMemory)
        wb2 = load_workbook(filePath)
        sheet = wb2.worksheets[0]
        max_column = sheet.max_column

        year_month = sheet.cell(1, 1).value
        year_month_day = str(year_month) + '01'
        date = datetime.datetime.strptime(year_month_day, "%Y%m%d").date()

        country = sheet.cell(2, 1).value.upper()

        for row in range(4, sheet.max_row + 1):
            sku = sheet.cell(row, 1).value
            head_shipping_unit_cost = sheet.cell(row, 3).value
            skuHeadShippingUnitCost = SkuHeadShippingUnitCost.objects.filter(sku = sku ,\
                                                                                    type = ('G', 'General') ,\
                                                                                    country = country ,\
                                                                                    date = date)
            if skuHeadShippingUnitCost.count():
                skuHeadShippingUnitCost = skuHeadShippingUnitCost.first()
                skuHeadShippingUnitCost.head_shipping_unit_cost = head_shipping_unit_cost
                skuHeadShippingUnitCost.save()
            else:
                (sku_purchasing_price, create) = SkuHeadShippingUnitCost.objects.get_or_create(sku = sku ,\
                                                                                    type = ('G', 'General') ,\
                                                                                    country = country ,\
                                                                                    head_shipping_unit_cost = head_shipping_unit_cost ,\
                                                                                    date = date)

    template = loader.get_template('update_all_product_head_shipping_unit_cost.html')
    context = {
        'form':UploadFileForm()
    }

    return HttpResponse(template.render(context, request))

def setting_up_sku_managed_by_sales_person(sku, sales_person_name, sales_assistant_name):
    skuManagedBySalesPerson = SkuManagedBySalesPerson.objects.filter(sku = sku).first()
    if skuManagedBySalesPerson:
        skuManagedBySalesPerson.sales_person_name = sales_person_name
        skuManagedBySalesPerson.sales_assistant_name = sales_assistant_name
        skuManagedBySalesPerson.save()
    else:
        (skuManagedBySalesPerson, created) = SkuManagedBySalesPerson.objects.get_or_create(sku = sku ,\
                                                                                        sales_person_name = sales_person_name, \
                                                                                        sales_assistant_name = sales_assistant_name)

def setting_up_sku_contributor(sku, proposer_name, designer_name):
    skuContributor = SkuContributor.objects.filter(sku = sku).first()
    if skuContributor:
        if proposer_name:
            skuContributor.proposer_name = proposer_name
        if designer_name:
            skuContributor.designer_name = designer_name
        skuContributor.save()
    else:
        if proposer_name and designer_name:
            (skuContributor, created) = SkuContributor.objects.get_or_create(sku = sku,proposer_name = proposer_name, designer_name = designer_name)


def initialize_sku_asset_liability(sku):
    initial = True
    date = datetime.date.today()
    fba_inventory = 0
    shenzhen_inventory = 0
    fbaInventory = FbaInventory.objects.filter(sku = sku)
    if fbaInventory.count():
        fba_inventory = fbaInventory.first().total_unit

    inventory = Inventory.objects.filter(warehouse_name = shenzhen_warehouse_name,sku = sku)
    if inventory.count():
        shenzhen_inventory = inventory.first().qty
    initial_inventory_quantity = fba_inventory + shenzhen_inventory
    greater = SkuPurchasingPrice.objects.filter(sku = sku, date__gte=date).order_by('date').first()
    if greater:
        unit_purchasing_price = greater.purchasing_price
    else:
        less = SkuPurchasingPrice.objects.filter(sku = sku, date__lte=date).order_by('-date').first()
        if less:
            unit_purchasing_price = less.purchasing_price
        else:
            unit_purchasing_price = 0.0
    greater = SkuHeadShippingUnitCost.objects.filter(sku = sku, type = ('G', 'General'), country = 'US', date__gte=date).order_by('date').first()
    if greater:
        head_shipping_unit_price = greater.head_shipping_unit_cost
    else:
        less = SkuHeadShippingUnitCost.objects.filter(sku = sku, type = ('G', 'General'), country = 'US', date__lte=date).order_by('-date').first()
        if less:
            head_shipping_unit_price = less.head_shipping_unit_cost
        else:
            head_shipping_unit_price = 0.0
    initial_other_cost = 0.0
    initial_inventory_value = (fba_inventory + shenzhen_inventory) * unit_purchasing_price
    initial_liabilities = fba_inventory * (unit_purchasing_price + head_shipping_unit_price) + shenzhen_inventory * unit_purchasing_price
    liabilities = initial_liabilities
    initial_investment = 0.0
    history_inventment = 0.0
    cash_amount = 0.0
    net_asset_amount = -liabilities

    #先看看给销售外角色用的资产负债表是否已存在，存在的话，就不用管,不存在，则要读取当前的库存数量，对于已经有库存的产品，设置要放在月初1号，不是1号，则不会建立资产负债表，对于没有库存的产品，可以建立
    skuAssetLiabilityTable = SkuAssetLiabilityTable.objects.filter(sku = sku, for_sales = False).first()
    if not skuAssetLiabilityTable :
        if (initial_inventory_value > 0 and date.day == 15) or initial_inventory_value == 0:
            skuAssetLiabilityTable = SkuAssetLiabilityTable(sku = sku ,\
                                                            liabilities = liabilities ,\
                                                            date = date ,\
                                                            initial_inventory_quantity = initial_inventory_quantity ,\
                                                            unit_purchasing_price = unit_purchasing_price ,\
                                                            initial_inventory_value = initial_inventory_value ,\
                                                            head_shipping_unit_price = head_shipping_unit_price ,\
                                                            initial_other_cost = initial_other_cost ,\
                                                            initial_liabilities = initial_liabilities ,\
                                                            initial = initial ,\
                                                            initial_investment = initial_investment ,\
                                                            history_inventment = history_inventment ,\
                                                            cash_amount = cash_amount ,\
                                                            net_asset_amount = net_asset_amount)
            skuAssetLiabilityTable.save()
    #再建给销售夜色的资产负债表
        skuAssetLiabilityTable = SkuAssetLiabilityTable(sku = sku ,\
                                                        for_sales = True ,\
                                                        liabilities = liabilities ,\
                                                        date = date ,\
                                                        initial_inventory_quantity = initial_inventory_quantity ,\
                                                        unit_purchasing_price = unit_purchasing_price ,\
                                                        initial_inventory_value = initial_inventory_value ,\
                                                        head_shipping_unit_price = head_shipping_unit_price ,\
                                                        initial_other_cost = initial_other_cost ,\
                                                        initial_liabilities = initial_liabilities ,\
                                                        initial = initial ,\
                                                        initial_investment = initial_investment ,\
                                                        history_inventment = history_inventment ,\
                                                        cash_amount = cash_amount ,\
                                                        net_asset_amount = net_asset_amount)
        skuAssetLiabilityTable.save()
    return

@login_required
def update_sales_person_and_managing_sku_list(request):
    if request.method == 'POST':
        fileInMemory = request.FILES['file'].read()
        filePath = BytesIO(fileInMemory)
        wb2 = load_workbook(filePath)
        sheet = wb2.worksheets[0]
        max_column = sheet.max_column

        # 检查title是否正确
        titles = sheet.cell(1, 1).value + sheet.cell(1, 2).value + sheet.cell(1, 3).value + sheet.cell(1, 4).value + sheet.cell(1, 5).value + sheet.cell(1, 6).value
        if titles != 'SKU销售名销售助理名选品师名设计师名初始化销售的财务报表否':
            return HttpResponse('请检查上传的模版')
        for row in range(2, sheet.max_row + 1):
            sku = sheet.cell(row, 1).value
            sales_person_name = sheet.cell(row, 2).value
            sales_assistant_name = sheet.cell(row, 3).value
            proposer_name = sheet.cell(row, 4).value
            designer_name = sheet.cell(row, 5).value
            initial = sheet.cell(row, 6).value
            if initial:
                initial = True
            else:
                initial = False
            setting_up_sku_managed_by_sales_person(sku, sales_person_name, sales_assistant_name)
            if proposer_name or designer_name:
                setting_up_sku_contributor(sku, proposer_name, designer_name)
            if initial:
                initialize_sku_asset_liability(sku)
    template = loader.get_template('update_sales_person_and_managing_sku_list.html')
    context = {
        'form':UploadFileForm()
    }

    return HttpResponse(template.render(context, request))

@login_required
def update_paid_purchase_order(request):
    if request.method == 'POST':
        fileInMemory = request.FILES['file'].read()
        filePath = BytesIO(fileInMemory)
        wb2 = load_workbook(filePath)
        sheet = wb2.worksheets[0]
        max_column = sheet.max_column

        po_date_column = 1
        po_number_column = 2
        sku_column = 5
        qty_column = 11
        transaction_amount_column = 13

        po_date_title = '采购日期'
        po_number_title = '采购单据号'
        sku_title = '商品编号'
        qty_title = '数量'
        transaction_amount_title = '采购金额'

        # 先检查表头对不对

        if sheet.cell(5, po_date_column).value == po_date_title and \
            sheet.cell(5, po_number_column).value == po_number_title and \
            sheet.cell(5, sku_column).value == sku_title and \
            sheet.cell(5, qty_column).value == qty_title and \
            sheet.cell(5, transaction_amount_column).value == transaction_amount_title :
            for row in range(6, sheet.max_row):
                po_date = sheet.cell(row, po_date_column).value
                po_date = datetime.datetime.strptime(po_date, "%Y-%m-%d").date()
                po_number = sheet.cell(row, po_number_column).value
                sku = sheet.cell(row, sku_column).value
                qty = sheet.cell(row, qty_column).value
                transaction_amount = sheet.cell(row, transaction_amount_column).value
                skuPurchaseOrder = SkuPurchaseOrder.objects.filter(sku = sku, po_number = po_number).first()
                if skuPurchaseOrder:
                   skuPurchaseOrder.date = po_date
                   skuPurchaseOrder.qty = qty
                   skuPurchaseOrder.transaction_amount = transaction_amount
                else:
                    skuPurchaseOrder = SkuPurchaseOrder(sku = sku, po_number = po_number ,\
                                                        date = po_date, qty = qty, transaction_amount = transaction_amount)
                skuPurchaseOrder.save()
        else:
            return HttpResponse('请检查上传的文件格式是否正确')
    template = loader.get_template('update_paid_purchase_order.html')
    context = {
        'form':UploadFileForm()
    }

    return HttpResponse(template.render(context, request))

def notFound2Zero(k, dict, *args):
    if args:
        if k in dict:
            return dict[k][args[0]]
        return 0.0
    else:
        if k in dict:
            return dict[k]
        return 0.0

@login_required
def update_sales_transaction(request):
    template = loader.get_template('update_sales_transaction.html')
    if request.method == 'POST':
        form = UploadTransactionAdCurrencyForm(request.POST, request.FILES)
        if form.is_valid():
            START_ROW = 2
            # PRODUCT_AD_COLUMN_NUM_DICT = {'sku':7, 'value': 13}
            CHECK_KEYWORD = {'sku': 'Advertised SKU', 'value': 'Spend'}
            CURRENCY_KEYWORD = {'US':'usd2cny', 'CA': 'cad2cny', 'AU': 'aud2cny' \
                                , 'AE': 'aed2cny', 'UK': 'gbp2cny' , 'DE': 'eur2cny' \
                                , 'FR': 'eur2cny', 'IT': 'eur2cny', 'ES': 'eur2cny' \
                                , 'SG': 'sgd2cny'}
            adCosts = {}
            if 'file2' in request.FILES:
                fileInMemory = request.FILES['file2'].read()
                filePath = BytesIO(fileInMemory)
                adCosts = readSkuValue(filePath, CHECK_KEYWORD, START_ROW)
            if 'file3' in request.FILES:
                fileInMemory = request.FILES['file3'].read()
                filePath = BytesIO(fileInMemory)
                brandAdCosts = readSkuValue(filePath, {'sku': 'Campaign Name', 'value': 'Spend'}, 2)
                totalBrandAdCost = sum([v for k,v in brandAdCosts.items()])
            else:
                totalBrandAdCost = 0.0
            country = 'US'
            file2InMemory = request.FILES['file1'].read()
            sales = readTransaction(file2InMemory)
            totalQuantity = sum([v['quantity'] for k, v in sales.items() if k])
            fixedCost = readFixedCost(file2InMemory)
            currencyRatesDict, currencyRateUpdateTime = readCurrencyRate(BytesIO(request.FILES['file4'].read()))
            currencyRate = currencyRatesDict[CURRENCY_KEYWORD[country]]
            skus = list(set(list(sales.keys()) + list(adCosts.keys())))
            if "" in skus:
                skus.remove("")
            skus.sort()
            start_date = datetime.datetime.strptime(request.POST['start_date'], "%Y-%m-%d").date()
            end_date = datetime.datetime.strptime(request.POST['end_date'], "%Y-%m-%d").date()
            for sku in skus:
                adCost = notFound2Zero(sku,adCosts) * currencyRate
                quantity = notFound2Zero(sku,sales,'quantity')
                brandAdCost = 0.0
                fixedCostPerSKU = 0.0
                if totalQuantity:
                    brandAdCost = float(quantity) / float(totalQuantity) * totalBrandAdCost
                    fixedCostPerSKU = float(quantity) / float(totalQuantity) * fixedCost
                adCost += brandAdCost
                amazonFee = -1.0 * (notFound2Zero(sku,sales,'AmazonFee')  + fixedCostPerSKU)
                skuManagedBySalesPerson = SkuManagedBySalesPerson.objects.filter(sku = sku).first()
                profitLossTable = ProfitLossTable.objects.filter(sku = sku, start_date = start_date, end_date = end_date).first()
                if not profitLossTable:
                    profitLossTable = ProfitLossTable(sku = sku, start_date = start_date, end_date = end_date, status = ('T', '交易订单录入'))
                profitLossTable.sales_amount = notFound2Zero(sku,sales,'sales')
                profitLossTable.sales_quantity = quantity
                profitLossTable.amazon_fee = amazonFee
                profitLossTable.ad_fee = adCost
                profitLossTable.country = (country, country)
                profitLossTable.currency_rate = currencyRate
                if skuManagedBySalesPerson:
                    sales_person_name = skuManagedBySalesPerson.sales_person_name
                    profitLossTable.sales_person_name = sales_person_name
                profitLossTable.save()
        else:
            context = {
                'form': form
            }
            return HttpResponse(template.render(context, request))
    context = {
        'form':UploadTransactionAdCurrencyForm()
    }
    return HttpResponse(template.render(context, request))

@login_required
def confirm_po_head_shipping(request):
    template = loader.get_template('confirm_po_head_shipping.html')
    profitLossTable = ProfitLossTable.objects.filter(status = ('T', '交易订单录入')).order_by('-end_date').first()
    no_need_to_confirm = False
    if profitLossTable:
        skuProfitLossTableDict = {}
        skuHeadShippingCostDict = {}
        total_po_amount = 0.0
        start_date = profitLossTable.start_date
        end_date = profitLossTable.end_date
        # 先统计头程运费
        fbaShipmentCosts = FbaShipmentCost.objects.filter(date__gte=start_date,date__lte=end_date)
        if fbaShipmentCosts.count():
            for fbaShipmentCost in fbaShipmentCosts:
                shipment_id = fbaShipmentCost.shipment_id
                cost_per_kg = fbaShipmentCost.cost_per_kg
                fbaShipments = FbaShipment.objects.filter(shipment_id = shipment_id)
                for fbaShipment in fbaShipments:
                    for shippedSkuQty in fbaShipment.shipped_sku_qties.all():
                        sku = shippedSkuQty.sku
                        qty = shippedSkuQty.qty
                        skuWeight = SkuWeight.objects.filter(sku=sku).first()
                        if skuWeight.heavy_or_light:
                            weight = skuWeight.real_weight
                        elif fbaShipmentCost.shipway == ('E', '快递'):
                            weight = skuWeight.converted_weight * 6.0 / 5.0
                        else:
                            weight = skuWeight.converted_weight
                        head_shipping_cost = weight * qty * cost_per_kg
                        if sku in skuHeadShippingCostDict:
                            skuHeadShippingCostDict[sku] += head_shipping_cost
                        else:
                            skuHeadShippingCostDict[sku] = head_shipping_cost
        skus_in_profitLossTables = set()
        for profitLossTable in ProfitLossTable.objects.filter(status = ('T', '交易订单录入')).order_by('-end_date').all():
            sku = profitLossTable.sku
            skus_in_profitLossTables.add(sku)
            skuProfitLossTableDict[sku] = {'profitLossTable': profitLossTable, 'total_sku_po_amount':0.0, 'total_sku_head_shipping_cost':0.0}

            skuPurchaseOrders = SkuPurchaseOrder.objects.filter(sku = sku, date__gte=start_date,date__lte=end_date)
            total_sku_po_amount = 0.0
            if skuPurchaseOrders.count():
                total_sku_po_amount =  skuPurchaseOrders.aggregate(Sum('transaction_amount'))['transaction_amount__sum']
            skuProfitLossTableDict[sku]['total_sku_po_amount'] = total_sku_po_amount
            total_po_amount += total_sku_po_amount
            if sku in skuHeadShippingCostDict:
                skuProfitLossTableDict[sku]['total_sku_head_shipping_cost'] = skuHeadShippingCostDict[sku]
        for skuPurchaseOrder in SkuPurchaseOrder.objects.filter(date__gte=start_date,date__lte=end_date).all():
            if skuPurchaseOrder.sku not in skus_in_profitLossTables:
                sku = skuPurchaseOrder.sku
                skuProfitLossTableDict[sku] = {'profitLossTable': profitLossTable, 'total_sku_po_amount':0.0, 'total_sku_head_shipping_cost':0.0}
                skuPurchaseOrders = SkuPurchaseOrder.objects.filter(sku = sku, date__gte=start_date,date__lte=end_date)
                total_sku_po_amount =  skuPurchaseOrders.aggregate(Sum('transaction_amount'))['transaction_amount__sum']
                skuProfitLossTableDict[sku]['total_sku_po_amount'] = total_sku_po_amount

                total_po_amount += total_sku_po_amount
                if sku in skuHeadShippingCostDict:
                    skuProfitLossTableDict[sku]['total_sku_head_shipping_cost'] = skuHeadShippingCostDict[sku]
                # skuManagedBySalesPerson = SkuManagedBySalesPerson.objects.filter(sku = sku).first()
                # profitLossTable = ProfitLossTable.objects.filter(sku = sku, start_date = start_date, end_date = end_date).first()
                # if not profitLossTable:
                #     profitLossTable = ProfitLossTable(sku = sku, start_date = start_date, end_date = end_date, status = ('T', '交易订单录入'))
                # profitLossTable.country = ('US','US')
                # if skuManagedBySalesPerson:
                #     sales_person = skuManagedBySalesPerson.sales_person
                #     profitLossTable.sales_person = sales_person
                #     profitLossTable.sales_person_bonus_fee_percent = skuManagedBySalesPerson.bonus_percent
                # profitLossTable.save()

    else:
        no_need_to_confirm = True
    if request.method == 'POST':
        form = ConfirmPoHeadShippingForm(request.POST)
        if form.is_valid():
            if request.POST['yes_or_no'] == 'yes':
                # 对有销售数据的损益表更新购货信息以及头程运费信息
                for profitLossTable in ProfitLossTable.objects.filter(status = ('T', '交易订单录入')).order_by('-end_date').all():
                    sku = profitLossTable.sku
                    profitLossTable.product_purchasing_fee = skuProfitLossTableDict[sku]['total_sku_po_amount']
                    if sku in skuHeadShippingCostDict:
                        profitLossTable.head_shipping_fee = skuHeadShippingCostDict[sku]
                    profitLossTable.status = ('P', '购货和头程运费确认')
                    profitLossTable.save()
                #将没有销售数据但是有购货订单的损益表建立了
                sku_in_po_not_in_profitLossTables = set()

                for skuPurchaseOrder in SkuPurchaseOrder.objects.filter(date__gte=start_date,date__lte=end_date).all():
                    if skuPurchaseOrder.sku not in skus_in_profitLossTables:
                        sku = skuPurchaseOrder.sku
                        sku_in_po_not_in_profitLossTables.add(sku)
                        skuPurchaseOrders = SkuPurchaseOrder.objects.filter(sku = sku, date__gte=start_date,date__lte=end_date)
                        total_sku_po_amount =  skuPurchaseOrders.aggregate(Sum('transaction_amount'))['transaction_amount__sum']
                        skuManagedBySalesPerson = SkuManagedBySalesPerson.objects.filter(sku = sku).first()
                        profitLossTable = ProfitLossTable.objects.filter(sku = sku, start_date = start_date, end_date = end_date).first()
                        if not profitLossTable:
                            profitLossTable = ProfitLossTable(sku = sku, start_date = start_date, end_date = end_date)
                        profitLossTable.country = ('US','US')
                        profitLossTable.status = ('P', '购货和头程运费确认')
                        if skuManagedBySalesPerson:
                            sales_person_name = skuManagedBySalesPerson.sales_person_name
                            profitLossTable.sales_person_name = sales_person_name
                        profitLossTable.product_purchasing_fee = total_sku_po_amount
                        if sku in skuHeadShippingCostDict:
                            profitLossTable.head_shipping_fee = skuHeadShippingCostDict[sku]
                        profitLossTable.save()
                #将没有销售数据没有有购货订单，但是有头程运费的的损益表建立了
                for sku, sku_head_shipping_cost in skuHeadShippingCostDict.items():
                    if sku not in sku_in_po_not_in_profitLossTables and sku not in skus_in_profitLossTables:
                        skuManagedBySalesPerson = SkuManagedBySalesPerson.objects.filter(sku = sku).first()
                        profitLossTable = ProfitLossTable.objects.filter(sku = sku, start_date = start_date, end_date = end_date).first()
                        if not profitLossTable:
                            profitLossTable = ProfitLossTable(sku = sku, start_date = start_date, end_date = end_date)
                        profitLossTable.country = ('US','US')
                        profitLossTable.status = ('P', '购货和头程运费确认')
                        if skuManagedBySalesPerson:
                            sales_person_name = skuManagedBySalesPerson.sales_person_name
                            profitLossTable.sales_person_name = sales_person_name
                        profitLossTable.head_shipping_fee = skuHeadShippingCostDict[sku]
                        profitLossTable.save()
                no_need_to_confirm = True
        else:
            context = {
                'start_date' : start_date ,\
                'end_date' : end_date ,\
                'po_amount':total_po_amount ,\
                'head_shipping_cost': sum(skuHeadShippingCostDict.values()),\
                'form': form
            }
            return HttpResponse(template.render(context, request))
    if no_need_to_confirm:
        template = loader.get_template('no_need_confirm_po_head_shipping.html')
        context = {}
    else:
        context = {
            'start_date' : start_date ,\
            'end_date' : end_date ,\
            'po_amount':total_po_amount ,\
            'head_shipping_cost': sum(skuHeadShippingCostDict.values()),\
            'form':ConfirmPoHeadShippingForm()
        }
    return HttpResponse(template.render(context, request))

@login_required
def get_sku_pl_al_table(request,sku):
    if request.user.groups.filter(name='head_of_sales').exists() or \
      SkuManagedBySalesPerson.objects.filter(sku = sku, sales_person_name = request.user.username).count() or \
      SkuManagedBySalesPerson.objects.filter(sku = sku, sales_assistant_name = request.user.username).count() or \
      SkuContributor.objects.filter(sku = sku, proposer_name = request.user.username).count() or \
      SkuContributor.objects.filter(sku = sku, designer_name = request.user.username).count():

        template = loader.get_template('get_sku_pl_al_table.html')
        profitLossTable = ProfitLossTable.objects.filter(sku = sku, status = ('P', '购货和头程运费确认')).order_by('-end_date').first()
        if profitLossTable:
            start_date = profitLossTable.start_date
            end_date = profitLossTable.end_date
        #先判断用户是哪类？1 运营，无该产品的其他贡献 2:运营，且有其他产品的贡献 3 非运营，但是有其他产品的贡献
        if SkuManagedBySalesPerson.objects.filter(sku = sku, sales_person_name = request.user.username).count() or \
            SkuManagedBySalesPerson.objects.filter(sku = sku, sales_assistant_name = request.user.username).count() and \
            SkuContributor.objects.filter(sku = sku, proposer_name = request.user.username).count() == 0 and \
            SkuContributor.objects.filter(sku = sku, designer_name = request.user.username).count() == 0:
            skuAssetLiabilityTables = SkuAssetLiabilityTable.objects.filter(sku = sku, for_sales = True, date__lt = end_date).order_by('-date').all()
            if skuAssetLiabilityTables.count() > 1:
                if skuAssetLiabilityTables[0].date == skuAssetLiabilityTables[1].date:
                    skuAssetLiabilityTableForSalesPerson = SkuAssetLiabilityTable.objects.filter(sku = sku, for_sales = True, date = skuAssetLiabilityTables[0].date, initial = True).order_by('-date').first()
                else:
                    skuAssetLiabilityTableForSalesPerson = SkuAssetLiabilityTable.objects.filter(sku = sku, for_sales = True, date__lt = end_date).order_by('-date').first()
            elif skuAssetLiabilityTables.count() == 1:
                skuAssetLiabilityTableForSalesPerson = SkuAssetLiabilityTable.objects.filter(sku = sku, for_sales = True, date__lt = end_date).order_by('-date').first()
            else:
                return HttpResponse('%s 还没有初始化资产负债表' %sku)
            user_type = 'SalesOnly'
        elif (SkuManagedBySalesPerson.objects.filter(sku = sku, sales_person_name = request.user.username).count() or \
            SkuManagedBySalesPerson.objects.filter(sku = sku, sales_assistant_name = request.user.username).count()) and \
            (SkuContributor.objects.filter(sku = sku, proposer_name = request.user.username).count() or \
            SkuContributor.objects.filter(sku = sku, designer_name = request.user.username).count()):
            skuAssetLiabilityTables = SkuAssetLiabilityTable.objects.filter(sku = sku, for_sales = True, date__lt = end_date).order_by('-date').all()
            if skuAssetLiabilityTables.count() > 1:
                if skuAssetLiabilityTables[0].date == skuAssetLiabilityTables[1].date:
                    skuAssetLiabilityTableForSalesPerson = SkuAssetLiabilityTable.objects.filter(sku = sku, for_sales = True, date = skuAssetLiabilityTables[0].date, initial = True).order_by('-date').first()
                else:
                    skuAssetLiabilityTableForSalesPerson = SkuAssetLiabilityTable.objects.filter(sku = sku, for_sales = True, date__lt = end_date).order_by('-date').first()
            elif skuAssetLiabilityTables.count() == 1:
                skuAssetLiabilityTableForSalesPerson = SkuAssetLiabilityTable.objects.filter(sku = sku, for_sales = True, date__lt = end_date).order_by('-date').first()
            else:
                return HttpResponse('%s 还没有初始化资产负债表' %sku)
            skuAssetLiabilityTable = SkuAssetLiabilityTable.objects.filter(sku = sku, for_sales = False, date__lt = end_date).order_by('-date').first()
            user_type = 'SalesAndProductContributor'
        elif (SkuManagedBySalesPerson.objects.filter(sku = sku, sales_person_name = request.user.username).count() == 0 and \
            SkuManagedBySalesPerson.objects.filter(sku = sku, sales_assistant_name = request.user.username).count() == 0) and \
            (SkuContributor.objects.filter(sku = sku, proposer_name = request.user.username).count() or \
            SkuContributor.objects.filter(sku = sku, designer_name = request.user.username).count()):
            skuAssetLiabilityTable = SkuAssetLiabilityTable.objects.filter(sku = sku, for_sales = False, date__lt = end_date).order_by('-date').first()
            user_type = 'ProductContributorOnly'

        skuAssetLiabilityTable = SkuAssetLiabilityTable.objects.filter(sku = sku, date__lt = end_date).order_by('-date').first()
        cash_before = skuAssetLiabilityTable.cash_amount
        po_head_shipping_cost = profitLossTable.product_purchasing_fee + profitLossTable.head_shipping_fee
        profit = (profitLossTable.sales_amount - profitLossTable.amazon_fee - profitLossTable.ad_fee) * profitLossTable.currency_rate - po_head_shipping_cost
        sales_amount_in_cny = profitLossTable.sales_amount * profitLossTable.currency_rate
        amazon_fee_in_cny = profitLossTable.amazon_fee * profitLossTable.currency_rate
        ad_fee_in_cny = profitLossTable.ad_fee * profitLossTable.currency_rate
        cash_after = cash_before + profit
        if cash_after < 0:
            liabilities_after = skuAssetLiabilityTable.liabilities - cash_after
            cash_after = 0.0
        else:
            if cash_after >= skuAssetLiabilityTable.liabilities:
                liabilities_after = 0.0
                cash_after = cash_after - skuAssetLiabilityTable.liabilities
            else:
                liabilities_after = skuAssetLiabilityTable.liabilities - cash_after
                cash_after = 0.0
        historyTodayProductSales = HistoryTodayProductSales.objects.filter(product__sku =sku).order_by('-date').first()
        sold_qty_average_7d = historyTodayProductSales.sold_qty_average_7d
        fba_shenzhen_inventory = historyTodayProductSales.fba_inventory.total_unit + check_shenzhen_inventory(sku)
        sold_out_day_number = int(float(fba_shenzhen_inventory) / sold_qty_average_7d)
        receive_money_per_unit_sold = (sales_amount_in_cny-amazon_fee_in_cny-ad_fee_in_cny) / float(profitLossTable.sales_quantity)
        gross_profit_sold_out = int(receive_money_per_unit_sold*fba_shenzhen_inventory - liabilities_after)
        liabilities_zero_day_number = 0
        roi = 0.0
        roi_annual = 0.0
        if liabilities_after > 0:
            liabilities_zero_day_number = int(liabilities_after/receive_money_per_unit_sold/sold_qty_average_7d)
            roi = gross_profit_sold_out / liabilities_after
            roi_annual = (1.0 + roi) ** (float(365) / float(sold_out_day_number)) - 1
            roi_commission_deducted = roi*0.8
            roi_commission_deducted_annual = float(365) / float(sold_out_day_number) * gross_profit_sold_out * 0.8 /liabilities_after #(1.0 + roi_commission_deducted) ** (float(365) / float(sold_out_day_number)) - 1
        context = {
            'sku':sku,
            'start_date': start_date,
            'end_date': end_date ,
            'liabilities_before': skuAssetLiabilityTable.liabilities,
            'liabilities': liabilities_after,
            'cash': cash_after ,
            'net_asset': cash_after - liabilities_after ,
            'profit': profit,
            'profitLossTable': profitLossTable,
            'sales_amount_in_cny':sales_amount_in_cny,
            'amazon_fee_in_cny':amazon_fee_in_cny,
            'ad_fee_in_cny':ad_fee_in_cny,
            'receive_money_per_unit_sold': receive_money_per_unit_sold,
            'fba_shenzhen_inventory' : fba_shenzhen_inventory ,
            'sold_qty_average_7d': sold_qty_average_7d,
            'sold_out_day_number': sold_out_day_number,
            'liabilities_zero_day_number':liabilities_zero_day_number,
            'gross_profit_sold_out': gross_profit_sold_out,
            'gross_profit_sold_out_5_percent': gross_profit_sold_out * 0.05,
            'roi': int(roi * 100),
            'roi_annual':int(roi_annual * 100) ,
            'roi_commission_deducted': int(roi_commission_deducted * 100),
            'roi_commission_deducted_annual':int(roi_commission_deducted_annual * 100),
        }

        return HttpResponse(template.render(context, request))
    else:
        return HttpResponse('你没有权限查看这个产品的财务报表')

def get_sku_list_pl_al_table(sku_list):
    profitLossTable = ProfitLossTable.objects.filter(status = ('P', '购货和头程运费确认'), sku__in = sku_list).order_by('-end_date').first()
    start_date = profitLossTable.start_date
    end_date = profitLossTable.end_date
    skuAssetLiabilityTables = SkuAssetLiabilityTable.objects.filter(for_sales=False, date__gte=start_date, sku__in = sku_list).all()
    liabilities_before_total = 0.0
    liabilities_after_total = 0.0
    cash_after_total = 0.0
    profit_total = 0.0
    sales_amount_in_cny_total =0.0
    amazon_fee_in_cny_total = 0.0
    ad_fee_in_cny_total = 0.0
    receive_money_total = 0.0
    fba_shenzhen_inventory_total = 0
    sold_qty_average_7d_total = 0.0
    liabilities_zero_day_number_total = 0
    gross_profit_sold_out_total = 0.0
    sold_qty_total = 0.0
    po_cost_total = 0.0
    head_shipping_cost_total = 0.0
    for skuAssetLiabilityTable in skuAssetLiabilityTables:
        liabilities_before_total += skuAssetLiabilityTable.liabilities
        cash_before = skuAssetLiabilityTable.cash_amount
        sku = skuAssetLiabilityTable.sku
        profitLossTable = ProfitLossTable.objects.filter(sku = sku, status = ('P', '购货和头程运费确认')).order_by('-end_date').first()
        if profitLossTable:
            sold_qty_total += profitLossTable.sales_quantity
            po_cost_total += profitLossTable.product_purchasing_fee
            head_shipping_cost_total += profitLossTable.head_shipping_fee
            po_head_shipping_cost = profitLossTable.product_purchasing_fee + profitLossTable.head_shipping_fee
            profit = (profitLossTable.sales_amount - profitLossTable.amazon_fee - profitLossTable.ad_fee) * profitLossTable.currency_rate - po_head_shipping_cost
            sales_amount_in_cny = profitLossTable.sales_amount * profitLossTable.currency_rate
            amazon_fee_in_cny = profitLossTable.amazon_fee * profitLossTable.currency_rate
            ad_fee_in_cny = profitLossTable.ad_fee * profitLossTable.currency_rate
            cash_after = cash_before + profit
            if cash_after < 0:
                liabilities_after = skuAssetLiabilityTable.liabilities - cash_after
                cash_after = 0.0
            else:
                if cash_after >= skuAssetLiabilityTable.liabilities:
                    liabilities_after = 0.0
                    cash_after = cash_after - skuAssetLiabilityTable.liabilities
                else:
                    liabilities_after = skuAssetLiabilityTable.liabilities - cash_after
                    cash_after = 0.0
            historyTodayProductSale = HistoryTodayProductSales.objects.filter(product__sku =sku).order_by('-date').first()
            if historyTodayProductSale:
                sold_qty_average_7d = historyTodayProductSale.sold_qty_average_7d
                fba_shenzhen_inventory = historyTodayProductSale.fba_inventory.total_unit + check_shenzhen_inventory(sku)
            else:
                sold_qty_average_7d = 0.0
                fba_shenzhen_inventory = 0.0 + check_shenzhen_inventory(sku)
            # sold_out_day_number = int(float(fba_shenzhen_inventory) / sold_qty_average_7d)

            if profitLossTable.sales_quantity:
                receive_money_by_this_sku = (sales_amount_in_cny-amazon_fee_in_cny-ad_fee_in_cny)
                receive_money_per_unit_sold = (sales_amount_in_cny-amazon_fee_in_cny-ad_fee_in_cny) / float(profitLossTable.sales_quantity)
            else:
                receive_money_by_this_sku = (sales_amount_in_cny-amazon_fee_in_cny-ad_fee_in_cny)
                receive_money_per_unit_sold = (sales_amount_in_cny-amazon_fee_in_cny-ad_fee_in_cny) / 100.0
            receive_money_total += receive_money_by_this_sku
            gross_profit_sold_out = int(receive_money_per_unit_sold*fba_shenzhen_inventory - liabilities_after)

            liabilities_after_total += liabilities_after
            cash_after_total += cash_after
            profit_total += profit
            sales_amount_in_cny_total += sales_amount_in_cny
            amazon_fee_in_cny_total += amazon_fee_in_cny
            ad_fee_in_cny_total += ad_fee_in_cny
            fba_shenzhen_inventory_total += fba_shenzhen_inventory
            sold_qty_average_7d_total += sold_qty_average_7d
            gross_profit_sold_out_total += gross_profit_sold_out
        else:
            liabilities_after_total += skuAssetLiabilityTable.liabilities
            cash_after_total += cash_before
    sold_out_day_number_total = int(float(fba_shenzhen_inventory_total) / sold_qty_average_7d_total)
    receive_money_per_unit_sold_average = receive_money_total / sold_qty_total
    if liabilities_after_total > 0:
        liabilities_zero_day_number_total = int(liabilities_after_total/receive_money_per_unit_sold_average/sold_qty_average_7d_total)
    gross_profit_sold_out_total = int(receive_money_per_unit_sold_average*fba_shenzhen_inventory_total - liabilities_after_total)

    context = {
        # 'sku':sku,
        'start_date': start_date,
        'end_date': end_date ,
        'liabilities_before': liabilities_before_total,
        'liabilities': liabilities_after_total,
        'cash': cash_after_total ,
        'net_asset': cash_after_total - liabilities_after_total ,
        'profit': profit_total,
        'sold_qty_total': sold_qty_total,
        'po_cost_total': po_cost_total,
        'head_shipping_cost_total': head_shipping_cost_total,
        # 'profitLossTable': profitLossTable,
        'sales_amount_in_cny':sales_amount_in_cny_total,
        'amazon_fee_in_cny':amazon_fee_in_cny_total,
        'ad_fee_in_cny':ad_fee_in_cny_total,
        'receive_money_per_unit_sold': receive_money_per_unit_sold_average,
        'fba_shenzhen_inventory' : fba_shenzhen_inventory_total ,
        'sold_qty_average_7d': sold_qty_average_7d_total,
        'sold_out_day_number': sold_out_day_number_total,
        'liabilities_zero_day_number':liabilities_zero_day_number_total,
        'gross_profit_sold_out': gross_profit_sold_out_total,
        'gross_profit_sold_out_5_percent': gross_profit_sold_out_total * 0.05
    }
    return context

@login_required
def finance_dashboard(request):
    template = loader.get_template('finance_dashboard.html')
    if request.user.groups.filter(name='sales').exists() or \
        request.user.groups.filter(name='designer').exists() or \
        request.user.groups.filter(name='proposer').exists():
        roles_sku_list_dict = {}
        if request.user.groups.filter(name='head_of_sales').exists():
            profitLossTable = ProfitLossTable.objects.filter(status = ('P', '购货和头程运费确认')).order_by('-end_date').first()
            start_date = profitLossTable.start_date
            end_date = profitLossTable.end_date
            skuAssetLiabilityTables = SkuAssetLiabilityTable.objects.filter(for_sales=False, date__gte=start_date).all()
            sku_list_as_head_of_sales = [skuAssetLiabilityTables.sku for skuAssetLiabilityTables in skuAssetLiabilityTables]
        sku_list_as_sales_person = [skuManagedBySalesPerson.sku for skuManagedBySalesPerson in SkuManagedBySalesPerson.objects.filter(sales_person_name=request.user.username).all()]
        sku_list_as_sales_assistant = [skuManagedBySalesPerson.sku for skuManagedBySalesPerson in SkuManagedBySalesPerson.objects.filter(sales_assistant_name=request.user.username).all()]
        sku_list_as_proposer = [skuContributor.sku for skuContributor in SkuContributor.objects.filter(proposer_name = request.user.username).all()]
        sku_list_as_designer = [skuContributor.sku for skuContributor in SkuContributor.objects.filter(designer_name = request.user.username).all()]
        if(len(sku_list_as_head_of_sales)):
            roles_sku_list_dict['head_of_sales'] = sku_list_as_head_of_sales
        if(len(sku_list_as_sales_person)):
            roles_sku_list_dict['sales_person'] = sku_list_as_sales_person
        if(len(sku_list_as_sales_assistant)):
            roles_sku_list_dict['sales_assistant'] = sku_list_as_sales_assistant
        if(len(sku_list_as_proposer)):
            roles_sku_list_dict['proposer'] = sku_list_as_proposer
        if(len(sku_list_as_designer)):
            roles_sku_list_dict['designer'] = sku_list_as_designer
        if request.method == 'POST':
            role = request.POST['role']
        else:
            role = next(iter(roles_sku_list_dict))
        context = get_sku_list_pl_al_table(roles_sku_list_dict[role])
        ROLE_NAMES_DICT = {'head_of_sales': '销售总监', 'sales_person': '销售', 'sales_assistant': '销售助理', 'proposer': '选品师', 'designer': '设计师'}
        context['role_name'] = ROLE_NAMES_DICT[role]
        context['roles_sku_list_dict'] = roles_sku_list_dict
    return HttpResponse(template.render(context, request))
